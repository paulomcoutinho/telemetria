# -*- coding: utf-8 -*-
"""
Módulo: widget_medidores.py
=============================
Aba de gestão de medidores de vazão (JanelaGestaoDados – aba 2).
 
Funcionalidades:
  - Busca por 6 critérios: Rótulo, Usuário, CNARH, Código UC,
    Operador, Sistema Hídrico;
  - Edição inline com UPDATE; campos de vazão/potência disparam
    diálogos de confirmação de unidade ao perder foco;
  - Campos inválidos destacados em vermelho;
  - Desativação lógica via sufixo '#' no rótulo;
  - Reativação em lote via DialogReativacao;
  - Exportação XLSX com formatação institucional (openpyxl).
 
Autor : SFI/ANA
Versão: 2.0 – Março/2026
"""
 
# ---------------------------------------------------------------------------
# Imports Qt – widgets, layout e utilitários de UI
# ---------------------------------------------------------------------------
from qgis.PyQt.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QComboBox,
    QFrame, QMessageBox, QSizePolicy, QAbstractItemView,
    QHeaderView, QListWidget, QListWidgetItem, QScrollArea,
    QDialog, QDialogButtonBox, QApplication,
    QDateEdit, QCheckBox, QGroupBox, QRadioButton, QButtonGroup,
)
from qgis.PyQt.QtCore import Qt, QDate
 
# ---------------------------------------------------------------------------
# Imports Python padrão
# ---------------------------------------------------------------------------
import os
import psycopg2
from datetime import datetime

# ---------------------------------------------------------------------------
# Diálogos auxiliares de confirmação de unidade e reativação
# ---------------------------------------------------------------------------
from .dialogo_unidade_vazao    import DialogoUnidadeVazao
from .dialogo_unidade_potencia import DialogoUnidadePotencia
from .dialogo_reativacao       import DialogReativacao


class WidgetMedidores(QWidget):
    """Aba de gestão de medidores de telemetria dentro da JanelaGestaoDados.

    Centraliza todas as operações de consulta e manutenção dos registros da
    tabela ``tb_intervencao``, oferecendo pesquisa por múltiplos critérios,
    edição inline de atributos, desativação lógica de medidores, reativação
    de medidores desativados e exportação da lista para XLSX.

    Fluxo principal de interação:
        1. **Busca**: o usuário seleciona o critério (Rótulo, Usuário,
           CNARH, Operador, Unidade de Automonitoramento ou Sistema Hídrico) e digita o termo;
           os resultados são exibidos em ``QTreeWidget`` com colunas de ID,
           Rótulo e Usuário.
        2. **Seleção**: ao clicar em um item da árvore, todos os campos de
           detalhe são preenchidos com os dados do medidor (interferência,
           CNARH, operador, vazão nominal, potência, tipo, modo de transmissão,
           coordenadas e data de cadastro).
        3. **Edição**: os campos editáveis ficam habilitados; ao alterar vazão
           nominal ou potência, os diálogos ``DialogoUnidadeVazao`` e
           ``DialogoUnidadePotencia`` são disparados automaticamente para
           confirmar a unidade de entrada e converter para m³/s ou kW,
           respectivamente.
        4. **Salvamento**: o botão "Salvar alterações" persiste o UPDATE em
           ``tb_intervencao``; campos inválidos são destacados em vermelho
           e o salvamento é bloqueado até correção.
        5. **Desativação / Reativação**: o botão "Desativar" anexa ``#`` ao
           rótulo do medidor; o botão "Reativar medidores" abre o
           ``DialogReativacao`` para reabilitar um ou mais registros
           desativados de uma só vez.
        6. **Exportação**: gera um XLSX formatado com cabeçalho azul,
           linhas alternadas e bordas finas, salvo em ``Downloads``.

    Flags de controle de conversão:
        vazao_ja_convertida (bool): Indica que o valor atual de vazão
            já foi convertido de m³/h para m³/s nesta sessão de edição.
        potencia_ja_convertida (bool): Indica que o valor atual de potência
            já foi convertido de cv para kW nesta sessão de edição.

    Attributes:
        conn (psycopg2.connection): Conexão ativa ao PostgreSQL em autocommit.
        usuario_logado (str | None): Usuário da sessão atual.
        medidor_atual (int | None): ID do medidor selecionado na árvore;
            ``None`` quando nenhum está carregado.
        combo_criterio (QComboBox): Seletor do critério de busca.
        input_busca (QLineEdit): Campo de texto para o termo de busca.
        tree_medidores (QTreeWidget): Listagem de resultados com colunas
            ID, Rótulo e Usuário.
        input_vazao (QLineEdit): Vazão nominal em m³/s; dispara conversão
            ao perder o foco.
        input_potencia (QLineEdit): Potência em kW; dispara conversão ao
            perder o foco.
    """
    
    def __init__(self, conexao, usuario=None):
        super().__init__()
        self.conn = conexao
        self.usuario_logado = usuario
        try:
            self.conn.autocommit = True
        except:
            pass        
        self.medidor_atual = None
        self.initUI()
        
        # Variáveis de controle para conversão
        self.vazao_ja_convertida = False
        self.potencia_ja_convertida = False
        self.ultimo_valor_vazao = None
        self.ultimo_valor_potencia = None        

        # Conectar eventos de saída dos campos
        self.input_vazao.editingFinished.connect(self.processar_vazao)
        self.input_potencia.editingFinished.connect(self.processar_potencia)        
        
    def initUI(self):
        """Configura a interface do widget de medidores."""
        # Layout principal com scroll area para garantir que tudo caiba
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(5, 5, 5, 5)
        
        # Criar um widget container para todo o conteúdo
        content_widget = QWidget()
        content_widget.setStyleSheet("""
            QWidget {
                background-color: white;
                border:1px solid #ccc;
                border-radius: 5px;
            }
        """)
        layout = QVBoxLayout(content_widget)
        
        # === ÁREA DE BUSCA COM AUTOCOMPLETAR ===
        busca_container = QWidget()
        busca_container.setStyleSheet("border: none;")
        busca_layout = QVBoxLayout(busca_container)
        busca_layout.setContentsMargins(0, 5, 0, 5)
        
        # Layout superior com critério e busca
        layout_busca_superior = QHBoxLayout()
        layout_busca_superior.setSpacing(10)
        
        label_criterio = QLabel("Buscar medidor:")
        label_criterio.setFixedWidth(80)
        label_criterio.setStyleSheet("border: none;")
        
        self.combo_criterio = QComboBox()
        self.combo_criterio.addItems(["Rótulo", "Usuário", "CNARH", "Operador", "UAM", "Sistema Hídrico"])
        self.combo_criterio.setFixedHeight(30)
        self.combo_criterio.setStyleSheet("""
            QComboBox {
                border:1px solid #ccc;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
        """)
        self.combo_criterio.currentIndexChanged.connect(self.atualizar_placeholder)
        
        self.input_busca = QLineEdit()
        self.input_busca.setPlaceholderText("Digite o termo de busca...")
        self.input_busca.setFixedHeight(30)
        self.input_busca.setStyleSheet("""
            QLineEdit {
                border:1px solid #ccc;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
        """)
        self.input_busca.textChanged.connect(self.buscar_medidores_autocomplete)
        self.input_busca.returnPressed.connect(self.buscar_medidores)

        btn_buscar = QPushButton("Buscar")
        btn_buscar.setFixedSize(80, 30)
        btn_buscar.setStyleSheet("""
            QPushButton {
                background-color: #5474b8;
                color: white;
                font-size: 11px;
                font-weight: bold;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #2050b8;
            }
        """)
        btn_buscar.clicked.connect(self.buscar_medidores)
        
        layout_busca_superior.addWidget(label_criterio)
        layout_busca_superior.addWidget(self.combo_criterio)
        layout_busca_superior.addWidget(self.input_busca)
        layout_busca_superior.addWidget(btn_buscar)
        
        # ComboBox para sugestões de autocompletar
        self.combo_sugestoes = QComboBox()
        self.combo_sugestoes.setFixedHeight(28)
        self.combo_sugestoes.setStyleSheet("""
            QComboBox {
                border: 1px solid #4CAF50;
                border-radius: 3px;
                padding: 3px;
                background-color: #f9fff9;
            }
        """)
        self.combo_sugestoes.setVisible(False)
        self.combo_sugestoes.activated.connect(self.selecionar_sugestao)
        
        busca_layout.addLayout(layout_busca_superior)
        busca_layout.addWidget(self.combo_sugestoes)
        
        layout.addWidget(busca_container)

        # Label de contagem de resultados
        self.lbl_contagem = QLabel("")
        self.lbl_contagem.setStyleSheet(
            "border: none; color: #5474b8; font-style: italic; font-size: 11px; padding: 0px 2px;"
        )
        self.lbl_contagem.setVisible(False)
        layout.addWidget(self.lbl_contagem)

        # Lista de resultados com tamanho fixo e bordas arredondadas
        self.lista_resultados = QListWidget()
        self.lista_resultados.setMinimumHeight(80)
        self.lista_resultados.setMaximumHeight(150)
        self.lista_resultados.setSelectionMode(QAbstractItemView.ExtendedSelection) # Permite seleção múltipla
        self.lista_resultados.setStyleSheet("""
            QListWidget {
                border:1px solid #ccc;
                border-radius: 3px;
                background-color: white;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:selected {
                background-color: #5474b8;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #e0f0ff;
            }
        """)
        # Substituímos itemClicked por itemSelectionChanged para lidar com múltipla seleção
        self.lista_resultados.itemSelectionChanged.connect(self.gerenciar_selecao_lista)
        layout.addWidget(self.lista_resultados)
        
        # Separador
        separador2 = QFrame()
        separador2.setFrameShape(QFrame.HLine)
        separador2.setFrameShadow(QFrame.Sunken)
        separador2.setStyleSheet("border: none; background-color: #ccc; max-height: 1px;")
        layout.addWidget(separador2)
        
        # === SEÇÃO DE DADOS DO MEDIDOR ===
        dados_container = QWidget()
        dados_container.setStyleSheet("border: none;")
        dados_layout = QVBoxLayout(dados_container)
        dados_layout.setSpacing(8)
        dados_layout.setContentsMargins(0, 5, 0, 5)
        
        dados_label = QLabel("Dados do medidor:")
        dados_label.setStyleSheet("font-weight: bold; color: #555; border: none;")
        dados_layout.addWidget(dados_label)
        
        # Grid para os campos principais
        grid_layout = QGridLayout()
        grid_layout.setSpacing(10)
        grid_layout.setContentsMargins(0, 5, 0, 5)
        
        # Estilos consistentes
        label_style = "border: none;"
        input_style = "border: 1px solid #ccc; border-radius: 3px; padding: 3px; background-color: white;"
        input_readonly_style = "border: 1px solid #ccc; border-radius: 3px; padding: 3px; background-color: #f5f5f5;"
        combo_style = """
            QComboBox {
                border: 1px solid #ccc;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
        """
        
        # Linha 1
        lbl_rotulo = QLabel("Rótulo:")
        lbl_rotulo.setStyleSheet(label_style)
        grid_layout.addWidget(lbl_rotulo, 0, 0)
        self.input_rotulo = QLineEdit()
        self.input_rotulo.setReadOnly(True)
        self.input_rotulo.setFixedHeight(28)
        self.input_rotulo.setStyleSheet(input_readonly_style)
        grid_layout.addWidget(self.input_rotulo, 0, 1)
        
        lbl_codigo_uc = QLabel("Código UC:")
        lbl_codigo_uc.setStyleSheet(label_style)
        grid_layout.addWidget(lbl_codigo_uc, 0, 2)
        self.input_codigo_uc = QLineEdit()
        self.input_codigo_uc.setFixedHeight(28)
        self.input_codigo_uc.setStyleSheet(input_style)
        grid_layout.addWidget(self.input_codigo_uc, 0, 3)
        
        # Linha 2
        lbl_vazao = QLabel("Vazão (m³/s):")
        lbl_vazao.setStyleSheet(label_style)
        grid_layout.addWidget(lbl_vazao, 1, 0)
        self.input_vazao = QLineEdit()
        self.input_vazao.setPlaceholderText("Digite valor em m³/s ou m³/h")        
        self.input_vazao.setFixedHeight(28)
        self.input_vazao.setStyleSheet(input_style)
        grid_layout.addWidget(self.input_vazao, 1, 1)
        
        lbl_potencia = QLabel("Potência (kW):")
        lbl_potencia.setStyleSheet(label_style)
        grid_layout.addWidget(lbl_potencia, 1, 2)
        self.input_potencia = QLineEdit()
        self.input_potencia.setPlaceholderText("Digite valor em cv ou kW")         
        self.input_potencia.setFixedHeight(28)
        self.input_potencia.setStyleSheet(input_style)
        grid_layout.addWidget(self.input_potencia, 1, 3)
        
        # Linha 3
        lbl_equipamento = QLabel("Equipamento:")
        lbl_equipamento.setStyleSheet(label_style)
        grid_layout.addWidget(lbl_equipamento, 2, 0)
        self.combo_equipamento = QComboBox()
        self.combo_equipamento.setFixedHeight(28)
        self.combo_equipamento.setStyleSheet(combo_style)
        self.carregar_equipamentos()
        grid_layout.addWidget(self.combo_equipamento, 2, 1)
        
        lbl_transmissao = QLabel("Transmissão:")
        lbl_transmissao.setStyleSheet(label_style)
        grid_layout.addWidget(lbl_transmissao, 2, 2)
        self.combo_transmissao = QComboBox()
        self.combo_transmissao.setFixedHeight(28)
        self.combo_transmissao.setStyleSheet(combo_style)
        self.carregar_modos_transmissao()
        grid_layout.addWidget(self.combo_transmissao, 2, 3)
        
        # Linha 4 - Operador (ocupando 2 colunas)
        lbl_operador = QLabel("Operador responsável:")
        lbl_operador.setStyleSheet(label_style)
        grid_layout.addWidget(lbl_operador, 3, 0)
        self.combo_operador_responsavel = QComboBox()
        self.combo_operador_responsavel.setEditable(True)
        self.combo_operador_responsavel.setFixedHeight(28)
        self.combo_operador_responsavel.setStyleSheet(combo_style)
        self.carregar_operadores_combo()
        grid_layout.addWidget(self.combo_operador_responsavel, 3, 1, 1, 3)
        
        dados_layout.addLayout(grid_layout)
        layout.addWidget(dados_container)
        
        # Separador
        separador3 = QFrame()
        separador3.setFrameShape(QFrame.HLine)
        separador3.setFrameShadow(QFrame.Sunken)
        separador3.setStyleSheet("border: none; background-color: #ccc; max-height: 1px;")
        layout.addWidget(separador3)
        
        # === SEÇÃO DE INFORMAÇÕES DE REFERÊNCIA ===
        info_container = QWidget()
        info_container.setStyleSheet("border: none;")
        info_layout = QVBoxLayout(info_container)
        info_layout.setSpacing(8)
        info_layout.setContentsMargins(0, 5, 0, 5)
        
        info_label = QLabel("Informações de referência:")
        info_label.setStyleSheet("font-weight: bold; color: #555; border: none;")
        info_layout.addWidget(info_label)
        
        # Grid para informações de referência
        info_grid = QGridLayout()
        info_grid.setSpacing(10)
        info_grid.setContentsMargins(0, 5, 0, 5)
        
        # Linha 1
        lbl_interferencia = QLabel("Interferência:")
        lbl_interferencia.setStyleSheet(label_style)
        info_grid.addWidget(lbl_interferencia, 0, 0)
        self.input_interferencia = QLineEdit()
        self.input_interferencia.setReadOnly(True)
        self.input_interferencia.setFixedHeight(28)
        self.input_interferencia.setStyleSheet(input_readonly_style)
        info_grid.addWidget(self.input_interferencia, 0, 1)
        
        lbl_usuario = QLabel("Usuário:")
        lbl_usuario.setStyleSheet(label_style)
        info_grid.addWidget(lbl_usuario, 0, 2)
        self.input_usuario = QLineEdit()
        self.input_usuario.setReadOnly(True)
        self.input_usuario.setFixedHeight(28)
        self.input_usuario.setStyleSheet(input_readonly_style)
        info_grid.addWidget(self.input_usuario, 0, 3)
        
        # Linha 2
        lbl_cnarh = QLabel("CNARH:")
        lbl_cnarh.setStyleSheet(label_style)
        info_grid.addWidget(lbl_cnarh, 1, 0)
        self.input_cnarh = QLineEdit()
        self.input_cnarh.setReadOnly(True)
        self.input_cnarh.setFixedHeight(28)
        self.input_cnarh.setStyleSheet(input_readonly_style)
        info_grid.addWidget(self.input_cnarh, 1, 1, 1, 3)
        
        info_layout.addLayout(info_grid)
        layout.addWidget(info_container)

        # === BOTÕES ===
        btn_container = QWidget()
        btn_container.setStyleSheet("border: none;")
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.setSpacing(15)
        btn_layout.setContentsMargins(0, 10, 0, 5)
        
        self.btn_salvar = QPushButton("Salvar alterações")
        self.btn_salvar.setFixedHeight(35)
        self.btn_salvar.setStyleSheet("""
            QPushButton {
                background-color: #175cc3;
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
            QPushButton:hover {
                background-color: #5474b8;
            }
        """)
        self.btn_salvar.clicked.connect(self.salvar_alteracoes)
        self.btn_salvar.setEnabled(False)

        # Botão Desativar (Novo)
        self.btn_desativar = QPushButton("Desativar medidor")
        self.btn_desativar.setFixedHeight(35)
        self.btn_desativar.setEnabled(False) # Inicia desabilitado
        self.btn_desativar.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }        
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.btn_desativar.clicked.connect(self.desativar_medidores)

        # --- NOVO BOTÃO REATIVAR ---
        self.btn_reativar = QPushButton("Reativar Medidor(es)")
        self.btn_reativar.setFixedHeight(35)
        self.btn_reativar.setEnabled(False)
        self.btn_reativar.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.btn_reativar.clicked.connect(self.abrir_dialogo_reativacao)

        self.btn_limpar = QPushButton("Limpar")
        self.btn_limpar.setFixedHeight(35)
        self.btn_limpar.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                font-size: 12px;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.btn_limpar.clicked.connect(self.limpar_campos)

        self.btn_exportar_medidores = QPushButton("Exportar medidores")
        self.btn_exportar_medidores.setFixedHeight(35)
        self.btn_exportar_medidores.setToolTip("Exportar lista completa de medidores cadastrados para Excel")
        self.btn_exportar_medidores.setStyleSheet("""
            QPushButton {
                background-color: #1d7a3a;
                color: white;
                font-size: 11px;
                font-weight: bold;
                padding: 8px 12px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                background-color: #155c2b;
            }
            QPushButton:pressed {
                background-color: #0f4320;
            }
        """)
        self.btn_exportar_medidores.clicked.connect(self.exportar_medidores_excel)

        # --- NOVO: Botão Ver no Mapa ---
        '''
        self.btn_ver_mapa = QPushButton("Ver no mapa")
        self.btn_ver_mapa.setFixedHeight(35)
        self.btn_ver_mapa.setVisible(False)
        self.btn_ver_mapa.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        self.btn_ver_mapa.clicked.connect(self.carregar_medidores_no_mapa)
        '''
        btn_layout.addWidget(self.btn_salvar)
        btn_layout.addWidget(self.btn_desativar)
        btn_layout.addWidget(self.btn_reativar)
        btn_layout.addWidget(self.btn_limpar)
        btn_layout.addWidget(self.btn_exportar_medidores)
        #btn_layout.addWidget(self.btn_ver_mapa)
        btn_layout.addStretch()
        
        layout.addWidget(btn_container)
        
        # Adicionar um stretch no final para melhor organização
        layout.addStretch(1)
        
        # Margens do conteúdo
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)
        
        # === SCROLL AREA ===
        # Criar scroll area para conteúdo
        scroll_area = QScrollArea()
        scroll_area.setWidget(content_widget)
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
        """)
        
        main_layout.addWidget(scroll_area)
        main_layout.setContentsMargins(8, 8, 8, 8)
        self.setLayout(main_layout)
        
        # Conectar sinais de alteração
        self.input_codigo_uc.textChanged.connect(self.habilitar_salvar)
        self.input_vazao.textChanged.connect(self.habilitar_salvar)
        self.input_potencia.textChanged.connect(self.habilitar_salvar)
        self.combo_equipamento.currentIndexChanged.connect(self.habilitar_salvar)
        self.combo_transmissao.currentIndexChanged.connect(self.habilitar_salvar)
        self.combo_operador_responsavel.currentIndexChanged.connect(self.habilitar_salvar)
        
        # Tamanho mínimo do widget
        self.setMinimumHeight(500)
        self.setMaximumHeight(800)
        
        # Política de tamanho
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # --- Lógica para usuário somente leitura ---
        if self.usuario_logado == "telemetria_ro":
            ro_style = "border: 1px solid #ccc; border-radius: 3px; padding: 3px; background-color: #f5f5f5;"
            self.input_codigo_uc.setReadOnly(True)
            self.input_codigo_uc.setStyleSheet(ro_style)
            self.input_vazao.setReadOnly(True)
            self.input_vazao.setStyleSheet(ro_style)
            self.input_potencia.setReadOnly(True)
            self.input_potencia.setStyleSheet(ro_style)
            self.combo_equipamento.setEnabled(False)
            self.combo_transmissao.setEnabled(False)
            self.combo_operador_responsavel.setEnabled(False)
            self.btn_salvar.hide()
            self.btn_desativar.hide()
            self.btn_reativar.hide()

    def carregar_medidores_no_mapa(self):
        """
        Slot para o botão 'Ver no mapa'.
        Funcionalidade a ser implementada: carregará as geometrias dos medidores listados.
        """
        # Placeholder: funcionalidade futura
        print("Funcionalidade 'Carregar no Mapa' acionada.")

    def gerenciar_selecao_lista(self):
        """
        Gerencia a seleção de itens na lista de resultados.
        Regras:
          - btn_desativar: habilitado apenas se TODOS os selecionados NÃO têm '#' no rótulo.
          - btn_reativar:  habilitado se ALGUM selecionado tem '#' no rótulo.
        """
        itens_selecionados = self.lista_resultados.selectedItems()
        qtd = len(itens_selecionados)

        if qtd == 0:
            self.btn_desativar.setEnabled(False)
            self.btn_reativar.setEnabled(False)
            self.limpar_campos_formulario()
            return

        # Verifica quais itens têm '#' no rótulo (parte antes do primeiro ' - ')
        tem_hash = []
        for item in itens_selecionados:
            rotulo = item.text().split(" - ")[0]
            tem_hash.append('#' in rotulo)

        algum_com_hash = any(tem_hash)
        todos_sem_hash = not algum_com_hash

        # btn_desativar só se todos estiverem sem '#'
        self.btn_desativar.setEnabled(todos_sem_hash)
        self.btn_desativar.setText(
            "Desativar medidor" if qtd == 1 else "Desativar Medidores"
        )

        # btn_reativar se ao menos um tiver '#'
        self.btn_reativar.setEnabled(algum_com_hash)

        # Formulário: preenche apenas na seleção simples
        if qtd == 1:
            self.carregar_dados_medidor(itens_selecionados[0])
        else:
            # limpar_campos_formulario força btn_reativar para False internamente,
            # por isso preservamos o estado correto e o restauramos em seguida.
            reativar_state = self.btn_reativar.isEnabled()
            self.limpar_campos_formulario()
            self.btn_reativar.setEnabled(reativar_state)

    def desativar_medidores(self):
        """Desativa um ou mais medidores selecionados na lista."""
        itens_selecionados = self.lista_resultados.selectedItems()
        
        if not itens_selecionados:
            return

        ids_selecionados = [item.data(Qt.UserRole) for item in itens_selecionados]
        qtd = len(ids_selecionados)

        texto_msg = "Deseja realmente desativar o medidor selecionado?" if qtd == 1 else \
                    f"Deseja realmente desativar os {qtd} medidores selecionados?"
        texto_msg += "\n\nIsso definirá o operador de telemetria como NULL e adicionará '#' ao rótulo."

        resposta = QMessageBox.question(
            self, 
            "Confirmar Desativação", 
            texto_msg,
            QMessageBox.Yes | QMessageBox.No, 
            QMessageBox.No
        )

        if resposta != QMessageBox.Yes:
            return

        try:
            cursor = self.conn.cursor()
            query = """
            UPDATE tb_intervencao 
            SET operador_telemetria = NULL, 
                rotulo = rotulo || '#'
            WHERE id = ANY(%s);
            """
            cursor.execute(query, (ids_selecionados,))
            self.conn.commit()
            
            QMessageBox.information(self, "Sucesso", f"{qtd} medidor(es) desativado(s) com sucesso!")
            
            self.limpar_campos()
            if self.input_busca.text().strip():
                self.buscar_medidores()
            
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Erro", f"Erro ao desativar medidores: {e}")
        finally:
            cursor.close()

    def limpar_campos_formulario(self):
        """
        Limpa apenas os campos de entrada de dados (inputs e combos),
        mantendo a lista de resultados intacta e medidor_atual como None.
        """
        self.medidor_atual = None
        self.input_rotulo.clear()
        self.input_codigo_uc.clear()
        self.input_vazao.clear()
        self.input_potencia.clear()
        self.combo_equipamento.setCurrentIndex(0)
        self.combo_transmissao.setCurrentIndex(0)
        self.combo_operador_responsavel.setCurrentIndex(0)
        self.input_interferencia.clear()
        self.input_usuario.clear()
        self.input_cnarh.clear()
        self.btn_salvar.setEnabled(False)        
        self.btn_reativar.setEnabled(False)        
        self.vazao_ja_convertida = False
        self.potencia_ja_convertida = False
        self.ultimo_valor_vazao = None
        self.ultimo_valor_potencia = None

        if hasattr(self, 'lbl_contagem'):
            self.lbl_contagem.setVisible(False)        
    
    def atualizar_placeholder(self):
        """Atualiza o placeholder do campo de busca conforme o critério selecionado."""
        criterio = self.combo_criterio.currentText()
        if criterio == "Rótulo":
            self.input_busca.setPlaceholderText("Digite rótulo do medidor...")        
        elif criterio == "Usuário":
            self.input_busca.setPlaceholderText("Digite nome do usuário...")
        elif criterio == "CNARH":
            self.input_busca.setPlaceholderText("Digite o CNARH...")
        elif criterio == "Operador":
            self.input_busca.setPlaceholderText("Digite nome do operador...")
        elif criterio == "UAM":
            self.input_busca.setPlaceholderText("Digite a Unidade de Automonitoramento...")            
        elif criterio == "Sistema Hídrico":
            self.input_busca.setPlaceholderText("Digite o Sistema Hídrico...")
        
    def buscar_medidores_autocomplete(self, texto):
        """
        Busca medidores em tempo real para preencher o autocompletar.
        Utiliza a view_ft_intervencao para buscas textuais e ft_sishidrico_buffer para espacial.
        """
        texto = texto.strip()
        if len(texto) < 2:
            self.combo_sugestoes.setVisible(False)
            return
            
        criterio = self.combo_criterio.currentText()
        cursor = None
        
        try:
            cursor = self.conn.cursor()
            termo_busca = f"%{texto}%"
            
            # Unificação: Utiliza view_ft_intervencao, ft_uam_buffer e ft_sishidrico_buffer para critérios textuais
            if criterio == "Rótulo":
                query = "SELECT DISTINCT rotulo_medidor FROM view_ft_intervencao WHERE LOWER(rotulo_medidor) LIKE LOWER(%s) ORDER BY rotulo_medidor LIMIT 10;"
            elif criterio == "Usuário":
                query = "SELECT DISTINCT nome_usuario FROM view_ft_intervencao WHERE LOWER(nome_usuario) LIKE LOWER(%s) ORDER BY nome_usuario LIMIT 10;"
            elif criterio == "CNARH":
                query = "SELECT DISTINCT nu_cnarh FROM view_ft_intervencao WHERE LOWER(nu_cnarh) LIKE LOWER(%s) ORDER BY nu_cnarh LIMIT 10;"
            elif criterio == "Operador":
                query = "SELECT DISTINCT nome_operador FROM view_ft_intervencao WHERE LOWER(nome_operador) LIKE LOWER(%s) ORDER BY nome_operador LIMIT 10;"
            elif criterio == "UAM":
                query = "SELECT DISTINCT nmautomonit FROM ft_uam_buffer WHERE LOWER(nmautomonit) LIKE LOWER(%s) ORDER BY nmautomonit LIMIT 10;"            
            elif criterio == "Sistema Hídrico":
                query = "SELECT DISTINCT bafnm FROM ft_sishidrico_buffer WHERE LOWER(bafnm) LIKE LOWER(%s) ORDER BY bafnm LIMIT 10;"
            else:
                return

            cursor.execute(query, (termo_busca,))
            resultados = cursor.fetchall()
            
            self.combo_sugestoes.clear()
            if resultados:
                for resultado in resultados:
                    valor = resultado[0]
                    if valor:
                        self.combo_sugestoes.addItem(valor)
                self.combo_sugestoes.setVisible(True)
            else:
                self.combo_sugestoes.setVisible(False)
                
        except Exception as e:
            print(f"Erro no autocompletar: {e}")
        finally:
            if cursor:
                cursor.close()
            
    def buscar_medidores(self):
        """
        Busca medidores com base no critério selecionado.
        
        Realiza buscas textuais na 'view_ft_intervencao' ou busca espacial 
        (intersecção) com 'ft_sishidrico_buffer' transformado para SRID 4326.
        Gerencia o cursor de espera e a visibilidade do botão 'Ver no Mapa'.
        """
        criterio = self.combo_criterio.currentText()
        termo = self.input_busca.text().strip()
        
        if not termo:
            QMessageBox.warning(self, "Campo vazio", "Digite um termo para busca.")
            return
        
        cursor = None
        try:
            cursor = self.conn.cursor()
            termo_busca = f"%{termo}%"
            
            # === Lógica de Busca ===
            if criterio == "Sistema Hídrico":
                # -- Busca Espacial --
                QApplication.setOverrideCursor(Qt.WaitCursor)
                
                query = """
                SELECT DISTINCT v.id, v.rotulo_medidor, v.nome_usuario, v.nome_operador,
                       v.nu_interferencia_cnarh
                FROM view_ft_intervencao v
                JOIN ft_sishidrico_buffer s 
                    ON ST_Intersects(v.geom, ST_Transform(s.geom, 4326))
                WHERE LOWER(s.bafnm) LIKE LOWER(%s)
                ORDER BY v.rotulo_medidor;
                """
                
                try:
                    cursor.execute(query, (termo_busca,))
                    resultados = cursor.fetchall()
                finally:
                    QApplication.restoreOverrideCursor()

            elif criterio == "UAM":
                # -- Busca Espacial --
                QApplication.setOverrideCursor(Qt.WaitCursor)
                
                query = """
                SELECT DISTINCT v.id, v.rotulo_medidor, v.nome_usuario, v.nome_operador,
                       v.nu_interferencia_cnarh
                FROM view_ft_intervencao v
                JOIN ft_uam_buffer s 
                    ON ST_Intersects(v.geom, ST_Transform(s.geom, 4326))
                WHERE LOWER(s.nmautomonit) LIKE LOWER(%s)
                ORDER BY v.rotulo_medidor;
                """
                
                try:
                    cursor.execute(query, (termo_busca,))
                    resultados = cursor.fetchall()
                finally:
                    QApplication.restoreOverrideCursor()
            
            elif criterio == "Rótulo":
                # === Lógica de Busca Unificada via view_ft_intervencao ===
                QApplication.setOverrideCursor(Qt.WaitCursor)
                
                try:
                    query = """
                    SELECT DISTINCT 
                        i.id,
                        i.rotulo as rotulo_medidor,
                        u.nome_usuario as nome_usuario,
                        o.nome as nome_operador,
                        u.codigo_interferencia as nu_interferencia_cnarh
                    FROM tb_intervencao i
                    LEFT JOIN tb_intervencao_interferencia a ON i.id = a.intervencao_id
                    LEFT JOIN tb_interferencia u ON a.interferencia_id = u.id
                    LEFT JOIN tb_operador_telemetria o ON i.operador_telemetria = o.id
                    WHERE LOWER(i.rotulo::text) LIKE LOWER(%s)
                    ORDER BY i.rotulo;
                    """
                    
                    cursor.execute(query, (termo_busca,))
                    resultados = cursor.fetchall()
                finally:
                    QApplication.restoreOverrideCursor()
            
            else:
                # -- Busca Textual Unificada (View) --
                base_query = """
                SELECT DISTINCT id, rotulo_medidor, nome_usuario, nome_operador,
                       nu_interferencia_cnarh
                FROM view_ft_intervencao
                """
                order_by = "ORDER BY rotulo_medidor;"                

                if criterio == "Usuário":
                    query = f"{base_query} WHERE LOWER(nome_usuario) LIKE LOWER(%s) {order_by}"
                elif criterio == "CNARH":
                    query = f"{base_query} WHERE LOWER(nu_cnarh) LIKE LOWER(%s) {order_by}"
                elif criterio == "Operador":
                    query = f"{base_query} WHERE LOWER(nome_operador) LIKE LOWER(%s) {order_by}"
                else:
                    return

                cursor.execute(query, (termo_busca,))
                resultados = cursor.fetchall()

            # === Atualização da Interface ===
            self.lista_resultados.clear()
            self.btn_desativar.setEnabled(False)
            self.btn_reativar.setEnabled(False)
            self.limpar_campos_formulario()
            
            if not resultados:
                self.lista_resultados.addItem("Nenhum resultado encontrado.")
                self.lbl_contagem.setVisible(False)
                return
            
            # Preenchimento da Lista
            for id_medidor, rotulo, nome_usuario, operador_nome, *_ in resultados:
                texto = f"{rotulo}"
                if nome_usuario:
                    texto += f" - Usuário: {nome_usuario}"
                if operador_nome:
                    texto += f" | Operador: {operador_nome}"
                item = QListWidgetItem(texto)
                item.setData(Qt.UserRole, id_medidor)
                self.lista_resultados.addItem(item)

            # Contagem idêntica ao Excel: interferências distintas via nu_interferencia_cnarh (r[4])
            nu_interf_distintas = len({r[4] for r in resultados if r[4]})
            self.lbl_contagem.setText(
                f"Total: {len(resultados)} medidor(es)  |  {nu_interf_distintas} interferência(s)"
            )
            self.lbl_contagem.setVisible(True)

            #nu_interf = len({item.text().split(" - ")[0] for item in
            #                 [self.lista_resultados.item(i)
            #                  for i in range(self.lista_resultados.count())]
            #                 if item.data(Qt.UserRole)})
            #self.lbl_contagem.setText(
            #    f"Total: {len(resultados)} medidor(es)  |  {nu_interf} interferência(s)"
            #)
            #self.lbl_contagem.setVisible(True)
                
        except Exception as e:
            # Garante que o cursor seja restaurado em caso de erro
            if criterio in ["Sistema Hídrico", "Rótulo"]:
                QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Erro", f"Erro na busca: {e}")
        finally:
            if cursor:
                cursor.close()

    def selecionar_sugestao(self, index):
        """Seleciona uma sugestão do combobox e preenche o campo de busca."""
        if index >= 0:
            texto = self.combo_sugestoes.itemText(index)
            self.input_busca.setText(texto)
            self.combo_sugestoes.setVisible(False)
            self.buscar_medidores()
            
    def buscar_medidores_por_operador_id(self, id_operador):
        """Busca medidores exclusivamente pelo ID do operador (para a funcionalidade 'Ver Vinculados')."""
        if not id_operador:
            return

        try:
            cursor = self.conn.cursor()
            
            # Utiliza a view_ft_intervencao assumindo que contenha id_operador_telemetria ou similar
            # Se a view não tiver o ID, usamos nome (menos preciso), mas assumimos aqui a estrutura ideal.
            # Caso a view não tenha o ID, manteríamos o join original.
            # Vamos manter a query original via join para garantir que funcione, 
            # pois a view_view_ft_intervencao pode não ter o ID do operador (apenas o nome).
            query = """
            SELECT i.rotulo, i.id, inf.nome_usuario, ot.nome as operador_nome
            FROM tb_intervencao i
            LEFT JOIN tb_intervencao_interferencia ii ON i.id = ii.intervencao_id
            LEFT JOIN tb_interferencia inf ON ii.interferencia_id = inf.id
            LEFT JOIN tb_operador_telemetria ot ON i.operador_telemetria = ot.id
            WHERE i.operador_telemetria = %s
            ORDER BY i.rotulo;
            """
                
            cursor.execute(query, (id_operador,))
            resultados = cursor.fetchall()
            
            self.lista_resultados.clear()
            self.btn_desativar.setEnabled(False)
            self.btn_reativar.setEnabled(False)
            self.limpar_campos_formulario()
            
            if not resultados:
                self.lista_resultados.addItem("Nenhum resultado encontrado.")
                self.btn_ver_mapa.setVisible(False)
                return
            '''
            # Lógica do botão Ver no Mapa
            qtd = len(resultados)
            self.btn_ver_mapa.setVisible(True)
            self.btn_ver_mapa.setText("Ver medidor no mapa" if qtd == 1 else "Ver medidores no mapa")
            '''    
            for rotulo, id_medidor, nome_usuario, operador_nome in resultados:
                texto = f"{rotulo}"
                if nome_usuario:
                    texto += f" - Usuário: {nome_usuario}"
                if operador_nome:
                    texto += f" | Operador: {operador_nome}"
                item = QListWidgetItem(texto)
                item.setData(Qt.UserRole, id_medidor)
                self.lista_resultados.addItem(item)
                self.lbl_contagem.setText(
                    f"Total: {len(resultados)} medidor(es)"
                )
                self.lbl_contagem.setVisible(True)                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro na busca: {e}")
        finally:
            cursor.close()
                   
    def carregar_dados_medidor(self, item):
        """Carrega os dados do medidor selecionado."""
        id_medidor = item.data(Qt.UserRole)
        if not id_medidor:
            return
            
        try:
            cursor = self.conn.cursor()
            
            # Consulta principal (mantida original pois carrega dados detalhados para edição)
            query = """
            SELECT 
                i.id, i.rotulo, i.vazao_nominal, i.potencia,
                i.tipo_medidor_id, i.modo_transmissao_id, i.operador_telemetria,
                inf.codigo_interferencia, inf.nome_usuario, inf.numero_cadastro,
                tm.descricao, mt.descricao, ot.nome, ot.email,
                ci.codigo_uc
            FROM tb_intervencao i
            LEFT JOIN tb_intervencao_interferencia ii ON i.id = ii.intervencao_id
            LEFT JOIN tb_interferencia inf ON ii.interferencia_id = inf.id
            LEFT JOIN tb_tipo_medidor tm ON i.tipo_medidor_id = tm.id
            LEFT JOIN tb_modo_transmissao mt ON i.modo_transmissao_id = mt.id
            LEFT JOIN tb_operador_telemetria ot ON i.operador_telemetria = ot.id
            LEFT JOIN tb_codigo_uc_intervencao ci ON i.id = ci.intervencao_id
            WHERE i.id = %s;
            """
            
            cursor.execute(query, (id_medidor,))
            medidor = cursor.fetchone()
            
            if medidor:
                self.medidor_atual = medidor
                
                # Preencher campos
                self.input_rotulo.setText(medidor[1] if medidor[1] else "")
                self.input_codigo_uc.setText(medidor[14] if medidor[14] else "")
                
                # AJUSTE: Exibir "0" se valor for None no banco
                self.input_vazao.setText(str(medidor[2]) if medidor[2] is not None else "0")
                self.input_potencia.setText(str(medidor[3]) if medidor[3] is not None else "0")
                
                # Configurar comboboxes
                tipo_medidor_id = medidor[4]
                modo_transmissao_id = medidor[5]
                operador_id = medidor[6]
                
                for i in range(self.combo_equipamento.count()):
                    if self.combo_equipamento.itemData(i) == tipo_medidor_id:
                        self.combo_equipamento.setCurrentIndex(i)
                        break
                        
                for i in range(self.combo_transmissao.count()):
                    if self.combo_transmissao.itemData(i) == modo_transmissao_id:
                        self.combo_transmissao.setCurrentIndex(i)
                        break
                
                self.carregar_operadores_combo()
                
                operador_selecionado = False
                for i in range(self.combo_operador_responsavel.count()):
                    if self.combo_operador_responsavel.itemData(i) == operador_id:
                        self.combo_operador_responsavel.setCurrentIndex(i)
                        operador_selecionado = True
                        break
                
                if not operador_selecionado and operador_id:
                    operador_nome = medidor[12] if medidor[12] else "Desconhecido"
                    operador_email = medidor[13] if medidor[13] else ""
                    display_text = f"{operador_nome} ({operador_email})" if operador_email else operador_nome
                    self.combo_operador_responsavel.insertItem(1, display_text, operador_id)
                    self.combo_operador_responsavel.setCurrentIndex(1)
                
                # Preencher campos de informações relacionadas
                self.input_interferencia.setText(medidor[7] if medidor[7] else "")
                self.input_usuario.setText(medidor[8] if medidor[8] else "")
                self.input_cnarh.setText(medidor[9] if medidor[9] else "")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar dados do medidor: {e}")
        finally:
            cursor.close()
            
        # Resetar flags de conversão ao carregar novo medidor
        self.vazao_ja_convertida = False
        self.potencia_ja_convertida = False
        self.ultimo_valor_vazao = self.input_vazao.text().replace(',', '.')
        self.ultimo_valor_potencia = self.input_potencia.text().replace(',', '.')        
                
    def carregar_equipamentos(self):
        """Carrega os equipamentos de medição."""
        try:
            cursor = self.conn.cursor()
            query = "SELECT id, descricao FROM tb_tipo_medidor ORDER BY descricao;"
            cursor.execute(query)
            equipamentos = cursor.fetchall()
            
            self.combo_equipamento.clear()
            self.combo_equipamento.addItem(" -- ", None)
            
            for id_equip, descricao in equipamentos:
                self.combo_equipamento.addItem(descricao, id_equip)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar equipamentos: {e}")
        finally:
            cursor.close()
            
    def carregar_modos_transmissao(self):
        """Carrega os modos de transmissão."""
        try:
            cursor = self.conn.cursor()
            query = "SELECT id, descricao FROM tb_modo_transmissao ORDER BY descricao;"
            cursor.execute(query)
            modos = cursor.fetchall()
            
            self.combo_transmissao.clear()
            self.combo_transmissao.addItem(" -- ", None)
            
            for id_modo, descricao in modos:
                self.combo_transmissao.addItem(descricao, id_modo)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar modos de transmissão: {e}")
        finally:
            cursor.close()
                       
    def carregar_operadores_combo(self):
        """Carrega a lista de operadores no combobox de operador responsável."""
        try:
            cursor = self.conn.cursor()
            query = "SELECT id, nome, email FROM tb_operador_telemetria ORDER BY nome ASC;"
            cursor.execute(query)
            operadores = cursor.fetchall()
            
            self.combo_operador_responsavel.clear()
            self.combo_operador_responsavel.addItem(" -- ", None)
            
            for id_operador, nome, email in operadores:
                display_text = f"{nome} ({email})" if email else nome
                self.combo_operador_responsavel.addItem(display_text, id_operador)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar operadores: {e}")
        finally:
            cursor.close()
            
    def habilitar_salvar(self):
        """Habilita o botão de salvar quando houver alterações."""
        if self.usuario_logado == "telemetria_ro":
            return

        if self.medidor_atual:
            codigo_uc_original = self.medidor_atual[14] or ""
            vazao_original = str(self.medidor_atual[2]) if self.medidor_atual[2] is not None else "0"
            potencia_original = str(self.medidor_atual[3]) if self.medidor_atual[3] is not None else "0"
            tipo_medidor_original = self.medidor_atual[4]
            modo_transmissao_original = self.medidor_atual[5]
            operador_original = self.medidor_atual[6]
            
            codigo_uc_atual = self.input_codigo_uc.text()
            vazao_atual = self.input_vazao.text()
            potencia_atual = self.input_potencia.text()
            tipo_medidor_atual = self.combo_equipamento.currentData()
            modo_transmissao_atual = self.combo_transmissao.currentData()
            operador_atual = self.combo_operador_responsavel.currentData()
            
            alterado = (codigo_uc_original != codigo_uc_atual or
                       vazao_original != vazao_atual or
                       potencia_original != potencia_atual or
                       tipo_medidor_original != tipo_medidor_atual or
                       modo_transmissao_original != modo_transmissao_atual or
                       operador_original != operador_atual)
            
            self.btn_salvar.setEnabled(alterado)

    def processar_vazao(self):
        """Processa a vazão assim que o usuário sair do campo."""
        if hasattr(self, '_processing_vazao') and self._processing_vazao:
            return

        vazao_texto = self.input_vazao.text().strip().replace(',', '.')
        
        if not vazao_texto:
            self.ultimo_valor_vazao = None
            return
        
        if self.ultimo_valor_vazao is not None and vazao_texto == self.ultimo_valor_vazao:
            return
        
        self._processing_vazao = True
        
        try:
            vazao_valor_original = float(vazao_texto)
            
            dialogo_vazao = DialogoUnidadeVazao(vazao_valor_original, self)
            
            if dialogo_vazao.exec_() == QDialog.Accepted:
                unidade_vazao = dialogo_vazao.get_unidade()
                
                self.input_vazao.blockSignals(True)
                
                if unidade_vazao == "m3h":
                    vazao_calculado = vazao_valor_original / 3600
                    if vazao_calculado >= 0.1:
                        vazao_novo = round(vazao_calculado, 2)
                        self.input_vazao.setText(f"{vazao_novo:.2f}")
                        vazao_formatada = f"{vazao_novo:.2f}"
                    else:
                        vazao_novo = round(vazao_calculado, 3)
                        self.input_vazao.setText(f"{vazao_novo:.3f}")
                        vazao_formatada = f"{vazao_novo:.3f}"
                    
                    self.ultimo_valor_vazao = self.input_vazao.text()
                    
                    QMessageBox.information(self, "Conversão Aplicada", 
                        f"Vazão convertida:\n{vazao_valor_original} m³/h → {vazao_formatada} m³/s")
                    
                elif unidade_vazao == "m3s":
                    if vazao_valor_original >= 0.1:
                        vazao_novo = round(vazao_valor_original, 2)
                        self.input_vazao.setText(f"{vazao_novo:.2f}")
                    else:
                        vazao_novo = round(vazao_valor_original, 3)
                        self.input_vazao.setText(f"{vazao_novo:.3f}")
                    
                    self.ultimo_valor_vazao = self.input_vazao.text()
                
                self.input_vazao.blockSignals(False)
            else:
                self.input_vazao.blockSignals(True)
                self.input_vazao.clear()
                self.input_vazao.blockSignals(False)
                self.ultimo_valor_vazao = None
                
        except ValueError:
            self.input_vazao.blockSignals(True)
            self.input_vazao.clear()
            self.input_vazao.blockSignals(False)
            self.ultimo_valor_vazao = None
            QMessageBox.warning(self, "Valor inválido", "A vazão deve ser um número válido.")
        finally:
            self._processing_vazao = False

    def processar_potencia(self):
        """Processa a potência assim que o usuário sair do campo."""
        if hasattr(self, '_processing_potencia') and self._processing_potencia:
            return

        potencia_texto = self.input_potencia.text().strip().replace(',', '.')
        
        if not potencia_texto:
            self.ultimo_valor_potencia = None
            return
        
        if self.ultimo_valor_potencia is not None and potencia_texto == self.ultimo_valor_potencia:
            return
        
        self._processing_potencia = True
        
        try:
            potencia_valor_original = float(potencia_texto)
            
            dialogo_potencia = DialogoUnidadePotencia(potencia_valor_original, self)
            
            if dialogo_potencia.exec_() == QDialog.Accepted:
                unidade_potencia = dialogo_potencia.get_unidade()
                
                self.input_potencia.blockSignals(True)
                
                if unidade_potencia == "cv":
                    potencia_novo = int(round(potencia_valor_original * 0.7355))
                    self.input_potencia.setText(str(potencia_novo))
                    self.ultimo_valor_potencia = str(potencia_novo)
                    
                    QMessageBox.information(self, "Conversão Aplicada",
                        f"Potência convertida:\n{potencia_valor_original} cv → {potencia_novo} kW")
                        
                elif unidade_potencia == "kw":
                    potencia_novo = int(round(potencia_valor_original))
                    self.input_potencia.setText(str(potencia_novo))
                    self.ultimo_valor_potencia = str(potencia_novo)
                
                self.input_potencia.blockSignals(False)
            else:
                self.input_potencia.blockSignals(True)
                self.input_potencia.clear()
                self.input_potencia.blockSignals(False)
                self.ultimo_valor_potencia = None
                
        except ValueError:
            self.input_potencia.blockSignals(True)
            self.input_potencia.clear()
            self.input_potencia.blockSignals(False)
            self.ultimo_valor_potencia = None
            QMessageBox.warning(self, "Valor inválido", "A potência deve ser um número válido.")
        finally:
            self._processing_potencia = False
        
    def salvar_alteracoes(self):
        """Salva as alterações nos dados do medidor."""
        if not self.medidor_atual:
            return
            
        id_medidor = self.medidor_atual[0]
        
        codigo_uc_novo = self.input_codigo_uc.text().strip()
        
        vazao_texto = self.input_vazao.text().strip().replace(',', '.')
        if not vazao_texto:
            QMessageBox.warning(self, "Campo Vazio", "Os campos vazão ou potência não podem ficar vazio")
            return
        try:
            vazao_novo = float(vazao_texto)
        except ValueError:
            QMessageBox.warning(self, "Valor inválido", "A vazão deve ser um número válido.")
            return
        
        potencia_texto = self.input_potencia.text().strip().replace(',', '.')
        if not potencia_texto:
            QMessageBox.warning(self, "Campo Vazio", "Os campos vazão ou potência não podem ficar vazio")
            return
        try:
            potencia_novo = int(round(float(potencia_texto)))
        except ValueError:
            QMessageBox.warning(self, "Valor inválido", "A potência deve ser um número válido.")
            return
        
        tipo_medidor_novo = self.combo_equipamento.currentData()
        modo_transmissao_novo = self.combo_transmissao.currentData()
        operador_novo = self.combo_operador_responsavel.currentData()
        
        if not tipo_medidor_novo or not modo_transmissao_novo:
            QMessageBox.warning(self, "Campos obrigatórios", "Selecione o equipamento e modo de transmissão.")
            return
            
        if not operador_novo:
            QMessageBox.warning(self, "Operador obrigatório", "Selecione um operador responsável.")
            return
            
        mensagem = "Confirme as alterações:\n\n"
        
        if (self.medidor_atual[14] or "") != codigo_uc_novo:
            mensagem += f"Código UC: {self.medidor_atual[14] or ''} → {codigo_uc_novo}\n"
            
        if str(self.medidor_atual[2] if self.medidor_atual[2] is not None else "0") != str(vazao_novo):
            mensagem += f"Vazão: {self.medidor_atual[2] or ''} → {vazao_novo} m³/s\n"
            
        if str(self.medidor_atual[3] if self.medidor_atual[3] is not None else "0") != str(potencia_novo):
            mensagem += f"Potência: {self.medidor_atual[3] or ''} → {potencia_novo} kW\n"
            
        try:
            cursor = self.conn.cursor()
            
            cursor.execute("SELECT descricao FROM tb_tipo_medidor WHERE id = %s", (self.medidor_atual[4],))
            tipo_medidor_original_desc = cursor.fetchone()
            tipo_medidor_original_desc = tipo_medidor_original_desc[0] if tipo_medidor_original_desc else "Desconhecido"
            
            cursor.execute("SELECT descricao FROM tb_modo_transmissao WHERE id = %s", (self.medidor_atual[5],))
            modo_transmissao_original_desc = cursor.fetchone()
            modo_transmissao_original_desc = modo_transmissao_original_desc[0] if modo_transmissao_original_desc else "Desconhecido"
            
            cursor.execute("SELECT descricao FROM tb_tipo_medidor WHERE id = %s", (tipo_medidor_novo,))
            tipo_medidor_novo_desc = cursor.fetchone()[0]
            
            cursor.execute("SELECT descricao FROM tb_modo_transmissao WHERE id = %s", (modo_transmissao_novo,))
            modo_transmissao_novo_desc = cursor.fetchone()[0]
            
            operador_original_nome = self.medidor_atual[12] if self.medidor_atual[12] else "Desconhecido"
            operador_original_email = self.medidor_atual[13] if self.medidor_atual[13] else ""
            
            cursor.execute("SELECT nome, email FROM tb_operador_telemetria WHERE id = %s", (operador_novo,))
            operador_novo_info = cursor.fetchone()
            operador_novo_nome = operador_novo_info[0] if operador_novo_info else "Desconhecido"
            operador_novo_email = operador_novo_info[1] if operador_novo_info and operador_novo_info[1] else ""
            
            if self.medidor_atual[4] != tipo_medidor_novo:
                mensagem += f"Equipamento: {tipo_medidor_original_desc} → {tipo_medidor_novo_desc}\n"
                
            if self.medidor_atual[5] != modo_transmissao_novo:
                mensagem += f"Transmissão: {modo_transmissao_original_desc} → {modo_transmissao_novo_desc}\n"
            
            if self.medidor_atual[6] != operador_novo:
                operador_original_display = f"{operador_original_nome}"
                if operador_original_email:
                    operador_original_display += f" ({operador_original_email})"
                
                operador_novo_display = f"{operador_novo_nome}"
                if operador_novo_email:
                    operador_novo_display += f" ({operador_novo_email})"
                
                mensagem += f"Operador responsável: {operador_original_display} → {operador_novo_display}\n"
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao obter descrições: {e}")
            return
        finally:
            cursor.close()
            
        resposta = QMessageBox.question(
            self, "Confirmar alterações", mensagem,
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if resposta != QMessageBox.Yes:
            return
            
        try:
            cursor = self.conn.cursor()
            
            query_intervencao = """
            UPDATE tb_intervencao 
            SET vazao_nominal = %s, potencia = %s, 
                tipo_medidor_id = %s, modo_transmissao_id = %s,
                operador_telemetria = %s
            WHERE id = %s;
            """
            cursor.execute(query_intervencao, (vazao_novo, potencia_novo, 
                                              tipo_medidor_novo, modo_transmissao_novo,
                                              operador_novo, id_medidor))
            
            if codigo_uc_novo:
                cursor.execute("SELECT 1 FROM tb_codigo_uc_intervencao WHERE intervencao_id = %s", (id_medidor,))
                if cursor.fetchone():
                    query_uc = "UPDATE tb_codigo_uc_intervencao SET codigo_uc = %s WHERE intervencao_id = %s"
                else:
                    query_uc = "INSERT INTO tb_codigo_uc_intervencao (codigo_uc, intervencao_id) VALUES (%s, %s)"
                cursor.execute(query_uc, (codigo_uc_novo, id_medidor))
            else:
                cursor.execute("""
                    UPDATE tb_codigo_uc_intervencao 
                    SET codigo_uc = NULL 
                    WHERE intervencao_id = %s
                """, (id_medidor,))
            
            try:
                codigo_interferencia = self.medidor_atual[7]

                if codigo_interferencia:
                    cursor.execute("""
                        SELECT DISTINCT ci.codigo_uc
                        FROM tb_codigo_uc_intervencao ci
                        JOIN tb_intervencao_interferencia ii ON ii.intervencao_id = ci.intervencao_id
                        JOIN tb_interferencia inf ON inf.id = ii.interferencia_id
                        WHERE inf.codigo_interferencia = %s AND ci.codigo_uc IS NOT NULL
                        ORDER BY ci.codigo_uc
                    """, (codigo_interferencia,))
                    ucs_distintos = [row[0] for row in cursor.fetchall()]
                    codigo_uc_concat = ', '.join(ucs_distintos) if ucs_distintos else None

                    cursor.execute("""
                        UPDATE tb_interferencia 
                        SET codigo_uc = %s
                        WHERE codigo_interferencia = %s
                    """, (codigo_uc_concat, codigo_interferencia))
                        
            except Exception as e_sync:
                print(f"Aviso: Não foi possível atualizar tb_interferencia: {e_sync}")
            
            self.conn.commit()
            
            QMessageBox.information(self, "Sucesso", "Dados atualizados com sucesso!")
            
            self.medidor_atual = list(self.medidor_atual)
            self.medidor_atual[2] = vazao_novo
            self.medidor_atual[3] = potencia_novo
            self.medidor_atual[4] = tipo_medidor_novo
            self.medidor_atual[5] = modo_transmissao_novo
            self.medidor_atual[6] = operador_novo
            self.medidor_atual[12] = operador_novo_nome
            self.medidor_atual[13] = operador_novo_email
            self.medidor_atual[14] = codigo_uc_novo if codigo_uc_novo else None
            
            self.vazao_ja_convertida = False
            self.potencia_ja_convertida = False
            self.ultimo_valor_vazao = None
            self.ultimo_valor_potencia = None
            
            self.habilitar_salvar()
            
            self.buscar_medidores()
            
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Erro", f"Erro ao salvar alterações: {e}")
        finally:
            cursor.close()
        
    def limpar_campos(self):
        """Limpa todos os campos do formulário e a lista de resultados."""
        self.medidor_atual = None
        self.lista_resultados.clear()
        self.btn_desativar.setEnabled(False)
        #self.btn_ver_mapa.setVisible(False)
        self.limpar_campos_formulario()
        self.input_busca.clear()

    def abrir_dialogo_reativacao(self):
        """
        Abre diálogo inline para reativar medidor(es) com '#' no rótulo.
        Fluxo:
          1. Se houver seleção mista (com e sem '#'), avisa e pede confirmação.
          2. Mostra diálogo com listbox de operadores para vinculação.
          3. Após escolha, exibe confirmação e executa UPDATE.
        """
        itens_selecionados = self.lista_resultados.selectedItems()

        # Filtra apenas os medidores que possuem '#' no rótulo
        itens_reativacao = [
            item for item in itens_selecionados
            if '#' in item.text().split(" - ")[0]
        ]

        if not itens_reativacao:
            return

        ids_reativar = [item.data(Qt.UserRole) for item in itens_reativacao]
        qtd = len(ids_reativar)
        qtd_ignorados = len(itens_selecionados) - qtd

        # Avisa o usuário caso parte da seleção não tenha '#' e seja ignorada
        if qtd_ignorados > 0:
            aviso = (
                f"Atenção: {qtd_ignorados} medidor(es) selecionado(s) não estão "
                f"desativados (sem '#' no rótulo) e serão ignorados.\n\n"
                f"Apenas {qtd} medidor(es) com '#' serão reativados.\n\n"
                "Deseja continuar?"
            )
            resposta_aviso = QMessageBox.warning(
                self,
                "Seleção parcial",
                aviso,
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if resposta_aviso != QMessageBox.Yes:
                return

        # --- Diálogo 1: escolha do operador ---
        dlg_operador = QDialog(self)
        dlg_operador.setWindowTitle("Reativar Medidor(es)")
        dlg_operador.setMinimumWidth(380)

        layout_dlg = QVBoxLayout(dlg_operador)
        layout_dlg.setSpacing(10)
        layout_dlg.setContentsMargins(16, 16, 16, 12)

        lbl = QLabel("Vincular ao operador:")
        lbl.setStyleSheet("font-weight: bold;")
        layout_dlg.addWidget(lbl)

        combo_op = QComboBox()
        combo_op.setFixedHeight(30)
        combo_op.setStyleSheet("""
            QComboBox {
                border: 1px solid #ccc;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
        """)

        # Carrega operadores direto do banco
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                "SELECT id, nome, email FROM tb_operador_telemetria ORDER BY nome ASC;"
            )
            operadores = cursor.fetchall()
            cursor.close()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar operadores: {e}")
            return

        combo_op.addItem(" -- selecione -- ", None)
        for id_op, nome, email in operadores:
            display = f"{nome} ({email})" if email else nome
            combo_op.addItem(display, id_op)

        layout_dlg.addWidget(combo_op)

        btn_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        btn_box.button(QDialogButtonBox.Ok).setText("Continuar")
        btn_box.accepted.connect(dlg_operador.accept)
        btn_box.rejected.connect(dlg_operador.reject)
        layout_dlg.addWidget(btn_box)

        if dlg_operador.exec_() != QDialog.Accepted:
            return

        operador_id   = combo_op.currentData()
        operador_nome = combo_op.currentText()

        if not operador_id:
            QMessageBox.warning(
                self, "Operador não selecionado",
                "Selecione um operador para continuar."
            )
            return

        # --- Diálogo 2: confirmação ---
        msg_confirmacao = (
            f"Deseja reativar {qtd} medidor(es) e vinculá-los ao operador:\n\n"
            f"  {operador_nome}\n\n"
            "Isso removerá o '#' do rótulo e definirá o operador de telemetria."
        )
        resposta = QMessageBox.question(
            self,
            "Confirmar Reativação",
            msg_confirmacao,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if resposta != QMessageBox.Yes:
            return

        # --- Executa UPDATE ---
        try:
            cursor = self.conn.cursor()
            query = """
                UPDATE tb_intervencao
                SET operador_telemetria = %s,
                    rotulo = REPLACE(rotulo, '#', '')
                WHERE id = ANY(%s);
            """
            cursor.execute(query, (operador_id, ids_reativar))
            self.conn.commit()
            cursor.close()

            QMessageBox.information(
                self, "Sucesso",
                f"{qtd} medidor(es) reativado(s) com sucesso!"
            )

            self.limpar_campos()
            if self.input_busca.text().strip():
                self.buscar_medidores()

        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Erro", f"Erro ao reativar medidores: {e}")

    def exportar_medidores_excel(self):
        """
        Ponto de entrada: exibe diálogo de escolha de escopo e delega à geração
        do Excel.
        Opções (checkboxes independentes — cada uma marcada gera um arquivo):
          • Todos cadastrados
          • Do método de busca  (habilitado apenas quando há busca ativa)
          • Por período de atividade
        """
        # ── Determina se há uma busca ativa com resultados reais ──────────────────
        criterio_atual = self.combo_criterio.currentText().strip()
        termo_atual    = self.input_busca.text().strip()
 
        _tem_item_real = (
            self.lista_resultados.count() > 0
            and self.lista_resultados.item(0).data(Qt.UserRole) is not None
        )
        tem_busca = bool(termo_atual and _tem_item_real)
 
        # Resolve nome completo do item selecionado no combo de sugestões
        _idx = self.combo_sugestoes.currentIndex()
        valor_busca = (
            self.combo_sugestoes.itemText(_idx).strip()
            if (self.combo_sugestoes.isVisible()
                and _idx >= 0
                and self.combo_sugestoes.itemText(_idx).strip())
            else termo_atual
        )
 
        # ── Diálogo ───────────────────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle("Exportar Medidores — Excel")
        dlg.setMinimumWidth(480)
        dlg.setStyleSheet("background-color: white;")
 
        layout = QVBoxLayout(dlg)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 18, 20, 16)
 
        lbl_titulo = QLabel("Selecione o(s) escopo(s) da exportação:")
        lbl_titulo.setStyleSheet("font-weight: bold; font-size: 12px; color: #2c3e50;")
        layout.addWidget(lbl_titulo)
 
        # ── Opção 1: Todos cadastrados ────────────────────────────────────────────
        chk_todos = QCheckBox("Todos cadastrados")
        chk_todos.setStyleSheet(
            "QCheckBox { font-weight: bold; font-size: 11px; color: #175cc3; spacing: 6px; }"
            "QCheckBox::indicator { width: 15px; height: 15px;"
            "  border: 1px solid #175cc3; border-radius: 3px; background: white; }"
            "QCheckBox::indicator:checked { background-color: #175cc3; border-color: #175cc3; }"
        )
        lbl_todos = QLabel("Lista completa de medidores cadastrados no sistema.")
        lbl_todos.setWordWrap(True)
        lbl_todos.setStyleSheet("font-size: 11px; color: #495057; margin-left: 22px;")
        layout.addWidget(chk_todos)
        layout.addWidget(lbl_todos)
 
        # ── Opção 2: Do método de busca ───────────────────────────────────────────
        if tem_busca:
            desc_busca = (
                f"Apenas medidores do(a) <b>{criterio_atual}</b> "
                f"<i>{valor_busca}</i>."
            )
        else:
            desc_busca = (
                "Realize uma busca (Rótulo, Usuário, CNARH, Operador, "
                "UAM ou Sistema Hídrico) para habilitar esta opção."
            )
 
        chk_busca = QCheckBox("Do método de busca")
        chk_busca.setEnabled(tem_busca)
        chk_busca.setChecked(tem_busca)   # pré-marcado quando há busca ativa
        chk_busca.setStyleSheet(
            "QCheckBox { font-weight: bold; font-size: 11px; color: #1d7a3a; spacing: 6px; }"
            "QCheckBox::indicator { width: 15px; height: 15px;"
            "  border: 1px solid #1d7a3a; border-radius: 3px; background: white; }"
            "QCheckBox::indicator:checked { background-color: #1d7a3a; border-color: #1d7a3a; }"
            "QCheckBox:disabled { color: #adb5bd; }"
            "QCheckBox::indicator:disabled { border-color: #ced4da; background: #f1f3f5; }"
        )
        lbl_busca = QLabel(desc_busca)
        lbl_busca.setWordWrap(True)
        lbl_busca.setStyleSheet("font-size: 11px; color: #495057; margin-left: 22px;")
        layout.addWidget(chk_busca)
        layout.addWidget(lbl_busca)
 
        # ── Separador ─────────────────────────────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #dee2e6; margin: 2px 0;")
        layout.addWidget(sep)
 
        # ── Opção 3: Por período de atividade ─────────────────────────────────────
        chk_periodo = QCheckBox("Por período de atividade")
        chk_periodo.setStyleSheet(
            "QCheckBox { font-weight: bold; font-size: 11px; color: #6f4e8a; spacing: 6px; }"
            "QCheckBox::indicator { width: 15px; height: 15px;"
            "  border: 1px solid #6f4e8a; border-radius: 3px; background: white; }"
            "QCheckBox::indicator:checked { background-color: #6f4e8a; border-color: #6f4e8a; }"
        )
        lbl_periodo_desc = QLabel(
            "Medidores ativos em uma data específica ou dentro de um intervalo."
        )
        lbl_periodo_desc.setWordWrap(True)
        lbl_periodo_desc.setStyleSheet("font-size: 11px; color: #495057; margin-left: 22px;")
        layout.addWidget(chk_periodo)
        layout.addWidget(lbl_periodo_desc)
 
        # Controles de data (visíveis mas desabilitados até chk_periodo ser marcado)
        linha_datas = QHBoxLayout()
        linha_datas.setSpacing(8)
        linha_datas.setContentsMargins(22, 0, 0, 0)
 
        lbl_de = QLabel("De:")
        lbl_de.setStyleSheet("font-size: 11px; color: #495057;")
        lbl_de.setFixedWidth(22)
 
        dte_inicio = QDateEdit()
        dte_inicio.setCalendarPopup(True)
        dte_inicio.setDate(QDate.currentDate().addMonths(-1))
        dte_inicio.setDisplayFormat("dd/MM/yyyy")
        dte_inicio.setFixedHeight(26)
        dte_inicio.setEnabled(False)
        dte_inicio.setStyleSheet(
            "QDateEdit { border: 1px solid #c3aed6; border-radius: 3px;"
            "  padding: 2px 6px; font-size: 11px; background: white; }"
            "QDateEdit:focus { border-color: #6f4e8a; }"
            "QDateEdit:disabled { background: #f1f3f5; color: #adb5bd; border-color: #dee2e6; }"
        )
 
        chk_intervalo = QCheckBox("até")
        chk_intervalo.setEnabled(False)
        chk_intervalo.setStyleSheet(
            "QCheckBox { font-size: 11px; color: #495057; spacing: 5px; }"
            "QCheckBox::indicator { width: 13px; height: 13px;"
            "  border: 1px solid #c3aed6; border-radius: 3px; background: white; }"
            "QCheckBox::indicator:checked { background-color: #6f4e8a; border-color: #6f4e8a; }"
            "QCheckBox:disabled { color: #adb5bd; }"
            "QCheckBox::indicator:disabled { border-color: #ced4da; background: #f1f3f5; }"
        )
 
        dte_fim = QDateEdit()
        dte_fim.setCalendarPopup(True)
        dte_fim.setDate(QDate.currentDate())
        dte_fim.setDisplayFormat("dd/MM/yyyy")
        dte_fim.setFixedHeight(26)
        dte_fim.setEnabled(False)
        dte_fim.setStyleSheet(
            "QDateEdit { border: 1px solid #c3aed6; border-radius: 3px;"
            "  padding: 2px 6px; font-size: 11px; background: white; }"
            "QDateEdit:focus { border-color: #6f4e8a; }"
            "QDateEdit:disabled { background: #f1f3f5; color: #adb5bd; border-color: #dee2e6; }"
        )
 
        # habilita controles de data quando chk_periodo é marcado
        def _toggle_periodo(marcado):
            dte_inicio.setEnabled(marcado)
            chk_intervalo.setEnabled(marcado)
            dte_fim.setEnabled(marcado and chk_intervalo.isChecked())
 
        chk_periodo.toggled.connect(_toggle_periodo)
        chk_intervalo.toggled.connect(
            lambda marcado: dte_fim.setEnabled(marcado and chk_periodo.isChecked())
        )
 
        linha_datas.addWidget(lbl_de)
        linha_datas.addWidget(dte_inicio)
        linha_datas.addWidget(chk_intervalo)
        linha_datas.addWidget(dte_fim)
        linha_datas.addStretch()
        layout.addLayout(linha_datas)
 
        # ── Botão único + Cancelar ────────────────────────────────────────────────
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.HLine)
        sep2.setStyleSheet("color: #dee2e6; margin: 4px 0;")
        layout.addWidget(sep2)
 
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
 
        btn_gerar = QPushButton("Gerar Excel")
        btn_gerar.setFixedHeight(34)
        btn_gerar.setStyleSheet(
            "QPushButton { background-color: #175cc3; color: white;"
            "  font-weight: bold; font-size: 11px; border-radius: 4px; padding: 6px 18px; }"
            "QPushButton:hover { background-color: #1249a3; }"
            "QPushButton:disabled { background-color: #adb5bd; color: #e9ecef; }"
        )
 
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setFixedHeight(34)
        btn_cancelar.setStyleSheet(
            "QPushButton { background-color: #6c757d; color: white;"
            "  font-size: 11px; border-radius: 4px; padding: 6px 14px; }"
            "QPushButton:hover { background-color: #5a6268; }"
        )
 
        # desabilita "Gerar Excel" quando nenhum checkbox estiver marcado
        def _atualizar_btn(*_):
            btn_gerar.setEnabled(
                chk_todos.isChecked()
                or chk_busca.isChecked()
                or chk_periodo.isChecked()
            )
 
        chk_todos.toggled.connect(_atualizar_btn)
        chk_busca.toggled.connect(_atualizar_btn)
        chk_periodo.toggled.connect(_atualizar_btn)
        _atualizar_btn()   # estado inicial
 
        btn_gerar.clicked.connect(dlg.accept)
        btn_cancelar.clicked.connect(dlg.reject)
 
        btn_row.addWidget(btn_gerar)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancelar)
        layout.addLayout(btn_row)
 
        if dlg.exec_() != QDialog.Accepted:
            return
 
        # ── Coleta parâmetros de período (usados em mais de um ramo) ─────────────
        data_inicio = dte_inicio.date().toPyDate() if chk_periodo.isChecked() else None
        data_fim    = (dte_fim.date().toPyDate()
                       if chk_periodo.isChecked() and chk_intervalo.isChecked()
                       else None)
 
        # ── Despacho ──────────────────────────────────────────────────────────────
        # "Todos cadastrados" é sempre independente — processa primeiro, sem perguntas
        if chk_todos.isChecked():
            self._gerar_excel_medidores()
 
        # Quando busca E período estão marcados simultaneamente, pergunta ao usuário
        # como o arquivo de período deve ser gerado.
        if chk_busca.isChecked() and chk_periodo.isChecked():
            fim_str      = data_fim.strftime('%d/%m/%Y') if data_fim else None
            periodo_label = (
                f"{data_inicio.strftime('%d/%m/%Y')} → {fim_str}"
                if fim_str
                else data_inicio.strftime('%d/%m/%Y')
            )

            # ── Diálogo customizado ───────────────────────────────────────────────
            dlg2 = QDialog(self)
            dlg2.setWindowTitle("Exportação por período")
            dlg2.setMinimumWidth(480)
            dlg2.setStyleSheet("""
                QDialog { background-color: white; }
                QLabel  { background: transparent; }
                QRadioButton {
                    font-size: 11px;
                    color: #212529;
                    background: transparent;
                    spacing: 8px;
                    padding: 4px 0px;
                }
                QRadioButton::indicator {
                    width: 15px;
                    height: 15px;
                    border-radius: 8px;
                    border: 2px solid #5474b8;
                    background-color: white;
                }
                QRadioButton::indicator:checked {
                    background-color: #5474b8;
                    border: 2px solid #5474b8;
                    image: none;
                }
                QRadioButton::indicator:hover {
                    border-color: #175cc3;
                }
            """)

            lay2 = QVBoxLayout(dlg2)
            lay2.setSpacing(14)
            lay2.setContentsMargins(20, 18, 20, 16)

            # Texto da indagação
            lbl_pergunta = QLabel(
                f"Como o <b>método de busca</b> e o <b>período de atividade</b> estão "
                f"marcados simultaneamente, como deseja gerar o arquivo de medidores "
                f"ativos para o período <i>({periodo_label})</i>?"
            )
            lbl_pergunta.setWordWrap(True)
            lbl_pergunta.setStyleSheet("color: #212529; font-size: 11px; background: transparent;")
            lay2.addWidget(lbl_pergunta)

            # Separador
            sep_top = QFrame()
            sep_top.setFrameShape(QFrame.HLine)
            sep_top.setStyleSheet("color: #dee2e6; margin: 2px 0;")
            lay2.addWidget(sep_top)

            # RadioButtons
            grp_botoes = QButtonGroup(dlg2)

            rb_independente = QRadioButton("Arquivo independente do método de busca")
            #rb_independente.setStyleSheet("font-size: 11px; color: #212529; background: transparent;")

            rb_vinculado = QRadioButton(
                f"Arquivo vinculado ao método de busca  ({criterio_atual}: {valor_busca})"
            )
            #rb_vinculado.setStyleSheet("font-size: 11px; color: #212529; background: transparent;")

            grp_botoes.addButton(rb_independente)
            grp_botoes.addButton(rb_vinculado)

            lay2.addWidget(rb_independente)
            lay2.addWidget(rb_vinculado)

            # Separador
            sep_bot = QFrame()
            sep_bot.setFrameShape(QFrame.HLine)
            sep_bot.setStyleSheet("color: #dee2e6; margin: 2px 0;")
            lay2.addWidget(sep_bot)

            # Botões Gerar Excel / Cancelar
            btn_row2 = QHBoxLayout()
            btn_row2.setSpacing(10)

            btn_gerar2 = QPushButton("Gerar Excel")
            btn_gerar2.setEnabled(False)
            btn_gerar2.setFixedHeight(34)
            btn_gerar2.setStyleSheet(
                "QPushButton { background-color: #175cc3; color: white;"
                "  font-weight: bold; font-size: 11px; border-radius: 4px; padding: 6px 18px; }"
                "QPushButton:hover { background-color: #1249a3; }"
                "QPushButton:disabled { background-color: #adb5bd; color: #e9ecef; }"
            )

            btn_cancelar2 = QPushButton("Cancelar")
            btn_cancelar2.setFixedHeight(34)
            btn_cancelar2.setStyleSheet(
                "QPushButton { background-color: #6c757d; color: white;"
                "  font-size: 11px; border-radius: 4px; padding: 6px 14px; }"
                "QPushButton:hover { background-color: #5a6268; }"
            )

            # Habilita "Gerar Excel" assim que qualquer radiobutton for selecionado
            grp_botoes.buttonClicked.connect(lambda _: btn_gerar2.setEnabled(True))

            btn_gerar2.clicked.connect(dlg2.accept)
            btn_cancelar2.clicked.connect(dlg2.reject)

            btn_row2.addWidget(btn_gerar2)
            btn_row2.addStretch()
            btn_row2.addWidget(btn_cancelar2)
            lay2.addLayout(btn_row2)

            # ── Executa e despacha ────────────────────────────────────────────────
            # Gera o arquivo de busca em qualquer caso (estava marcado)
            self._gerar_excel_medidores(criterio=criterio_atual, termo=valor_busca)

            if dlg2.exec_() == QDialog.Accepted:
                if rb_independente.isChecked():
                    self._gerar_excel_medidores_periodo(data_inicio, data_fim)
                else:
                    self._gerar_excel_medidores_periodo(
                        data_inicio, data_fim,
                        criterio=criterio_atual, termo=valor_busca
                    )

        # Busca individual (sem período marcado)
        if chk_busca.isChecked() and not chk_periodo.isChecked():
            self._gerar_excel_medidores(criterio=criterio_atual, termo=valor_busca)

        # Período individual (sem busca marcada)
        if chk_periodo.isChecked() and not chk_busca.isChecked():
            self._gerar_excel_medidores_periodo(data_inicio, data_fim)                    
 
    def _gerar_excel_medidores(self, criterio=None, termo=None):
        """
        Gera o arquivo Excel de medidores.
 
        criterio / termo = None  → todos os medidores cadastrados.
        criterio / termo = str   → apenas os medidores retornados pelo critério
                                   de busca, usando a mesma lógica SQL de
                                   buscar_medidores():
                                   • "Rótulo"         – JOIN direto em tb_intervencao
                                   • "Usuário"        – view_ft_intervencao / nome_usuario
                                   • "CNARH"          – view_ft_intervencao / nu_cnarh
                                   • "Operador"       – view_ft_intervencao / nome_operador
                                   • "UAM"            – ST_Intersects c/ ft_uam_buffer / nmautomonit
                                   • "Sistema Hídrico"– ST_Intersects c/ ft_sishidrico_buffer / bafnm
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self, "Biblioteca ausente",
                "A biblioteca 'openpyxl' não está instalada.\n"
                "Instale-a via: pip install openpyxl"
            )
            return
 
        import os, sys, re
        from datetime import datetime
 
        # ── SELECT padrão de colunas para o Excel ─────────────────────────────────
        _cols_excel = """
            SELECT DISTINCT
                   v.id
                 , v.nome_usuario
                 , v.nome_operador
                 , v.nu_interferencia_cnarh
                 , v.nu_cnarh
                 , v.rotulo_medidor
                 , v.vazao_nominal
                 , v.potencia
                 , v.tipo_medidor
                 , v.modo_transmissao
            FROM public.view_ft_intervencao v
        """
 
        # ── Consultar banco ───────────────────────────────────────────────────────
        cursor = None
        try:
            cursor = self.conn.cursor()
            termo_busca = f"%{termo}%" if termo else None
 
            if not criterio or not termo:
                # ── Todos os medidores ────────────────────────────────────────────
                cursor.execute(f"""
                    {_cols_excel}
                    ORDER BY v.nome_usuario, v.rotulo_medidor
                """)
 
            elif criterio == "Rótulo":
                # Idêntico a buscar_medidores(): JOIN direto em tb_intervencao
                cursor.execute("""
                    SELECT DISTINCT
                           i.id
                         , inf.nome_usuario
                         , o.nome        AS nome_operador
                         , inf.codigo_interferencia AS nu_interferencia_cnarh
                         , inf.numero_cadastro      AS nu_cnarh
                         , i.rotulo      AS rotulo_medidor
                         , i.vazao_nominal
                         , i.potencia
                         , tm.descricao  AS tipo_medidor
                         , mt.descricao  AS modo_transmissao
                    FROM tb_intervencao i
                    LEFT JOIN tb_intervencao_interferencia a  ON i.id = a.intervencao_id
                    LEFT JOIN tb_interferencia              inf ON a.interferencia_id = inf.id
                    LEFT JOIN tb_operador_telemetria        o   ON i.operador_telemetria = o.id
                    LEFT JOIN tb_tipo_medidor               tm  ON i.tipo_medidor_id = tm.id
                    LEFT JOIN tb_modo_transmissao           mt  ON i.modo_transmissao_id = mt.id
                    WHERE LOWER(i.rotulo::text) LIKE LOWER(%s)
                    ORDER BY i.rotulo
                """, (termo_busca,))
 
            elif criterio == "Usuário":
                cursor.execute(f"""
                    {_cols_excel}
                    WHERE LOWER(v.nome_usuario) LIKE LOWER(%s)
                    ORDER BY v.nome_usuario, v.rotulo_medidor
                """, (termo_busca,))
 
            elif criterio == "CNARH":
                cursor.execute(f"""
                    {_cols_excel}
                    WHERE LOWER(v.nu_cnarh) LIKE LOWER(%s)
                    ORDER BY v.nome_usuario, v.rotulo_medidor
                """, (termo_busca,))
 
            elif criterio == "Operador":
                cursor.execute(f"""
                    {_cols_excel}
                    WHERE LOWER(v.nome_operador) LIKE LOWER(%s)
                    ORDER BY v.nome_usuario, v.rotulo_medidor
                """, (termo_busca,))
 
            elif criterio == "UAM":
                # Idêntico a buscar_medidores(): ST_Intersects com ft_uam_buffer
                QApplication.setOverrideCursor(Qt.WaitCursor)
                try:
                    cursor.execute(f"""
                        {_cols_excel}
                        JOIN ft_uam_buffer s
                            ON ST_Intersects(v.geom, ST_Transform(s.geom, 4326))
                        WHERE LOWER(s.nmautomonit) LIKE LOWER(%s)
                        ORDER BY v.nome_usuario, v.rotulo_medidor
                    """, (termo_busca,))
                finally:
                    QApplication.restoreOverrideCursor()
 
            elif criterio == "Sistema Hídrico":
                # Idêntico a buscar_medidores(): ST_Intersects com ft_sishidrico_buffer
                QApplication.setOverrideCursor(Qt.WaitCursor)
                try:
                    cursor.execute(f"""
                        {_cols_excel}
                        JOIN ft_sishidrico_buffer s
                            ON ST_Intersects(v.geom, ST_Transform(s.geom, 4326))
                        WHERE LOWER(s.bafnm) LIKE LOWER(%s)
                        ORDER BY v.nome_usuario, v.rotulo_medidor
                    """, (termo_busca,))
                finally:
                    QApplication.restoreOverrideCursor()
 
            else:
                QMessageBox.warning(self, "Critério inválido",
                                    f"Critério de busca não reconhecido: '{criterio}'.")
                return
 
            linhas = cursor.fetchall()
 
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao consultar medidores:\n{e}")
            return
        finally:
            if cursor:
                cursor.close()
 
        if not linhas:
            QMessageBox.information(self, "Aviso", "Nenhum medidor encontrado para os critérios selecionados.")
            return
 
        # ── helper fora do loop ───────────────────────────────────────────────────
        def _fmt(v, decimais=4):
            if v is None:
                return "—"
            try:
                return round(float(v), decimais)
            except Exception:
                return str(v)
 
        # ── Título e sufixo do nome de arquivo ────────────────────────────────────
        def _sanitizar(texto):
            """Remove/substitui caracteres inválidos para nome de arquivo."""
            import unicodedata
            # Normaliza acentos (ex: "Ç" → "C", "é" → "e")
            txt = unicodedata.normalize('NFKD', texto)
            txt = ''.join(c for c in txt if not unicodedata.combining(c))
            # Substitui espaços e caracteres especiais por '_'
            txt = re.sub(r'[^A-Za-z0-9_\-]', '_', txt)
            # Colapsa múltiplos '_' consecutivos
            txt = re.sub(r'_+', '_', txt).strip('_')
            return txt.upper()
 
        if criterio and termo:
            titulo_xlsx = (
                f"Lista de Medidores de Telemetria Cadastrados – "
                f"{criterio} {termo}"
            )
            sufixo_arquivo = f"{_sanitizar(criterio)}_{_sanitizar(termo)}"
        else:
            titulo_xlsx    = "Lista de Medidores de Telemetria Cadastrados"
            sufixo_arquivo = "TODOS"
 
        # ── Nome e caminho do arquivo ─────────────────────────────────────────────
        ts           = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"MEDIDORES_{sufixo_arquivo}_{ts}.xlsx"
        downloads = (
            os.path.join(os.environ.get('USERPROFILE', ''), 'Downloads')
            if sys.platform == 'win32'
            else os.path.join(os.path.expanduser('~'), 'Downloads')
        )
        caminho = os.path.join(downloads, nome_arquivo)
 
        if os.path.exists(caminho):
            resp = QMessageBox.question(
                self, "Arquivo Existente",
                f"'{nome_arquivo}' já existe.\nDeseja substituir?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if resp == QMessageBox.No:
                return
 
        # ── Estilos (idênticos ao padrão institucional) ───────────────────────────
        fill_azul   = PatternFill("solid", fgColor="175cc3")
        fill_alt    = PatternFill("solid", fgColor="eaf2ff")
        font_branca = Font(bold=True, size=10, color="ffffff")
        font_titulo = Font(bold=True, size=13, color="175cc3")
        font_sub    = Font(size=10, italic=True, color="495057")
        font_normal = Font(size=10)
        ali_centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ali_esq     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        borda       = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
 
        # ── Workbook / aba ────────────────────────────────────────────────────────
        # Contorno para bug openpyxl + lxml no Python 3.12/QGIS 3.3x (Windows):
        # copy(DEFAULT_FONT) → to_tree() → lxml.etree.Element causa access violation.
        import sys as _sys
        _lxml_mod   = _sys.modules.pop('lxml',        None)
        _lxml_etree = _sys.modules.pop('lxml.etree',  None)
        try:
            wb = Workbook()
        finally:
            if _lxml_mod   is not None: _sys.modules['lxml']       = _lxml_mod
            if _lxml_etree is not None: _sys.modules['lxml.etree'] = _lxml_etree
        ws = wb.active
        ws.title = "Medidores"
 
        colunas  = ["ID", "Usuário", "Operador", "Nº Interferência CNARH",
                    "Nº CNARH", "Rótulo do Medidor", "Vazão Nominal (m³/s)",
                    "Potência (kW)", "Tipo de Medidor", "Modo de Transmissão"]
        larguras = [6, 36, 30, 24, 18, 24, 22, 16, 22, 22]
        n_cols   = len(colunas)
        col_last = get_column_letter(n_cols)
 
        # Linha 1 – título (inclui critério+termo quando é busca filtrada)
        ws.merge_cells(f"A1:{col_last}1")
        ws["A1"]           = titulo_xlsx
        ws["A1"].font      = font_titulo
        ws["A1"].alignment = ali_centro
        ws.row_dimensions[1].height = 22
 
        # Linha 2 – subtítulo com data/hora e total
        ws.merge_cells(f"A2:{col_last}2")
        nu_interf_distintas = len({r[3] for r in linhas if r[3]})
        ws["A2"]           = (
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}"
            f"  –  Total: {len(linhas)} medidor(es)  |  {nu_interf_distintas} interferência(s)"
        )
        ws["A2"].font      = font_sub
        ws["A2"].alignment = ali_centro
        ws.row_dimensions[2].height = 18
 
        # Linha 3 – cabeçalho da tabela
        for ci, cab in enumerate(colunas, 1):
            cell           = ws.cell(row=3, column=ci, value=cab)
            cell.font      = font_branca
            cell.fill      = fill_azul
            cell.alignment = ali_centro
            cell.border    = borda
        ws.row_dimensions[3].height = 20
 
        # Linhas de dados
        for idx, row_data in enumerate(linhas, start=4):
            (id_med, nome_usu, nome_op, nu_interf, nu_cnarh,
             rotulo, vazao, potencia, tipo, transmissao) = row_data
 
            valores = [
                id_med,
                nome_usu    or "—",
                nome_op     or "—",
                nu_interf   or "—",
                nu_cnarh    or "—",
                rotulo      or "—",
                _fmt(vazao),
                _fmt(potencia, 2),
                tipo        or "—",
                transmissao or "—",
            ]
 
            fill_linha = fill_alt if (idx % 2 == 0) else None
 
            for ci, val in enumerate(valores, 1):
                cell        = ws.cell(row=idx, column=ci, value=val)
                cell.font   = font_normal
                cell.border = borda
                if fill_linha:
                    cell.fill = fill_linha
                cell.alignment = ali_centro if ci in (1, 7, 8) else ali_esq
                if ci in (7, 8) and isinstance(val, float):
                    cell.number_format = '#,##0.0000' if ci == 7 else '#,##0.00'
 
        # Larguras e freeze
        for ci, larg in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(ci)].width = larg
        ws.freeze_panes = "A4"
 
        # Rodapé
        rodape_row = len(linhas) + 5
        ws.merge_cells(f"A{rodape_row}:{col_last}{rodape_row}")
        cell_rod           = ws.cell(row=rodape_row, column=1,
                                     value="Sistema DURH Diária por Telemetria (SFI/ANA) – Relatório gerado automaticamente")
        cell_rod.font      = Font(size=8, italic=True, color="888888")
        cell_rod.alignment = ali_centro
 
        # ── Salvar ────────────────────────────────────────────────────────────────
        try:
            wb.save(caminho)
        except Exception as e:
            QMessageBox.critical(self, "Erro ao salvar", f"Não foi possível salvar o arquivo:\n{e}")
            return
 
        QMessageBox.information(
            self, "Exportação concluída",
            f"Arquivo exportado com sucesso!\n\n{caminho}",
            QMessageBox.Ok
        )
 
    def _gerar_excel_medidores_periodo(self, data_inicio, data_fim=None,
                                       criterio=None, termo=None):
        """
        Gera Excel de medidores ativos em uma data específica ou dentro de um
        intervalo, reutilizando o SELECT base do relatório de período de atividade.

        Parâmetros
        ----------
        data_inicio : datetime.date
            Data de referência (ou início do intervalo).
        data_fim : datetime.date | None
            Fim do intervalo. Quando None, aplica apenas data_inicio como
            ponto de corte (medidores que iniciaram atividade até aquela data).
        criterio : str | None
            Critério de busca ("Rótulo", "Usuário", "CNARH", "Operador",
            "UAM", "Sistema Hídrico").  None = sem filtro adicional.
        termo : str | None
            Valor do critério de busca.  None = sem filtro adicional.

        Filtro de período aplicado
        --------------------------
        • Data única  → ``MIN(data) FILTER (consumo > 0) <= data_inicio``
        • Intervalo   → ``MIN(data) FILTER (consumo > 0) <= data_fim``
                         AND ``MAX(data) FILTER (consumo > 0) >= data_inicio``

        Colunas exportadas
        ------------------
        CNARH | Interferência(s) | Usuário/Empreendimento | Nome Empreendimento |
        Rótulo do Medidor | UAM | Data Início Atividade | Total Dias c/ Dados |
        Total Dias c/ Dados no Período (apenas quando há intervalo de datas) |
        Total Dias s/ Dados no Período | Dias Sequenciais s/ Dados |
        Última Data c/ Dado
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self, "Biblioteca ausente",
                "A biblioteca 'openpyxl' não está instalada.\n"
                "Instale-a via: pip install openpyxl"
            )
            return
 
        import os, sys, re
        from datetime import datetime
        import unicodedata
 
        # ── Helpers ───────────────────────────────────────────────────────────────
        def _sanitizar(texto):
            txt = unicodedata.normalize('NFKD', texto)
            txt = ''.join(c for c in txt if not unicodedata.combining(c))
            txt = re.sub(r'[^A-Za-z0-9_\\-]', '_', txt)
            txt = re.sub(r'_+', '_', txt).strip('_')
            return txt.upper()
 
        # ── Monta filtro de período para o HAVING / WHERE interno ─────────────────
        # O SELECT base agrupa por (nu_cnarh, usuario, rotulo) e as métricas
        # de data ficam em metricas_telemetria.  Aplicamos o filtro na CTE
        # metricas_telemetria via WHERE na tb_telemetria_intervencao_diaria.
        #
        # Data única  : data_inicio_ativ <= data_ref
        # Intervalo   : data_inicio_ativ <= data_fim  AND  data_fim_ativ >= data_inicio
        #
        # "data_inicio_ativ" = MIN(t.data) FILTER (consumo > 0)
        # "data_fim_ativ"    = MAX(t.data) FILTER (consumo > 0)
 
        # ── Monta filtro adicional de busca (JOIN / WHERE extra) ──────────────────
        # Segue a mesma lógica do _gerar_excel_medidores().
        # Para UAM e Sistema Hídrico precisamos de um JOIN lateral extra na CTE
        # interferencias_base.
 
        # A CTE principal é construída dinamicamente para suportar os filtros.
 
        termo_like = f"%{termo}%" if termo else None
 
        # ── Constrói a query ──────────────────────────────────────────────────────
        # Parte fixa: interferencias_base com filtros padrão + filtro de busca
        # Parte variável: cláusula de período no HAVING final
 
        # Filtros fixos de qualidade do dado
        _where_base = """
            _vuoir.rotulo_intervencao_medidor !~~ '%%999%%'
            AND _vuoir.rotulo_intervencao_medidor !~~ '%%VERDE GRANDE%%'
            AND _vuoir.rotulo_intervencao_medidor !~~ '%%#'
            AND _vuoir.longitude IS NOT NULL
            AND _vuoir.latitude IS NOT NULL
        """
 
        # Filtro de busca adicional na CTE interferencias_base
        _join_extra   = ""
        _where_busca  = ""
        _params_busca = []
 
        if criterio and termo:
            if criterio == "Rótulo":
                _where_busca = "AND LOWER(rotulo_intervencao_medidor) LIKE LOWER(%s)"
                _params_busca = [termo_like]
            elif criterio == "Usuário":
                _where_busca = "AND LOWER(usuario) LIKE LOWER(%s)"
                _params_busca = [termo_like]
            elif criterio == "CNARH":
                _where_busca = "AND LOWER(nu_cnarh::text) LIKE LOWER(%s)"
                _params_busca = [termo_like]
            elif criterio == "Operador":
                # Operador não existe em view_usuario_operador_id_rotulo diretamente;
                # juntamos com tb_intervencao/tb_operador_telemetria via intervencao_id.
                _join_extra = """
                    JOIN public.tb_intervencao          _ti  ON _ti.id = intervencao_id
                    JOIN public.tb_operador_telemetria  _top ON _top.id = _ti.operador_telemetria
                """
                _where_busca = "AND LOWER(_top.nome) LIKE LOWER(%s)"
                _params_busca = [termo_like]
            elif criterio == "UAM":
                _join_extra = """
                    JOIN public.ft_uam_buffer _uam
                        ON ST_Intersects(
                            ST_Transform(ST_SetSRID(
                                ST_MakePoint(_vuoir.longitude, _vuoir.latitude), 4326), 4674),
                            _uam.geom
                        )
                """
                _where_busca = "AND LOWER(_uam.nmautomonit) LIKE LOWER(%s)"
                _params_busca = [termo_like]
            elif criterio == "Sistema Hídrico":
                _join_extra = """
                    JOIN public.ft_sishidrico_buffer _sh
                        ON ST_Intersects(
                            ST_Transform(ST_SetSRID(
                                ST_MakePoint(_vuoir.longitude, _vuoir.latitude), 4326), 4674),
                            _sh.geom
                        )
                """
                _where_busca = "AND LOWER(_sh.bafnm) LIKE LOWER(%s)"
                _params_busca = [termo_like]
 
        # Filtro de período no HAVING — usa as colunas propagadas até uam_priorizado
        # (data_inicio e data_fim são aliases definidos em dados_empreendimento,
        #  herdados via d.* em uam_priorizado; mt.* não está mais acessível aqui).
        if data_fim:
            _having_periodo = """
                HAVING
                    MIN(data_inicio) IS NOT NULL
                    AND MIN(data_inicio) <= %(p_fim)s
                    AND MAX(data_fim)    >= %(p_ini)s
            """
        else:
            _having_periodo = """
                HAVING
                    MIN(data_inicio) IS NOT NULL
                    AND MIN(data_inicio) <= %(p_ini)s
            """
 
        _params_periodo = {'p_ini': data_inicio, 'p_fim': data_fim}
 
        # ── Query completa ────────────────────────────────────────────────────────
        sql = f"""
        WITH interferencias_base AS (
            SELECT
                _vuoir.nu_cnarh,
                _vuoir.nu_interferencia_cnarh,
                _vuoir.usuario,
                _vuoir.rotulo_intervencao_medidor,
                _vuoir.longitude,
                _vuoir.latitude,
                _vuoir.intervencao_id
            FROM public.view_usuario_operador_id_rotulo _vuoir
            {_join_extra}
            WHERE {_where_base}
              {_where_busca}
        ),
        metricas_telemetria AS (
            -- Agrega métricas brutas por medidor dentro do período solicitado
            SELECT
                t.intervencao_id,
                MIN(t.data) FILTER (WHERE t.consumo_diario > 0) AS data_inicio,
                MAX(t.data) FILTER (WHERE t.consumo_diario > 0) AS data_fim,
                COUNT(DISTINCT t.data) FILTER (WHERE t.consumo_diario > 0) AS nu_dias,
                -- Dias com consumo dentro do intervalo solicitado pelo usuário
                COUNT(DISTINCT t.data) FILTER (
                    WHERE t.consumo_diario > 0
                      AND t.data >= %(mp_ini)s
                      AND t.data <= %(mp_fim)s
                ) AS nu_dias_periodo,
                -- Última data com dado no período
                MAX(t.data) FILTER (
                    WHERE t.consumo_diario > 0
                      AND t.data >= %(mp_ini)s
                      AND t.data <= %(mp_fim)s
                ) AS ultima_data_dado,
                -- Total de dias sem dado no período (calendário - dias com dado)
                (
                    (%(mp_fim)s::date - %(mp_ini)s::date + 1)
                    - COUNT(DISTINCT t.data) FILTER (
                        WHERE t.consumo_diario > 0
                          AND t.data >= %(mp_ini)s
                          AND t.data <= %(mp_fim)s
                    )
                ) AS nu_dias_sem_dados
            FROM public.tb_telemetria_intervencao_diaria t
            WHERE EXISTS (
                SELECT 1 FROM interferencias_base i
                WHERE i.intervencao_id = t.intervencao_id
            )
            GROUP BY t.intervencao_id
        ),
        -- Sequência máxima de dias consecutivos SEM dado por medidor no período
        dias_com_dado AS (
            -- Materializa os dias COM dado de cada medidor no período
            SELECT DISTINCT
                t.intervencao_id,
                t.data::date AS dia
            FROM public.tb_telemetria_intervencao_diaria t
            WHERE t.consumo_diario > 0
              AND t.data >= %(mp_ini)s
              AND t.data <= %(mp_fim)s
              AND EXISTS (
                  SELECT 1 FROM interferencias_base i
                  WHERE i.intervencao_id = t.intervencao_id
              )
        ),
        calendario_periodo AS (
            -- Série de todos os dias do período para cada medidor presente
            SELECT
                m.intervencao_id,
                g.dia::date AS dia
            FROM metricas_telemetria m
            CROSS JOIN LATERAL generate_series(
                %(mp_ini)s::date,
                %(mp_fim)s::date,
                INTERVAL '1 day'
            ) AS g(dia)
        ),
        dias_sem_dado AS (
            -- Dias do calendário que NÃO têm leitura com consumo
            SELECT
                c.intervencao_id,
                c.dia,
                -- Island: date - integer = date; dias consecutivos caem no mesmo grp
                c.dia - (ROW_NUMBER() OVER (
                    PARTITION BY c.intervencao_id
                    ORDER BY c.dia
                ))::int AS grp
            FROM calendario_periodo c
            LEFT JOIN dias_com_dado d
                ON d.intervencao_id = c.intervencao_id AND d.dia = c.dia
            WHERE d.dia IS NULL
        ),
        seq_max AS (
            -- Maior sequência consecutiva de dias sem dado por medidor
            SELECT
                intervencao_id,
                MAX(cnt) AS max_dias_seq_sem_dados
            FROM (
                SELECT intervencao_id, grp, COUNT(*) AS cnt
                FROM dias_sem_dado
                GROUP BY intervencao_id, grp
            ) sub
            GROUP BY intervencao_id
        ),
        dados_empreendimento AS (
            SELECT
                i.*,
                mt.data_inicio,
                mt.data_fim,
                mt.nu_dias,
                mt.nu_dias_periodo,
                mt.nu_dias_sem_dados,
                mt.ultima_data_dado,
                COALESCE(sq.max_dias_seq_sem_dados, 0) AS max_dias_seq_sem_dados,
                CASE
                    WHEN b.nome_empreendimento IS NOT NULL THEN b.nome_empreendimento
                    WHEN c.nome_empreendimento IS NOT NULL THEN c.nome_empreendimento
                    ELSE sfi.nome_empreendimento
                END AS nome_empreendimento,
                COALESCE(b.longitude, c.longitude) AS lon_emp,
                COALESCE(b.latitude,  c.latitude)  AS lat_emp,
                i.longitude AS lon_a,
                i.latitude  AS lat_a,
                b.nmautomonit AS nmautomonit_b,
                (b.codigo_interferencia IS NOT NULL) AS tem_mapserver,
                (c.codigo_interferencia IS NOT NULL) AS tem_captacao
            FROM interferencias_base i
            LEFT JOIN metricas_telemetria mt
                ON mt.intervencao_id = i.intervencao_id
            LEFT JOIN seq_max sq
                ON sq.intervencao_id = i.intervencao_id
            LEFT JOIN public.tb_mapserver_obrigatoriedade b
                ON b.codigo_interferencia = i.nu_interferencia_cnarh::integer
            LEFT JOIN public.tb_captacao_obrigatorios c
                ON c.codigo_interferencia = i.nu_interferencia_cnarh::integer
            LEFT JOIN public.tb_mv_sfi_cnarh40 sfi
                ON sfi.codigo_interferencia = i.nu_interferencia_cnarh::integer
               AND b.codigo_interferencia IS NULL
               AND c.codigo_interferencia IS NULL
        ),
        uam_priorizado AS (
            SELECT
                d.*,
                CASE
                    WHEN d.nmautomonit_b IS NOT NULL     THEN d.nmautomonit_b
                    WHEN uam_emp.nmautomonit IS NOT NULL THEN uam_emp.nmautomonit
                    WHEN uam_a.nmautomonit IS NOT NULL   THEN uam_a.nmautomonit
                    ELSE NULL
                END AS nome_uam
            FROM dados_empreendimento d
            LEFT JOIN LATERAL (
                SELECT nmautomonit
                FROM public.ft_uam_buffer
                WHERE d.lon_emp IS NOT NULL AND d.lat_emp IS NOT NULL
                  AND ST_Intersects(
                      ST_Transform(ST_SetSRID(
                          ST_MakePoint(d.lon_emp, d.lat_emp), 4326), 4674),
                      geom
                  )
                LIMIT 1
            ) uam_emp ON (d.tem_mapserver OR d.tem_captacao)
            LEFT JOIN LATERAL (
                SELECT nmautomonit
                FROM public.ft_uam_buffer
                WHERE (d.nmautomonit_b IS NULL AND uam_emp.nmautomonit IS NULL)
                  AND ST_Intersects(
                      ST_Transform(ST_SetSRID(
                          ST_MakePoint(d.lon_a, d.lat_a), 4326), 4674),
                      geom
                  )
                LIMIT 1
            ) uam_a ON TRUE
        )
        SELECT
            nu_cnarh,
            STRING_AGG(DISTINCT nu_interferencia_cnarh::text, ','
                       ORDER BY nu_interferencia_cnarh::text) AS codigo_interferencia,
            usuario                  AS nome_usuario,
            MAX(nome_empreendimento) AS nome_empreendimento,
            rotulo_intervencao_medidor AS rotulo_medidor,
            MAX(nome_uam)            AS nome_uam,
            MIN(data_inicio)         AS data_inicio_atividade,
            SUM(nu_dias)             AS total_dias,
            -- Dias com consumo dentro do intervalo solicitado
            SUM(nu_dias_periodo)     AS total_dias_periodo,
            -- Novas métricas de ausência de dado
            SUM(nu_dias_sem_dados)        AS total_dias_sem_dados,
            MAX(max_dias_seq_sem_dados)   AS max_seq_dias_sem_dados,
            MAX(ultima_data_dado)         AS ultima_data_dado
        FROM uam_priorizado
        GROUP BY nu_cnarh, usuario, rotulo_intervencao_medidor
        {_having_periodo}
        ORDER BY rotulo_intervencao_medidor
        """
 
        # ── Monta parâmetros como dicionário nomeado (psycopg2 pyformat) ───────────
        # Todos os placeholders da query usam %(nome)s; os parâmetros de busca
        # posicionais (%s) são convertidos para %(busca_0)s … %(busca_N)s.
        _mp_fim = data_fim if data_fim else data_inicio

        # Converte placeholders posicionais da busca para nomeados
        params_finais: dict = {}
        for i, val in enumerate(_params_busca):
            placeholder_novo = f"%(busca_{i})s"
            sql = sql.replace("%s", placeholder_novo, 1)
            params_finais[f"busca_{i}"] = val

        # Parâmetros de período — nomeados, usados N vezes na query sem problema
        params_finais["mp_ini"] = data_inicio
        params_finais["mp_fim"] = _mp_fim
        params_finais["p_ini"]  = data_inicio
        params_finais["p_fim"]  = data_fim  # None quando data única (HAVING ignora)
 
        # ── Executa ───────────────────────────────────────────────────────────────
        cursor = None
        linhas = []
        try:
            QApplication.setOverrideCursor(Qt.WaitCursor)
            cursor = self.conn.cursor()
            cursor.execute(sql, params_finais)
            linhas = cursor.fetchall()
        except Exception as e:
            QMessageBox.critical(self, "Erro na consulta",
                                 f"Erro ao consultar medidores por período:\n{e}")
            return
        finally:
            QApplication.restoreOverrideCursor()
            if cursor:
                cursor.close()
 
        if not linhas:
            d_ini_str = data_inicio.strftime('%d/%m/%Y')
            d_fim_str = data_fim.strftime('%d/%m/%Y') if data_fim else None
            msg_periodo = (
                f"de {d_ini_str} a {d_fim_str}" if d_fim_str
                else f"em {d_ini_str}"
            )
            QMessageBox.information(
                self, "Sem resultados",
                f"Nenhum medidor com atividade registrada {msg_periodo}."
            )
            return
 
        # ── Nome e caminho do arquivo ─────────────────────────────────────────────
        d_ini_str = data_inicio.strftime('%Y%m%d')
        d_fim_str = data_fim.strftime('%Y%m%d') if data_fim else None
        sufixo_periodo = (
            f"DE_{d_ini_str}_A_{d_fim_str}" if d_fim_str
            else f"EM_{d_ini_str}"
        )
        sufixo_busca = (
            f"_{_sanitizar(criterio)}_{_sanitizar(termo)}"
            if (criterio and termo)
            else ""
        )
        ts           = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"MEDIDORES_PERIODO_{sufixo_periodo}{sufixo_busca}_GER{ts}.xlsx"
        downloads = (
            os.path.join(os.environ.get('USERPROFILE', ''), 'Downloads')
            if sys.platform == 'win32'
            else os.path.join(os.path.expanduser('~'), 'Downloads')
        )
        caminho = os.path.join(downloads, nome_arquivo)
 
        if os.path.exists(caminho):
            resp = QMessageBox.question(
                self, "Arquivo Existente",
                f"'{nome_arquivo}' já existe.\nDeseja substituir?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if resp == QMessageBox.No:
                return
 
        # ── Título do Excel ───────────────────────────────────────────────────────
        d_ini_lbl = data_inicio.strftime('%d/%m/%Y')
        d_fim_lbl = data_fim.strftime('%d/%m/%Y') if data_fim else None
        periodo_lbl = (
            f"de {d_ini_lbl} a {d_fim_lbl}" if d_fim_lbl
            else f"em {d_ini_lbl}"
        )
        if criterio and termo:
            titulo_xlsx = (
                f"Medidores Ativos {periodo_lbl}"
                f" – {criterio}: {termo}"
            )
        else:
            titulo_xlsx = f"Medidores Ativos {periodo_lbl}"
 
        # ── Estilos institucionais ────────────────────────────────────────────────
        fill_roxo   = PatternFill("solid", fgColor="6f4e8a")   # cabeçalho período
        fill_alt    = PatternFill("solid", fgColor="f3eeff")   # linha alternada
        font_branca = Font(bold=True, size=10, color="ffffff")
        font_titulo = Font(bold=True, size=13, color="6f4e8a")
        font_sub    = Font(size=10, italic=True, color="495057")
        font_normal = Font(size=10)
        ali_centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ali_esq     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        borda       = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
 
        # ── Workbook / workaround lxml ────────────────────────────────────────────
        import sys as _sys
        _lxml_mod   = _sys.modules.pop('lxml',        None)
        _lxml_etree = _sys.modules.pop('lxml.etree',  None)
        try:
            wb = Workbook()
        finally:
            if _lxml_mod   is not None: _sys.modules['lxml']       = _lxml_mod
            if _lxml_etree is not None: _sys.modules['lxml.etree'] = _lxml_etree
 
        ws = wb.active
        ws.title = "Medidores por Período"
 
        colunas  = [
            "Nº CNARH",
            "Interferência(s) CNARH",
            "Usuário / Outorgado",
            "Nome do Empreendimento",
            "Rótulo do Medidor",
            "UAM",
            "Início Atividade",
            "Total Dias c/ Dados",
        ]
        # Coluna de dias no período apenas quando há intervalo de datas
        if data_fim:
            colunas.append("Total Dias c/ Dados no Período")
        # Novas colunas — sempre presentes
        colunas += [
            "Total Dias s/ Dados no Período",
            "Dias Sequenciais s/ Dados",
            "Última Data c/ Dado",
        ]
        larguras = [18, 26, 36, 36, 26, 28, 16, 18]
        if data_fim:
            larguras.append(26)
        larguras += [26, 22, 18]
 
        n_cols   = len(colunas)
        col_last = get_column_letter(n_cols)
 
        # Linha 1 – título
        ws.merge_cells(f"A1:{col_last}1")
        ws["A1"]           = titulo_xlsx
        ws["A1"].font      = font_titulo
        ws["A1"].alignment = ali_centro
        ws.row_dimensions[1].height = 22
 
        # Linha 2 – subtítulo
        ws.merge_cells(f"A2:{col_last}2")
        # Conta conjuntos de interferências distintos ("x,y" conta como 1 conjunto)
        nu_interf_distintas = len({r[1] for r in linhas if r[1]})
        ws["A2"] = (
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}"
            f"  –  Total: {len(linhas)} medidor(es)  |  {nu_interf_distintas} interferência(s)"
        )
        ws["A2"].font      = font_sub
        ws["A2"].alignment = ali_centro
        ws.row_dimensions[2].height = 18
 
        # Linha 3 – cabeçalho (roxo institucional para distinguir do relatório geral)
        for ci, cab in enumerate(colunas, 1):
            cell           = ws.cell(row=3, column=ci, value=cab)
            cell.font      = font_branca
            cell.fill      = fill_roxo
            cell.alignment = ali_centro
            cell.border    = borda
        ws.row_dimensions[3].height = 20
 
        # Linhas de dados
        # SELECT: 0=nu_cnarh 1=cod_interf 2=nome_usu 3=nome_emp 4=rotulo
        #   5=nome_uam 6=dt_inicio 7=total_dias 8=total_dias_periodo
        #   9=total_dias_sem_dados 10=max_seq_dias_sem_dados 11=ultima_data_dado
        for idx, row_data in enumerate(linhas, start=4):
            (nu_cnarh, cod_interf, nome_usu, nome_emp,
             rotulo, nome_uam, dt_inicio, total_dias,
             total_dias_periodo, total_dias_sem_dados,
             max_seq_dias_sem_dados, ultima_data_dado) = row_data[:12]

            dt_inicio_str = (
                dt_inicio.strftime('%d/%m/%Y')
                if dt_inicio else "—"
            )
            ultima_data_str = (
                ultima_data_dado.strftime('%d/%m/%Y')
                if ultima_data_dado else "—"
            )

            valores = [
                nu_cnarh    or "—",
                cod_interf  or "—",
                nome_usu    or "—",
                nome_emp    or "—",
                rotulo      or "—",
                nome_uam    or "—",
                dt_inicio_str,
                total_dias         if total_dias         is not None else "—",
            ]
            if data_fim:
                valores.append(
                    total_dias_periodo if total_dias_periodo is not None else "—"
                )
            # Novas colunas — sempre adicionadas ao final
            valores += [
                total_dias_sem_dados   if total_dias_sem_dados   is not None else "—",
                max_seq_dias_sem_dados if max_seq_dias_sem_dados is not None else "—",
                ultima_data_str,
            ]

            fill_linha = fill_alt if (idx % 2 == 0) else None

            # Colunas centralizadas: CNARH(1), Início(7), Total Dias c/ Dados(8),
            # + Período(9 se data_fim), + 3 novas (sempre as 3 últimas)
            n_val = len(valores)
            cols_centro = {1, 7, 8, n_val - 2, n_val - 1, n_val}
            if data_fim:
                cols_centro.add(9)

            for ci, val in enumerate(valores, 1):
                cell        = ws.cell(row=idx, column=ci, value=val)
                cell.font   = font_normal
                cell.border = borda
                if fill_linha:
                    cell.fill = fill_linha
                cell.alignment = ali_centro if ci in cols_centro else ali_esq

        # Larguras e freeze
        for ci, larg in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(ci)].width = larg
        ws.freeze_panes = "A4"
 
        # Rodapé
        rodape_row = len(linhas) + 5
        ws.merge_cells(f"A{rodape_row}:{col_last}{rodape_row}")
        cell_rod           = ws.cell(row=rodape_row, column=1,
                                     value="Sistema DURH Diária por Telemetria (SFI/ANA) – Relatório gerado automaticamente")
        cell_rod.font      = Font(size=8, italic=True, color="888888")
        cell_rod.alignment = ali_centro
 
        # ── Salvar ────────────────────────────────────────────────────────────────
        try:
            wb.save(caminho)
        except Exception as e:
            QMessageBox.critical(self, "Erro ao salvar",
                                 f"Não foi possível salvar o arquivo:\n{e}")
            return
 
        QMessageBox.information(
            self, "Exportação concluída",
            f"Arquivo exportado com sucesso!\n\n{caminho}",
            QMessageBox.Ok
        )