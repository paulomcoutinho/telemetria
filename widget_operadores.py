# -*- coding: utf-8 -*-
"""
Módulo: widget_operadores.py
==============================
Aba de gestão de operadores de telemetria (JanelaGestaoDados – aba 1).

Funcionalidades:
  - Busca por nome via QComboBox editável;
  - Edição inline de Nome, CPF/CNPJ e E-mail com UPDATE direto;
  - Exclusão com confirmação (DELETE);
  - Navegação para a aba de medidores;
  - Exportação XLSX com formatação institucional (openpyxl):
      todos os operadores OU somente os com transmissão ativa.

Autor : SFI/ANA
Versão: 2.0 – Março/2026
"""

# ---------------------------------------------------------------------------
# Imports Qt – widgets, layout e utilitários de UI
# ---------------------------------------------------------------------------
from qgis.PyQt.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QComboBox,
    QFrame, QMessageBox, QSizePolicy,
    QGridLayout, QDialog, QCheckBox,
)
from qgis.PyQt.QtCore import Qt, QTimer
from qgis.PyQt.QtGui import QRegExpValidator

# ---------------------------------------------------------------------------
# Imports Python padrão
# ---------------------------------------------------------------------------
import os
import psycopg2
from datetime import datetime
import re

# ---------------------------------------------------------------------------
# Exportação Excel (opcional)
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False

# ---------------------------------------------------------------------------
# Sistema de design centralizado
# ---------------------------------------------------------------------------
try:
    from . import ui_tema
except ImportError:
    import ui_tema


class WidgetOperadores(QWidget):
    """Aba de gestão de operadores de telemetria dentro da JanelaGestaoDados.

    Permite ao usuário autenticado consultar, editar e excluir registros de
    operadores na tabela ``tb_operador_telemetria``, além de exportar a lista
    completa (ou apenas os com transmissão ativa) para um arquivo XLSX
    formatado na pasta de Downloads do sistema operacional.

    O fluxo de uso típico é:
        1. O usuário digita ou seleciona um operador no ``QComboBox`` editável;
        2. Ao confirmar a seleção, os campos ID, Nome, CPF/CNPJ, E-mail e
           Data de Cadastro são preenchidos automaticamente;
        3. O usuário edita os campos desejados e clica em "Salvar alterações"
           para persistir o UPDATE na base;
        4. Alternativamente, pode acionar "Apagar" para remover o registro
           após confirmação, ou "Exportar XLSX" para gerar o relatório.

    O botão "Ir para Medidores" ativa a aba de medidores na janela-pai
    (``JanelaGestaoDados``) sem recriar a instância, navegando diretamente
    para o índice correto do ``QTabWidget``.

    Attributes:
        conn (psycopg2.connection): Conexão ativa ao PostgreSQL em autocommit.
        usuario_logado (str | None): Usuário da sessão; utilizado para
            registros de auditoria ou restrições futuras.
        parent_window (JanelaGestaoDados | None): Referência à janela-pai,
            usada para troca programática de abas.
        operador_atual (int | None): ID do operador atualmente selecionado
            no combo de busca; ``None`` quando nenhum está carregado.
        combo_operadores (QComboBox): Campo editável e pesquisável para
            seleção de operadores pelo nome.
        input_id (QLineEdit): Exibe o ID do registro (somente leitura).
        input_nome (QLineEdit): Nome do operador, editável.
        input_numero_cadastro (QLineEdit): CPF, CNPJ ou código CNARH,
            limitado a 15 caracteres.
        input_email (QLineEdit): E-mail do operador, editável.
        input_data (QLineEdit): Data de cadastro (somente leitura).
        btn_salvar (QPushButton): Habilitado somente quando um operador
            está carregado; persiste as alterações via UPDATE.
        btn_apagar (QPushButton): Remove o registro após confirmação do
            usuário.
    """
    
    def __init__(self, conexao, usuario=None, parent_window=None):
        super().__init__()
        self.conn = conexao
        self.usuario_logado = usuario
        self.parent_window = parent_window # Referência para a janela principal para trocar de aba
        try:
            self.conn.autocommit = True
        except:
            pass        
        self.operador_atual = None
        self.initUI()
        
    def initUI(self):
        """Configura a interface do widget de operadores."""
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(5, 5, 5, 5)
        
        # Criar um widget container para todo o conteúdo
        content_widget = QWidget()
        content_widget.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 5px;
            }
        """)
        layout = QVBoxLayout(content_widget)
        
        
        # Busca compacta
        busca_widget = QWidget()
        busca_widget.setStyleSheet("border: none;")
        busca_layout = QHBoxLayout(busca_widget)
        busca_layout.setContentsMargins(0, 5, 0, 5)
        busca_layout.setSpacing(5)
        
        label_busca = QLabel("Buscar:")
        label_busca.setFixedWidth(50)
        label_busca.setStyleSheet("border: none;")
        
        self.combo_operadores = QComboBox()
        self.combo_operadores.setEditable(True)
        self.combo_operadores.setFixedHeight(30)
        self.combo_operadores.setStyleSheet("border: 1px solid #ccc;")
        
        btn_carregar = QPushButton("↻")
        btn_carregar.setToolTip("Recarregar lista")
        btn_carregar.setFixedSize(30, 30)
        btn_carregar.setStyleSheet("""
            QPushButton {
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """)
        btn_carregar.clicked.connect(self.carregar_lista_operadores)
        
        busca_layout.addWidget(label_busca)
        busca_layout.addWidget(self.combo_operadores)
        busca_layout.addWidget(btn_carregar)
        layout.addWidget(busca_widget)
        
        # Campos de dados em grid compacto
        dados_widget = QWidget()
        dados_widget.setStyleSheet("border: none;")
        dados_layout = QGridLayout(dados_widget)
        dados_layout.setSpacing(6)
        dados_layout.setContentsMargins(0, 5, 0, 5)
        
        # Labels com estilo consistente
        label_style = "border: none;"
        input_style = "border: 1px solid #ccc; padding: 2px;"
        
        # Linha 1
        lbl_id = QLabel("ID:")
        lbl_id.setStyleSheet(label_style)
        dados_layout.addWidget(lbl_id, 0, 0)
        self.input_id = QLineEdit()
        self.input_id.setReadOnly(True)
        self.input_id.setFixedHeight(30)
        self.input_id.setStyleSheet(input_style + "background-color: #f5f5f5;")
        dados_layout.addWidget(self.input_id, 0, 1)
        
        lbl_nome = QLabel("Nome:")
        lbl_nome.setStyleSheet(label_style)
        dados_layout.addWidget(lbl_nome, 0, 2)
        self.input_nome = QLineEdit()
        self.input_nome.setFixedHeight(30)
        self.input_nome.setStyleSheet(input_style)
        dados_layout.addWidget(self.input_nome, 0, 3)
        
        # Linha 2
        lbl_cpf = QLabel("CPF/CNPJ:")
        lbl_cpf.setStyleSheet(label_style)
        dados_layout.addWidget(lbl_cpf, 1, 0)
        self.input_numero_cadastro = QLineEdit()
        self.input_numero_cadastro.setMaxLength(15)
        self.input_numero_cadastro.setFixedHeight(30)
        self.input_numero_cadastro.setStyleSheet(input_style)
        dados_layout.addWidget(self.input_numero_cadastro, 1, 1)
        
        lbl_email = QLabel("E-mail:")
        lbl_email.setStyleSheet(label_style)
        dados_layout.addWidget(lbl_email, 1, 2)
        self.input_email = QLineEdit()
        self.input_email.setFixedHeight(30)
        self.input_email.setStyleSheet(input_style)
        dados_layout.addWidget(self.input_email, 1, 3)
        
        # Linha 3
        lbl_data = QLabel("Data:")
        lbl_data.setStyleSheet(label_style)
        dados_layout.addWidget(lbl_data, 2, 0)
        self.input_data = QLineEdit()
        self.input_data.setReadOnly(True)
        self.input_data.setFixedHeight(30)
        self.input_data.setStyleSheet(input_style + "background-color: #f5f5f5;")
        dados_layout.addWidget(self.input_data, 2, 1, 1, 3)  # Ocupa 3 colunas
        
        layout.addWidget(dados_widget)
        
        # Botões compactos
        botoes_widget = QWidget()
        botoes_widget.setStyleSheet("border: none;")
        botoes_layout = QHBoxLayout(botoes_widget)
        botoes_layout.setContentsMargins(0, 5, 0, 0)
        botoes_layout.setSpacing(10)
        
        self.btn_salvar = QPushButton("Salvar alterações")
        self.btn_salvar.setFixedHeight(35)
        self.btn_salvar.setStyleSheet("""
            QPushButton {
                background-color: #175cc3;
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
            QPushButton:hover {
                background-color: #5474b8;
            }
        """)
        self.btn_salvar.clicked.connect(self.salvar_alteracoes)
        self.btn_salvar.setEnabled(False)

        # NOVO: Botão Apagar
        self.btn_apagar = QPushButton("Apagar")
        self.btn_apagar.setFixedHeight(35)
        self.btn_apagar.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.btn_apagar.clicked.connect(self.apagar_operador)
        self.btn_apagar.setEnabled(False)

        # NOVO: Botão Ver Medidores Vinculados
        self.btn_ver_medidores = QPushButton("Ver Medidores Vinculados")
        self.btn_ver_medidores.setFixedHeight(35)
        # Estilo idêntico ao botão buscar da aba Medidores (#5474b8)
        self.btn_ver_medidores.setStyleSheet("""
            QPushButton {
                background-color: #5474b8;
                color: white;
                font-size: 11px;
                font-weight: bold;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #2050b8;
            }
        """)
        self.btn_ver_medidores.setVisible(False) # Invisível até ocorrer a busca
        self.btn_ver_medidores.clicked.connect(self.ir_para_medidores_vinculados)

        self.btn_limpar = QPushButton("Limpar")
        self.btn_limpar.setFixedHeight(35)
        self.btn_limpar.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                font-size: 12px;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.btn_limpar.clicked.connect(self.limpar_campos)

        self.btn_exportar_operadores = QPushButton("Exportar operadores cadastrados")
        self.btn_exportar_operadores.setFixedHeight(35)
        self.btn_exportar_operadores.setToolTip("Exportar lista completa de operadores cadastrados para Excel")
        self.btn_exportar_operadores.setStyleSheet("""
            QPushButton {
                background-color: #1d7a3a;
                color: white;
                font-size: 11px;
                font-weight: bold;
                padding: 8px 12px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                background-color: #155c2b;
            }
            QPushButton:pressed {
                background-color: #0f4320;
            }
        """)
        self.btn_exportar_operadores.clicked.connect(self.exportar_operadores_excel)
                
        botoes_layout.addWidget(self.btn_salvar)
        botoes_layout.addWidget(self.btn_apagar)
        botoes_layout.addWidget(self.btn_ver_medidores) # Adicionado na ordem
        botoes_layout.addWidget(self.btn_limpar)
        botoes_layout.addWidget(self.btn_exportar_operadores)
        botoes_layout.addStretch()
        
        layout.addWidget(botoes_widget)
        
        # Margens do conteúdo
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)
        
        # Adicionar o container ao layout principal
        main_layout.addWidget(content_widget)
        main_layout.setContentsMargins(8, 8, 8, 8)
        
        # Conectar sinais
        self.combo_operadores.currentIndexChanged.connect(self.carregar_dados_operador)
        self.input_nome.textChanged.connect(self.habilitar_salvar)
        self.input_email.textChanged.connect(self.habilitar_salvar)
        self.input_numero_cadastro.textChanged.connect(self.habilitar_salvar)
        
        # Carregar lista inicial
        self.carregar_lista_operadores()
        
        # Política de tamanho
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.setLayout(main_layout)

        # --- ADICIONADO: Lógica para usuário somente leitura ---
        if self.usuario_logado == "telemetria_ro":
            # Define o estilo visual de campos de leitura
            ro_style = "border: 1px solid #ccc; padding: 2px; background-color: #f5f5f5;"
            
            # Bloqueia campos editáveis e aplica o estilo
            self.input_nome.setReadOnly(True)
            self.input_nome.setStyleSheet(ro_style)
            
            self.input_numero_cadastro.setReadOnly(True)
            self.input_numero_cadastro.setStyleSheet(ro_style)
            
            self.input_email.setReadOnly(True)
            self.input_email.setStyleSheet(ro_style)
            
            self.combo_operadores.setEditable(False) # Permite seleção mas não digitação
            
            # Esconde botões Salvar e Apagar
            self.btn_salvar.hide()
            self.btn_apagar.hide()
            
    def ir_para_medidores_vinculados(self):
        """Muda para a aba de medidores e busca exclusivamente pelo ID do operador."""
        if not self.operador_atual or not self.parent_window:
            return
            
        # Obtém o ID do operador atual
        id_operador = self.operador_atual[0]
        
        # Acessa o widget de medidores através da janela principal
        widget_medidores = self.parent_window.widget_medidores
        
        # Limpa a UI de busca anterior para não confundir o usuário
        widget_medidores.input_busca.clear()
        
        # Chama o método específico que utiliza o ID para busca precisa
        widget_medidores.buscar_medidores_por_operador_id(id_operador)
        
        # Muda para a aba de medidores (índice 1)
        self.parent_window.tabs.setCurrentIndex(1)

    def carregar_lista_operadores(self):
        """Carrega a lista de operadores no combobox."""
        try:
            cursor = self.conn.cursor()
            query = "SELECT id, nome, email FROM tb_operador_telemetria ORDER BY nome ASC;"
            cursor.execute(query)
            operadores = cursor.fetchall()
            
            self.combo_operadores.clear()
            self.combo_operadores.addItem("-- Digite ou selecione um operador --", None)
            
            for id_operador, nome, email in operadores:
                self.combo_operadores.addItem(f"{nome} ({email})", id_operador)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar operadores: {e}")
        finally:
            cursor.close()
            
    def carregar_dados_operador(self, index):
        """Carrega os dados do operador selecionado."""
        if index <= 0:
            self.limpar_campos()
            return
            
        id_operador = self.combo_operadores.itemData(index)
        if not id_operador:
            return
            
        try:
            cursor = self.conn.cursor()
            query = """
            SELECT id, nome, numero_cadastro, email, data 
            FROM tb_operador_telemetria 
            WHERE id = %s;
            """
            cursor.execute(query, (id_operador,))
            operador = cursor.fetchone()
            
            if operador:
                self.operador_atual = operador
                self.input_id.setText(str(operador[0]))
                self.input_nome.setText(operador[1] if operador[1] else "")
                self.input_numero_cadastro.setText(operador[2] if operador[2] else "")
                self.input_email.setText(operador[3] if operador[3] else "")
                self.input_data.setText(str(operador[4]) if operador[4] else "")
                
                # Habilitar botão apagar quando houver operador carregado
                if self.usuario_logado != "telemetria_ro":
                    self.btn_apagar.setEnabled(True)
                
                # O botão de ver medidores é visível para todos.
                self.btn_ver_medidores.setVisible(True)
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar dados do operador: {e}")
        finally:
            cursor.close()
            
    def habilitar_salvar(self):
        """Habilita o botão de salvar quando houver alterações."""
        # Se for usuário RO, nunca habilita o salvar (embora o botão esteja oculto)
        if self.usuario_logado == "telemetria_ro":
            return

        if self.operador_atual:
            nome_original = self.operador_atual[1] or ""
            email_original = self.operador_atual[3] or ""
            cadastro_original = self.operador_atual[2] or ""
            
            nome_atual = self.input_nome.text()
            email_atual = self.input_email.text()
            cadastro_atual = self.input_numero_cadastro.text()
            
            alterado = (nome_original != nome_atual or 
                       email_original != email_atual or
                       cadastro_original != cadastro_atual)
            
            self.btn_salvar.setEnabled(alterado)
    
    def apagar_operador(self):
        """Apaga o operador selecionado após validações."""
        if not self.operador_atual:
            QMessageBox.warning(self, "Aviso", "Nenhum operador selecionado para apagar.")
            return
        
        id_operador = self.operador_atual[0]
        nome_operador = self.operador_atual[1] or "Sem nome"
        
        try:
            cursor = self.conn.cursor()
            
            # Verificar se há medidores vinculados com rótulos que NÃO sejam LIKE '999%'
            query_verificacao = """
            SELECT rotulo 
            FROM tb_intervencao 
            WHERE operador_telemetria = %s 
            AND rotulo NOT LIKE '999%%'
            ORDER BY rotulo;
            """
            cursor.execute(query_verificacao, (id_operador,))
            rotulos_vinculados = cursor.fetchall()
            
            if rotulos_vinculados:
                # Existem medidores ativos vinculados
                lista_rotulos = [rotulo[0] for rotulo in rotulos_vinculados]
                rotulos_str = ", ".join(lista_rotulos)
                
                mensagem = (
                    f"Não é possível apagar o operador '{nome_operador}'.\n\n"
                    f"Existem {len(lista_rotulos)} medidor(es) ativo(s) vinculado(s) "
                    f"a este operador:\n\n{rotulos_str}\n\n"
                    f"Para prosseguir com a exclusão, acesse a aba 'Medidores cadastrados' para:\n"
                    f"1. Alterar o operador responsável desses medidores, OU\n"
                    f"2. Desativar esses medidores vinculados\n\n"
                    f"Após realizar uma dessas ações, retorne a esta aba para apagar o operador."
                )
                
                QMessageBox.warning(self, "Operador possui medidores vinculados", mensagem)
                return
            
            # Se chegou aqui, não há impedimentos
            # Confirmação final
            resposta = QMessageBox.question(
                self,
                "Confirmar exclusão",
                f"Tem certeza que deseja apagar o operador '{nome_operador}'?\n\n"
                f"Esta ação não poderá ser desfeita.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if resposta != QMessageBox.Yes:
                return
            
            # Executar a exclusão em duas etapas:
            # 1. Primeiro, setar como NULL o operador_telemetria na tb_intervencao
            query_update = """
            UPDATE tb_intervencao 
            SET operador_telemetria = NULL,
                rotulo = rotulo || '#' 
            WHERE operador_telemetria = %s;
            """
            cursor.execute(query_update, (id_operador,))
            
            # 2. Depois, apagar o operador da tb_operador_telemetria
            query_delete = "DELETE FROM tb_operador_telemetria WHERE id = %s;"
            cursor.execute(query_delete, (id_operador,))
            
            self.conn.commit()
            
            QMessageBox.information(
                self, 
                "Sucesso", 
                f"Operador '{nome_operador}' foi apagado com sucesso!"
            )
            
            # Limpar campos e recarregar lista
            self.limpar_campos()
            self.carregar_lista_operadores()
            
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Erro", f"Erro ao apagar operador: {e}")
        finally:
            cursor.close()
            
    def salvar_alteracoes(self):
        """Salva as alterações nos dados do operador."""
        if not self.operador_atual:
            return
            
        id_operador = self.operador_atual[0]
        nome_original = self.operador_atual[1] or ""
        email_original = self.operador_atual[3] or ""
        cadastro_original = self.operador_atual[2] or ""
        
        nome_novo = self.input_nome.text().strip()
        email_novo = self.input_email.text().strip()
        cadastro_novo = self.input_numero_cadastro.text().strip()
        
        # Validar e-mail
        if email_novo and not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email_novo):
            QMessageBox.warning(self, "E-mail inválido", "O formato do e-mail é inválido.")
            return
            
        # Preparar mensagem de confirmação
        mensagem = "Confirme as alterações:\n\n"
        if nome_original != nome_novo:
            mensagem += f"Nome: {nome_original} → {nome_novo}\n"
        if email_original != email_novo:
            mensagem += f"E-mail: {email_original} → {email_novo}\n"
        if cadastro_original != cadastro_novo:
            mensagem += f"CPF/CNPJ/CNARH: {cadastro_original} → {cadastro_novo}\n"
            
        resposta = QMessageBox.question(
            self, "Confirmar alterações", mensagem,
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if resposta != QMessageBox.Yes:
            return
            
        # Salvar no banco
        try:
            cursor = self.conn.cursor()
            query = """
            UPDATE tb_operador_telemetria 
            SET nome = %s, email = %s, numero_cadastro = %s
            WHERE id = %s;
            """
            cursor.execute(query, (nome_novo, email_novo, cadastro_novo, id_operador))
            self.conn.commit()
            
            QMessageBox.information(self, "Sucesso", "Dados atualizados com sucesso!")
            
            # Atualizar dados locais
            self.operador_atual = (id_operador, nome_novo, cadastro_novo, email_novo, self.operador_atual[4])
            self.habilitar_salvar()
            
            # Atualizar combobox
            self.carregar_lista_operadores()
            
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Erro", f"Erro ao salvar alterações: {e}")
        finally:
            cursor.close()
            
    def limpar_campos(self):
        """Limpa todos os campos do formulário."""
        self.operador_atual = None
        self.input_id.clear()
        self.input_nome.clear()
        self.input_numero_cadastro.clear()
        self.input_email.clear()
        self.input_data.clear()
        self.btn_salvar.setEnabled(False)
        self.btn_apagar.setEnabled(False)
        self.btn_ver_medidores.setVisible(False) # Esconde o botão ao limpar
        self.combo_operadores.setCurrentIndex(0)

    def exportar_operadores_excel(self):
        """
        Ponto de entrada: exibe diálogo de escolha de escopo e delega à geração
        do Excel.
        Opções (checkboxes independentes — cada uma marcada gera um arquivo):
          • Todos cadastrados
          • Com transmissão ativa
        """
        # ── Diálogo ───────────────────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle("Exportar Operadores — Excel")
        dlg.setMinimumWidth(420)
        dlg.setStyleSheet("background-color: white;")

        layout = QVBoxLayout(dlg)
        layout.setSpacing(10)
        layout.setContentsMargins(20, 18, 20, 16)

        lbl_titulo = QLabel("Selecione o(s) escopo(s) da exportação:")
        lbl_titulo.setStyleSheet("font-weight: bold; font-size: 12px; color: #2c3e50;")
        layout.addWidget(lbl_titulo)

        # ── Opção 1: Todos cadastrados ────────────────────────────────────────────
        chk_todos = QCheckBox("Todos cadastrados")
        chk_todos.setStyleSheet(
            "QCheckBox { font-weight: bold; font-size: 11px; color: #175cc3; spacing: 6px; }"
            "QCheckBox::indicator { width: 15px; height: 15px;"
            "  border: 1px solid #175cc3; border-radius: 3px; background: white; }"
            "QCheckBox::indicator:checked { background-color: #175cc3; border-color: #175cc3; }"
        )
        lbl_todos = QLabel("Lista completa de operadores cadastrados no sistema.")
        lbl_todos.setWordWrap(True)
        lbl_todos.setStyleSheet("font-size: 11px; color: #495057; margin-left: 22px;")
        layout.addWidget(chk_todos)
        layout.addWidget(lbl_todos)

        # ── Opção 2: Com transmissão ativa ────────────────────────────────────────
        chk_ativos = QCheckBox("Com transmissão ativa")
        chk_ativos.setStyleSheet(
            "QCheckBox { font-weight: bold; font-size: 11px; color: #1d7a3a; spacing: 6px; }"
            "QCheckBox::indicator { width: 15px; height: 15px;"
            "  border: 1px solid #1d7a3a; border-radius: 3px; background: white; }"
            "QCheckBox::indicator:checked { background-color: #1d7a3a; border-color: #1d7a3a; }"
        )
        lbl_ativos = QLabel(
            "Apenas operadores que já transmitiram dados de medidores reais."
        )
        lbl_ativos.setWordWrap(True)
        lbl_ativos.setStyleSheet("font-size: 11px; color: #495057; margin-left: 22px;")
        layout.addWidget(chk_ativos)
        layout.addWidget(lbl_ativos)

        # ── Separador + botões ────────────────────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #dee2e6; margin: 4px 0;")
        layout.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        btn_gerar = QPushButton("Gerar Excel")
        btn_gerar.setFixedHeight(34)
        btn_gerar.setEnabled(False)
        btn_gerar.setStyleSheet(
            "QPushButton { background-color: #175cc3; color: white;"
            "  font-weight: bold; font-size: 11px; border-radius: 4px; padding: 6px 18px; }"
            "QPushButton:hover { background-color: #1249a3; }"
            "QPushButton:disabled { background-color: #adb5bd; color: #e9ecef; }"
        )

        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setFixedHeight(34)
        btn_cancelar.setStyleSheet(
            "QPushButton { background-color: #6c757d; color: white;"
            "  font-size: 11px; border-radius: 4px; padding: 6px 14px; }"
            "QPushButton:hover { background-color: #5a6268; }"
        )

        # habilita/desabilita "Gerar Excel" conforme checkboxes
        def _atualizar_btn(*_):
            btn_gerar.setEnabled(chk_todos.isChecked() or chk_ativos.isChecked())

        chk_todos.toggled.connect(_atualizar_btn)
        chk_ativos.toggled.connect(_atualizar_btn)

        btn_gerar.clicked.connect(dlg.accept)
        btn_cancelar.clicked.connect(dlg.reject)

        btn_row.addWidget(btn_gerar)
        btn_row.addStretch()
        btn_row.addWidget(btn_cancelar)
        layout.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        # ── Despacho: gera um arquivo por opção marcada ───────────────────────────
        if chk_todos.isChecked():
            self._gerar_excel_operadores(somente_ativos=False)

        if chk_ativos.isChecked():
            self._gerar_excel_operadores(somente_ativos=True)

    def _gerar_excel_operadores(self, somente_ativos=False):
        """
        Gera o arquivo Excel de operadores.
        somente_ativos=False  → todos os cadastrados.
        somente_ativos=True   → apenas operadores com consumo_diario > 0
                                 em medidores válidos (sem 999, sem VERDE GRANDE,
                                 sem operador 162, com coordenadas).
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self, "Biblioteca ausente",
                "A biblioteca 'openpyxl' não está instalada.\n"
                "Instale-a via: pip install openpyxl"
            )
            return

        import os, sys
        from datetime import datetime

        # ── Consultar banco ───────────────────────────────────────────────────────
        try:
            cursor = self.conn.cursor()

            if not somente_ativos:
                cursor.execute("""
                    SELECT row_number() OVER (ORDER BY nome)::integer AS id
                         , nome
                         , email
                         , numero_cadastro
                         , CASE WHEN numero_cadastro IS NULL
                                THEN 'Outorgado'::text
                                ELSE 'Não outorgado'
                           END AS tipo_operador
                         , data AS data_cadastro
                    FROM public.tb_operador_telemetria
                """)
            else:
                cursor.execute("""
                    SELECT row_number() OVER (ORDER BY ot.nome)::integer AS id
                         , ot.nome
                         , ot.email
                         , ot.numero_cadastro
                         , CASE WHEN ot.numero_cadastro IS NULL
                                THEN 'Outorgado'::text
                                ELSE 'Não outorgado'
                           END AS tipo_operador
                         , ot.data AS data_cadastro

                         -- Rótulos dos medidores de teste (contém '999')
                         , (
                               SELECT STRING_AGG(DISTINCT i2.rotulo::text, ', '
                                                 ORDER BY i2.rotulo::text)
                               FROM public.tb_intervencao i2
                               --JOIN public.tb_telemetria_intervencao_diaria ttid2
                                 --ON ttid2.intervencao_id = i2.id
                               WHERE i2.operador_telemetria = ot.id
                                 AND i2.rotulo::text LIKE '%%999%%'
                                 --AND ttid2.consumo_diario > 0
                           ) AS ds_medidores_teste

                         -- Rótulos dos medidores reais
                         , (
                               SELECT STRING_AGG(DISTINCT i2.rotulo::text, ', '
                                                 ORDER BY i2.rotulo::text)
                               FROM public.tb_intervencao i2
                               JOIN public.tb_telemetria_intervencao_diaria ttid2
                                 ON ttid2.intervencao_id = i2.id
                               WHERE i2.operador_telemetria = ot.id
                                 AND i2.rotulo::text NOT LIKE '%%999%%'
                                 AND i2.rotulo::text NOT LIKE 'VERDE GRANDE%%'
                                 AND i2.rotulo::text NOT LIKE '%%#'
                                 AND ttid2.consumo_diario > 0
                           ) AS ds_medidores_reais

                         -- CNARHs (numero_cadastro de tb_interferencia) distintos dos medidores reais
                         , (
                               SELECT STRING_AGG(DISTINCT inf2.numero_cadastro::text, ', '
                                                 ORDER BY inf2.numero_cadastro::text)
                               FROM public.tb_intervencao i2
                               JOIN public.tb_telemetria_intervencao_diaria ttid2
                                 ON ttid2.intervencao_id = i2.id
                               JOIN public.tb_intervencao_interferencia ii2
                                 ON ii2.intervencao_id = i2.id
                               JOIN public.tb_interferencia inf2
                                 ON inf2.id = ii2.interferencia_id
                               WHERE i2.operador_telemetria = ot.id
                                 AND i2.rotulo::text NOT LIKE '%%999%%'
                                 AND i2.rotulo::text NOT LIKE 'VERDE GRANDE%%'
                                 AND i2.rotulo::text NOT LIKE '%%#'
                                 AND inf2.numero_cadastro NOT LIKE 'TESTE'
                                 AND ttid2.consumo_diario > 0
                           ) AS cnarhs_reais

                         -- Última data com dado transmitido (medidores reais)
                         , (
                               SELECT MAX(ttid2.data)
                               FROM public.tb_intervencao i2
                               JOIN public.tb_telemetria_intervencao_diaria ttid2
                                 ON ttid2.intervencao_id = i2.id
                               WHERE i2.operador_telemetria = ot.id
                                 AND i2.rotulo::text NOT LIKE '%%999%%'
                                 AND i2.rotulo::text NOT LIKE 'VERDE GRANDE%%'
                                 AND i2.rotulo::text NOT LIKE '%%#'
                                 AND ttid2.consumo_diario > 0
                           ) AS ultima_data

                    FROM public.tb_operador_telemetria ot
                    WHERE ot.id IN (
                        SELECT DISTINCT i.operador_telemetria
                        FROM public.tb_intervencao i
                        JOIN public.tb_telemetria_intervencao_diaria ttid
                          ON ttid.intervencao_id = i.id
                        WHERE i.rotulo::text NOT LIKE '%%999%%'
                          AND i.rotulo::text NOT LIKE 'VERDE GRANDE%%'
                          AND i.rotulo::text NOT LIKE '%%#'
                          AND i.longitude IS NOT NULL
                          AND i.latitude  IS NOT NULL
                          AND ttid.consumo_diario > 0
                    )
                """)

            linhas = cursor.fetchall()
            cursor.close()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao consultar operadores:\n{e}")
            return

        if not linhas:
            QMessageBox.information(self, "Aviso", "Nenhum operador encontrado para os critérios selecionados.")
            return

        # ── Nome e caminho do arquivo ─────────────────────────────────────────────
        ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
        sufixo     = "COM_TRANSMISSAO" if somente_ativos else "TODOS"
        nome_arquivo = f"OPERADORES_{sufixo}_{ts}.xlsx"
        downloads = (
            os.path.join(os.environ.get('USERPROFILE', ''), 'Downloads')
            if sys.platform == 'win32'
            else os.path.join(os.path.expanduser('~'), 'Downloads')
        )
        caminho = os.path.join(downloads, nome_arquivo)

        if os.path.exists(caminho):
            resp = QMessageBox.question(
                self, "Arquivo Existente",
                f"'{nome_arquivo}' já existe.\nDeseja substituir?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if resp == QMessageBox.No:
                return

        # ── Estilos (idênticos ao padrão JanelaGraficosMedidor) ──────────────────
        fill_azul   = PatternFill("solid", fgColor="175cc3")
        fill_alt    = PatternFill("solid", fgColor="eaf2ff")
        font_branca = Font(bold=True, size=10, color="ffffff")
        font_titulo = Font(bold=True, size=13, color="175cc3")
        font_sub    = Font(size=10, italic=True, color="495057")
        font_normal = Font(size=10)
        ali_centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ali_esq     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        borda       = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        # ── Workbook / aba ────────────────────────────────────────────────────────
        # Contorno para bug openpyxl + lxml no Python 3.12/QGIS 3.3x (Windows):
        # copy(DEFAULT_FONT) → to_tree() → lxml.etree.Element causa access violation.
        # Solução: mascarar temporariamente o lxml para que o openpyxl use o
        # xml.etree.ElementTree puro do Python, que não apresenta o problema.
        import sys as _sys
        _lxml_mod  = _sys.modules.pop('lxml',        None)
        _lxml_etree = _sys.modules.pop('lxml.etree', None)
        try:
            wb = Workbook()
        finally:
            if _lxml_mod   is not None: _sys.modules['lxml']        = _lxml_mod
            if _lxml_etree is not None: _sys.modules['lxml.etree']  = _lxml_etree
        ws = wb.active
        ws.title = "Operadores"

        if somente_ativos:
            larguras = [6, 36, 34, 26, 16, 16, 26, 30, 35, 18]
            colunas  = ["ID", "Nome", "E-mail", "Nº Cadastro / CPF / CNPJ",
                        "Operador", "Data de Cadastro",
                        "Medidores de Teste", "Medidores Reais",
                        "CNARHs c/ Medidores Reais", "Última Data c/ Dado"]
        else:
            larguras = [6, 40, 38, 22, 26, 18]
            colunas  = ["ID", "Nome", "E-mail", "Nº Cadastro / CPF / CNPJ",
                        "Operador", "Data de Cadastro"]
        n_cols   = len(colunas)
        col_last = get_column_letter(n_cols)

        titulo_xlsx = (
            "Operadores com Transmissão Ativa de Dados"
            if somente_ativos
            else "Lista de Operadores de Telemetria Cadastrados"
        )

        # Linha 1 – título
        ws.merge_cells(f"A1:{col_last}1")
        ws["A1"]           = titulo_xlsx
        ws["A1"].font      = font_titulo
        ws["A1"].alignment = ali_centro
        ws.row_dimensions[1].height = 22

        # Linha 2 – subtítulo
        ws.merge_cells(f"A2:{col_last}2")
        ws["A2"]           = (f"Gerado em: {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}"
                              f"  –  Total: {len(linhas)} operador(es)")
        ws["A2"].font      = font_sub
        ws["A2"].alignment = ali_centro
        ws.row_dimensions[2].height = 18

        # Linha 3 – cabeçalho
        for ci, cab in enumerate(colunas, 1):
            cell           = ws.cell(row=3, column=ci, value=cab)
            cell.font      = font_branca
            cell.fill      = fill_azul
            cell.alignment = ali_centro
            cell.border    = borda
        ws.row_dimensions[3].height = 20

        # Linhas de dados
        for idx, row_data in enumerate(linhas, start=4):
            if somente_ativos:
                id_num, nome, email, num_cad, tipo, data_cad, \
                    ds_teste, ds_reais, cnarhs, ultima_data = row_data
            else:
                id_num, nome, email, num_cad, tipo, data_cad = row_data
                ds_teste = ds_reais = cnarhs = ultima_data = None

            if data_cad:
                try:
                    data_str = data_cad.strftime("%d/%m/%Y") if hasattr(data_cad, 'strftime') else str(data_cad)
                except Exception:
                    data_str = str(data_cad)
            else:
                data_str = "—"

            if ultima_data:
                try:
                    ultima_data_str = ultima_data.strftime("%d/%m/%Y") if hasattr(ultima_data, 'strftime') else str(ultima_data)
                except Exception:
                    ultima_data_str = str(ultima_data)
            else:
                ultima_data_str = "—"

            if somente_ativos:
                valores = [id_num, nome or "—", email or "—",
                           num_cad or "—", tipo or "—", data_str,
                           ds_teste or "—", ds_reais or "—",
                           cnarhs or "—", ultima_data_str]
            else:
                valores = [id_num, nome or "—", email or "—",
                           num_cad or "—", tipo or "—", data_str]
            fill_linha = fill_alt if (idx % 2 == 0) else None

            for ci, val in enumerate(valores, 1):
                cell           = ws.cell(row=idx, column=ci, value=val)
                cell.font      = font_normal
                cell.border    = borda
                if fill_linha:
                    cell.fill  = fill_linha
                cell.alignment = ali_centro if ci == 1 else ali_esq

        # Larguras e freeze
        for ci, larg in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(ci)].width = larg
        ws.freeze_panes = "A4"

        # Rodapé
        rodape_row = len(linhas) + 5
        ws.merge_cells(f"A{rodape_row}:{col_last}{rodape_row}")
        cell_rod           = ws.cell(row=rodape_row, column=1,
                                     value="Sistema DURH Diária por Telemetria (SFI/ANA) – Relatório gerado automaticamente pelo plugin QGIS")
        cell_rod.font      = Font(size=8, italic=True, color="888888")
        cell_rod.alignment = ali_centro

        # ── Salvar ────────────────────────────────────────────────────────────────
        try:
            wb.save(caminho)
        except Exception as e:
            QMessageBox.critical(self, "Erro ao salvar", f"Não foi possível salvar o arquivo:\n{e}")
            return

        QMessageBox.information(
            self, "Exportação concluída",
            f"Arquivo exportado com sucesso!\n\n{caminho}",
            QMessageBox.Ok
        )