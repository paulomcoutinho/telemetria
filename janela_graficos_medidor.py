# -*- coding: utf-8 -*-
"""
Módulo: janela_graficos_medidor.py
====================================
Janela de visualização gráfica de consumo hídrico via Matplotlib embutido
(FigureCanvasQTAgg).

Apresenta dois painéis temporais:
  - Gráfico mensal : volume consumido por mês no ano corrente, com tooltips
                     interativos e suporte a barras empilhadas/sobrepostas;
  - Gráfico diário : consumo dia a dia para o mês selecionado, com linha de
                     volume outorgado e marcação visual de anomalias.

Recursos adicionais:
  - Legenda interativa (ocultar/exibir séries por clique);
  - Exportação XLSX com formatação institucional;
  - Relatório TXT de leituras de 15 minutos para a data selecionada;
  - Botão "Ver no Mapa": minimiza esta janela e a JanelaMonitoramento anterior;
  - Tamanho adaptado: 1200×900 (múltiplas interferências) / 1000×750 (simples).

Dependência opcional: matplotlib ≥ 3.6 (MATPLOTLIB_DISPONIVEL).

Autor : SFI/ANA
Versão: 2.0 – Março/2026
"""

# ---------------------------------------------------------------------------
# Imports Qt – widgets, layout e utilitários de UI
# ---------------------------------------------------------------------------
from qgis.PyQt.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QComboBox, QFrame,
    QMessageBox, QSizePolicy, QScrollArea,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QDateEdit, QAbstractItemView,
    QTabWidget,
    QDialog,
    QListWidget,
    QListWidgetItem,
    QCheckBox,
    QDesktopWidget,
    QApplication,
    QToolButton,    
)
from qgis.PyQt.QtCore import Qt, QDate, QTimer, QUrl, QSize
from qgis.PyQt.QtGui import QColor, QDesktopServices, QFont, QIcon

# ---------------------------------------------------------------------------
# Imports QGIS core / canvas
# ---------------------------------------------------------------------------
from qgis.core import (
    QgsProject,
    QgsCoordinateTransform,
    QgsCoordinateReferenceSystem,
    QgsPointXY,
    QgsRectangle,
    QgsRasterLayer,
    QgsLayerTreeLayer,
    QgsDataSourceUri,
    QgsVectorLayer,
    QgsMarkerSymbol,
    QgsSingleSymbolRenderer,
    QgsPalLayerSettings,
    QgsTextFormat,
    QgsTextBufferSettings,
    QgsVectorLayerSimpleLabeling,
    QgsUnitTypes,
    QgsAuthMethodConfig,
    QgsApplication,
    QgsSettings,    
)
from qgis.utils import iface

from qgis.gui import QgsMapToolIdentify

# ---------------------------------------------------------------------------
# Imports Python padrão
# ---------------------------------------------------------------------------
import os
import psycopg2
import math
from datetime import datetime, date
import calendar
import sys
import requests

# ---------------------------------------------------------------------------
# Matplotlib (opcional – gráficos)
# ---------------------------------------------------------------------------
try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.colors import to_rgb
    from matplotlib.patches import Patch
    from matplotlib.lines import Line2D
    MATPLOTLIB_DISPONIVEL = True
except ImportError:
    MATPLOTLIB_DISPONIVEL = False
    print("AVISO: A biblioteca 'matplotlib' não foi encontrada. A janela de gráficos pode não funcionar.")    

# ---------------------------------------------------------------------------
# Exportação Excel (opcional)
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False

# ---------------------------------------------------------------------------
# Sistema de design centralizado
# ---------------------------------------------------------------------------
try:
    from . import ui_tema
except ImportError:
    import ui_tema


# ---------------------------------------------------------------------------
# Módulo filho instanciado por esta janela
# ---------------------------------------------------------------------------
from .janela_monitoramento_detalhes import JanelaMonitoramentoDetalhes


class JanelaGraficosMedidor(QWidget):
    """Janela de visualização gráfica de consumo e vazão dos medidores selecionados.

    Renderiza séries temporais de dados de telemetria em dois painéis
    Matplotlib embutidos (mensal e diário), com suporte a múltiplos medidores
    e agregação por interferência. É instanciada a partir de
    ``JanelaMonitoramento`` e recebe todos os IDs e metadados dos medidores
    selecionados para construir os gráficos.

    Funcionalidades principais:

        **Gráfico mensal**: exibe o volume consumido por mês (m³) ao longo
        do ano corrente, com tooltips interativos ao passar o cursor sobre
        as barras. Em modo de agregação de múltiplas interferências, as
        séries são empilhadas ou sobrepostas conforme a escolha do usuário.

        **Gráfico diário**: exibe o consumo dia a dia para o mês selecionado
        pelo usuário, com identificação de anomalias (leituras com saltos
        negativos ou valores acima do limite físico baseado na vazão nominal
        × fator de segurança). Dias com anomalias são marcados visualmente.

        **Legenda interativa**: clicando em uma entrada da legenda, o usuário
        oculta/exibe a série correspondente nos dois gráficos simultaneamente.

        **Volume outorgado**: linha de referência horizontal no gráfico diário,
        representando o volume outorgado mensal dividido pelo número de dias;
        pode ser ocultada individualmente.

        **Exportação**: botões de exportação para Excel (XLSX com formatação
        institucional, cabeçalho azul e linhas alternadas) e relatório TXT
        com dados a cada 15 minutos para a data selecionada.

        **Ver no Mapa**: minimiza esta janela (e a ``JanelaMonitoramento``
        anterior, se visível) e centraliza o QGIS Canvas na localização
        do medidor selecionado.

    O tamanho da janela é adaptado automaticamente: janelas maiores
    (1200 × 900) para agregações com múltiplas interferências, e tamanho
    padrão (1000 × 750) para seleções simples.

    Attributes:
        janela_anterior (JanelaMonitoramento): Referência à janela de busca;
            restaurada ao fechar ou retornar.
        conn (psycopg2.connection): Conexão ativa ao PostgreSQL.
        lista_ids_selecionados (list[int]): IDs dos medidores incluídos
            na visualização.
        lista_dados_selecionados (list[tuple]): Metadados completos de cada
            medidor (interferência, rótulo, usuário, operador, CNARH, etc.).
        usuario_logado (str | None): Usuário autenticado na sessão.
        senha_conexao (str | None): Credencial para conexões auxiliares.
        nome_usuario (str | None): Nome de exibição do usuário, usado no
            título da janela.
        modo_agregacao (str | None): Modo de combinação das séries quando
            há múltiplos medidores (ex.: ``"soma"``).
        criterio_busca (str | None): Critério utilizado na busca de origem
            (``"CNARH"``, ``"Usuário"`` ou ``"Sistema Hídrico"``).
        termo_busca (str | None): Termo digitado na busca de origem.
        nome_completo (str | None): Nome completo do usuário/sistema buscado.
        eh_multipla_interferencia (bool): ``True`` quando os medidores
            selecionados pertencem a mais de uma interferência CNARH.
        codigos_interf (set[str]): Conjunto dos códigos de interferência
            distintos presentes na seleção.
        is_selecao_total (bool): ``True`` quando todos os medidores do
            resultado foram selecionados.
        colunas_ocultas_mensal (list): Índices das séries ocultadas no
            gráfico mensal via legenda interativa.
        colunas_ocultas_diario (list): Índices das séries ocultadas no
            gráfico diário via legenda interativa.
        medidores_ocultos_grafico (set): Conjunto de rótulos de medidores
            com visibilidade desativada pelo usuário.
        volume_outorgado_oculto (bool): ``True`` quando a linha de volume
            outorgado foi ocultada no gráfico diário.
        eh_empilhado_mensal (bool): Indica modo de barras empilhadas no
            gráfico mensal.
        eh_empilhado_diario (bool): Indica modo de barras empilhadas no
            gráfico diário.
        dados_originais_mensal (Any): Cache do dataset mensal carregado;
            preservado para permitir replotagem sem nova consulta ao banco.
        dados_originais_diario (Any): Cache do dataset diário carregado.
    """
    
    def __init__(self, janela_anterior, conexao, lista_ids_medidores, lista_dados_medidores, 
                 usuario=None, senha=None, nome_usuario=None, modo_agregacao=None, 
                 criterio_busca=None, termo_busca=None, nome_completo=None,
                 eh_multipla_interferencia=False, codigos_interf=None, is_selecao_total=False,
                 ano_inicial=None, mes_inicial=None):
        super().__init__()
        self.janela_anterior = janela_anterior
        self.conn = conexao
        self.lista_ids_selecionados = lista_ids_medidores
        self.lista_dados_selecionados = lista_dados_medidores
        self.usuario_logado = usuario
        self.senha_conexao = senha
        self.nome_usuario = nome_usuario
        self.modo_agregacao = modo_agregacao
        self.criterio_busca = criterio_busca  
        self.termo_busca = termo_busca
        self.nome_completo = nome_completo
        self.eh_multipla_interferencia = eh_multipla_interferencia
        self.codigos_interf = codigos_interf if codigos_interf else set()
        self.modo_agregacao_atual = modo_agregacao
        self.is_selecao_total = is_selecao_total
        self.ano_inicial = ano_inicial
        self.mes_inicial = mes_inicial        
        
        self.colunas_ocultas_mensal = []  
        self.colunas_ocultas_diario = []
        self._janelas_abertas = []        
        self.medidores_ocultos_grafico = set()
        self.volume_outorgado_oculto = False        
        self.eh_empilhado_mensal = False
        self.eh_empilhado_diario = False
        self.tooltip_data_mensal = {}
        self.tooltip_data_diario = {}
        
        self.dados_originais_mensal = None
        self.dados_originais_diario = None

        self.configurar_titulo_janela() 
           
        # === AUMENTAR TAMANHO DA JANELA ===
        # Verificar se é agregação com múltiplas interferências
        codigos_interf = set([d[5] for d in self.lista_dados_selecionados])
        eh_multipla_interferencia = len(codigos_interf) > 1
        
        if eh_multipla_interferencia and self.modo_agregacao:
            # Janela maior para agregações (mais espaço para o gráfico)
            self.setGeometry(100, 50, 1200, 900)  # Aumentado de 1000x750 para 1200x900
        else:
            # Tamanho padrão para casos simples
            self.setGeometry(100, 100, 1000, 750)
        
        self.center()
        
        try:
            self.primary_color = ui_tema.StyleConfig.PRIMARY_COLOR
            self.secondary_color = ui_tema.StyleConfig.SECONDARY_COLOR
            self.text_color = ui_tema.StyleConfig.TEXT_DARK
            self.bg_color = ui_tema.StyleConfig.BACKGROUND_WHITE
            self.border_color = ui_tema.StyleConfig.BORDER_COLOR
        except:
            self.primary_color = "#175cc3"
            self.secondary_color = "#5474b8"
            self.text_color = "#343a40"
            self.bg_color = "#ffffff"
            self.border_color = "#dee2e6"

        self.cid_hover_mensal = None
        self.cid_hover_diario = None
            
        self.initUI()
        
        if self.ano_inicial is not None:
            self.combo_ano_mensal.setCurrentText(str(self.ano_inicial))
            
    def initUI(self):
        """Configura a interface da janela de gráficos."""
        try:
            ui_tema.aplicar_tema_arredondado(self)
        except:
            pass

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(10)
        
        # === CABEÇALHO COM INFORMAÇÕES DO MEDIDOR ===
        header_container = QWidget()
        header_container.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                padding: 15px;
            }
        """)
        header_layout = QVBoxLayout(header_container)

        # Configurar cabeçalho baseado no modo de agregação atual
        if self.eh_multipla_interferencia:
            self.configurar_cabecalho_multipla_interferencia(header_layout, self.codigos_interf)
        else:
            self.configurar_cabecalho_padrao(header_layout)

        main_layout.addWidget(header_container)
        
        # === ABAS DE GRÁFICOS ===
        self.tabs_graficos = QTabWidget()
        self.tabs_graficos.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #dee2e6;
                border-radius: 5px;
                background-color: white;
                margin-top: 5px;
            }
            QTabBar::tab {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: white;
                border-bottom: 2px solid #175cc3;
                font-weight: bold;
            }
        """)
        
        # --- Aba 1: Consumo Mensal ---
        self.tab_mensal = QWidget()
        layout_tab_mensal = QVBoxLayout(self.tab_mensal)
        layout_tab_mensal.setSpacing(10)

        # === SUB-ABAS DE AGREGAÇÃO (somente para múltiplas interferências) ===
        if self.eh_multipla_interferencia:
            self.subtabs_mensal = QTabWidget()
            self.subtabs_mensal.setStyleSheet("""
                QTabWidget::pane {
                    border: 1px solid #dee2e6;
                    background-color: #f8f9fa;
                }
                QTabBar::tab {
                    background-color: #e9ecef;
                    padding: 6px 12px;
                    font-size: 11px;
                }
                QTabBar::tab:selected {
                    background-color: white;
                    border-bottom: 2px solid #28a745;
                }
            """)

            titulo_subaba_total = "Totalizado" if self.is_selecao_total else "Total"
            
            # Criar sub-aba para cada critério de agregação
            self.criar_subaba_agregacao(self.subtabs_mensal, "criterio_busca",
                                        titulo_subaba_total, self.nome_completo)
            self.criar_subaba_agregacao(self.subtabs_mensal, "interferencia",
                                        "Por interferência", None)
            
            # Conectar mudança de sub-aba para atualizar gráfico
            self.subtabs_mensal.currentChanged.connect(self.on_subtab_mensal_changed)
            
            layout_tab_mensal.addWidget(self.subtabs_mensal)
        else:
            # Sem sub-abas para caso simples
            self.subtabs_mensal = None

        # INSTRUÇÃO VISUAL
        instrucao_mensal = QLabel("Clique na coluna do mês desejado para visualizar calendário mensal")
        instrucao_mensal.setStyleSheet("""
            QLabel {
                background-color: #e3f2fd;
                color: #1976d2;
                padding: 8px;
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }
        """)
        instrucao_mensal.setAlignment(Qt.AlignCenter)
        layout_tab_mensal.addWidget(instrucao_mensal)

        # SELETOR DE ANO
        seletor_mensal_container = QWidget()
        layout_seletor_mensal = QHBoxLayout(seletor_mensal_container)
        layout_seletor_mensal.addWidget(QLabel("Selecione o ano:"))

        self.combo_ano_mensal = QComboBox()
        anos = list(range(2020, QDate.currentDate().year() + 2))
        for ano in anos:
            self.combo_ano_mensal.addItem(str(ano))
        self.combo_ano_mensal.setCurrentText(str(QDate.currentDate().year()))
        self.combo_ano_mensal.currentIndexChanged.connect(self.atualizar_grafico_mensal)

        layout_seletor_mensal.addWidget(self.combo_ano_mensal)

        # BOTÃO EXPORTAR PNG
        self.btn_exportar_png_mensal = QPushButton("Exportar como imagem")
        self.btn_exportar_png_mensal.setStyleSheet(f"""
            QPushButton {{
                background-color: white;
                color: {self.primary_color};
                border: 1px solid {self.primary_color};
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #f0f0f0;
            }}
        """)
        self.btn_exportar_png_mensal.setToolTip("Baixar gráfico mensal como PNG")
        self.btn_exportar_png_mensal.clicked.connect(lambda: self.exportar_grafico_png("mensal"))
        
        layout_seletor_mensal.addWidget(self.btn_exportar_png_mensal)        
        layout_seletor_mensal.addStretch()

        layout_tab_mensal.addWidget(seletor_mensal_container)

        # GRÁFICO MENSAL
        if MATPLOTLIB_DISPONIVEL:
            if self.eh_multipla_interferencia:
                figsize = (10, 6)
            else:
                figsize = (8, 4)
                
            fig_mensal = Figure(figsize=figsize, dpi=100)
            self.canvas_mensal = FigureCanvas(fig_mensal)
            self.canvas_mensal.axes = fig_mensal.add_subplot(111)
            self.canvas_mensal.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            
            self.canvas_mensal.mpl_connect('button_press_event', self.on_click_mensal)
            self.canvas_mensal.mpl_connect('motion_notify_event', self.on_hover_mensal)
            self.canvas_mensal.mpl_connect('pick_event', self.on_pick_legenda_mensal)
                       
            self.annot_mensal = self.canvas_mensal.axes.annotate(
                "", xy=(0,0), xytext=(20,20), textcoords="offset points",
                bbox=dict(boxstyle="round", fc="w"),
                arrowprops=dict(arrowstyle="->")
            )
            self.annot_mensal.set_visible(False)
        else:
            self.canvas_mensal = QWidget()

        layout_tab_mensal.addWidget(self.canvas_mensal, stretch=10)
        
        # --- Aba 2: Consumo Diário ---
        self.tab_diario = QWidget()
        layout_tab_diario = QVBoxLayout(self.tab_diario)
        layout_tab_diario.setSpacing(10)
        
        # === SUB-ABAS DE AGREGAÇÃO (somente para múltiplas interferências) ===
        if self.eh_multipla_interferencia:
            self.subtabs_diario = QTabWidget()
            self.subtabs_diario.setStyleSheet("""
                QTabWidget::pane {
                    border: 1px solid #dee2e6;
                    background-color: #f8f9fa;
                }
                QTabBar::tab {
                    background-color: #e9ecef;
                    padding: 6px 12px;
                    font-size: 11px;
                }
                QTabBar::tab:selected {
                    background-color: white;
                    border-bottom: 2px solid #28a745;
                }
            """)

            titulo_subaba_total = "Totalizado" if self.is_selecao_total else "Total"
            
            # Criar sub-aba para cada critério de agregação
            self.criar_subaba_agregacao(self.subtabs_diario, "criterio_busca",
                                        titulo_subaba_total, self.nome_completo)
            self.criar_subaba_agregacao(self.subtabs_diario, "interferencia",
                                        "Por interferência", None)
            
            # Conectar mudança de sub-aba
            self.subtabs_diario.currentChanged.connect(self.on_subtab_diario_changed)
            
            layout_tab_diario.addWidget(self.subtabs_diario)
        else:
            self.subtabs_diario = None

        # INSTRUÇÃO VISUAL
        instrucao_diario = QLabel("Clique na coluna do dia desejado para visualizar dados a cada 15 minutos")
        instrucao_diario.setStyleSheet("""
            QLabel {
                background-color: #e3f2fd;
                color: #1976d2;
                padding: 8px;
                border-radius: 4px;
                font-size: 11px;
                font-weight: bold;
            }
        """)
        instrucao_diario.setAlignment(Qt.AlignCenter)
        layout_tab_diario.addWidget(instrucao_diario)
        
        # SELETOR DE MÊS/ANO
        seletor_container = QWidget()
        layout_seletor = QHBoxLayout(seletor_container)
        layout_seletor.addWidget(QLabel("Selecione o mês/ano:"))
        
        self.combo_ano_graf = QComboBox()
        for ano in anos:
            self.combo_ano_graf.addItem(str(ano))
        self.combo_ano_graf.setCurrentText(str(QDate.currentDate().year()))
        self.combo_ano_graf.currentIndexChanged.connect(self.atualizar_grafico_diario)
        
        self.combo_mes_graf = QComboBox()
        self.combo_mes_graf.addItems(["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"])
        self.combo_mes_graf.setCurrentIndex(QDate.currentDate().month() - 1)
        self.combo_mes_graf.currentIndexChanged.connect(self.atualizar_grafico_diario)
        
        layout_seletor.addWidget(self.combo_mes_graf)
        layout_seletor.addWidget(self.combo_ano_graf)

        # BOTÃO EXPORTAR PNG
        self.btn_exportar_png_diario = QPushButton("Exportar como imagem")
        self.btn_exportar_png_diario.setStyleSheet(f"""
            QPushButton {{
                background-color: white;
                color: {self.primary_color};
                border: 1px solid {self.primary_color};
                border-radius: 8px;
                padding: 8px 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #f0f0f0;
            }}
        """)
        self.btn_exportar_png_diario.setToolTip("Baixar gráfico diário como PNG")
        self.btn_exportar_png_diario.clicked.connect(lambda: self.exportar_grafico_png("diario"))
        layout_seletor.addWidget(self.btn_exportar_png_diario)
        
        layout_seletor.addStretch()
        
        layout_tab_diario.addWidget(seletor_container)

        # GRÁFICO DIÁRIO
        if MATPLOTLIB_DISPONIVEL:
            if self.eh_multipla_interferencia:
                figsize = (10, 6)
            else:
                figsize = (8, 4)
                
            fig_diario = Figure(figsize=figsize, dpi=100)
            self.canvas_diario = FigureCanvas(fig_diario)
            self.canvas_diario.axes = fig_diario.add_subplot(111)
            self.canvas_diario.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            
            self.canvas_diario.mpl_connect('button_press_event', self.on_click_diario)
            self.canvas_diario.mpl_connect('motion_notify_event', self.on_hover_diario)
            self.canvas_diario.mpl_connect('pick_event', self.on_pick_legenda_diario)            
            
            self.annot_diario = self.canvas_diario.axes.annotate(
                "", xy=(0,0), xytext=(20,20), textcoords="offset points",
                bbox=dict(boxstyle="round", fc="w"),
                arrowprops=dict(arrowstyle="->")
            )
            self.annot_diario.set_visible(False)
        else:
            self.canvas_diario = QWidget()
        
        layout_tab_diario.addWidget(self.canvas_diario, stretch=10)
        
        # ADICIONAR ABAS PRINCIPAIS
        self.tabs_graficos.addTab(self.tab_mensal, "Consumo Mensal")
        self.tabs_graficos.addTab(self.tab_diario, "Consumo Diário")
        self.tabs_graficos.currentChanged.connect(self.on_tab_change)
        
        main_layout.addWidget(self.tabs_graficos, stretch=10)
        
        # === RODAPÉ (Botões) ===
        botoes_container = QWidget()
        botoes_container.setStyleSheet(f"""
            QWidget {{
                background-color: white;
                border: 1px solid {self.border_color};
                border-radius: 5px;
                padding: 15px;
            }}
        """)
        botoes_layout = QHBoxLayout(botoes_container)
        botoes_layout.setSpacing(15)

        # Botões...
        self.btn_controle_colunas = QPushButton("Controle de Colunas")
        self.btn_controle_colunas.setStyleSheet(f"""
            QPushButton {{
                background-color: #6c757d;
                color: white;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #5a6268;
            }}
        """)
        self.btn_controle_colunas.clicked.connect(self.mostrar_dialogo_controle_colunas)
        
        self.btn_restaurar_todas = QPushButton("Restaurar Todas")
        self.btn_restaurar_todas.setEnabled(False) 
        self.btn_restaurar_todas.setStyleSheet(f"""
            QPushButton {{
                background-color: #28a745;
                color: white;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:disabled {{
                background-color: #cccccc;
            }}
            QPushButton:hover {{
                background-color: #218838;
            }}
        """)
        self.btn_restaurar_todas.clicked.connect(self.restaurar_todas_colunas)
        
        botoes_layout.addWidget(self.btn_controle_colunas)
        botoes_layout.addWidget(self.btn_restaurar_todas)
        botoes_layout.addStretch()
               
        self.btn_excel = QPushButton("Exportar Excel")
        self.btn_excel.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.primary_color};
                color: white;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {self.secondary_color};
            }}
        """)
        self.btn_excel.clicked.connect(self.exportar_excel)
        
        self.btn_mapa = QPushButton("Ver no Mapa")
        self.btn_mapa.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.primary_color};
                color: white;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {self.secondary_color};
            }}
        """)
        self.btn_mapa.clicked.connect(self.ver_no_mapa)
        
        botoes_layout.addWidget(self.btn_excel)
        botoes_layout.addWidget(self.btn_mapa)
        botoes_layout.addStretch()
        
        btn_voltar = QPushButton("Fechar")
        btn_voltar.setStyleSheet(f"""
            QPushButton {{
                background-color: white;
                color: {self.primary_color};
                border: 1px solid {self.primary_color};
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #f0f0f0;
            }}
        """)
        btn_voltar.clicked.connect(self.close)
        botoes_layout.addWidget(btn_voltar)
        
        main_layout.addWidget(botoes_container)
        
        self.setLayout(main_layout)
        
        # Carregar gráficos iniciais
        self.atualizar_grafico_mensal()        
        self.atualizar_grafico_diario()
    
    def configurar_titulo_janela(self):
        """Configura o título da janela conforme regras de negócio."""
        if self.eh_multipla_interferencia:
            if self.modo_agregacao == "criterio_busca":
                if self.criterio_busca == "Sistema Hídrico":
                    display = self.nome_completo if self.nome_completo else self.termo_busca
                    self.setWindowTitle(f"Estatísticas - DURH Diária por Telemetria - Sistema Hídrico: {display}")
                elif self.criterio_busca == "UAM":
                    display = self.nome_completo if self.nome_completo else self.termo_busca
                    self.setWindowTitle(f"Estatísticas - DURH Diária por Telemetria - UAM: {display}")
                elif self.criterio_busca == "CNARH":
                    display = self.nome_completo if self.nome_completo else self.termo_busca
                    self.setWindowTitle(f"Estatísticas - DURH Diária por Telemetria - CNARH: {display}")
                else:  # Usuário
                    display = self.nome_completo if self.nome_completo else self.termo_busca
                    self.setWindowTitle(f"Estatísticas - DURH Diária por Telemetria - Usuário: {display}")
            else:  # interferencia
                self.setWindowTitle(f"🔀 Estatísticas - Por Interferência ({len(self.codigos_interf)} interferências)")
        else:
            if len(self.lista_ids_selecionados) > 1:
                self.setWindowTitle(f"Estatísticas - DURH Diária por Telemetria - (Múltiplos: {len(self.lista_ids_selecionados)})")
            else:
                self.setWindowTitle(f"Estatísticas - DURH Diária por Telemetria - {self.lista_dados_selecionados[0][1]}")

    def configurar_cabecalho_multipla_interferencia(self, header_layout, codigos_interf):
        """Configura cabeçalho simplificado para múltiplas interferências."""
        while header_layout.count():
            item = header_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        # SEMPRE mostrar o critério de busca, independente do modo de agregação
        if self.criterio_busca == "Sistema Hídrico":
            display = self.nome_completo if self.nome_completo else self.termo_busca
            titulo_texto = f"🌊 Sistema Hídrico: {display}"
        elif self.criterio_busca == "UAM":
            display = self.nome_completo if self.nome_completo else self.termo_busca
            titulo_texto = f"🌊 UAM: {display}"            
        elif self.criterio_busca == "CNARH":
            display = self.nome_completo if self.nome_completo else self.termo_busca
            titulo_texto = f"📋 CNARH: {display}"
        else:  # Usuário
            display = self.nome_completo if self.nome_completo else self.termo_busca
            titulo_texto = f"👤 Usuário: {display}"
        
        lbl_titulo = QLabel(titulo_texto)
        lbl_titulo.setStyleSheet("""
            font-size: 16px; 
            font-weight: bold; 
            color: #175cc3;
            margin-bottom: 8px;
        """)
        lbl_titulo.setWordWrap(True)
        header_layout.addWidget(lbl_titulo)
        
        # Sub-cabeçalho com códigos das interferências
        codigos_lista = sorted(list(codigos_interf))
        codigos_texto = ", ".join([str(c) for c in codigos_lista])
        
        lbl_interferencias = QLabel(f"Interferências: {codigos_texto}")
        lbl_interferencias.setStyleSheet("""
            font-size: 12px; 
            color: #495057;
            padding: 8px;
            background-color: #f8f9fa;
            border-radius: 3px;
            border: 1px solid #dee2e6;
        """)
        lbl_interferencias.setWordWrap(True)
        header_layout.addWidget(lbl_interferencias)
        
    def configurar_cabecalho_padrao(self, header_layout):
        """Configura cabeçalho padrão para seleção única ou mesma interferência."""
        # ADICIONAR NOME DO USUÁRIO NO TOPO (se disponível)
        if self.nome_usuario and self.nome_usuario != "Não informado":
            lbl_usuario = QLabel(f"👤 Usuário: {self.nome_usuario}")
            try:
                lbl_usuario.setStyleSheet(f"""
                    font-size: 13px; 
                    font-weight: bold; 
                    color: {ui_tema.StyleConfig.SECONDARY_COLOR};
                    margin-bottom: 8px;
                    padding-bottom: 8px;
                    border-bottom: 1px solid #e9ecef;
                """)
            except NameError:
                lbl_usuario.setStyleSheet("""
                    font-size: 13px; 
                    font-weight: bold; 
                    color: #5474b8;
                    margin-bottom: 8px;
                    padding-bottom: 8px;
                    border-bottom: 1px solid #e9ecef;
                """)
            lbl_usuario.setWordWrap(True)
            header_layout.addWidget(lbl_usuario)
        
        # BUSCAR CNARH, CÓDIGO DA INTERFERÊNCIA E LISTAR TODOS OS RÓTULOS
        primeiro_medidor = self.lista_dados_selecionados[0]
        cnarh = primeiro_medidor[4]
        codigo_interf = primeiro_medidor[5]
        
        # Coletar todos os rótulos
        lista_rotulos = [medidor[1] for medidor in self.lista_dados_selecionados]
        
        # MONTAR TEXTO DO TÍTULO: CNARH | INTERFERÊNCIA | MEDIDOR(ES)
        texto_titulo = f"CNARH: {cnarh}" if cnarh and cnarh != "Não informado" else "CNARH: Não informado"
        texto_titulo += " | "
        texto_titulo += f"Interferência: {codigo_interf}" if codigo_interf and codigo_interf != "Não informado" else "Interferência: Não informada"
        texto_titulo += " | "
        
        if len(lista_rotulos) == 1:
            texto_titulo += f"Medidor: {lista_rotulos[0]}"
        else:
            rotulos_texto = ", ".join(lista_rotulos)
            texto_titulo += f"Medidores: {rotulos_texto}"
        
        titulo = QLabel(texto_titulo)
        try:
            titulo.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {ui_tema.StyleConfig.PRIMARY_COLOR};")
        except NameError:
            titulo.setStyleSheet("font-size: 16px; font-weight: bold; color: #175cc3;")
        
        titulo.setWordWrap(True)
        header_layout.addWidget(titulo)

    def formatar_br(self, valor):
        """Converte número float para string no padrão brasileiro (1.000,00)."""
        if valor is None:
            return "0,00"
        # 1. Formata com separador padrão (1,234.56)
        # 2. Substitui ponto por arroba temporariamente (1,234@56)
        # 3. Troca vírgula por ponto (1.234@56)
        # 4. Troca arroba por vírgula (1.234,56)
        return f"{valor:,.2f}".replace('.', '@').replace(',', '.').replace('@', ',')

    def mostrar_dialogo_controle_colunas(self):
        """Mostra diálogo com totalização de volumes (múltiplos medidores)."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Controle de Colunas do Gráfico")
        dialog.setModal(True)
        dialog.setMinimumWidth(500)
        
        layout = QVBoxLayout(dialog)
        
        # Determinar qual gráfico está ativo
        aba_atual = self.tabs_graficos.currentIndex()
        
        if aba_atual == 0:  # Gráfico Mensal
            # 1. Busca dados brutos (com repetição de medidores)
            dados_brutos = self.get_data_mensal()
            tipo = "mensal"
            
            if not dados_brutos:
                QMessageBox.information(self, "Sem dados", "Não há dados disponíveis para o gráfico mensal.")
                return
                
            # 2. Totalizar por Mês
            dados_agregados = {} # {mes: total_consumo}
            
            for row in dados_brutos:
                mes = int(row[0])
                val = float(row[1]) if row[1] else 0.0
                if mes in dados_agregados:
                    dados_agregados[mes] += val
                else:
                    dados_agregados[mes] = val
            
            # 3. Converter para lista ordenada para exibição
            lista_exibir = sorted(dados_agregados.items()) # [(1, total), (2, total)...]
            
            nomes_meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                          "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
            
            lista_widget = QListWidget()
            lista_widget.setSelectionMode(QAbstractItemView.MultiSelection)
            
            for mes, total in lista_exibir:
                mes_nome = nomes_meses[mes-1]
                # Usa o formatador BR
                item_text = f"{mes_nome}: {self.formatar_br(total)} m³"
                item = QListWidgetItem(item_text)
                
                # O índice para ocultar é o absoluto (0 a 11)
                indice_absoluto = mes - 1
                
                if indice_absoluto in self.colunas_ocultas_mensal:
                    item.setCheckState(Qt.Checked)
                    item.setBackground(QColor(255, 200, 200))
                    item.setForeground(QColor(100, 100, 100))
                else:
                    item.setCheckState(Qt.Unchecked)
                    
                # Guarda o índice absoluto no UserRole para facilitar
                item.setData(Qt.UserRole, indice_absoluto)
                lista_widget.addItem(item)
                
        else:  # Gráfico Diário
            dados_brutos = self.get_data_diario()
            tipo = "diario"
            
            if not dados_brutos:
                QMessageBox.information(self, "Sem dados", "Não há dados disponíveis para o gráfico diário.")
                return
            
            # 1. Totalizar por DIA
            dados_agregados = {}
            
            for row in dados_brutos:
                dia = int(row[0])
                val = float(row[1]) if row[1] else 0.0
                if dia in dados_agregados:
                    dados_agregados[dia] += val
                else:
                    dados_agregados[dia] = val
            
            # 2. Converter para lista ordenada
            lista_exibir = sorted(dados_agregados.items())
            
            mes_selecionado = self.combo_mes_graf.currentIndex() + 1
            ano_selecionado = int(self.combo_ano_graf.currentText())
            
            lista_widget = QListWidget()
            lista_widget.setSelectionMode(QAbstractItemView.MultiSelection)
            
            for dia, total in lista_exibir:
                data = QDate(ano_selecionado, mes_selecionado, dia)
                if data.isValid():
                    data_str = data.toString("dd/MM/yyyy")
                else:
                    data_str = f"Dia {dia}"
                    
                item_text = f"{data_str}: {self.formatar_br(total)} m³"
                item = QListWidgetItem(item_text)
                
                # O índice relativo (posição na lista de dados totais)
                # Para simplificar, usamos a posição na lista ordenada 'lista_exibir'
                idx_relativo = lista_exibir.index((dia, total))
                item.setData(Qt.UserRole, idx_relativo)
                
                if idx_relativo in self.colunas_ocultas_diario:
                    item.setCheckState(Qt.Checked)
                    item.setBackground(QColor(255, 200, 200))
                    item.setForeground(QColor(100, 100, 100))
                else:
                    item.setCheckState(Qt.Unchecked)
                    
                lista_widget.addItem(item)
        
        layout.addWidget(QLabel("Selecione a(s) coluna(s) para ocultar/restaurar:"))
        layout.addWidget(lista_widget)
        
        # Botões do diálogo
        btn_layout = QHBoxLayout()
        
        btn_ocultar = QPushButton("Ocultar Selecionada(s)")
        btn_ocultar.clicked.connect(lambda: self.processar_selecao_colunas(
            lista_widget, tipo, ocultar=True))
        btn_ocultar.clicked.connect(dialog.close)
        
        btn_restaurar = QPushButton("Restaurar Selecionada(s)")
        btn_restaurar.setEnabled(False)

        btn_restaurar.setStyleSheet("""
            QPushButton {
                background-color: #5474b8;
                color: white;
                border-radius: 8px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
            QPushButton:hover {
                background-color: #2050b8;
            }
        """)

        # Função para verificar se há itens marcados e habilitar botão
        def verificar_selecao():
            tem_selecao = False
            for i in range(lista_widget.count()):
                if lista_widget.item(i).checkState() == Qt.Checked:
                    tem_selecao = True
                    break
            btn_restaurar.setEnabled(tem_selecao)
            
        # Conectar sinal de mudança de item na lista
        lista_widget.itemChanged.connect(verificar_selecao)
        
        # Verificação inicial caso a lista já abra com itens marcados
        verificar_selecao()
        
        btn_restaurar.clicked.connect(lambda: self.processar_selecao_colunas(
            lista_widget, tipo, ocultar=False))
        btn_restaurar.clicked.connect(dialog.close)
        
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.clicked.connect(dialog.close)
        
        btn_layout.addWidget(btn_ocultar)
        btn_layout.addWidget(btn_restaurar)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancelar)
        
        layout.addLayout(btn_layout)
        dialog.exec_()

    def calcular_escala_y(self, dados_consumo, volumes_outorgados=None):
        """Calcula escala Y adequada considerando consumo e volume outorgado."""
        max_consumo = max(dados_consumo) if dados_consumo else 0
        
        max_outorgado = 0
        if volumes_outorgados:
            max_outorgado = max(volumes_outorgados)
        
        # O máximo absoluto é o maior entre consumo e outorgado
        max_absoluto = max(max_consumo, max_outorgado)
        
        if max_absoluto == 0:
            return 100  # Valor padrão se não houver dados
        
        # Adicionar 10% de margem superior
        return max_absoluto * 1.1

    def criar_subaba_agregacao(self, tab_widget, modo, titulo, nome_completo):
        """Cria uma sub-aba para modo de agregação."""
        # Widget container para a sub-aba (vazio, apenas para identificação)
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)  # Sem espaçamento
        
        # Adicionar stretch para evitar warnings de layout vazio
        layout.addStretch()        
                
        # Adicionar ao tab widget
        tab_widget.addTab(container, titulo)
        
        # Armazenar referência ao modo para uso posterior
        container.modo_agregacao = modo

    def on_subtab_mensal_changed(self, index):
        """Chamado quando muda a sub-aba de agregação no mensal."""
        if not self.subtabs_mensal:
            return
        
        # Obter modo da sub-aba selecionada
        container = self.subtabs_mensal.widget(index)
        if hasattr(container, 'modo_agregacao'):
            novo_modo = container.modo_agregacao
            
            if novo_modo != self.modo_agregacao_atual:
                self.modo_agregacao_atual = novo_modo
                self.modo_agregacao = novo_modo
                
                # Atualizar cabeçalho
                self.atualizar_cabecalho()
                
                # Recarregar gráfico
                self.atualizar_grafico_mensal()

    def on_subtab_diario_changed(self, index):
        """Chamado quando muda a sub-aba de agregação no diário."""
        if not self.subtabs_diario:
            return
        
        container = self.subtabs_diario.widget(index)
        if hasattr(container, 'modo_agregacao'):
            novo_modo = container.modo_agregacao
            
            if novo_modo != self.modo_agregacao_atual:
                self.modo_agregacao_atual = novo_modo
                self.modo_agregacao = novo_modo
                
                # Atualizar cabeçalho
                self.atualizar_cabecalho()
                
                # Recarregar gráfico
                self.atualizar_grafico_diario()

    def atualizar_cabecalho(self):
        """Atualiza o cabeçalho quando muda o modo de agregação."""
        # Encontrar o header_layout
        for i in range(self.layout().count()):
            item = self.layout().itemAt(i)
            if item.widget():
                widget = item.widget()
                if isinstance(widget, QWidget) and widget.layout():
                    # Verificar se é o container do cabeçalho
                    if widget.styleSheet() and "background-color: white" in widget.styleSheet():
                        header_layout = widget.layout()
                        # Limpar e recriar
                        while header_layout.count():
                            child = header_layout.takeAt(0)
                            if child.widget():
                                child.widget().deleteLater()
                        
                        # Recriar cabeçalho
                        if self.eh_multipla_interferencia:
                            self.configurar_cabecalho_multipla_interferencia(header_layout, self.codigos_interf)
                        else:
                            self.configurar_cabecalho_padrao(header_layout)
                        
                        # Atualizar título da janela
                        self.configurar_titulo_janela()
                        break
    
    def processar_selecao_colunas(self, lista_widget, tipo, ocultar=True):
        """Processa a seleção de colunas para ocultar ou restaurar."""
        indices_selecionados = []
        
        for i in range(lista_widget.count()):
            item = lista_widget.item(i)
            if item.checkState() == Qt.Checked:
                indice = item.data(Qt.UserRole)
                indices_selecionados.append(indice)
        
        if not indices_selecionados:
            QMessageBox.information(self, "Aviso", "Nenhuma coluna selecionada.")
            return
        
        if tipo == "mensal":
            if ocultar:
                # Adicionar às colunas ocultas (evitando duplicatas)
                for idx in indices_selecionados:
                    if idx not in self.colunas_ocultas_mensal:
                        self.colunas_ocultas_mensal.append(idx)
            else:
                # Remover das colunas ocultas
                for idx in indices_selecionados:
                    if idx in self.colunas_ocultas_mensal:
                        self.colunas_ocultas_mensal.remove(idx)
            
            # Atualizar gráfico
            self.atualizar_grafico_mensal()
            
        else:  # diario
            if ocultar:
                # Adicionar às colunas ocultas (evitando duplicatas)
                for idx in indices_selecionados:
                    if idx not in self.colunas_ocultas_diario:
                        self.colunas_ocultas_diario.append(idx)
            else:
                # Remover das colunas ocultas
                for idx in indices_selecionados:
                    if idx in self.colunas_ocultas_diario:
                        self.colunas_ocultas_diario.remove(idx)
            
            # Atualizar gráfico
            self.atualizar_grafico_diario()
        
        # Lógica para habilitar/desabilitar o botão "Restaurar Todas" da janela principal
        # Habilita se houver colunas ocultas em qualquer um dos gráficos
        tem_ocultas = (len(self.colunas_ocultas_mensal) > 0 or 
                       len(self.colunas_ocultas_diario) > 0)
        self.btn_restaurar_todas.setEnabled(tem_ocultas)
    
    def restaurar_todas_colunas(self):
        """Restaura todas as colunas ocultas."""
        resposta = QMessageBox.question(
            self, 
            "Restaurar Todas as Colunas",
            "Deseja restaurar todas as colunas ocultas?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if resposta == QMessageBox.Yes:
            self.colunas_ocultas_mensal.clear()
            self.colunas_ocultas_diario.clear()
            
            # Atualizar ambos os gráficos
            self.atualizar_grafico_mensal()
            self.atualizar_grafico_diario()
            
            # Desabilita o botão "Restaurar Todas" novamente
            self.btn_restaurar_todas.setEnabled(False)
            
            QMessageBox.information(self, "Sucesso", "Todas as colunas foram restauradas.")

    def on_tab_change(self, index):
        """Sincroniza o ano entre as abas e recarrega o gráfico alvo."""
        # Aba 0: Consumo Mensal
        # Aba 1: Consumo Diário
        
        try:
            # Verifica qual aba está sendo acessada
            if index == 1:  # Aba Diário selecionada
                # Sincronizar o Ano da aba Diária com o Ano da aba Mensal
                if hasattr(self, 'combo_ano_mensal') and hasattr(self, 'combo_ano_graf'):
                    ano_mensal = self.combo_ano_mensal.currentText()
                    if self.combo_ano_graf.currentText() != ano_mensal:
                        self.combo_ano_graf.setCurrentText(ano_mensal)
                
                # === NOVO: Resetar sub-aba para "Totalizado" ===
                if hasattr(self, 'subtabs_diario') and self.subtabs_diario:
                    self.subtabs_diario.setCurrentIndex(0)  # Primeira sub-aba (Totalizado)
                    # Atualizar modo de agregação
                    container = self.subtabs_diario.widget(0)
                    if hasattr(container, 'modo_agregacao'):
                        self.modo_agregacao_atual = container.modo_agregacao
                        self.modo_agregacao = container.modo_agregacao
                        self.atualizar_cabecalho()
                
                # Atualizar gráfico Diário
                self.atualizar_grafico_diario()
                
            elif index == 0:  # Aba Mensal selecionada
                # Sincronizar o Ano da aba Mensal com o Ano da aba Diária
                if hasattr(self, 'combo_ano_graf') and hasattr(self, 'combo_ano_mensal'):
                    ano_diario = self.combo_ano_graf.currentText()
                    if self.combo_ano_mensal.currentText() != ano_diario:
                        self.combo_ano_mensal.setCurrentText(ano_diario)
                
                # === NOVO: Resetar sub-aba para "Totalizado" ===
                if hasattr(self, 'subtabs_mensal') and self.subtabs_mensal:
                    self.subtabs_mensal.setCurrentIndex(0)  # Primeira sub-aba (Totalizado)
                    # Atualizar modo de agregação
                    container = self.subtabs_mensal.widget(0)
                    if hasattr(container, 'modo_agregacao'):
                        self.modo_agregacao_atual = container.modo_agregacao
                        self.modo_agregacao = container.modo_agregacao
                        self.atualizar_cabecalho()
                
                # Atualizar gráfico Mensal
                self.atualizar_grafico_mensal()
        
        except Exception as e:
            print(f"Erro ao mudar de aba: {e}")
            import traceback
            traceback.print_exc()

    def get_data_mensal(self):
        """Busca dados totalizados por mês com suporte a agregação."""
        if hasattr(self, 'combo_ano_mensal'):
            ano = int(self.combo_ano_mensal.currentText())
        else:
            ano = int(self.combo_ano_graf.currentText())
        
        try:
            cursor = self.conn.cursor()
            
            # Verificar se é múltiplo e se há modo de agregação
            e_multiplo = len(self.lista_ids_selecionados) > 1
            codigos_interf = set([d[5] for d in self.lista_dados_selecionados])
            eh_multipla_interferencia = len(codigos_interf) > 1
            
            if eh_multipla_interferencia and self.modo_agregacao:
                # Regra f) - Agregação conforme seleção do usuário
                if self.modo_agregacao == "interferencia":
                    # Agrupar por mês E código de interferência
                    query = """
                    SELECT EXTRACT(MONTH FROM DATE(data)) as mes, 
                           SUM(consumo_diario) as consumo, 
                           intervencao_id,
                           (SELECT nu_interferencia_cnarh FROM view_ft_intervencao 
                            WHERE id = tb_telemetria_intervencao_diaria.intervencao_id LIMIT 1) as cod_interf
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s 
                    AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY mes, intervencao_id
                    ORDER BY mes, intervencao_id;
                    """
                else:
                    # Agregar por critério de busca (tudo junto)
                    query = """
                    SELECT EXTRACT(MONTH FROM DATE(data)) as mes, 
                           SUM(consumo_diario) as consumo, 
                           0 as intervencao_id  -- 0 indica agregação total
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s 
                    AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY mes
                    ORDER BY mes;
                    """
            elif e_multiplo:
                # Query original para múltiplos mesma interferência
                query = """
                SELECT EXTRACT(MONTH FROM DATE(data)) as mes, 
                       SUM(consumo_diario) as consumo, 
                       intervencao_id
                FROM tb_telemetria_intervencao_diaria
                WHERE intervencao_id IN %s 
                AND EXTRACT(YEAR FROM DATE(data)) = %s
                GROUP BY mes, intervencao_id
                ORDER BY mes, intervencao_id;
                """
            else:
                # Query para único medidor
                query = """
                SELECT EXTRACT(MONTH FROM DATE(data)) as mes, 
                       SUM(consumo_diario) as consumo, 
                       MAX(intervencao_id) as id_medidor
                FROM tb_telemetria_intervencao_diaria
                WHERE intervencao_id IN %s 
                AND EXTRACT(YEAR FROM DATE(data)) = %s
                GROUP BY mes
                ORDER BY mes;
                """
            
            tuple_ids = tuple(self.lista_ids_selecionados)
            cursor.execute(query, (tuple_ids, ano))
            dados = cursor.fetchall()
            cursor.close()
            return dados
        except Exception as e:
            print(f"Erro ao buscar dados mensais: {e}")
            return []

    def get_data_diario(self):
        """Busca dados totalizados por dia com suporte a agregação."""
        mes = self.combo_mes_graf.currentIndex() + 1
        ano = int(self.combo_ano_graf.currentText())
        
        try:
            cursor = self.conn.cursor()
            e_multiplo = len(self.lista_ids_selecionados) > 1
            codigos_interf = set([d[5] for d in self.lista_dados_selecionados])
            eh_multipla_interferencia = len(codigos_interf) > 1
            
            if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                # Agrupar por dia E medidor (para posterior agregação por interferência)
                query = """
                SELECT EXTRACT(DAY FROM DATE(data)) as dia, 
                       SUM(consumo_diario) as consumo, 
                       intervencao_id
                FROM tb_telemetria_intervencao_diaria
                WHERE intervencao_id IN %s 
                AND EXTRACT(MONTH FROM DATE(data)) = %s 
                AND EXTRACT(YEAR FROM DATE(data)) = %s
                GROUP BY dia, intervencao_id
                ORDER BY dia, intervencao_id;
                """
            elif eh_multipla_interferencia and self.modo_agregacao == "criterio_busca":
                # Agrupar por dia apenas (soma tudo)
                query = """
                SELECT EXTRACT(DAY FROM DATE(data)) as dia, 
                       SUM(consumo_diario) as consumo, 
                       0 as intervencao_id
                FROM tb_telemetria_intervencao_diaria
                WHERE intervencao_id IN %s 
                AND EXTRACT(MONTH FROM DATE(data)) = %s 
                AND EXTRACT(YEAR FROM DATE(data)) = %s
                GROUP BY dia
                ORDER BY dia;
                """
            elif e_multiplo:
                # Query original para múltiplos mesma interferência
                query = """
                SELECT EXTRACT(DAY FROM DATE(data)) as dia, 
                       SUM(consumo_diario) as consumo, 
                       intervencao_id
                FROM tb_telemetria_intervencao_diaria
                WHERE intervencao_id IN %s 
                AND EXTRACT(MONTH FROM DATE(data)) = %s 
                AND EXTRACT(YEAR FROM DATE(data)) = %s
                GROUP BY dia, intervencao_id
                ORDER BY dia, intervencao_id;
                """
            else:
                # Query para único medidor
                query = """
                SELECT EXTRACT(DAY FROM DATE(data)) as dia, 
                       SUM(consumo_diario) as consumo, 
                       MAX(intervencao_id) as id_medidor
                FROM tb_telemetria_intervencao_diaria
                WHERE intervencao_id IN %s 
                AND EXTRACT(MONTH FROM DATE(data)) = %s 
                AND EXTRACT(YEAR FROM DATE(data)) = %s
                GROUP BY dia
                ORDER BY dia;
                """
            
            tuple_ids = tuple(self.lista_ids_selecionados)
            cursor.execute(query, (tuple_ids, mes, ano))
            dados = cursor.fetchall()
            cursor.close()
            return dados
        except Exception as e:
            print(f"Erro ao buscar dados diários: {e}")
            return []
              
    def get_volumes_outorgados(self, codigo_interferencia):
        """
        Busca volumes mensais outorgados da tabela view_volume_outorgado com base no código de interferência.
        Suporta múltiplas interferências concatenadas (ex: "1606998, 1645972").
        Retorna lista de 12 valores [jan, fev, ..., dez] com a SOMA dos volumes ou None se não encontrar dados.
        """
        if not codigo_interferencia or codigo_interferencia == "Não informado":
            return None
        
        try:
            cursor = self.conn.cursor()
            
            # === IDENTIFICAR SE HÁ MÚLTIPLAS INTERFERÊNCIAS ===
            # Verificar se o código contém vírgulas (múltiplas interferências)
            if ',' in str(codigo_interferencia):
                # Separar os códigos (pode ter espaços após as vírgulas)
                codigos_lista = [cod.strip() for cod in str(codigo_interferencia).split(',')]
            else:
                # Única interferência
                codigos_lista = [str(codigo_interferencia).strip()]
            
            # === BUSCAR VOLUMES PARA CADA INTERFERÊNCIA ===
            volumes_totais = [0.0] * 12  # Inicializar com 12 zeros (jan a dez)
            interferencias_encontradas = 0
            
            for cod_interf in codigos_lista:
                # Pular se vazio ou inválido
                if not cod_interf or cod_interf == "Não informado":
                    continue
                
                query = """
                    SELECT vol_jan, vol_fev, vol_mar, vol_abr, vol_mai, vol_jun,
                           vol_jul, vol_ago, vol_set, vol_out, vol_nov, vol_dez
                    FROM view_volume_outorgado
                    WHERE codigo_interferencia = %s
                    LIMIT 1;
                """
                cursor.execute(query, (cod_interf,))
                resultado = cursor.fetchone()
                
                if resultado:
                    interferencias_encontradas += 1
                    # Somar os volumes desta interferência aos totais
                    for i, vol in enumerate(resultado):
                        if vol is not None:
                            volumes_totais[i] += float(vol)
            
            cursor.close()
            
            # Verificar se encontrou pelo menos uma interferência
            if interferencias_encontradas > 0:
                print(f"[INFO] Volumes outorgados somados de {interferencias_encontradas} interferência(s)")
                return volumes_totais
            else:
                print(f"[INFO] Volume outorgado não encontrado para interferência(s): {codigo_interferencia}")
                return None
                
        except Exception as e:
            print(f"[INFO] Erro ao buscar volume outorgado para interferência {codigo_interferencia}: {e}")
            return None

    def atualizar_grafico_mensal(self):
        """Atualiza gráfico mensal com suporte a agregação por interferência ou critério de busca."""
        if not MATPLOTLIB_DISPONIVEL:
            return
        
        try:
            self.canvas_mensal.figure.clear()
            ax = self.canvas_mensal.figure.add_subplot(111)
            
            dados_brutos = self.get_data_mensal()
            
            if len(self.lista_ids_selecionados) >= 1 and dados_brutos:
                
                # Verificar modo de agregação
                codigos_interf = set([d[5] for d in self.lista_dados_selecionados])
                eh_multipla_interferencia = len(codigos_interf) > 1
                
                # Configurar mapeamento de rótulos conforme modo de agregação
                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    # === MODO POR INTERFERÊNCIA: Criar uma série para cada código de interferência ===
                    
                    # Mapear cada código de interferência para um ID representativo e cor
                    mapa_interferencia_para_id = {}  # {cod_interf: med_id_representativo}
                    mapa_rotulos = {}  # {med_id: "Interf. XXXX"}
                    cores_por_interferencia = {}
                    
                    for i, d in enumerate(self.lista_dados_selecionados):
                        med_id = d[0]
                        cod_interf = d[5]
                        if cod_interf not in mapa_interferencia_para_id:
                            mapa_interferencia_para_id[cod_interf] = med_id
                            mapa_rotulos[med_id] = f"Interf. {cod_interf}"
                            cores_por_interferencia[cod_interf] = med_id
                    
                    # IDs únicos representativos (um por interferência)
                    ids_todos = list(mapa_interferencia_para_id.values())
                    
                    # Agregar dados por interferência e mês
                    dados_por_interferencia = {}  # {cod_interf: {mes: consumo_total}}
                    
                    for row in dados_brutos:
                        mes = int(row[0])
                        consumo = float(row[1]) if row[1] else 0.0
                        med_id = row[2]
                        
                        # Encontrar código de interferência deste medidor
                        cod_interf = None
                        for d in self.lista_dados_selecionados:
                            if d[0] == med_id:
                                cod_interf = d[5]
                                break
                        
                        if cod_interf:
                            if cod_interf not in dados_por_interferencia:
                                dados_por_interferencia[cod_interf] = {}
                            if mes not in dados_por_interferencia[cod_interf]:
                                dados_por_interferencia[cod_interf][mes] = 0.0
                            dados_por_interferencia[cod_interf][mes] += consumo
                    
                    # Reconstruir dados_brutos no formato (mes, consumo, med_id_representativo)
                    dados_brutos = []
                    for cod_interf, meses in dados_por_interferencia.items():
                        med_id_rep = mapa_interferencia_para_id[cod_interf]
                        for mes, consumo in sorted(meses.items()):
                            dados_brutos.append((mes, consumo, med_id_rep))
                    
                    # Flag para indicar que queremos barras empilhadas
                    self.eh_empilhado_mensal = True
                    
                elif eh_multipla_interferencia and self.modo_agregacao == "criterio_busca":
                    # === MODO POR SH/CNARH/USUÁRIO (select all) ou TOTAL (seleção parcial) ===
                    if self.is_selecao_total:
                        # Seleção total: usar o critério de busca como rótulo
                        mapa_rotulos = {0: f"Total {self.criterio_busca}"}
                    else:
                        # Seleção parcial manual: rótulo genérico
                        mapa_rotulos = {0: "Total interferências"}
                    ids_todos = [0]

                    # Agregar todos os dados por mês
                    dados_agregados = {}
                    for row in dados_brutos:
                        mes     = int(row[0])
                        consumo = float(row[1]) if row[1] else 0.0
                        if mes not in dados_agregados:
                            dados_agregados[mes] = 0.0
                        dados_agregados[mes] += consumo

                    dados_brutos = [(mes, consumo_total, 0)
                                    for mes, consumo_total in sorted(dados_agregados.items())]
                    self.eh_empilhado_mensal = False
                    
                else:
                    # === MODO PADRÃO: Comportamento original ===
                    mapa_rotulos = {d[0]: d[1] for d in self.lista_dados_selecionados}
                    ids_todos = self.lista_ids_selecionados
                    self.eh_empilhado_mensal = len(ids_todos) > 1
                
                # Aplicar filtro de medidores/ocultos
                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    # No modo por interferência, ocultar por código de interferência
                    ids_visiveis = []
                    for med_id in ids_todos:
                        # Encontrar código de interferência deste med_id
                        cod_interf = None
                        for d in self.lista_dados_selecionados:
                            if d[0] == med_id:
                                cod_interf = d[5]
                                break
                        if cod_interf and cod_interf not in self.medidores_ocultos_grafico:
                            ids_visiveis.append(med_id)
                else:
                    ids_visiveis = [mid for mid in ids_todos if mid not in self.medidores_ocultos_grafico]
                
                # === BUSCAR VOLUME OUTORGADO ===
                if eh_multipla_interferencia:
                    volumes_outorgados_lista = []
                    for cod_interf in codigos_interf:
                        vol = self.get_volumes_outorgados(cod_interf)
                        if vol:
                            volumes_outorgados_lista.append(vol)
                    
                    if volumes_outorgados_lista:
                        volumes_outorgados_raw = [0.0] * 12
                        for vol in volumes_outorgados_lista:
                            for i in range(12):
                                volumes_outorgados_raw[i] += vol[i]
                    else:
                        volumes_outorgados_raw = None
                else:
                    codigo_interf = self.lista_dados_selecionados[0][5] if self.lista_dados_selecionados else None
                    volumes_outorgados_raw = self.get_volumes_outorgados(codigo_interf)
                
                tem_volume_outorgado = volumes_outorgados_raw is not None
                
                # Verificar se TUDO está oculto
                tudo_oculto = (not ids_visiveis) and (self.volume_outorgado_oculto or not tem_volume_outorgado)
                
                # Coletar todos os meses disponíveis
                todos_meses = sorted(list(set(int(d[0]) for d in dados_brutos)))
                
                # Preparar matriz de dados: dados_matrix[med_id][idx_mes] = valor
                dados_matrix = {}
                for med_id in ids_visiveis:
                    dados_matrix[med_id] = [0.0] * len(todos_meses)
                
                for row in dados_brutos:
                    mes = int(row[0])
                    val = float(row[1]) if row[1] else 0.0
                    med_id = row[2]
                    
                    if mes in todos_meses and med_id in dados_matrix:
                        idx = todos_meses.index(mes)
                        dados_matrix[med_id][idx] = val
                
                # Aplicar filtro de colunas ocultas
                meses_plot = []
                indices_originais_plot = []
                for i, mes in enumerate(todos_meses):
                    idx_absoluto = mes - 1
                    if idx_absoluto not in self.colunas_ocultas_mensal:
                        meses_plot.append(mes)
                        indices_originais_plot.append(i)
                
                if not meses_plot:
                    ax.text(0.5, 0.5, "Sem dados visíveis", ha='center', transform=ax.transAxes)
                    ax.set_xlim(0, 1)
                    ax.set_ylim(0, 1)
                    ax.axis('off')
                    self.canvas_mensal.draw()
                    return
                
                import matplotlib.pyplot as plt
                from matplotlib.patches import Patch
                from matplotlib.lines import Line2D
                cmap = plt.get_cmap('tab10')
                
                # Configurar cores
                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    # Cores por interferência
                    cores_map = {}
                    for i, med_id in enumerate(ids_todos):
                        cores_map[med_id] = cmap(i % 10)
                elif eh_multipla_interferencia and self.modo_agregacao == "criterio_busca":
                    cores_map = {0: cmap(0)}
                else:
                    cores_map = {mid: cmap(i % 10) for i, mid in enumerate(ids_todos)}
                
                bottoms = [0.0] * len(meses_plot)
                self.tooltip_data_mensal = {}
                self.alturas_totais_mensal = [0.0] * len(meses_plot)
                handles_visiveis = []
                labels_visiveis = []
                
                # === PLOTAR BARRAS ===
                for i, med_id in enumerate(ids_visiveis):
                    vals = [dados_matrix[med_id][idx] for idx in indices_originais_plot]
                    cor = cores_map.get(med_id, cmap(0))
                    
                    # === CORREÇÃO: Registrar dados para tooltip EM TODOS OS MODOS ===
                    for pos_idx, val in enumerate(vals):
                        if pos_idx not in self.tooltip_data_mensal:
                            self.tooltip_data_mensal[pos_idx] = {}
                        self.tooltip_data_mensal[pos_idx][med_id] = val
                        self.alturas_totais_mensal[pos_idx] += val
                    
                    label = mapa_rotulos.get(med_id, f"ID {med_id}")
                    
                    # Plotar barras empilhadas (bottoms acumula)
                    ax.bar(range(len(meses_plot)), vals, bottom=bottoms,
                           color=cor, label=label,
                           alpha=0.8, edgecolor='white', linewidth=0.5, picker=True)
                    
                    # Atualizar bottoms para próxima série (empilhamento)
                    bottoms = [b + v for b, v in zip(bottoms, vals)]
                    
                    handles_visiveis.append(ax.patches[-1])
                    labels_visiveis.append(label)
                
                # Plotar volume outorgado
                if tem_volume_outorgado and not self.volume_outorgado_oculto:
                    volumes_plot = []
                    for mes in meses_plot:
                        volumes_plot.append(volumes_outorgados_raw[mes - 1] if mes <= 12 else 0.0)
                    
                    ax.plot(range(len(meses_plot)), volumes_plot,
                           'ro-', linewidth=2, markersize=6, zorder=10)
                    self.volumes_outorgados_mensal = volumes_plot
                    
                    if not ids_visiveis:
                        self.alturas_totais_mensal = volumes_plot.copy()
                else:
                    self.volumes_outorgados_mensal = None
                
                # === CRIAR LEGENDA ===

                # IDs cujo consumo total no período visível é > 0
                ids_com_consumo_mensal = {
                    med_id
                    for med_id, vals in dados_matrix.items()
                    if sum(vals[idx] for idx in indices_originais_plot) > 0
                }

                handles_legenda = []
                labels_legenda  = []

                for med_id in ids_todos:
                    # Ocultar da legenda se consumo == 0 no modo "por interferência"
                    if (eh_multipla_interferencia
                            and self.modo_agregacao == "interferencia"
                            and med_id not in ids_com_consumo_mensal):
                        continue

                    if med_id in ids_visiveis:
                        idx_v = ids_visiveis.index(med_id)
                        handles_legenda.append(handles_visiveis[idx_v])
                        labels_legenda.append(mapa_rotulos.get(med_id, f"ID {med_id}"))
                    else:
                        cor   = cores_map.get(med_id, cmap(0))
                        patch = Patch(facecolor=cor, alpha=0.3)
                        handles_legenda.append(patch)
                        labels_legenda.append(mapa_rotulos.get(med_id, f"ID {med_id}"))

                # Adicionar volume outorgado
                if tem_volume_outorgado:
                    handle_outorgado = Line2D([0], [0], color='red', marker='o',
                                             linestyle='-', linewidth=2, markersize=6)
                    handles_legenda.append(handle_outorgado)
                    labels_legenda.append('Volume Outorgado')

                if handles_legenda:
                    self.legend_mensal = ax.legend(handles_legenda, labels_legenda,
                                                   fontsize=8, loc='upper left')

                    # Estilizar textos
                    ids_na_legenda = [
                        mid for mid in ids_todos
                        if not (eh_multipla_interferencia
                                and self.modo_agregacao == "interferencia"
                                and mid not in ids_com_consumo_mensal)
                    ]
                    for i, text in enumerate(self.legend_mensal.get_texts()):
                        if i < len(ids_na_legenda):
                            med_id_ref  = ids_na_legenda[i]
                            esta_oculto = False
                            if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                                cod_interf_ref = None
                                for d in self.lista_dados_selecionados:
                                    if d[0] == med_id_ref:
                                        cod_interf_ref = d[5]
                                        break
                                esta_oculto = cod_interf_ref in self.medidores_ocultos_grafico
                            else:
                                esta_oculto = med_id_ref in self.medidores_ocultos_grafico

                            if esta_oculto:
                                text.set_color('#999999')
                                text.set_fontstyle('italic')
                            else:
                                text.set_color('black')
                                text.set_fontstyle('normal')
                        elif i == len(ids_na_legenda) and tem_volume_outorgado:
                            if self.volume_outorgado_oculto:
                                text.set_color('#999999')
                                text.set_fontstyle('italic')
                            else:
                                text.set_color('black')
                                text.set_fontstyle('normal')

                    for text in self.legend_mensal.get_texts():
                        text.set_picker(True)
                else:
                    self.legend_mensal = None
                
                # Configuração final
                nomes_meses = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", 
                              "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
                ax.set_xticks(range(len(meses_plot)))
                ax.set_xticklabels([nomes_meses[m - 1] for m in meses_plot])
                ax.set_ylabel("Consumo (m³)")

                # Calcular escala Y automática
                altura_maxima = max(self.alturas_totais_mensal) if self.alturas_totais_mensal else 0
                ymax = self.calcular_escala_y(self.alturas_totais_mensal, 
                                              self.volumes_outorgados_mensal if hasattr(self, 'volumes_outorgados_mensal') else None)

                ax.set_ylim(0, ymax)  # Usar escala calculada em vez de apenas bottom=0

                # Título conforme modo de agregação
                ano_titulo = self.combo_ano_mensal.currentText() if hasattr(self, 'combo_ano_mensal') else self.combo_ano_graf.currentText()

                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    titulo = f"Consumo Mensal por Interferência - {ano_titulo}"
                elif eh_multipla_interferencia and self.modo_agregacao == "criterio_busca":
                    if self.is_selecao_total:
                        titulo = f"Consumo Mensal Total - {self.criterio_busca} - {ano_titulo}"
                    else:
                        titulo = f"Consumo Mensal Total - {ano_titulo}"
                else:
                    titulo = f"Consumo Mensal - {ano_titulo}"

                ax.set_title(titulo, fontsize=10, fontweight='bold', color='#175cc3')
                ax.set_ylim(bottom=0)
                
                if tudo_oculto:
                    ax.text(0.5, 0.5, "Todos os elementos ocultos.\nClique na legenda para restaurar.",
                           ha='center', va='center', transform=ax.transAxes,
                           fontsize=11, color='#6c757d', style='italic',
                           bbox=dict(boxstyle='round', facecolor='#f8f9fa', alpha=0.8))
                
                self.bars_mensal = ax.patches
                self.ax_mensal = ax
                self.x_mensal = meses_plot
                
            else:
                ax.text(0.5, 0.5, "Sem dados", ha='center', va='center', transform=ax.transAxes)
                ax.set_xlim(0, 1)
                ax.set_ylim(0, 1)
                ax.axis('off')
                self.bars_mensal = []
                self.eh_empilhado_mensal = False
                self.volumes_outorgados_mensal = None
            
            self.canvas_mensal.figure.tight_layout()
            self.canvas_mensal.draw()
            
            # Recriar annotation
            self.annot_mensal = ax.annotate("", xy=(0, 0), xytext=(0, 0), textcoords="offset points",
                                           bbox=dict(boxstyle="round,pad=0.5"),
                                           arrowprops=dict(arrowstyle="->"))
            self.annot_mensal.set_visible(False)
            
            # Gerenciar callbacks
            if hasattr(self, 'cid_hover_mensal') and self.cid_hover_mensal is not None:
                self.canvas_mensal.mpl_disconnect(self.cid_hover_mensal)
            self.cid_hover_mensal = self.canvas_mensal.mpl_connect(
                'motion_notify_event', self.on_hover_mensal
            )
            
            if hasattr(self, 'legend_mensal') and self.legend_mensal is not None:
                self.canvas_mensal.mpl_connect('pick_event', self.on_pick_legenda_mensal)
                
        except Exception as e:
            print(f"[ERRO] Falha ao atualizar gráfico mensal: {e}")
            import traceback
            traceback.print_exc()
        
    def on_hover_mensal(self, event):
        """Gerencia tooltip com posicionamento inteligente."""
        if not hasattr(self, 'annot_mensal'): 
            return
        if not hasattr(self, 'ax_mensal') or event.inaxes != self.ax_mensal:
            if self.annot_mensal.get_visible():
                self.annot_mensal.set_visible(False)
                self.canvas_mensal.draw_idle()
            return
        
        found = False
        if not hasattr(self, 'bars_mensal') or not self.bars_mensal:
            return
        
        xlim = self.ax_mensal.get_xlim()
        x_min, x_max = xlim[0], xlim[1]
        
        for i, bar in enumerate(self.bars_mensal):
            contido, _ = bar.contains(event)
            if contido:
                x_pos = bar.get_x() + bar.get_width() / 2
                col_idx = int(round(x_pos))
                
                if col_idx < 0 or col_idx >= len(self.alturas_totais_mensal):
                    continue
                
                altura_total = self.alturas_totais_mensal[col_idx]
                y_pos = altura_total / 2
                
                pos_rel = (x_pos - x_min) / (x_max - x_min)
                offset_x = -70 if pos_rel > 0.85 else (20 if pos_rel < 0.15 else 0)
                
                mes = self.x_mensal[col_idx] if col_idx < len(self.x_mensal) else 1
                nomes_meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                mes_nome = nomes_meses[int(mes)-1] if 1 <= int(mes) <= 12 else f"Mês {int(mes)}"
                
                # === TOOLTIP PARA MODO TOTALIZADO (criterio_busca) ===
                if self.modo_agregacao == "criterio_busca":
                    consumo = altura_total

                    try:
                        txt_val = self.formatar_br(consumo)
                    except:
                        txt_val = f"{consumo:,.2f}"

                    if self.is_selecao_total:
                        # Seleção total: mostrar o critério de busca
                        if self.criterio_busca == "Sistema Hídrico":
                            nome_criterio = self.nome_completo if self.nome_completo else "Sistema Hídrico"
                        elif self.criterio_busca == "CNARH":
                            nome_criterio = f"CNARH: {self.nome_completo if self.nome_completo else self.termo_busca}"
                        else:
                            nome_criterio = f"Usuário: {self.nome_completo if self.nome_completo else self.termo_busca}"
                    else:
                        # Seleção parcial: rótulo genérico
                        nome_criterio = "Total interferências"

                    texto_tooltip = f"{mes_nome}\n{nome_criterio}\n{txt_val} m³"
                    
                    # Adicionar volume outorgado se disponível
                    if (hasattr(self, 'volumes_outorgados_mensal') and 
                        self.volumes_outorgados_mensal and 
                        not self.volume_outorgado_oculto):
                        vol_out = self.volumes_outorgados_mensal[col_idx]
                        if vol_out > 0:
                            try:
                                txt_out = self.formatar_br(vol_out)
                                texto_tooltip += f"\nOutorgado: {txt_out} m³"
                            except:
                                pass
                    
                    self.annot_mensal.set_text(texto_tooltip)
                    self.annot_mensal.set_fontsize(9)
                    found = True
                
                # === TOOLTIP PARA MODO POR INTERFERÊNCIA ===
                elif self.modo_agregacao == "interferencia":
                    if col_idx in self.tooltip_data_mensal:
                        dados_col = self.tooltip_data_mensal[col_idx]
                        total = sum(dados_col.values())
                        
                        try:
                            texto_total = self.formatar_br(total)
                            texto_linhas = f"{mes_nome}\nTotal: {texto_total} m³\n----------------\n"
                            
                            # Mostrar por interferência
                            for med_id, val in dados_col.items():
                                cod_interf = None
                                for d in self.lista_dados_selecionados:
                                    if d[0] == med_id:
                                        cod_interf = d[5]
                                        break
                                
                                rotulo = f"Interf. {cod_interf}" if cod_interf else f"ID {med_id}"
                                texto_val = self.formatar_br(val)
                                texto_linhas += f"{rotulo}: {texto_val} m³\n"
                            
                            # Adicionar volume outorgado
                            if (hasattr(self, 'volumes_outorgados_mensal') and 
                                self.volumes_outorgados_mensal and 
                                not self.volume_outorgado_oculto):
                                vol_out = self.volumes_outorgados_mensal[col_idx]
                                if vol_out > 0:
                                    try:
                                        txt_out = self.formatar_br(vol_out)
                                        texto_linhas += f"\nOutorgado: {txt_out} m³"
                                    except:
                                        pass
                            
                            self.annot_mensal.set_text(texto_linhas)
                            self.annot_mensal.set_fontsize(8)
                            found = True
                        except:
                            pass
                
                # === TOOLTIP PADRÃO (múltiplos medidores mesma interferência) ===
                elif len(self.lista_ids_selecionados) > 1:
                    if col_idx in self.tooltip_data_mensal:
                        dados_col = self.tooltip_data_mensal[col_idx]
                        total = sum(dados_col.values())
                        mapa_rotulos = {d[0]: d[1] for d in self.lista_dados_selecionados}
                        
                        try:
                            texto_total = self.formatar_br(total)
                            texto_linhas = f"{mes_nome}\nTotal: {texto_total} m³\n----------------\n"
                            for med_id, val in dados_col.items():
                                rotulo = mapa_rotulos.get(med_id, f"ID {med_id}")
                                texto_val = self.formatar_br(val)
                                texto_linhas += f"{rotulo}: {texto_val} m³\n"
                            
                            if (hasattr(self, 'volumes_outorgados_mensal') and 
                                self.volumes_outorgados_mensal and 
                                not self.volume_outorgado_oculto):
                                vol_out = self.volumes_outorgados_mensal[col_idx]
                                if vol_out > 0:
                                    try:
                                        txt_out = self.formatar_br(vol_out)
                                        texto_linhas += f"\nOutorgado: {txt_out} m³"
                                    except:
                                        pass
                            
                            self.annot_mensal.set_text(texto_linhas)
                            self.annot_mensal.set_fontsize(8)
                            found = True
                        except:
                            pass
                
                # === TOOLTIP MEDIDOR ÚNICO ===
                else:
                    consumo = altura_total
                    
                    try:
                        txt_val = self.formatar_br(consumo)
                    except:
                        txt_val = f"{consumo:,.2f}"
                    
                    texto_tooltip = f"{mes_nome}\n{txt_val} m³"
                    
                    if (hasattr(self, 'volumes_outorgados_mensal') and 
                        self.volumes_outorgados_mensal and 
                        not self.volume_outorgado_oculto):
                        vol_out = self.volumes_outorgados_mensal[col_idx]
                        if vol_out > 0:
                            try:
                                txt_out = self.formatar_br(vol_out)
                                texto_tooltip += f"\nOutorgado: {txt_out} m³"
                            except:
                                pass
                    
                    self.annot_mensal.set_text(texto_tooltip)
                    self.annot_mensal.set_fontsize(9)
                    found = True
                
                if found:
                    self.annot_mensal.xy = (x_pos, y_pos)
                    self.annot_mensal.set_position((offset_x, 15))
                    
                    bbox = self.annot_mensal.get_bbox_patch()
                    bbox.set_boxstyle("round,pad=0.5")
                    bbox.set_facecolor('#222222')
                    bbox.set_edgecolor('white')
                    bbox.set_linewidth(1)
                    bbox.set_alpha(0.95)
                    self.annot_mensal.set_color('white')
                    if hasattr(self.annot_mensal, 'arrow_patch'):
                        self.annot_mensal.arrow_patch.set_visible(False)
                    self.annot_mensal.set_visible(True)
                    self.canvas_mensal.draw_idle()
                    break
        
        if not found and self.annot_mensal.get_visible():
            self.annot_mensal.set_visible(False)
            self.canvas_mensal.draw_idle()
        
    def atualizar_grafico_diario(self):
        """Atualiza gráfico diário com suporte a agregação por interferência ou critério de busca."""
        if not MATPLOTLIB_DISPONIVEL:
            return
        
        try:
            self.canvas_diario.figure.clear()
            ax = self.canvas_diario.figure.add_subplot(111)
            
            dados_brutos = self.get_data_diario()
            
            if len(self.lista_ids_selecionados) >= 1 and dados_brutos:
                
                # Verificar modo de agregação
                codigos_interf = set([d[5] for d in self.lista_dados_selecionados])
                eh_multipla_interferencia = len(codigos_interf) > 1
                
                # Configurar mapeamento de rótulos conforme modo de agregação
                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    # === MODO POR INTERFERÊNCIA: Criar uma série para cada código de interferência ===
                    
                    # Mapear cada código de interferência para um ID representativo e cor
                    mapa_interferencia_para_id = {}  # {cod_interf: med_id_representativo}
                    mapa_rotulos = {}  # {med_id: "Interf. XXXX"}
                    cores_por_interferencia = {}
                    
                    for i, d in enumerate(self.lista_dados_selecionados):
                        med_id = d[0]
                        cod_interf = d[5]
                        if cod_interf not in mapa_interferencia_para_id:
                            mapa_interferencia_para_id[cod_interf] = med_id
                            mapa_rotulos[med_id] = f"Interf. {cod_interf}"
                            cores_por_interferencia[cod_interf] = med_id
                    
                    # IDs únicos representativos (um por interferência)
                    ids_todos = list(mapa_interferencia_para_id.values())
                    
                    # Agregar dados por interferência e dia
                    dados_por_interferencia = {}  # {cod_interf: {dia: consumo_total}}
                    
                    for row in dados_brutos:
                        dia = int(row[0])
                        consumo = float(row[1]) if row[1] else 0.0
                        med_id = row[2]
                        
                        # Encontrar código de interferência deste medidor
                        cod_interf = None
                        for d in self.lista_dados_selecionados:
                            if d[0] == med_id:
                                cod_interf = d[5]
                                break
                        
                        if cod_interf:
                            if cod_interf not in dados_por_interferencia:
                                dados_por_interferencia[cod_interf] = {}
                            if dia not in dados_por_interferencia[cod_interf]:
                                dados_por_interferencia[cod_interf][dia] = 0.0
                            dados_por_interferencia[cod_interf][dia] += consumo
                    
                    # Reconstruir dados_brutos no formato (dia, consumo, med_id_representativo)
                    dados_brutos = []
                    for cod_interf, dias in dados_por_interferencia.items():
                        med_id_rep = mapa_interferencia_para_id[cod_interf]
                        for dia, consumo in sorted(dias.items()):
                            dados_brutos.append((dia, consumo, med_id_rep))
                    
                    # Flag para indicar que queremos barras empilhadas
                    self.eh_empilhado_diario = True
                    
                elif eh_multipla_interferencia and self.modo_agregacao == "criterio_busca":
                    # === MODO POR SH/CNARH/USUÁRIO (select all) ou TOTAL (seleção parcial) ===
                    if self.is_selecao_total:
                        mapa_rotulos = {0: f"Total {self.criterio_busca}"}
                    else:
                        mapa_rotulos = {0: "Total interferências"}
                    ids_todos = [0]

                    # Agregar todos os dados por dia
                    dados_agregados = {}
                    for row in dados_brutos:
                        dia     = int(row[0])
                        consumo = float(row[1]) if row[1] else 0.0
                        if dia not in dados_agregados:
                            dados_agregados[dia] = 0.0
                        dados_agregados[dia] += consumo

                    dados_brutos = [(dia, consumo_total, 0)
                                    for dia, consumo_total in sorted(dados_agregados.items())]
                    self.eh_empilhado_diario = False  # Uma única série, não precisa empilhar
                    
                else:
                    # === MODO PADRÃO: Comportamento original ===
                    mapa_rotulos = {d[0]: d[1] for d in self.lista_dados_selecionados}
                    ids_todos = self.lista_ids_selecionados
                    self.eh_empilhado_diario = len(ids_todos) > 1
                
                # Aplicar filtro de medidores/ocultos
                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    # No modo por interferência, ocultar por código de interferência
                    ids_visiveis = []
                    for med_id in ids_todos:
                        # Encontrar código de interferência deste med_id
                        cod_interf = None
                        for d in self.lista_dados_selecionados:
                            if d[0] == med_id:
                                cod_interf = d[5]
                                break
                        if cod_interf and cod_interf not in self.medidores_ocultos_grafico:
                            ids_visiveis.append(med_id)
                else:
                    ids_visiveis = [mid for mid in ids_todos if mid not in self.medidores_ocultos_grafico]
                
                # Verificar se TUDO está oculto
                tudo_oculto = not ids_visiveis
                
                # Coletar todos os dias disponíveis
                todos_dias = sorted(list(set(int(d[0]) for d in dados_brutos)))
                
                # Preparar matriz de dados: dados_matrix[med_id][idx_dia] = valor
                dados_matrix = {}
                for med_id in ids_visiveis:
                    dados_matrix[med_id] = [0.0] * len(todos_dias)
                
                for row in dados_brutos:
                    dia = int(row[0])
                    val = float(row[1]) if row[1] else 0.0
                    med_id = row[2]
                    
                    if dia in todos_dias and med_id in dados_matrix:
                        idx = todos_dias.index(dia)
                        dados_matrix[med_id][idx] = val
                
                # Aplicar filtro de colunas ocultas
                dias_plot = []
                indices_originais_plot = []
                for i, dia in enumerate(todos_dias):
                    if i not in self.colunas_ocultas_diario:
                        dias_plot.append(dia)
                        indices_originais_plot.append(i)
                
                if not dias_plot:
                    ax.text(0.5, 0.5, "Sem dados", ha='center', transform=ax.transAxes)
                    ax.set_xlim(0, 1)
                    ax.set_ylim(0, 1)
                    ax.axis('off')
                    self.canvas_diario.draw()
                    return
                
                import matplotlib.pyplot as plt
                from matplotlib.patches import Patch
                cmap = plt.get_cmap('tab10')
                
                # Configurar cores
                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    # Cores por interferência
                    cores_map = {}
                    for i, med_id in enumerate(ids_todos):
                        cores_map[med_id] = cmap(i % 10)
                elif eh_multipla_interferencia and self.modo_agregacao == "criterio_busca":
                    cores_map = {0: cmap(0)}
                else:
                    cores_map = {mid: cmap(i % 10) for i, mid in enumerate(ids_todos)}
                
                bottoms = [0.0] * len(dias_plot)
                self.tooltip_data_diario = {}
                self.alturas_totais_diario = [0.0] * len(dias_plot)
                handles_visiveis = []
                labels_visiveis = []
                
                # === PLOTAR BARRAS ===
                for i, med_id in enumerate(ids_visiveis):
                    vals = [dados_matrix[med_id][idx] for idx in indices_originais_plot]
                    cor = cores_map.get(med_id, cmap(0))
                    
                    # Registrar dados para tooltip
                    for pos_idx, val in enumerate(vals):
                        if pos_idx not in self.tooltip_data_diario:
                            self.tooltip_data_diario[pos_idx] = {}
                        self.tooltip_data_diario[pos_idx][med_id] = val
                        self.alturas_totais_diario[pos_idx] += val
                    
                    label = mapa_rotulos.get(med_id, f"ID {med_id}")
                    
                    # Plotar barras empilhadas (bottoms acumula)
                    ax.bar(range(len(dias_plot)), vals, bottom=bottoms,
                           color=cor, label=label,
                           alpha=0.8, edgecolor='white', linewidth=0.5, picker=True)
                    
                    # Atualizar bottoms para próxima série (empilhamento)
                    bottoms = [b + v for b, v in zip(bottoms, vals)]
                    
                    handles_visiveis.append(ax.patches[-1])
                    labels_visiveis.append(label)
                
                # === CRIAR LEGENDA UNIFICADA ===

                # IDs cujo consumo total no período visível é > 0
                ids_com_consumo_diario = {
                    med_id
                    for med_id, vals in dados_matrix.items()
                    if sum(vals[idx] for idx in indices_originais_plot) > 0
                }

                handles_legenda = []
                labels_legenda  = []

                for med_id in ids_todos:
                    # Ocultar da legenda se consumo == 0 no modo "por interferência"
                    if (eh_multipla_interferencia
                            and self.modo_agregacao == "interferencia"
                            and med_id not in ids_com_consumo_diario):
                        continue

                    if med_id in ids_visiveis:
                        idx_v = ids_visiveis.index(med_id)
                        handles_legenda.append(handles_visiveis[idx_v])
                        labels_legenda.append(mapa_rotulos.get(med_id, f"ID {med_id}"))
                    else:
                        cor   = cores_map.get(med_id, cmap(0))
                        patch = Patch(facecolor=cor, alpha=0.3)
                        handles_legenda.append(patch)
                        labels_legenda.append(mapa_rotulos.get(med_id, f"ID {med_id}"))

                if handles_legenda:
                    self.legend_diario = ax.legend(handles_legenda, labels_legenda,
                                                   fontsize=7, loc='upper left')

                    # Estilizar textos
                    ids_na_legenda = [
                        mid for mid in ids_todos
                        if not (eh_multipla_interferencia
                                and self.modo_agregacao == "interferencia"
                                and mid not in ids_com_consumo_diario)
                    ]
                    for i, text in enumerate(self.legend_diario.get_texts()):
                        if i >= len(ids_na_legenda):
                            break
                        med_id_ref  = ids_na_legenda[i]
                        esta_oculto = False
                        if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                            cod_interf_ref = None
                            for d in self.lista_dados_selecionados:
                                if d[0] == med_id_ref:
                                    cod_interf_ref = d[5]
                                    break
                            esta_oculto = cod_interf_ref in self.medidores_ocultos_grafico
                        else:
                            esta_oculto = med_id_ref in self.medidores_ocultos_grafico

                        if esta_oculto:
                            text.set_color('#999999')
                            text.set_fontstyle('italic')
                        else:
                            text.set_color('black')
                            text.set_fontstyle('normal')

                    for text in self.legend_diario.get_texts():
                        text.set_picker(True)
                else:
                    self.legend_diario = None
                
                # === CONFIGURAÇÃO FINAL DO GRÁFICO ===
                ax.set_xticks(range(len(dias_plot)))
                ax.set_xticklabels([str(d) for d in dias_plot], rotation=45, ha='right', fontsize=8)
                ax.set_ylabel("Consumo (m³)")

                # Calcular escala Y automática
                ymax = self.calcular_escala_y(self.alturas_totais_diario, None)  # Diário não tem volume outorgado plotado

                ax.set_ylim(0, ymax)

                mes_selecionado = self.combo_mes_graf.currentText()
                ano_selecionado = self.combo_ano_graf.currentText()

                # Título conforme modo de agregação
                if eh_multipla_interferencia and self.modo_agregacao == "interferencia":
                    titulo = f"Consumo Diário por Interferência - {mes_selecionado} {ano_selecionado}"
                elif eh_multipla_interferencia and self.modo_agregacao == "criterio_busca":
                    if self.is_selecao_total:
                        titulo = f"Consumo Diário Total - {self.criterio_busca} - {mes_selecionado} {ano_selecionado}"
                    else:
                        titulo = f"Consumo Diário Total - {mes_selecionado}/{ano_selecionado}"
                else:
                    titulo = f"Consumo Diário - {mes_selecionado} {ano_selecionado}"

                ax.set_title(titulo, fontsize=10, fontweight='bold', color='#175cc3')
                ax.set_ylim(bottom=0)
                
                # Mensagem informativa quando tudo oculto
                if tudo_oculto:
                    ax.text(0.5, 0.5, "Todos os medidores ocultos.\nClique na legenda para restaurar.",
                           ha='center', va='center', transform=ax.transAxes,
                           fontsize=11, color='#6c757d', style='italic',
                           bbox=dict(boxstyle='round', facecolor='#f8f9fa', alpha=0.8))
                
                self.bars_diario = ax.patches
                self.ax_diario = ax
                self.x_diario = dias_plot
                
            else:
                ax.text(0.5, 0.5, "Sem dados", ha='center', va='center', transform=ax.transAxes)
                ax.set_xlim(0, 1)
                ax.set_ylim(0, 1)
                ax.axis('off')
                self.bars_diario = []
                self.eh_empilhado_diario = False
            
            self.canvas_diario.figure.tight_layout()
            self.canvas_diario.draw()
            
            # === RECRIAR ANNOTATION ===
            self.annot_diario = ax.annotate("", xy=(0, 0), xytext=(0, 0), textcoords="offset points",
                                           bbox=dict(boxstyle="round,pad=0.5"),
                                           arrowprops=dict(arrowstyle="->"))
            self.annot_diario.set_visible(False)
            
            # === GERENCIAR CALLBACKS ===
            if hasattr(self, 'cid_hover_diario') and self.cid_hover_diario is not None:
                self.canvas_diario.mpl_disconnect(self.cid_hover_diario)
            self.cid_hover_diario = self.canvas_diario.mpl_connect(
                'motion_notify_event', self.on_hover_diario
            )
            
            if hasattr(self, 'legend_diario') and self.legend_diario is not None:
                self.canvas_diario.mpl_connect('pick_event', self.on_pick_legenda_diario)
                
        except Exception as e:
            print(f"[ERRO] Falha ao atualizar gráfico diário: {e}")
            import traceback
            traceback.print_exc()
            
    def on_hover_diario(self, event):
        """Gerencia tooltip para gráfico diário."""
        if not hasattr(self, 'annot_diario'):
            return
        if not hasattr(self, 'ax_diario') or event.inaxes != self.ax_diario:
            if self.annot_diario.get_visible():
                self.annot_diario.set_visible(False)
                self.canvas_diario.draw_idle()
            return
        
        found = False
        if not hasattr(self, 'bars_diario') or not self.bars_diario:
            return
        
        xlim = self.ax_diario.get_xlim()
        x_min, x_max = xlim[0], xlim[1]
        
        for i, bar in enumerate(self.bars_diario):
            contido, _ = bar.contains(event)
            if contido:
                x_pos = bar.get_x() + bar.get_width() / 2
                col_idx = int(round(x_pos))
                
                if col_idx < 0 or col_idx >= len(self.alturas_totais_diario):
                    continue
                
                altura_total = self.alturas_totais_diario[col_idx]
                y_pos = altura_total / 2
                
                pos_rel = (x_pos - x_min) / (x_max - x_min)
                offset_x = -130 if pos_rel > 0.85 else (20 if pos_rel < 0.15 else 0)
                
                dia = self.x_diario[col_idx] if col_idx < len(self.x_diario) else 1
                
                # === TOOLTIP PARA MODO TOTALIZADO ===
                if self.modo_agregacao == "criterio_busca":
                    consumo = altura_total

                    try:
                        txt_val = self.formatar_br(consumo)
                    except:
                        txt_val = f"{consumo:,.2f}"

                    if self.is_selecao_total:
                        if self.criterio_busca == "Sistema Hídrico":
                            nome_criterio = self.nome_completo if self.nome_completo else "Sistema Hídrico"
                        elif self.criterio_busca == "CNARH":
                            nome_criterio = f"CNARH: {self.nome_completo if self.nome_completo else self.termo_busca}"
                        else:
                            nome_criterio = f"Usuário: {self.nome_completo if self.nome_completo else self.termo_busca}"
                    else:
                        nome_criterio = "Total interferências"

                    texto_tooltip = f"Dia {int(dia)}\n{nome_criterio}\n{txt_val} m³"
                    
                    self.annot_diario.set_text(texto_tooltip)
                    self.annot_diario.set_fontsize(9)
                    found = True
                
                # === TOOLTIP PARA MODO POR INTERFERÊNCIA ===
                elif self.modo_agregacao == "interferencia":
                    if col_idx in self.tooltip_data_diario:
                        dados_col = self.tooltip_data_diario[col_idx]
                        total = sum(dados_col.values())
                        
                        try:
                            texto_total = self.formatar_br(total)
                            texto_linhas = f"Dia {int(dia)}\nTotal: {texto_total} m³\n----------------\n"
                            
                            for med_id, val in dados_col.items():
                                cod_interf = None
                                for d in self.lista_dados_selecionados:
                                    if d[0] == med_id:
                                        cod_interf = d[5]
                                        break
                                
                                rotulo = f"Interf. {cod_interf}" if cod_interf else f"ID {med_id}"
                                texto_val = self.formatar_br(val)
                                texto_linhas += f"{rotulo}: {texto_val} m³\n"
                            
                            self.annot_diario.set_text(texto_linhas)
                            self.annot_diario.set_fontsize(8)
                            found = True
                        except:
                            pass
                
                # === TOOLTIP PADRÃO ===
                elif len(self.lista_ids_selecionados) > 1:
                    if col_idx in self.tooltip_data_diario:
                        dados_col = self.tooltip_data_diario[col_idx]
                        total = sum(dados_col.values())
                        mapa_rotulos = {d[0]: d[1] for d in self.lista_dados_selecionados}
                        
                        try:
                            texto_total = self.formatar_br(total)
                            texto_linhas = f"Dia {int(dia)}\nTotal: {texto_total} m³\n----------------\n"
                            for med_id, val in dados_col.items():
                                rotulo = mapa_rotulos.get(med_id, f"ID {med_id}")
                                texto_val = self.formatar_br(val)
                                texto_linhas += f"{rotulo}: {texto_val} m³\n"
                            
                            self.annot_diario.set_text(texto_linhas)
                            self.annot_diario.set_fontsize(8)
                            found = True
                        except:
                            pass
                
                # === MEDIDOR ÚNICO ===
                else:
                    consumo = altura_total
                    
                    try:
                        txt_val = self.formatar_br(consumo)
                    except:
                        txt_val = f"{consumo:,.2f}"
                    
                    self.annot_diario.set_text(f"Dia {int(dia)}\n{txt_val} m³")
                    self.annot_diario.set_fontsize(9)
                    found = True
                
                if found:
                    self.annot_diario.xy = (x_pos, y_pos)
                    self.annot_diario.set_position((offset_x, 15))
                    
                    bbox = self.annot_diario.get_bbox_patch()
                    bbox.set_boxstyle("round,pad=0.5")
                    bbox.set_facecolor('#222222')
                    bbox.set_edgecolor('white')
                    bbox.set_linewidth(1)
                    bbox.set_alpha(0.95)
                    self.annot_diario.set_color('white')
                    if hasattr(self.annot_diario, 'arrow_patch'):
                        self.annot_diario.arrow_patch.set_visible(False)
                    self.annot_diario.set_visible(True)
                    self.canvas_diario.draw_idle()
                    break
        
        if not found and self.annot_diario.get_visible():
            self.annot_diario.set_visible(False)
            self.canvas_diario.draw_idle()
        
    def get_id_from_label(self, label_text, ids_todos, mapa_rotulos):
        """Helper para encontrar o ID de um medidor pelo texto da legenda."""
        # Inverso do mapa de rotulos
        mapa_rotulos_inverso = {v: k for k, v in mapa_rotulos.items()}
        # Tenta achar pelo label direto ou formatado
        if label_text in mapa_rotulos_inverso:
            return mapa_rotulos_inverso[label_text]
        return ids_todos[0] # Fallback   
        
    def on_click_mensal(self, event):
        """Detecta clique em coluna do gráfico mensal e abre calendário."""
        if not hasattr(self, 'bars_mensal') or not self.bars_mensal:
            return
        if not hasattr(self, 'ax_mensal') or event.inaxes != self.ax_mensal:
            return
        
        for bar in self.bars_mensal:
            contido, _ = bar.contains(event)
            if contido:
                # Calcula o índice da coluna clicada (0, 1, 2...)
                x_pos = bar.get_x() + bar.get_width() / 2
                index = int(round(x_pos))
                
                # Obtém o mês real da lista self.x_mensal
                if 0 <= index < len(self.x_mensal):
                    mes_clicado = int(self.x_mensal[index])
                    ano_selecionado = int(self.combo_ano_mensal.currentText())
                    self.abrir_janela_detalhes(mes=mes_clicado, ano=ano_selecionado)
                    break

    def on_click_diario(self, event):
        """Detecta clique em coluna do gráfico diário e abre dados de 15min."""
        if not hasattr(self, 'bars_diario') or not self.bars_diario:
            return
        if not hasattr(self, 'ax_diario') or event.inaxes != self.ax_diario:
            return
        
        for bar in self.bars_diario:
            contido, _ = bar.contains(event)
            if contido:
                x_pos = bar.get_x() + bar.get_width() / 2
                index = int(round(x_pos))
                
                if 0 <= index < len(self.x_diario):
                    dia_clicado = int(self.x_diario[index])
                    mes_selecionado = self.combo_mes_graf.currentIndex() + 1
                    ano_selecionado = int(self.combo_ano_graf.currentText())
                    self.abrir_janela_detalhes(dia=dia_clicado, mes=mes_selecionado, ano=ano_selecionado)
                    break
                    
    def get_data_mensal_filtrado(self):
        """Retorna dados mensais filtrados (com colunas ocultas removidas)."""
        dados = self.get_data_mensal()
        
        if not dados:
            return []
        
        # Filtrar colunas ocultas
        dados_filtrados = []
        for i, d in enumerate(dados):
            if i not in self.colunas_ocultas_mensal:
                dados_filtrados.append(d)
        
        return dados_filtrados
    
    def get_data_diario_filtrado(self):
        """Retorna dados diários filtrados (com colunas ocultas removidas)."""
        dados = self.get_data_diario()
        
        if not dados:
            return []
        
        # Filtrar colunas ocultas
        dados_filtrados = []
        for i, d in enumerate(dados):
            if i not in self.colunas_ocultas_diario:
                dados_filtrados.append(d)
        
        return dados_filtrados

    def on_pick_legenda_mensal(self, event):
        """Clique na legenda do gráfico mensal."""
        if event.mouseevent.button != 1:
            return
        if not hasattr(self, 'legend_mensal') or self.legend_mensal is None:
            return
        if event.artist not in self.legend_mensal.get_texts():
            return

        texto_clicado = event.artist.get_text()

        # ── sub-aba "Total": só Volume Outorgado é clicável
        if self.modo_agregacao == "criterio_busca":
            if texto_clicado == "Volume Outorgado":
                self.volume_outorgado_oculto = not self.volume_outorgado_oculto
                self.atualizar_grafico_mensal()
            return

        # ── Volume Outorgado (qualquer outro modo)
        if texto_clicado == "Volume Outorgado":
            self.volume_outorgado_oculto = not self.volume_outorgado_oculto
            self.atualizar_grafico_mensal()
            return

        # ── sub-aba "Por Interferência"
        if self.modo_agregacao == "interferencia":
            # A legenda lista UMA entrada por interferência na ordem de codigos_interf.
            # Recuperar o cod_interf pelo índice do texto clicado na legenda.
            texts           = self.legend_mensal.get_texts()
            # excluir o último item se for "Volume Outorgado"
            texts_dados     = [t for t in texts if t.get_text() != "Volume Outorgado"]
            if event.artist not in texts_dados:
                return
            index           = texts_dados.index(event.artist)
            # ids_todos foi construído na ordem de mapa_interferencia_para_id,
            # que segue a ordem de iteração de lista_dados_selecionados.
            # Reconstruir a mesma lista de cod_interf ordenada.
            codigos_ordenados = []
            vistos            = set()
            for d in self.lista_dados_selecionados:
                cod = d[5]
                if cod not in vistos:
                    codigos_ordenados.append(cod)
                    vistos.add(cod)

            if index >= len(codigos_ordenados):
                return
            cod_alvo = codigos_ordenados[index]

            if cod_alvo in self.medidores_ocultos_grafico:
                self.medidores_ocultos_grafico.remove(cod_alvo)
            else:
                self.medidores_ocultos_grafico.add(cod_alvo)

            self.atualizar_grafico_mensal()
            return

        # ── modo padrão (mesma interferência / único medidor)
        if len(self.lista_ids_selecionados) > 1:
            texts = self.legend_mensal.get_texts()
            texts_dados = [t for t in texts if t.get_text() != "Volume Outorgado"]
            if event.artist not in texts_dados:
                return
            index = texts_dados.index(event.artist)
            if index < len(self.lista_ids_selecionados):
                med_id = self.lista_ids_selecionados[index]
                if med_id in self.medidores_ocultos_grafico:
                    self.medidores_ocultos_grafico.remove(med_id)
                else:
                    self.medidores_ocultos_grafico.add(med_id)
            self.atualizar_grafico_mensal()

    def on_pick_legenda_diario(self, event):
        """Clique na legenda do gráfico diário."""
        if event.mouseevent.button != 1:
            return
        if not hasattr(self, 'legend_diario') or self.legend_diario is None:
            return
        if event.artist not in self.legend_diario.get_texts():
            return

        texto_clicado = event.artist.get_text()

        # ── sub-aba "Total": nada é clicável
        if self.modo_agregacao == "criterio_busca":
            return

        # ── sub-aba "Por Interferência"
        if self.modo_agregacao == "interferencia":
            texts       = self.legend_diario.get_texts()
            if event.artist not in texts:
                return
            index       = texts.index(event.artist)

            codigos_ordenados = []
            vistos            = set()
            for d in self.lista_dados_selecionados:
                cod = d[5]
                if cod not in vistos:
                    codigos_ordenados.append(cod)
                    vistos.add(cod)

            if index >= len(codigos_ordenados):
                return
            cod_alvo = codigos_ordenados[index]

            if cod_alvo in self.medidores_ocultos_grafico:
                self.medidores_ocultos_grafico.remove(cod_alvo)
            else:
                self.medidores_ocultos_grafico.add(cod_alvo)

            self.atualizar_grafico_diario()
            return

        # ── modo padrão (mesma interferência / único medidor)
        texts = self.legend_diario.get_texts()
        if event.artist not in texts:
            return
        index = texts.index(event.artist)
        if index < len(self.lista_ids_selecionados):
            med_id = self.lista_ids_selecionados[index]
            if med_id in self.medidores_ocultos_grafico:
                self.medidores_ocultos_grafico.remove(med_id)
            else:
                self.medidores_ocultos_grafico.add(med_id)
        self.atualizar_grafico_diario()
        
    def abrir_janela_detalhes(self, dia=None, mes=None, ano=None):
        """Abre a janela de detalhes configurada para a data específica."""
        QApplication.setOverrideCursor(Qt.WaitCursor)

        try:
            self.janela_detalhes = JanelaMonitoramentoDetalhes(
                self,
                self.conn,
                self.lista_ids_selecionados,
                self.lista_dados_selecionados,
                self.usuario_logado,
                criterio_busca = self.criterio_busca,
                nome_completo = self.nome_completo,
                eh_multipla_interferencia = self.eh_multipla_interferencia,
                codigos_interf = self.codigos_interf,
                is_selecao_total = self.is_selecao_total,
            )

            if dia is not None:
                data = QDate(ano, mes, dia)
                self.janela_detalhes.date_edit.setDate(data)
                self.janela_detalhes.combo_ano.setCurrentText(str(ano))
                self.janela_detalhes.combo_mes.setCurrentIndex(mes - 1)
                self.janela_detalhes.tabs_monitoramento.setCurrentIndex(1)
                self.janela_detalhes.atualizar_dados_15min()
            else:
                self.janela_detalhes.combo_mes.setCurrentIndex(mes - 1)
                self.janela_detalhes.combo_ano.setCurrentText(str(ano))
                self.janela_detalhes.tabs_monitoramento.setCurrentIndex(0)
                self.janela_detalhes.atualizar_calendario()

            self.janela_detalhes.show()
            self._janelas_abertas.append(self.janela_detalhes)            

        finally:
            QApplication.restoreOverrideCursor()

    def exportar_grafico_png(self, tipo=None):
            """
            Exporta o gráfico atual (mensal ou diário) como imagem PNG.
            Salva automaticamente na pasta Downloads do usuário.

            Nomenclatura:
            ─ Única interferência (comportamento original):
                {cnarh}_{interferencia}_{rotulos}_{consumo_mensal|consumo_diario}_{YYYY|YYYYmmm}.png

            ─ Múltiplas interferências, sub-aba "Total":
                TOTAL_{valor_busca}_{consumo_mensal|consumo_diario}_{YYYY|YYYYmmm}.png

            ─ Múltiplas interferências, sub-aba "Por interferência":
                Interferencias_{valor_busca}_{consumo_mensal|consumo_diario}_{YYYY|YYYYmmm}.png
            """
            if not MATPLOTLIB_DISPONIVEL:
                QMessageBox.warning(self, "Aviso",
                    "Matplotlib não está disponível. Não é possível exportar o gráfico.")
                return

            meses_abrev = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
                           "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

            # ── Determinar tipo (mensal / diário) ────────────────────────────────
            if tipo is None:
                tipo = "mensal" if self.tabs_graficos.currentIndex() == 0 else "diario"

            if tipo == "mensal":
                canvas = self.canvas_mensal
                ano    = self.combo_ano_mensal.currentText()
                sufixo_data = ano                                      # YYYY
                sufixo_tipo = "consumo_mensal"
            else:
                canvas = self.canvas_diario
                mes_idx     = self.combo_mes_graf.currentIndex()
                ano         = self.combo_ano_graf.currentText()
                sufixo_data = f"{ano}{meses_abrev[mes_idx]}"          # YYYYmmm
                sufixo_tipo = "consumo_diario"

            if not hasattr(canvas, 'figure') or canvas.figure is None:
                QMessageBox.warning(self, "Aviso",
                    f"Não há gráfico de consumo {tipo} disponível para exportação.")
                return

            # ── Helper de sanitização ────────────────────────────────────────────
            def _san(txt, maxlen=50):
                return "".join(
                    c if c.isalnum() or c in ('-', '_') else '_'
                    for c in str(txt)
                )[:maxlen]

            # ── Montar nome do arquivo ───────────────────────────────────────────
            codigos_interf       = set(d[5] for d in self.lista_dados_selecionados)
            eh_multipla_interf   = len(codigos_interf) > 1

            if eh_multipla_interf:
                # Valor do termo de busca (nome completo quando disponível)
                valor_busca = _san(self.nome_completo or self.termo_busca or "sem_valor", 60)

                # Prefixo conforme sub-aba ativa
                if self.modo_agregacao == "interferencia":
                    prefixo = "Interferencias"
                else:                                   # "criterio_busca" → Total
                    prefixo = "TOTAL"

                nome_arquivo = f"{prefixo}_{valor_busca}_{sufixo_tipo}_{sufixo_data}.png"

            else:
                # ── Nomenclatura original (única interferência) ──────────────────
                mapa_rotulos  = {d[0]: d[1] for d in self.lista_dados_selecionados}
                ids_medidores = self.lista_ids_selecionados

                primeiro_medidor      = self.lista_dados_selecionados[0]
                cnarh                 = primeiro_medidor[4]
                codigo_interferencia  = primeiro_medidor[5]

                if len(ids_medidores) == 1:
                    rotulos_nome = _san(
                        mapa_rotulos.get(ids_medidores[0], f"ID_{ids_medidores[0]}"), 30
                    )
                else:
                    rotulos_lista      = [mapa_rotulos.get(m, f"ID_{m}") for m in ids_medidores]
                    rotulos_sanitizados = [_san(r) for r in rotulos_lista]
                    rotulos_nome       = "_".join(rotulos_sanitizados)[:50]

                cnarh_limpo          = _san(cnarh, 15)
                interferencia_limpa  = _san(codigo_interferencia, 20)
                nome_arquivo = (
                    f"{cnarh_limpo}_{interferencia_limpa}_{rotulos_nome}"
                    f"_{sufixo_tipo}_{sufixo_data}.png"
                )

            # ── Caminho de destino ───────────────────────────────────────────────
            import os, sys
            downloads_path = (
                os.path.join(os.environ['USERPROFILE'], 'Downloads')
                if sys.platform == 'win32'
                else os.path.join(os.path.expanduser('~'), 'Downloads')
            )
            caminho_completo = os.path.join(downloads_path, nome_arquivo)

            # ── Verificar sobreescrita ───────────────────────────────────────────
            if os.path.exists(caminho_completo):
                resposta = QMessageBox.question(
                    self, "Arquivo Existente",
                    f"O arquivo '{nome_arquivo}' já existe na pasta Downloads.\nDeseja substituir?",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if resposta == QMessageBox.No:
                    QMessageBox.information(self, "Cancelado", "Exportação cancelada pelo usuário.")
                    return

            # ── Salvar ───────────────────────────────────────────────────────────
            try:
                canvas.figure.savefig(
                    caminho_completo,
                    dpi=300,
                    bbox_inches='tight',
                    facecolor='white',
                    edgecolor='none'
                )
                QMessageBox.information(
                    self, "Gráfico Exportado",
                    f"Gráfico salvo com sucesso!\n\n"
                    f"Arquivo: {nome_arquivo}\n"
                    f"Local: {downloads_path}\n"
                    f"Resolução: 300 DPI"
                )
            except Exception as e:
                QMessageBox.critical(
                    self, "Erro na Exportação",
                    f"Não foi possível salvar o gráfico:\n\n{str(e)}"
                )
                import traceback
                traceback.print_exc()
        
    def exportar_excel(self):
        """
        Ponto de entrada do botão 'Exportar Excel'.

        • Caso simples (mesma interferência ou único medidor):
            comportamento original – gera um único arquivo e exibe mensagem de sucesso.
        • Caso múltiplas interferências:
            exibe diálogo de seleção com checkboxes e depois gera
            um arquivo por item marcado.
        """
        codigos_interf       = set(d[5] for d in self.lista_dados_selecionados)
        eh_multipla_interf   = len(codigos_interf) > 1

        if not eh_multipla_interf:
            # ── comportamento original ────────────────────────────────────────────
            caminho = self._exportar_excel_por_interferencia(
                ids_medidores  = self.lista_ids_selecionados,
                dados_medidores= self.lista_dados_selecionados,
            )

            # Exibir mensagem de sucesso somente se o arquivo foi gerado
            # (retorno None indica que o usuário cancelou a substituição)
            if caminho:
                import os
                downloads_path = os.path.dirname(caminho)
                nome_arquivo   = os.path.basename(caminho)
                QMessageBox.information(
                    self,
                    "Relatório Exportado",
                    f"Relatório salvo com sucesso!\n\n"
                    f"Arquivo: {nome_arquivo}\n"
                    f"Local: {downloads_path}"
                )
            return

        # ── diálogo de seleção para múltiplas interferências ─────────────────────
        dialog = QDialog(self)
        dialog.setWindowTitle("Exportar Excel – Escolha o que exportar")
        dialog.setModal(True)
        dialog.setMinimumWidth(480)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        lbl_info = QLabel(
            "Selecione os dados que deseja exportar.\n"
            "Será gerado um arquivo Excel para cada item marcado."
        )
        lbl_info.setWordWrap(True)
        lbl_info.setStyleSheet("font-size: 12px; color: #495057;")
        layout.addWidget(lbl_info)

        frame_checks = QFrame()
        frame_checks.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 6px;
                padding: 10px;
            }
        """)
        frame_layout = QVBoxLayout(frame_checks)
        frame_layout.setSpacing(8)

        checkboxes = []

        if self.criterio_busca == "Sistema Hídrico":
            label_total = f"🌊 Sistema Hídrico: {self.nome_completo or self.termo_busca}"
        elif self.criterio_busca == "CNARH":
            label_total = f"📋 CNARH: {self.nome_completo or self.termo_busca}"
        else:
            label_total = f"👤 Usuário: {self.nome_completo or self.termo_busca}"

        label_total += "  (dados totalizados)"

        cb_total = QCheckBox(label_total)
        cb_total.setChecked(True)
        cb_total.setStyleSheet("font-size: 11px; font-weight: bold; color: #175cc3;")
        frame_layout.addWidget(cb_total)
        checkboxes.append((cb_total, "total", None))

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #dee2e6;")
        frame_layout.addWidget(sep)

        grupos = {}
        for d in self.lista_dados_selecionados:
            cod = d[5]
            grupos.setdefault(cod, []).append(d)

        for cod_interf, med_list in sorted(grupos.items()):
            rotulos = ", ".join(d[1] for d in med_list)
            cb_text = f"Interferência {cod_interf}  –  {rotulos}"
            cb = QCheckBox(cb_text)
            cb.setChecked(True)
            cb.setStyleSheet("font-size: 11px; color: #343a40;")
            frame_layout.addWidget(cb)
            checkboxes.append((cb, "interferencia", cod_interf))

        layout.addWidget(frame_checks)

        btn_sel_todos = QPushButton("Desmarcar todos")
        btn_sel_todos.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #175cc3;
                border: 1px solid #175cc3;
                border-radius: 5px;
                padding: 6px 14px;
                font-size: 11px;
            }
            QPushButton:hover { background-color: #e9f0fb; }
        """)

        def _toggle_todos():
            todos_marcados = all(cb.isChecked() for cb, _, __ in checkboxes)
            for cb, _, __ in checkboxes:
                cb.setChecked(not todos_marcados)
            btn_sel_todos.setText(
                "☐  Desmarcar todos" if todos_marcados else "Selecionar todos")

        btn_sel_todos.clicked.connect(_toggle_todos)
        layout.addWidget(btn_sel_todos, alignment=Qt.AlignLeft)

        btn_row = QHBoxLayout()
        btn_row.addStretch()

        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #6c757d; color: white;
                border-radius: 6px; padding: 8px 20px; font-weight: bold;
            }
            QPushButton:hover { background-color: #5a6268; }
        """)
        btn_cancelar.clicked.connect(dialog.reject)

        btn_exportar = QPushButton("Exportar")
        btn_exportar.setStyleSheet("""
            QPushButton {
                background-color: #175cc3; color: white;
                border-radius: 6px; padding: 8px 20px; font-weight: bold;
            }
            QPushButton:hover { background-color: #5474b8; }
        """)
        btn_exportar.clicked.connect(dialog.accept)

        btn_row.addWidget(btn_cancelar)
        btn_row.addWidget(btn_exportar)
        layout.addLayout(btn_row)

        if dialog.exec_() != QDialog.Accepted:
            return

        selecionados = [(tipo, payload)
                        for cb, tipo, payload in checkboxes if cb.isChecked()]

        if not selecionados:
            QMessageBox.information(self, "Aviso", "Nenhum item selecionado para exportar.")
            return

        QApplication.setOverrideCursor(Qt.WaitCursor)
        arquivos_gerados = []
        erros            = []

        try:
            for tipo, payload in selecionados:
                try:
                    if tipo == "total":
                        arq = self._exportar_excel_totalizado()
                    else:
                        cod_interf   = payload
                        ids_interf   = [d[0] for d in self.lista_dados_selecionados
                                        if d[5] == cod_interf]
                        dados_interf = [d   for d in self.lista_dados_selecionados
                                        if d[5] == cod_interf]
                        arq = self._exportar_excel_por_interferencia(
                            ids_medidores  = ids_interf,
                            dados_medidores= dados_interf,
                        )
                    if arq:
                        arquivos_gerados.append(arq)
                except Exception as e:
                    erros.append(f"{tipo} / {payload}: {e}")
        finally:
            QApplication.restoreOverrideCursor()

        if arquivos_gerados:
            import os
            downloads_path = os.path.dirname(arquivos_gerados[0])
            lista = "\n  • ".join(os.path.basename(a) for a in arquivos_gerados)
            msg   = (f"{len(arquivos_gerados)} arquivo(s) gerado(s) em:\n"
                     f"{downloads_path}\n\n  • {lista}")
            if erros:
                msg += f"\n\n⚠️ Erros:\n" + "\n".join(erros)
            QMessageBox.information(self, "Exportação Concluída", msg)
        else:
            QMessageBox.critical(self, "Erro", "Nenhum arquivo foi gerado.\n\n" +
                                 "\n".join(erros))

    def _exportar_excel_totalizado(self):
        """
        Gera o arquivo Excel com dados totalizados de todas as interferências.
        Estrutura: aba MENSAL + uma aba por mês (JAN…DEZ).
        Retorna o caminho completo do arquivo gerado.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import calendar, os, sys
        from datetime import datetime

        ano_selecionado = (int(self.combo_ano_mensal.currentText())
                           if hasattr(self, 'combo_ano_mensal')
                           else datetime.now().year)

        # ── nome do arquivo ───────────────────────────────────────────────────────
        def _san(txt, maxlen=30):
            return "".join(c if c.isalnum() or c in ('-','_') else '_'
                           for c in str(txt))[:maxlen]

        criterio_tag = _san(self.criterio_busca or "TOTAL", 20)
        valor_tag    = _san(self.nome_completo or self.termo_busca or "sem_valor", 100)
        nome_arquivo = f"TOTAL_{criterio_tag}_{valor_tag}_{ano_selecionado}.xlsx"

        downloads    = (os.path.join(os.environ['USERPROFILE'], 'Downloads')
                        if sys.platform == 'win32'
                        else os.path.join(os.path.expanduser('~'), 'Downloads'))
        caminho      = os.path.join(downloads, nome_arquivo)

        if os.path.exists(caminho):
            resp = QMessageBox.question(
                self, "Arquivo Existente",
                f"'{nome_arquivo}' já existe.\nDeseja substituir?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if resp == QMessageBox.No:
                return None

        # ── estilos ───────────────────────────────────────────────────────────────
        fill_azul    = PatternFill("solid", fgColor="175cc3")
        font_branca  = Font(bold=True, size=10, color="ffffff")
        font_titulo  = Font(bold=True, size=12, color="175cc3")
        ali_centro   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borda        = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'),  bottom=Side(style='thin'))
        fill_verm    = PatternFill("solid", fgColor="ffcccc")
        fill_amar    = PatternFill("solid", fgColor="ffffcc")
        font_leg     = Font(size=9, italic=True, color="555555")

        # ── cabeçalho textual do relatório ────────────────────────────────────────
        if self.criterio_busca == "Sistema Hídrico":
            linha1_txt = f"Sistema Hídrico: {self.nome_completo or self.termo_busca}"
        elif self.criterio_busca == "CNARH":
            linha1_txt = f"CNARH: {self.nome_completo or self.termo_busca}"
        else:
            linha1_txt = f"Usuário: {self.nome_completo or self.termo_busca}"

        codigos_txt  = ", ".join(sorted(str(c) for c in self.codigos_interf))
        linha2_txt   = f"Interferências: {codigos_txt}"

        # ── helper: estimar altura de linha para conteúdo com quebra ─────────────
        def _altura_linha_auto(texto, largura_cols_chars, altura_base=15, altura_por_linha=15):
            """
            Estima a altura necessária (em pontos) para um texto com wrap_text=True.
            largura_cols_chars: largura total aproximada em caracteres das colunas mescladas.
            """
            if not texto:
                return altura_base
            # Quebras de linha explícitas no texto
            linhas_explicitas = texto.split('\n')
            total_linhas = 0
            for trecho in linhas_explicitas:
                if largura_cols_chars > 0:
                    # Quantas linhas o trecho ocupa dado o espaço disponível
                    total_linhas += max(1, -(-len(trecho) // largura_cols_chars))  # ceil division
                else:
                    total_linhas += 1
            return max(altura_base, total_linhas * altura_por_linha)

        # ── volumes outorgados totalizados (soma de todas as interferências) ───────
        vols_out_total = [0.0] * 12
        for cod in self.codigos_interf:
            v = self.get_volumes_outorgados(cod)
            if v:
                for i in range(12):
                    vols_out_total[i] += v[i]
        tem_outorgado = any(x > 0 for x in vols_out_total)

        # ── buscar dados totalizados ──────────────────────────────────────────────
        def _dados_mensais_total(ano):
            try:
                cur = self.conn.cursor()
                cur.execute("""
                    SELECT EXTRACT(MONTH FROM DATE(data))::int AS mes,
                           SUM(consumo_diario) AS consumo
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s
                      AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY mes ORDER BY mes;
                """, (tuple(self.lista_ids_selecionados), ano))
                rows = cur.fetchall()
                cur.close()
                return {int(r[0]): float(r[1]) for r in rows if r[1] is not None}
            except Exception as e:
                print(f"[ERRO] dados mensais total: {e}")
                return {}

        def _dados_diarios_total(mes, ano):
            try:
                cur = self.conn.cursor()
                cur.execute("""
                    SELECT EXTRACT(DAY FROM DATE(data))::int AS dia,
                           SUM(consumo_diario) AS consumo
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s
                      AND EXTRACT(MONTH FROM DATE(data)) = %s
                      AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY dia ORDER BY dia;
                """, (tuple(self.lista_ids_selecionados), mes, ano))
                rows = cur.fetchall()
                cur.close()
                return {int(r[0]): float(r[1]) for r in rows if r[1] is not None}
            except Exception as e:
                print(f"[ERRO] dados diários total: {e}")
                return {}

        # ── workbook ──────────────────────────────────────────────────────────────
        wb          = Workbook()
        nomes_curto = ["JAN","FEV","MAR","ABR","MAI","JUN",
                       "JUL","AGO","SET","OUT","NOV","DEZ"]
        nomes_longo = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                       "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        dados_mens  = _dados_mensais_total(ano_selecionado)

        # ── largura aproximada das colunas mescladas na linha 2 (A+B+C ou A+B) ───
        # Usada para calcular a altura automática da linha 2
        LARGURA_COL_A = 14
        LARGURA_COL_B = 22
        LARGURA_COL_C = 30
        largura_linha2 = LARGURA_COL_A + LARGURA_COL_B + (LARGURA_COL_C if tem_outorgado else 0)

        # ── ABA MENSAL ────────────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Mensal"

        n_cols = 3 if tem_outorgado else 2
        col_ltr_last = get_column_letter(n_cols)

        ws.merge_cells(f'A1:{col_ltr_last}1')
        ws['A1'] = linha1_txt
        ws['A1'].font      = font_titulo
        ws['A1'].alignment = ali_centro

        # CORREÇÃO a): altura automática da linha 2 conforme conteúdo
        ws.merge_cells(f'A2:{col_ltr_last}2')
        ws['A2'] = linha2_txt
        ws['A2'].font      = Font(size=10, italic=True, color="495057")
        ws['A2'].alignment = ali_centro
        ws.row_dimensions[2].height = _altura_linha_auto(linha2_txt, largura_linha2)

        ws.merge_cells(f'A3:{col_ltr_last}3')
        ws['A3'] = f"Consumo Mensal Totalizado – {ano_selecionado}"
        ws['A3'].font      = Font(bold=True, size=11, color="2c3e50")
        ws['A3'].alignment = ali_centro

        cabecalhos = ["Mês", "Consumo Total (m³)"]
        if tem_outorgado:
            cabecalhos.append("Total Outorgado (m³) – soma das interf.")
        for ci, cab in enumerate(cabecalhos, 1):
            cell        = ws.cell(row=4, column=ci, value=cab)
            cell.font   = font_branca
            cell.fill   = fill_azul
            cell.alignment = ali_centro
            cell.border = borda

        alerta_mensal = False
        for mi, mes_num in enumerate(range(1, 13), 5):
            consumo  = dados_mens.get(mes_num, 0.0)
            out_val  = vols_out_total[mes_num - 1] if tem_outorgado else 0.0

            c_mes    = ws.cell(row=mi, column=1, value=nomes_longo[mes_num-1])
            c_cons   = ws.cell(row=mi, column=2, value=consumo)
            c_mes.alignment  = ali_centro;  c_mes.border  = borda
            c_cons.number_format = '#,##0.00'
            c_cons.alignment = ali_centro;  c_cons.border = borda

            if tem_outorgado:
                c_out = ws.cell(row=mi, column=3, value=out_val)
                c_out.number_format = '#,##0.00'
                c_out.alignment = ali_centro; c_out.border = borda
                if consumo > out_val > 0:
                    c_cons.fill = fill_verm
                    c_out.fill  = fill_amar
                    alerta_mensal = True

        ws.column_dimensions['A'].width = LARGURA_COL_A
        ws.column_dimensions['B'].width = LARGURA_COL_B
        if tem_outorgado:
            ws.column_dimensions['C'].width = LARGURA_COL_C

        # CORREÇÃO b): legenda sem mesclar — "Acima do total outorgado" só em B,
        #              "Alerta" só em C (sem mesclar com colunas adjacentes)
        if alerta_mensal:
            ul = 18
            c_b = ws.cell(row=ul, column=2, value="Acima do total outorgado")
            c_b.fill      = fill_verm
            c_b.font      = font_leg
            c_b.alignment = ali_centro
            c_b.border    = borda
            if tem_outorgado:
                c_c = ws.cell(row=ul, column=3, value="Alerta")
                c_c.fill      = fill_amar
                c_c.font      = font_leg
                c_c.alignment = ali_centro
                c_c.border    = borda

        # ── ABAS MENSAIS (JAN … DEZ) ──────────────────────────────────────────────
        for mi, mes_num in enumerate(range(1, 13)):
            ws_m        = wb.create_sheet(title=nomes_curto[mi])
            dias_no_mes = calendar.monthrange(ano_selecionado, mes_num)[1]
            dados_dia   = _dados_diarios_total(mes_num, ano_selecionado)
            out_mensal  = vols_out_total[mes_num - 1] if tem_outorgado else 0.0

            n_cols_m    = 3 if tem_outorgado else 2
            col_ltr_m   = get_column_letter(n_cols_m)

            ws_m.merge_cells(f'A1:{col_ltr_m}1')
            ws_m['A1'] = linha1_txt
            ws_m['A1'].font = font_titulo; ws_m['A1'].alignment = ali_centro

            # CORREÇÃO a): altura automática da linha 2 em cada aba mensal
            ws_m.merge_cells(f'A2:{col_ltr_m}2')
            ws_m['A2'] = linha2_txt
            ws_m['A2'].font = Font(size=10, italic=True, color="495057")
            ws_m['A2'].alignment = ali_centro
            ws_m.row_dimensions[2].height = _altura_linha_auto(linha2_txt, largura_linha2)

            ws_m.merge_cells(f'A3:{col_ltr_m}3')
            ws_m['A3'] = (f"Consumo Diário Totalizado – "
                          f"{nomes_longo[mi]} {ano_selecionado}")
            ws_m['A3'].font = Font(bold=True, size=11, color="2c3e50")
            ws_m['A3'].alignment = ali_centro

            cabs_m = ["Dia", "Consumo Total (m³)"]
            if tem_outorgado:
                cabs_m.append("Outorgado Diário (m³)")
            for ci, cab in enumerate(cabs_m, 1):
                cell = ws_m.cell(row=4, column=ci, value=cab)
                cell.font = font_branca; cell.fill = fill_azul
                cell.alignment = ali_centro; cell.border = borda

            alerta_diario = False
            out_dia_val   = out_mensal / dias_no_mes if out_mensal > 0 else 0.0

            for dia in range(1, dias_no_mes + 1):
                row      = 4 + dia
                consumo  = dados_dia.get(dia, 0.0)

                c_dia  = ws_m.cell(row=row, column=1, value=dia)
                c_cons = ws_m.cell(row=row, column=2, value=consumo)
                c_dia.alignment  = ali_centro; c_dia.border  = borda
                c_cons.number_format = '#,##0.00'
                c_cons.alignment = ali_centro; c_cons.border = borda

                if tem_outorgado:
                    c_out = ws_m.cell(row=row, column=3, value=out_dia_val)
                    c_out.number_format = '#,##0.00'
                    c_out.alignment = ali_centro; c_out.border = borda
                    if consumo > out_dia_val > 0:
                        c_cons.fill = fill_verm
                        c_out.fill  = fill_amar
                        alerta_diario = True

            ws_m.column_dimensions['A'].width = 8
            ws_m.column_dimensions['B'].width = LARGURA_COL_B
            if tem_outorgado:
                ws_m.column_dimensions['C'].width = LARGURA_COL_C

            # CORREÇÃO b): legenda sem mesclar nas abas mensais
            # ul = 4 (cabeçalho) + dias_no_mes (dados) + 1 (linha em branco) + 1 (legenda)
            if alerta_diario:
                ul = 6 + dias_no_mes
                c_b = ws_m.cell(row=ul, column=2, value="Acima do total outorgado")
                c_b.fill      = fill_verm
                c_b.font      = font_leg
                c_b.alignment = ali_centro
                c_b.border    = borda
                if tem_outorgado:
                    c_c = ws_m.cell(row=ul, column=3, value="Alerta")
                    c_c.fill      = fill_amar
                    c_c.font      = font_leg
                    c_c.alignment = ali_centro
                    c_c.border    = borda

        wb.save(caminho)
        return caminho
            
    def _exportar_excel_por_interferencia(self, ids_medidores, dados_medidores):
        """
        Gera o arquivo Excel para um conjunto de medidores de UMA interferência
        (ou para o caso simples de único medidor / mesma interferência).
        Idêntico ao exportar_excel() original, parametrizado.
        Retorna o caminho completo do arquivo gerado.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import calendar, os, sys
        from datetime import datetime

        ano_selecionado = (int(self.combo_ano_mensal.currentText())
                           if hasattr(self, 'combo_ano_mensal')
                           else datetime.now().year)

        mapa_rotulos = {d[0]: d[1] for d in dados_medidores}
        cnarh        = dados_medidores[0][4] if dados_medidores else "NAO_INFORMADO"
        cod_interf   = dados_medidores[0][5] if dados_medidores else "NAO_INFORMADO"

        if len(ids_medidores) == 1:
            texto_medidores = mapa_rotulos.get(ids_medidores[0], f"ID {ids_medidores[0]}")
        else:
            texto_medidores = " + ".join(mapa_rotulos.get(i, f"ID {i}")
                                         for i in ids_medidores)

        volumes_outorgados = self.get_volumes_outorgados(cod_interf)

        # ── nome do arquivo ───────────────────────────────────────────────────────
        def _san(txt, maxlen=25):
            return "".join(c if c.isalnum() or c in ('-','_') else '_'
                           for c in str(txt))[:maxlen]

        rotulos_lista = [mapa_rotulos.get(i, f"ID_{i}") for i in ids_medidores]
        if len(rotulos_lista) == 1:
            rotulos_tag = _san(rotulos_lista[0])
        else:
            prefixo = os.path.commonprefix([_san(r) for r in rotulos_lista])
            rotulos_tag = (prefixo.rstrip('_') if len(prefixo) > 3
                           else "_".join(_san(r) for r in rotulos_lista[:3]))[:30]

        nome_arquivo = (f"{_san(cnarh,15)}_{_san(cod_interf,15)}"
                        f"_{rotulos_tag}_{ano_selecionado}.xlsx")
        downloads    = (os.path.join(os.environ['USERPROFILE'], 'Downloads')
                        if sys.platform == 'win32'
                        else os.path.join(os.path.expanduser('~'), 'Downloads'))
        caminho      = os.path.join(downloads, nome_arquivo)

        if os.path.exists(caminho):
            resp = QMessageBox.question(
                self, "Arquivo Existente",
                f"'{nome_arquivo}' já existe.\nDeseja substituir?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if resp == QMessageBox.No:
                return None

        # ── estilos ───────────────────────────────────────────────────────────────
        fill_azul   = PatternFill("solid", fgColor="175cc3")
        font_branca = Font(bold=True, size=10, color="ffffff")
        font_titulo = Font(bold=True, size=12, color="175cc3")
        ali_centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borda       = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'),  bottom=Side(style='thin'))
        fill_verm   = PatternFill("solid", fgColor="ffcccc")
        fill_amar   = PatternFill("solid", fgColor="ffffcc")
        font_leg    = Font(size=9, italic=True, color="555555")

        eh_unico    = len(ids_medidores) == 1

        nomes_curto = ["JAN","FEV","MAR","ABR","MAI","JUN",
                       "JUL","AGO","SET","OUT","NOV","DEZ"]
        nomes_longo = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                       "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        # ── dados mensais e diários ───────────────────────────────────────────────
        dados_mens_brutos = self.get_data_mensal_completa(ano_selecionado)
        dados_mens_brutos = [r for r in dados_mens_brutos if r[2] in ids_medidores]

        dados_mensais = {mes: {} for mes in range(1, 13)}
        for row in dados_mens_brutos:
            mes = int(row[0]); val = float(row[1]) if row[1] else 0.0; mid = row[2]
            dados_mensais[mes][mid] = val

        dados_diarios_por_mes = {}
        for mes_num in range(1, 13):
            brutos = self.get_data_diario_completa(mes_num, ano_selecionado)
            brutos = [r for r in brutos if r[2] in ids_medidores]
            dados_dia = {dia: {} for dia in range(1, 32)}
            for row in brutos:
                dia = int(row[0]); val = float(row[1]) if row[1] else 0.0; mid = row[2]
                dados_dia[dia][mid] = val
            dados_diarios_por_mes[mes_num] = dados_dia

        # ── workbook ──────────────────────────────────────────────────────────────
        wb          = Workbook()
        linha1_txt  = (f"CNARH: {cnarh} | Interferência: {cod_interf} "
                       f"| Medidor(es): {texto_medidores}")

        # ── ABA MENSAL ────────────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Mensal"

        cabs = ["Mês"] + [mapa_rotulos.get(i, f"ID {i}") for i in ids_medidores]
        if not eh_unico:
            cabs.append("Total")
        cabs.append("Total Outorgado")
        n_cols = len(cabs)

        ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
        ws['A1'] = linha1_txt
        ws['A1'].font = font_titulo; ws['A1'].alignment = ali_centro

        ws.merge_cells(f'A2:{get_column_letter(n_cols)}2')
        ws['A2'] = f"Consumo Mensal – {ano_selecionado}"
        ws['A2'].font = Font(bold=True, size=11, color="2c3e50")
        ws['A2'].alignment = ali_centro

        for ci, cab in enumerate(cabs, 1):
            cell = ws.cell(row=3, column=ci, value=cab)
            cell.font = font_branca; cell.fill = fill_azul
            cell.alignment = ali_centro; cell.border = borda

        alerta_mensal = False
        for mi, mes_num in enumerate(range(1, 13), 4):
            c_mes = ws.cell(row=mi, column=1, value=nomes_longo[mes_num-1])
            c_mes.alignment = ali_centro; c_mes.border = borda

            total_mes = 0.0; col_cur = 2
            for mid in ids_medidores:
                val  = dados_mensais.get(mes_num, {}).get(mid, 0.0)
                cell = ws.cell(row=mi, column=col_cur, value=val)
                cell.number_format = '#,##0.00'
                cell.alignment = ali_centro; cell.border = borda
                total_mes += val; col_cur += 1

            if not eh_unico:
                ct = ws.cell(row=mi, column=col_cur, value=total_mes)
                ct.number_format = '#,##0.00'
                ct.alignment = ali_centro; ct.border = borda
                col_cur += 1
            else:
                ct = ws.cell(row=mi, column=2)

            out_val = volumes_outorgados[mes_num-1] if volumes_outorgados else 0.0
            c_out   = ws.cell(row=mi, column=col_cur, value=out_val)
            c_out.number_format = '#,##0.00'
            c_out.alignment = ali_centro; c_out.border = borda

            if total_mes > out_val > 0:
                ct.fill = fill_verm; c_out.fill = fill_amar
                alerta_mensal = True

        ws.column_dimensions['A'].width = 15
        for ci in range(2, n_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 18

        # Legenda aba Mensal: sem mesclar, com linha em branco antes
        if alerta_mensal:
            ul = 17  # 3 (cabeçalho) + 12 (meses) + 1 (em branco) + 1 (legenda)
            c_b = ws.cell(row=ul, column=2, value="Acima do total outorgado")
            c_b.fill      = fill_verm
            c_b.font      = font_leg
            c_b.alignment = ali_centro
            c_b.border    = borda
            c_c = ws.cell(row=ul, column=3, value="Alerta")
            c_c.fill      = fill_amar
            c_c.font      = font_leg
            c_c.alignment = ali_centro
            c_c.border    = borda

        # ── ABAS MENSAIS ──────────────────────────────────────────────────────────
        cabs_d = ["Dia"] + [mapa_rotulos.get(i, f"ID {i}") for i in ids_medidores]
        if not eh_unico:
            cabs_d.append("Total")
        cabs_d.append("Total Outorgado")
        n_cols_d = len(cabs_d)

        for mi, mes_num in enumerate(range(1, 13)):
            ws_m        = wb.create_sheet(title=nomes_curto[mi])
            dias_no_mes = calendar.monthrange(ano_selecionado, mes_num)[1]
            dados_dia   = dados_diarios_por_mes.get(mes_num, {})
            out_mensal  = volumes_outorgados[mes_num-1] if volumes_outorgados else 0.0
            out_dia_val = out_mensal / dias_no_mes if out_mensal > 0 else 0.0

            ws_m.merge_cells(f'A1:{get_column_letter(n_cols_d)}1')
            ws_m['A1'] = linha1_txt
            ws_m['A1'].font = font_titulo; ws_m['A1'].alignment = ali_centro

            ws_m.merge_cells(f'A2:{get_column_letter(n_cols_d)}2')
            ws_m['A2'] = (f"Consumo Diário – "
                          f"{nomes_longo[mi]} {ano_selecionado}")
            ws_m['A2'].font = Font(bold=True, size=11, color="2c3e50")
            ws_m['A2'].alignment = ali_centro

            for ci, cab in enumerate(cabs_d, 1):
                cell = ws_m.cell(row=3, column=ci, value=cab)
                cell.font = font_branca; cell.fill = fill_azul
                cell.alignment = ali_centro; cell.border = borda

            alerta_diario = False
            for dia in range(1, dias_no_mes + 1):
                row = 3 + dia
                c_dia = ws_m.cell(row=row, column=1, value=dia)
                c_dia.alignment = ali_centro; c_dia.border = borda

                total_dia = 0.0; col_cur = 2
                for mid in ids_medidores:
                    val  = dados_dia.get(dia, {}).get(mid, 0.0)
                    cell = ws_m.cell(row=row, column=col_cur, value=val)
                    cell.number_format = '#,##0.00'
                    cell.alignment = ali_centro; cell.border = borda
                    total_dia += val; col_cur += 1

                if not eh_unico:
                    ct = ws_m.cell(row=row, column=col_cur, value=total_dia)
                    ct.number_format = '#,##0.00'
                    ct.alignment = ali_centro; ct.border = borda
                    col_cur += 1
                else:
                    ct = ws_m.cell(row=row, column=2)

                c_out = ws_m.cell(row=row, column=col_cur, value=out_dia_val)
                c_out.number_format = '#,##0.00'
                c_out.alignment = ali_centro; c_out.border = borda

                if total_dia > out_dia_val > 0:
                    ct.fill = fill_verm; c_out.fill = fill_amar
                    alerta_diario = True

            ws_m.column_dimensions['A'].width = 8
            for ci in range(2, n_cols_d + 1):
                ws_m.column_dimensions[get_column_letter(ci)].width = 18

            # Legenda abas mensais: sem mesclar, com linha em branco antes
            # ul = 3 (cabeçalho) + dias_no_mes (dados) + 1 (em branco) + 1 (legenda)
            if alerta_diario:
                ul = 5 + dias_no_mes
                c_b = ws_m.cell(row=ul, column=2, value="Acima do total outorgado")
                c_b.fill      = fill_verm
                c_b.font      = font_leg
                c_b.alignment = ali_centro
                c_b.border    = borda
                c_c = ws_m.cell(row=ul, column=3, value="Alerta")
                c_c.fill      = fill_amar
                c_c.font      = font_leg
                c_c.alignment = ali_centro
                c_c.border    = borda

        wb.save(caminho)
        return caminho
   
    def get_data_mensal_completa(self, ano):
        """Busca dados mensais COMPLETOS (ignorando colunas ocultas) para exportação."""
        try:
            cursor = self.conn.cursor()
            e_multiplo = len(self.lista_ids_selecionados) > 1
            if e_multiplo:
                query = """
                    SELECT EXTRACT(MONTH FROM DATE(data)) as mes,
                           SUM(consumo_diario) as consumo,
                           intervencao_id
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s
                      AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY mes, intervencao_id
                    ORDER BY mes, intervencao_id;
                """
            else:
                query = """
                    SELECT EXTRACT(MONTH FROM DATE(data)) as mes,
                           SUM(consumo_diario) as consumo,
                           MAX(intervencao_id) as id_medidor
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s
                      AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY mes
                    ORDER BY mes;
                """
            tuple_ids = tuple(self.lista_ids_selecionados)
            cursor.execute(query, (tuple_ids, ano))
            dados = cursor.fetchall()
            cursor.close()
            return dados
        except Exception as e:
            print(f"Erro ao buscar dados mensais completos: {e}")
            return []

    def get_data_diario_completa(self, mes, ano):
        """Busca dados diários COMPLETOS (ignorando colunas ocultas) para exportação."""
        try:
            cursor = self.conn.cursor()
            e_multiplo = len(self.lista_ids_selecionados) > 1
            if e_multiplo:
                query = """
                    SELECT EXTRACT(DAY FROM DATE(data)) as dia,
                           SUM(consumo_diario) as consumo,
                           intervencao_id
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s
                      AND EXTRACT(MONTH FROM DATE(data)) = %s
                      AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY dia, intervencao_id
                    ORDER BY dia, intervencao_id;
                """
            else:
                query = """
                    SELECT EXTRACT(DAY FROM DATE(data)) as dia,
                           SUM(consumo_diario) as consumo,
                           MAX(intervencao_id) as id_medidor
                    FROM tb_telemetria_intervencao_diaria
                    WHERE intervencao_id IN %s
                      AND EXTRACT(MONTH FROM DATE(data)) = %s
                      AND EXTRACT(YEAR FROM DATE(data)) = %s
                    GROUP BY dia
                    ORDER BY dia;
                """
            tuple_ids = tuple(self.lista_ids_selecionados)
            cursor.execute(query, (tuple_ids, mes, ano))
            dados = cursor.fetchall()
            cursor.close()
            return dados
        except Exception as e:
            print(f"Erro ao buscar dados diários completos: {e}")
            return []

    def ver_no_mapa(self):
            """Orquestra a exibição dos medidores no mapa com ordem de camadas garantida."""
            from PyQt5.QtCore import Qt
            from PyQt5.QtWidgets import QApplication

            try:
                QApplication.setOverrideCursor(Qt.WaitCursor)
                self.adicionar_esri_world_imagery()
                self.adicionar_google_satellite()
                self.adicionar_openstreetmap()
                self.adicionar_webservice_uam()
                QApplication.restoreOverrideCursor()

                self.adicionar_webservice_car()

                QApplication.setOverrideCursor(Qt.WaitCursor)
                self.adicionar_webservice_intcnarhfed()
                self.adicionar_webservice_obrigat()
                self.adicionar_intervencoes()
                QApplication.restoreOverrideCursor()

                # Zoom final após pequeno delay
                from qgis.PyQt.QtCore import QTimer
                QTimer.singleShot(400, self.exibir_medidores_no_canvas)

                QMessageBox.information(self, "Mapa",
                    "Medidor(es) carregado(s) no mapa!\n\n"
                    "Clique no ponto para ver todos os detalhes\n"
                    "e informações completas das intervenções.")

                self.showMinimized()

                if isinstance(self.janela_anterior, JanelaMonitoramento) and self.janela_anterior.isVisible():
                    self.janela_anterior.showMinimized()

            except Exception as e:
                QApplication.restoreOverrideCursor()
                print(f"❌ Erro em ver_no_mapa: {e}")
                import traceback
                traceback.print_exc()
                
    def adicionar_openstreetmap(self):
        """Adiciona o OpenStreetMap como alternativa de mapa base."""
        try:
            # ADICIONAR VERIFICAÇÃO AQUI (Igual ao ESRI)
            layer_name = "OpenStreetMap"
            layers = QgsProject.instance().mapLayersByName(layer_name)
            
            if layers:
                # Se já existe, apenas ativa e sai
                iface.mapCanvas().setCurrentLayer(layers[0])
                return

            # Se não existe, prossegue para criar
            url_with_params = (
                "type=xyz&url=https://tile.openstreetmap.org/{z}/{x}/{y}.png"
                "&zmin=0&zmax=19"
                "&crs=EPSG:3857"
            )
            
            osm_layer = QgsRasterLayer(url_with_params, layer_name, "wms")
            
            if not osm_layer.isValid():
                QMessageBox.warning(None, "Erro", "Falha ao carregar OpenStreetMap. Verifique sua conexão com a internet.")
                return
            
            QgsProject.instance().addMapLayer(osm_layer)
            
        except Exception as e:
            QMessageBox.critical(None, "Erro", f"Erro ao carregar OpenStreetMap: {str(e)}")

    def adicionar_esri_world_imagery(self):
        """Adiciona camada de satélite ESRI."""
        try:
            layers = QgsProject.instance().mapLayersByName("ESRI World Imagery")         
            if layers:
                iface.mapCanvas().setCurrentLayer(layers[0])
                return

            uri = "type=xyz&url=https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}&zmax=19&zmin=0"
            rlayer = QgsRasterLayer(uri, "ESRI World Imagery", "wms")         
            if rlayer.isValid():
                QgsProject.instance().addMapLayer(rlayer)
        except Exception as e:
            print(f"Erro ao adicionar ESRI: {e}")

    def adicionar_google_satellite(self):
        """Adiciona camada de satélite da Google."""
        try:
            layers = QgsProject.instance().mapLayersByName("Google Satellite Hybrid")
            if layers:
                iface.mapCanvas().setCurrentLayer(layers[0])
                return

            uri = "type=xyz&url=https://mt1.google.com/vt/lyrs%3Dy%26x%3D%7Bx%7D%26y%3D%7By%7D%26z%3D%7Bz%7D"
            rlayer = QgsRasterLayer(uri, "Google Satellite Hybrid", "wms")
            if rlayer.isValid():
                QgsProject.instance().addMapLayer(rlayer)
        except Exception as e:
            print(f"Erro ao adicionar Google: {e}")

    def adicionar_webservice_uam(self):
        """Adiciona camada de ESRI MapServer das Unidades de Automonitoramento."""
        try:
            # Verificar se a camada já existe
            layers = QgsProject.instance().mapLayersByName("Unidades de Automonitoramento (SFI/ANA - ArcGIS MapServer)")
            if layers:
                layer = layers[0]
                # Garantir invisibilidade
                layer_node = QgsProject.instance().layerTreeRoot().findLayer(layer.id())
                if layer_node:
                    layer_node.setItemVisibilityChecked(False)
                return

            # URI para ArcGIS FeatureServer
            uri = "url='https://portal1.snirh.gov.br/server/rest/services/SFI/Unidades_Automonitoramento_UGRH_MRS/MapServer' crs='EPSG:4326' format='PNG32' layer='0' "
            
            rlayer = QgsRasterLayer(uri, "Unidades de Automonitoramento (SFI/ANA - ArcGIS MapServer)", "arcgismapserver")
            
            if rlayer.isValid():
                QgsProject.instance().addMapLayer(rlayer, False)
                
                root = QgsProject.instance().layerTreeRoot()

                novo_no = QgsLayerTreeLayer(rlayer)

                novo_no.setItemVisibilityChecked(False)
                novo_no.setExpanded(False)

                root.insertChildNode(0, novo_no)                    
            else:
                QMessageBox.warning(self, "Atenção", 
                    "Falha ao carregar a camada 'Unidades de Automonitoramento'.\n"
                    "Verifique sua conexão com a internet ou se o serviço está disponível.")
                print(f"Erro na camada: {rlayer.error().message()}")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao adicionar ESRI MapServer:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def adicionar_webservice_obrigat(self):
        """Adiciona camada de ESRI MapServer de Obrigatoriedade do Automonitoramento."""
        try:
            # Verificar se a camada já existe
            layers = QgsProject.instance().mapLayersByName("Obrigatoriedade Automonitoramento (SFI/ANA - ArcGIS MapServer)")
            if layers:
                layer = layers[0]
                # Garantir invisibilidade
                layer_node = QgsProject.instance().layerTreeRoot().findLayer(layer.id())
                if layer_node:
                    layer_node.setItemVisibilityChecked(False)
                return

            # URI para ArcGIS FeatureServer
            uri = "url='https://portal1.snirh.gov.br/server/rest/services/SFI/Obrigatoriedade_Automonitoramento_DW_v5/MapServer' crs='EPSG:4326' format='PNG32' layer='0' "
            
            rlayer = QgsRasterLayer(uri, "Obrigatoriedade Automonitoramento (SFI/ANA - ArcGIS MapServer)", "arcgismapserver")
            
            if rlayer.isValid():
                QgsProject.instance().addMapLayer(rlayer, False)
                
                root = QgsProject.instance().layerTreeRoot()

                novo_no = QgsLayerTreeLayer(rlayer)

                novo_no.setItemVisibilityChecked(False)
                novo_no.setExpanded(False)

                root.insertChildNode(0, novo_no)                    
            else:
                QMessageBox.warning(self, "Atenção", 
                    "Falha ao carregar a camada 'Obrigatoriedade Automonitoramento'.\n"
                    "Verifique sua conexão com a internet ou se o serviço está disponível.")
                print(f"Erro na camada: {rlayer.error().message()}")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao adicionar ESRI MapServer:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def adicionar_webservice_car(self):
            """
            Adiciona camada de ESRI MapServer do CAR DOMÍNIO PRIVADO (protegido por LGPD).
            Requer autenticação institucional ANA.
            """
            try:
                # Verificar se a camada já existe
                layers = QgsProject.instance().mapLayersByName("CAR Domínio Privado (SFI/ANA - Protegido)")
                if layers:
                    layer = layers[0]
                    layer_node = QgsProject.instance().layerTreeRoot().findLayer(layer.id())
                    if layer_node:
                        layer_node.setItemVisibilityChecked(False)
                    return

                # === AUTENTICAÇÃO ANA - FLUXO LGPD ===

                # 1. Aviso LGPD obrigatório
                mensagem_lgpd = (
                    "<h3>⚠️ Acesso a Dados Sensíveis</h3>"
                    "<p>Este serviço contém dados do <b>Cadastro Ambiental Rural (CAR)</b> "
                    "classificados como <b>sensíveis</b> conforme a <b>LGPD</b>.</p>"
                    "<p><b>Responsabilidades do usuário:</b></p>"
                    "<ul>"
                    "<li>✓ Uso exclusivo para fins institucionais</li>"
                    "<li>✓ Não compartilhamento de credenciais</li>"
                    "<li>✓ Responsabilidade pelo uso indevido</li>"
                    "</ul>"
                    "<p>Deseja prosseguir com a autenticação?</p>"
                )

                from qgis.PyQt.QtWidgets import QMessageBox
                reply = QMessageBox.question(
                    self, "Termo de Responsabilidade - LGPD",
                    mensagem_lgpd,
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if reply != QMessageBox.Yes:
                    return

                # 2. Solicitar credenciais institucionais (COM BOTÃO DE OLHO)
                from qgis.PyQt.QtWidgets import (
                    QDialog, QVBoxLayout, QLabel, QLineEdit,
                    QPushButton, QCheckBox, QHBoxLayout, QWidget, QToolButton
                )
                from qgis.PyQt.QtCore import Qt
                from qgis.PyQt.QtGui import QIcon, QPixmap
                from qgis.core import QgsSettings
                import requests
                from datetime import datetime, timedelta

                settings = QgsSettings()
                ultimo_usuario = settings.value("MeuPlugin/ANA/ultimo_usuario", "")

                portal_url = "https://portal1.snirh.gov.br/ana"
                server_url = "https://portal1.snirh.gov.br/server"

                dialog = QDialog(self)
                dialog.setWindowTitle("Autenticação ANA - CAR Protegido")
                dialog.setMinimumWidth(400)

                layout = QVBoxLayout()

                # Header
                header = QLabel("🔒 ACESSO RESTRITO - Dados CAR Protegidos")
                header.setStyleSheet("""
                    QLabel {
                        background-color: #d9534f; color: white;
                        padding: 15px; font-weight: bold; border-radius: 5px;
                    }
                """)
                header.setAlignment(Qt.AlignCenter)
                layout.addWidget(header)

                info = QLabel("Utilize suas credenciais institucionais (mesmo login do Portal ANA)")
                info.setWordWrap(True)
                info.setStyleSheet("padding: 10px; color: #555;")
                layout.addWidget(info)

                # Label de erro (inicialmente oculto)
                lbl_erro = QLabel("")
                lbl_erro.setWordWrap(True)
                lbl_erro.setStyleSheet("""
                    QLabel {
                        color: #d9534f; font-weight: bold;
                        padding: 6px; background-color: #fdf2f2;
                        border: 1px solid #f5c6cb; border-radius: 4px;
                    }
                """)
                lbl_erro.setAlignment(Qt.AlignCenter)
                lbl_erro.setVisible(False)
                layout.addWidget(lbl_erro)

                # === CAMPO USUÁRIO ===
                linha_usuario = QWidget()
                layout_usuario = QHBoxLayout(linha_usuario)
                layout_usuario.setContentsMargins(0, 0, 0, 0)
                layout_usuario.setSpacing(10)

                lbl_usuario = QLabel("Usuário:")
                lbl_usuario.setStyleSheet("font-weight: bold; color: #495057;")
                lbl_usuario.setFixedWidth(70)

                input_usuario = QLineEdit()
                input_usuario.setPlaceholderText("nome.sobrenome@ana.gov.br")
                if ultimo_usuario:
                    input_usuario.setText(ultimo_usuario)
                input_usuario.setFixedHeight(35)

                layout_usuario.addWidget(lbl_usuario)
                layout_usuario.addWidget(input_usuario)
                layout.addWidget(linha_usuario)

                # === CAMPO SENHA COM BOTÃO DE OLHO ===
                linha_senha = QWidget()
                layout_senha = QHBoxLayout(linha_senha)
                layout_senha.setContentsMargins(0, 0, 0, 0)
                layout_senha.setSpacing(10)

                lbl_senha = QLabel("Senha:")
                lbl_senha.setStyleSheet("font-weight: bold; color: #495057;")
                lbl_senha.setFixedWidth(70)

                input_senha = QLineEdit()
                input_senha.setEchoMode(QLineEdit.Password)
                input_senha.setPlaceholderText("Digite a senha")
                input_senha.setFixedHeight(35)

                toggle_senha_btn = QToolButton()
                toggle_senha_btn.setCursor(Qt.PointingHandCursor)
                toggle_senha_btn.setFixedSize(35, 35)
                toggle_senha_btn.setIconSize(QSize(18, 18))
                toggle_senha_btn.setStyleSheet("""
                    QToolButton {
                        border: 1px solid #ccc;
                        border-radius: 8px;
                        background-color: #f8f9fa;
                    }
                    QToolButton:hover {
                        background-color: #e2e6ea;
                        border: 1px solid #adb5bd;
                    }
                """)

                senha_visivel = False

                def atualizar_icone():
                    if senha_visivel:
                        toggle_senha_btn.setIcon(QIcon(':/images/themes/default/mActionShowAllLayers.svg'))
                    else:
                        toggle_senha_btn.setIcon(QIcon(':/images/themes/default/mActionHideAllLayers.svg'))

                def toggle_visibilidade():
                    nonlocal senha_visivel
                    senha_visivel = not senha_visivel
                    input_senha.setEchoMode(QLineEdit.Normal if senha_visivel else QLineEdit.Password)
                    atualizar_icone()

                toggle_senha_btn.clicked.connect(toggle_visibilidade)
                atualizar_icone()

                layout_senha.addWidget(lbl_senha)
                layout_senha.addWidget(input_senha)
                layout_senha.addWidget(toggle_senha_btn)
                layout.addWidget(linha_senha)

                # Checkbox lembrar usuário
                check_lembrar = QCheckBox("Lembrar meu usuário (não a senha)")
                check_lembrar.setChecked(bool(ultimo_usuario))
                check_lembrar.setStyleSheet("""
                    QCheckBox {
                        color: #666666; font-size: 12px; padding: 5px 0;
                    }
                    QCheckBox::indicator {
                        width: 16px; height: 16px;
                        border: 1px solid #adb5bd; border-radius: 4px;
                        background-color: white;
                    }
                    QCheckBox::indicator:checked {
                        background-color: #175cc3; border: 1px solid #175cc3;
                    }
                """)
                layout.addWidget(check_lembrar)

                layout.addStretch(1)

                btn_layout = QHBoxLayout()
                btn_layout.setSpacing(10)

                btn_ok = QPushButton("🔐 Entrar")
                btn_ok.setDefault(True)
                btn_ok.setStyleSheet("""
                    QPushButton {
                        background-color: #5cb85c; color: white;
                        border-radius: 8px; padding: 10px 20px; font-weight: bold;
                    }
                    QPushButton:hover { background-color: #4cae4c; }
                """)

                btn_cancelar = QPushButton("Cancelar")
                btn_cancelar.setStyleSheet("""
                    QPushButton {
                        background-color: transparent; color: #6c757d;
                        border: 1px solid #6c757d; border-radius: 8px;
                        padding: 10px 20px; font-weight: bold;
                    }
                    QPushButton:hover { background-color: #f0f0f0; }
                """)
                btn_cancelar.clicked.connect(dialog.reject)

                btn_ok.clicked.connect(dialog.accept)

                btn_layout.addWidget(btn_ok)
                btn_layout.addWidget(btn_cancelar)
                layout.addLayout(btn_layout)

                dialog.setLayout(layout)

                if ultimo_usuario:
                    input_senha.setFocus()
                else:
                    input_usuario.setFocus()

                # 3. Loop de autenticação — permite repetição em caso de erro
                token = None
                expira = None
                horas = 0
                minutos = 0
                while True:
                    if dialog.exec_() != QDialog.Accepted:
                        return

                    usuario = input_usuario.text().strip()
                    senha = input_senha.text()

                    if not usuario or not senha:
                        lbl_erro.setText("⚠ Informe usuário e senha.")
                        lbl_erro.setVisible(True)
                        input_senha.clear()
                        input_senha.setFocus()
                        continue

                    # Salvar/remover usuário
                    if check_lembrar.isChecked():
                        settings.setValue("MeuPlugin/ANA/ultimo_usuario", usuario)
                    else:
                        settings.remove("MeuPlugin/ANA/ultimo_usuario")

                    try:
                        # PASSO 1: Token do Portal com referer do servidor
                        url_token_portal = f"{portal_url}/sharing/rest/generateToken"
                        resp_portal = requests.post(
                            url_token_portal,
                            data={
                                'username': usuario,
                                'password': senha,
                                'client': 'referer',
                                'referer': server_url,
                                'expiration': 120,
                                'f': 'json'
                            },
                            verify=True,
                            timeout=30
                        )
                        resp_portal.raise_for_status()
                        resultado_portal = resp_portal.json()

                        if 'token' not in resultado_portal:
                            erro = resultado_portal.get('error', {})
                            codigo = erro.get('code', 'desconhecido')
                            msg = erro.get('message', 'Usuário ou senha incorretos.')
                            lbl_erro.setText(f"⚠ Erro {codigo}: {msg}")
                            lbl_erro.setVisible(True)
                            input_senha.clear()
                            input_senha.setFocus()
                            continue

                        token = resultado_portal['token']
                        expires_ms = resultado_portal.get('expires', 0)
                        if expires_ms:
                            expira = datetime.fromtimestamp(expires_ms / 1000)
                            tempo_restante = expira - datetime.now()
                            horas, resto = divmod(int(tempo_restante.total_seconds()), 3600)
                            minutos, _ = divmod(resto, 60)

                        break  # autenticação bem-sucedida

                    except requests.exceptions.SSLError:
                        lbl_erro.setText("⚠ Erro de certificado SSL. Contate a TI.")
                        lbl_erro.setVisible(True)
                        continue
                    except requests.exceptions.ConnectionError:
                        lbl_erro.setText("⚠ Sem conexão com o Portal ANA.")
                        lbl_erro.setVisible(True)
                        continue
                    except Exception as e:
                        lbl_erro.setText(f"⚠ Erro de conexão: {str(e)}")
                        lbl_erro.setVisible(True)
                        continue

                # 4. Criar camada — ampulheta apenas após autenticação confirmada
                QApplication.setOverrideCursor(Qt.WaitCursor)
                try:
                    from qgis.core import QgsAuthMethodConfig, QgsApplication

                    servico_url = "https://portal1.snirh.gov.br/server/rest/services/SFI/car_dominio_privado_v2/MapServer"

                    # Limpar auth configs anteriores do CAR
                    auth_manager = QgsApplication.authManager()
                    configs = auth_manager.availableAuthMethodConfigs()
                    for cfg_id, cfg in configs.items():
                        if cfg.name() == "CAR_ANA_temp":
                            auth_manager.removeAuthenticationConfig(cfg_id)

                    # Criar configuração de autenticação com header Referer
                    auth_cfg = QgsAuthMethodConfig()
                    auth_cfg.setName("CAR_ANA_temp")
                    auth_cfg.setMethod("EsriToken")
                    auth_cfg.setConfig("token", token)
                    auth_cfg.setConfig("referer", server_url)

                    auth_manager.storeAuthenticationConfig(auth_cfg)
                    auth_id = auth_cfg.id()

                    uri = f"url='{servico_url}' crs='EPSG:4326' format='PNG32' authcfg='{auth_id}' "

                    rlayer = QgsRasterLayer(
                        uri,
                        "CAR Domínio Privado (SFI/ANA - Protegido)",
                        "arcgismapserver"
                    )

                    if rlayer.isValid():
                        rlayer.setAttribution("© ANA - Dados CAR protegidos por LGPD. Acesso restrito.")
                        rlayer.setAttributionUrl("https://portal1.snirh.gov.br/ana")

                        QgsProject.instance().addMapLayer(rlayer, False)

                        root = QgsProject.instance().layerTreeRoot()
                        novo_no = QgsLayerTreeLayer(rlayer)
                        novo_no.setItemVisibilityChecked(False)
                        novo_no.setExpanded(False)
                        root.insertChildNode(0, novo_no)

                        msg_sucesso = (
                            f"<b>✓ CAR Domínio Privado carregado!</b><br><br>"
                            f"<b>Usuário:</b> {usuario}<br>"
                            f"<b>Sessão válida por:</b> {horas}h {minutos}min<br>"
                            f"<b>Expira em:</b> {expira.strftime('%d/%m/%Y %H:%M') if expira else 'N/A'}<br><br>"
                            f"<small>Os dados são protegidos por LGPD. Use com responsabilidade.</small>"
                        )
                        QApplication.restoreOverrideCursor()
                        QMessageBox.information(self, "Acesso Autorizado", msg_sucesso)

                    else:
                        QApplication.restoreOverrideCursor()
                        erro = rlayer.error().summary()
                        QMessageBox.critical(
                            self,
                            "Erro ao Carregar CAR",
                            f"Não foi possível carregar o serviço CAR Privado:<br><br>{erro}"
                        )
                        print(f"Erro na camada CAR: {erro}")

                finally:
                    QApplication.restoreOverrideCursor()

            except Exception as e:
                QApplication.restoreOverrideCursor()
                QMessageBox.critical(self, "Erro", f"Erro ao adicionar CAR:\n{str(e)}")
                import traceback
                traceback.print_exc() 
            
    def adicionar_webservice_intcnarhfed(self):
        """Adiciona camada de ESRI MapServer do CAR."""
        try:
            # Verificar se a camada já existe
            layers = QgsProject.instance().mapLayersByName("Interferência CNARH - Federal (SFI/ANA - ArcGIS MapServer)")
            if layers:
                layer = layers[0]
                # Garantir invisibilidade
                layer_node = QgsProject.instance().layerTreeRoot().findLayer(layer.id())
                if layer_node:
                    layer_node.setItemVisibilityChecked(False)
                return

            # URI para ArcGIS FeatureServer
            uri = "url='https://portal1.snirh.gov.br/server/rest/services/SFI/Interferências_CNARH/MapServer' crs='EPSG:4326' format='PNG32' layer='0' "
            
            rlayer = QgsRasterLayer(uri, "Interferência CNARH - Federal (SFI/ANA - ArcGIS MapServer", "arcgismapserver")
            
            if rlayer.isValid():
                QgsProject.instance().addMapLayer(rlayer, False)
                
                root = QgsProject.instance().layerTreeRoot()

                novo_no = QgsLayerTreeLayer(rlayer)
                
                novo_no.setItemVisibilityChecked(False)
                novo_no.setExpanded(False)
                
                root.insertChildNode(0, novo_no)                    
            else:
                QMessageBox.warning(self, "Atenção", 
                    "Falha ao carregar a camada 'Interferência CNARH - Federal'.\n"
                    "Verifique sua conexão com a internet ou se o serviço está disponível.")
                print(f"Erro na camada: {rlayer.error().message()}")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao adicionar ESRI MapServer:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def adicionar_intervencoes(self):
        """Adiciona a camada de intervenções usando a senha armazenada."""
        try:
            # === RECUPERAÇÃO DE PARÂMETROS ===
            dsn_params = self.conn.get_dsn_parameters()
            host = dsn_params.get('host')
            dbname = dsn_params.get('dbname')
            user = dsn_params.get('user', self.usuario_logado)
            port = dsn_params.get('port', '5432')
            
            # Usa a senha que foi passada no construtor
            password = self.senha_conexao

            # Validação de segurança
            if not password:
                QMessageBox.critical(
                    self, "Erro de Autenticação", 
                    "A senha não foi transmitida. Verifique se a janela foi aberta corretamente pelo Login."
                )
                return

            print(f"🔌 [Carregar Camada] Conectando como: {user}")

            # === DEFINIÇÃO DINÂMICA DO NOME DA CAMADA ===
            qtd_medidores = len(self.lista_ids_selecionados)
            if qtd_medidores > 1:
                layer_name = "Medidores"  # Plural
            else:
                layer_name = "Medidor"    # Singular

            # === LIMPEZA ROBUSTA DE CAMADAS ANTIGAS ===
            # Selecionei 1 medidor (criou "Medidor"). Agora seleciono 2 (vai criar "Medidores").
            # Se não removermos o "Medidor", ficará duplicado no painel.
            # Portanto, tentamos remover "Medidor" e "Medidores" antes de criar.
            nomes_remover = ["Medidor", "Medidores"]
            for nome in nomes_remover:
                layers = QgsProject.instance().mapLayersByName(nome)
                if layers:
                    old_layer = layers[0]
                    QgsProject.instance().removeMapLayer(old_layer.id())
                    print(f"🔄 Removida camada antiga '{nome}' para atualização.")
            
            ids_str = ",".join(str(id) for id in self.lista_ids_selecionados)
            
            uri = QgsDataSourceUri()
            uri.setConnection(host, port, dbname, user, password)
            
            query = f"""
            SELECT 
                id, nome_usuario, nome_operador, nu_interferencia_cnarh, nu_cnarh,
                rotulo_medidor, vazao_nominal, potencia, tipo_medidor, modo_transmissao,
                codigo_uc, geom
            FROM view_ft_intervencao
            WHERE id IN ({ids_str})
            """
            
            uri.setDataSource("", f"({query})", "geom", "", "id")
            uri_string = uri.uri(False)
            print(f"🔗 URI gerada: {uri_string}")
            
            layer = QgsVectorLayer(uri.uri(), layer_name, "postgres")
            
            if layer.isValid():
                print(f"✅ Camada '{layer_name}' carregada com sucesso! Features: {layer.featureCount()}")
                
                # --- CONFIGURAÇÕES VISUAIS (Simbologia e Labels) ---
                
                # 1. Símbolo
                symbol = QgsMarkerSymbol.createSimple({
                    'name': 'circle', 'color': 'red', 'size': '6',
                    'outline_style': 'solid', 'outline_color': 'black', 'outline_width': '1.0'
                })
                renderer = QgsSingleSymbolRenderer(symbol)
                layer.setRenderer(renderer)
                
                # 2. Labels
                label_settings = QgsPalLayerSettings()
                label_expression = """
                CASE 
                    WHEN "nu_interferencia_cnarh" IS NOT NULL AND "nu_interferencia_cnarh" != ''
                    THEN "nu_interferencia_cnarh" || ' - ' || "rotulo_medidor"
                    ELSE "rotulo_medidor"
                END
                """
                label_settings.fieldName = label_expression
                label_settings.isExpression = True
                
                text_format = QgsTextFormat()
                text_format.setFont(QFont("Arial", 10, QFont.Bold))
                text_format.setSize(10)
                text_format.setColor(QColor("black"))
                
                buffer_settings = QgsTextBufferSettings()
                buffer_settings.setEnabled(True)
                buffer_settings.setSize(1.5)
                buffer_settings.setColor(QColor("white"))
                buffer_settings.setOpacity(1.0)
                text_format.setBuffer(buffer_settings)
                
                label_settings.placement = QgsPalLayerSettings.AroundPoint
                label_settings.dist = 2.0
                label_settings.quadOffset = QgsPalLayerSettings.QuadrantBelow
                label_settings.scaleMin = 1000
                label_settings.scaleMax = 500000
                label_settings.setFormat(text_format)
                
                layer.setLabelsEnabled(True)
                layer.setLabeling(QgsVectorLayerSimpleLabeling(label_settings))
                
                # 3. Tooltips e Display Expression
                layer.setDisplayExpression("""
                '<b>Intervenção:</b> ' || coalesce("rotulo_medidor", 'Não informado') || '<br>' ||
                '<b>CNARH:</b> ' || coalesce("nu_interferencia_cnarh", 'Não informado') || '<br>' ||
                '<b>Usuário:</b> ' || coalesce("nome_usuario", 'Não informado') || '<br>' ||
                '<b>Operador:</b> ' || coalesce("nome_operador", 'Não informado')
                """)
                
                layer.setMapTipTemplate("""
                    <html>
                    <head>
                        <style>
                            body { font-family: Arial; font-size: 11px; }
                            .titulo { font-weight: bold; color: #175cc3; margin-bottom: 5px; }
                            .campo { margin-bottom: 3px; }
                            .rotulo { font-weight: bold; color: #333; }
                            .valor { color: #666; }
                        </style>
                    </head>
                    <body>
                        <div class="titulo">INFORMAÇÕES DA INTERVENÇÃO</div>
                        <div class="campo"><span class="rotulo">ID:</span> <span class="valor">[% "id" %]</span></div>
                        <div class="campo"><span class="rotulo">Rótulo:</span> <span class="valor">[% "rotulo_medidor" %]</span></div>
                        <div class="campo"><span class="rotulo">Interferência CNARH:</span> <span class="valor">[% "nu_interferencia_cnarh" %]</span></div>
                        <div class="campo"><span class="rotulo">CNARH:</span> <span class="valor">[% "nu_cnarh" %]</span></div>
                        <div class="campo"><span class="rotulo">Usuário:</span> <span class="valor">[% "nome_usuario" %]</span></div>
                        <div class="campo"><span class="rotulo">Operador:</span> <span class="valor">[% "nome_operador" %]</span></div>
                        <div class="campo"><span class="rotulo">Vazão Nominal:</span> <span class="valor">[% "vazao_nominal" %] m³/s</span></div>
                        <div class="campo"><span class="rotulo">Potência:</span> <span class="valor">[% "potencia" %]</span></div>
                        <div class="campo"><span class="rotulo">Tipo Medidor:</span> <span class="valor">[% "tipo_medidor" %]</span></div>
                        <div class="campo"><span class="rotulo">Modo Transmissão:</span> <span class="valor">[% "modo_transmissao" %]</span></div>
                        <div class="campo"><span class="rotulo">Código UC:</span> <span class="valor">[% "codigo_uc" %]</span></div>
                    </body>
                    </html>
                    """)
                    
                # Adicionar ao Projeto
                QgsProject.instance().addMapLayer(layer, False)
                
                # Garantir visibilidade
                root = QgsProject.instance().layerTreeRoot()
                layer_node = root.insertChildNode(0, QgsLayerTreeLayer(layer))
                if layer_node:
                    layer_node.setItemVisibilityChecked(True)
                
                self.configurar_identificacao(layer)
                    
            else:
                error_msg = layer.error().message() if layer.error() else "Erro desconhecido"
                print(f"❌ ERRO - Camada inválida: {error_msg}")
                # Opcional: tentar método alternativo aqui se desejado
                        
        except Exception as e:
            import traceback
            print(f"❌ ERRO DETALHADO:\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Erro", f"Erro ao carregar camada: {str(e)}")

    def configurar_identificacao(self, layer):
        """Configura a ferramenta de identificação para mostrar TODAS as features no ponto."""
        try:
            from qgis.gui import QgsMapToolIdentify
            from qgis.core import QgsFeature
            from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QDialog, QVBoxLayout, QTextEdit
            from PyQt5.QtCore import Qt
            
            # Criar uma ferramenta de identificação personalizada
            class IdentificacaoTool(QgsMapToolIdentify):
                def __init__(self, canvas, layer):
                    QgsMapToolIdentify.__init__(self, canvas)
                    self.canvas = canvas
                    self.layer = layer
                    
                def canvasReleaseEvent(self, event):
                    # Identificar features no ponto clicado
                    # TopDownAll garante que pegamos todas, não só a de cima
                    results = self.identify(event.x(), event.y(), [self.layer], 
                                           QgsMapToolIdentify.TopDownAll)
                    
                    if results:
                        # AJUSTE: Coletar TODAS as features encontradas
                        features_encontradas = [result.mFeature for result in results]
                        
                        # Mostrar diálogo com todas as features
                        self.mostrar_detalhes_feature(features_encontradas)
                
                def mostrar_detalhes_feature(self, lista_features):
                    """Mostra diálogo com todas as features encontradas no ponto."""
                    qtd = len(lista_features)
                    dialog = QDialog(self.canvas.parent())
                    dialog.setWindowTitle(f"Detalhes das Intervenções ({qtd})")
                    dialog.setMinimumWidth(800)
                    dialog.setMinimumHeight(700)
                    
                    layout = QVBoxLayout(dialog)
                    
                    # Criar widget de texto formatado
                    text_widget = QTextEdit()
                    text_widget.setReadOnly(True)
                    
                    # Construir HTML concatenando todas as features
                    html_completo = "<html><head>" + self.get_css_estilo() + "</head><body>"
                    
                    for i, feature in enumerate(lista_features):
                        # Cabeçalho separador para cada intervenção
                        html_completo += f"<div class='header'>📋 INTERVENÇÃO #{i+1}</div>"
                        
                        # HTML da feature atual
                        html_completo += self.gerar_html_feature(feature)
                        
                        # Linha divisória se houver mais features
                        if i < qtd - 1:
                            html_completo += "<hr style='border-top: 3px dashed #dee2e6; margin: 25px 0;'/>"
                    
                    html_completo += "</body></html>"
                    text_widget.setHtml(html_completo)
                    
                    layout.addWidget(text_widget)
                    dialog.exec_()
                
                def get_css_estilo(self):
                    """Retorna o CSS comum para formatação."""
                    return """
                    <style>
                        body { 
                            font-family: 'Segoe UI', Arial, sans-serif; 
                            font-size: 12px; 
                            margin: 15px;
                            color: #333;
                        }
                        .header {
                            background-color: #175cc3;
                            color: white;
                            padding: 12px;
                            border-radius: 5px;
                            margin-bottom: 15px;
                            font-size: 16px;
                            font-weight: bold;
                            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        }
                        .table {
                            width: 100%;
                            border-collapse: collapse;
                            margin-bottom: 10px;
                        }
                        .table th {
                            background-color: #f8f9fa;
                            color: #175cc3;
                            text-align: left;
                            padding: 10px;
                            border: 1px solid #dee2e6;
                            font-weight: bold;
                        }
                        .table td {
                            padding: 8px;
                            border: 1px solid #dee2e6;
                            vertical-align: top;
                        }
                        .table tr:nth-child(even) {
                            background-color: #f8f9fa;
                        }
                        .valor-nulo {
                            color: #999;
                            font-style: italic;
                        }
                        .secao {
                            margin-top: 15px;
                            padding: 8px;
                            background-color: #e9ecef;
                            border-left: 4px solid #175cc3;
                            font-weight: bold;
                            color: #495057;
                            font-size: 13px;
                        }
                        .campo-destaque {
                            font-weight: bold;
                            color: #175cc3;
                        }
                    </style>
                    """

                def gerar_html_feature(self, feature):
                    """Gera o HTML (tabelas) para UMA feature específica."""
                    fields = feature.fields()
                    
                    html = ""
                    
                    campos_identificacao = ['id', 'rotulo_medidor', 'nu_interferencia_cnarh', 'nu_cnarh']
                    campos_pessoais = ['nome_usuario', 'nome_operador']
                    campos_tecnicos = ['vazao_nominal', 'potencia', 'tipo_medidor', 'modo_transmissao']
                    campos_outros = ['codigo_uc']
                    
                    # Seção 1
                    html += "<div class='secao'>📋 IDENTIFICAÇÃO</div>"
                    html += "<table class='table'>"
                    for field_name in campos_identificacao:
                        if field_name in fields.names():
                            valor = feature[field_name]
                            html += f"""
                            <tr>
                                <td class='campo-destaque'>{self.formatar_nome_campo(field_name)}</td>
                                <td>{self.formatar_valor(valor)}</td>
                            </tr>
                            """
                    html += "</table>"
                    
                    # Seção 2
                    html += "<div class='secao'>👤 INFORMAÇÕES PESSOAIS</div>"
                    html += "<table class='table'>"
                    for field_name in campos_pessoais:
                        if field_name in fields.names():
                            valor = feature[field_name]
                            html += f"""
                            <tr>
                                <td class='campo-destaque'>{self.formatar_nome_campo(field_name)}</td>
                                <td>{self.formatar_valor(valor)}</td>
                            </tr>
                            """
                    html += "</table>"
                    
                    # Seção 3
                    html += "<div class='secao'>⚙️ INFORMAÇÕES TÉCNICAS</div>"
                    html += "<table class='table'>"
                    for field_name in campos_tecnicos:
                        if field_name in fields.names():
                            valor = feature[field_name]
                            html += f"""
                            <tr>
                                <td class='campo-destaque'>{self.formatar_nome_campo(field_name)}</td>
                                <td>{self.formatar_valor(valor)}</td>
                            </tr>
                            """
                    html += "</table>"
                    
                    # Seção 4
                    html += "<div class='secao'>📄 OUTRAS INFORMAÇÕES</div>"
                    html += "<table class='table'>"
                    for field_name in campos_outros:
                        if field_name in fields.names():
                            valor = feature[field_name]
                            html += f"""
                            <tr>
                                <td class='campo-destaque'>{self.formatar_nome_campo(field_name)}</td>
                                <td>{self.formatar_valor(valor)}</td>
                            </tr>
                            """
                    html += "</table>"
                    
                    # Coordenadas
                    if feature.hasGeometry():
                        geom = feature.geometry()
                        if not geom.isNull():
                            centroid = geom.centroid().asPoint()
                            html += f"""
                            <div class='secao'>📍 COORDENADAS GEOGRÁFICAS</div>
                            <table class='table'>
                                <tr>
                                    <td class='campo-destaque'>Longitude (X)</td>
                                    <td>{centroid.x():.6f}</td>
                                </tr>
                                <tr>
                                    <td class='campo-destaque'>Latitude (Y)</td>
                                    <td>{centroid.y():.6f}</td>
                                </tr>
                            </table>
                            """
                    return html

                def formatar_nome_campo(self, field_name):
                    """Formata o nome do campo para exibição."""
                    formatacoes = {
                        'id': 'ID',
                        'nome_usuario': 'Nome do Usuário',
                        'nome_operador': 'Nome do Operador',
                        'nu_interferencia_cnarh': 'Interferência CNARH',
                        'nu_cnarh': 'Número CNARH',
                        'rotulo_medidor': 'Rótulo do Medidor',
                        'vazao_nominal': 'Vazão Nominal (m³/s)',
                        'potencia': 'Potência Bomba (KW)',
                        'tipo_medidor': 'Tipo de Medidor',
                        'modo_transmissao': 'Modo de Transmissão',
                        'codigo_uc': 'Código da U.C. de Energia'
                    }
                    return formatacoes.get(field_name, field_name.replace('_', ' ').title())
                
                def formatar_valor(self, valor):
                    """Formata o valor para exibição."""
                    if valor is None or (isinstance(valor, str) and valor.strip() == ''):
                        return "<span class='valor-nulo'>Não informado</span>"
                    
                    if isinstance(valor, (int, float)):
                        if isinstance(valor, float):
                            return f"{valor:.3f}"
                        return str(valor)
                    
                    return str(valor)
            
            # Configurar a ferramenta no canvas
            canvas = iface.mapCanvas()
            self.identify_tool = IdentificacaoTool(canvas, layer)
            canvas.setMapTool(self.identify_tool)
            
            print(f"🔍 Ferramenta de identificação configurada (Modo Múltiplos Pontos).")
            
        except Exception as e:
            print(f"❌ Erro ao configurar identificação: {e}")
            import traceback
            traceback.print_exc()
    
    def configurar_simbologia_e_labels(self, layer):
        """Configura símbolos e labels da camada de forma centralizada."""
        try:
            # 1. CONFIGURAR SÍMBOLO
            symbol = QgsMarkerSymbol.createSimple({
                'name': 'circle',
                'color': '255,0,0,255',  # Vermelho RGB
                'size': '5',
                'outline_style': 'solid',
                'outline_color': '0,0,0,255',  # Preto
                'outline_width': '0.8'
            })
            renderer = QgsSingleSymbolRenderer(symbol)
            layer.setRenderer(renderer)
            
            # 2. CONFIGURAR LABELS
            label_settings = QgsPalLayerSettings()
            
            # Expressão para o label - ajustar nomes dos campos se necessário
            label_expression = """
            CASE 
                WHEN "nu_interferencia_cnarh" IS NOT NULL AND "nu_interferencia_cnarh" != ''
                THEN "nu_interferencia_cnarh" || ' - ' || "rotulo_medidor"
                ELSE "rotulo_medidor"
            END
            """
            
            label_settings.fieldName = label_expression
            label_settings.isExpression = True
            
            # Formato do texto
            text_format = QgsTextFormat()
            text_format.setFont(QFont("Arial", 11, QFont.Bold))
            text_format.setSize(11)
            text_format.setSizeUnit(QgsUnitTypes.RenderPoints)
            text_format.setColor(QColor(0, 0, 0))  # Preto
            
            # Buffer (borda branca)
            buffer_settings = QgsTextBufferSettings()
            buffer_settings.setEnabled(True)
            buffer_settings.setSize(2.0)
            buffer_settings.setSizeUnit(QgsUnitTypes.RenderMillimeters)
            buffer_settings.setColor(QColor(255, 255, 255))  # Branco
            buffer_settings.setOpacity(1.0)
            text_format.setBuffer(buffer_settings)
            
            label_settings.setFormat(text_format)
            
            # Posicionamento
            label_settings.placement = QgsPalLayerSettings.OverPoint
            label_settings.dist = 0
            label_settings.offsetType = QgsPalLayerSettings.FromPoint
            label_settings.quadOffset = QgsPalLayerSettings.QuadrantBelow
            label_settings.yOffset = 3.0
            label_settings.offsetUnits = QgsUnitTypes.RenderMillimeters
            
            # IMPORTANTE: Remover limites de escala para sempre mostrar
            label_settings.scaleVisibility = False
            
            # Prioridade e obstáculos
            label_settings.priority = 10  # Alta prioridade
            label_settings.obstacle = False  # Não bloquear outros labels
            
            # Habilitar labels
            layer.setLabelsEnabled(True)
            layer.setLabeling(QgsVectorLayerSimpleLabeling(label_settings))
            
            # Forçar refresh do renderer
            layer.triggerRepaint()
            
            print("Simbologia e labels configurados com sucesso!")
            print(f"- Símbolo: círculo vermelho 5pt com borda preta")
            print(f"- Labels: sempre visíveis, texto preto com borda branca")
            
        except Exception as e:
            print(f"Erro ao configurar simbologia/labels: {e}")
            import traceback
            traceback.print_exc()

    def exibir_medidores_no_canvas(self):
        """Dá zoom para cobrir todos os medidores selecionados e configura escala 1:50.000."""
        try:           
            # Tentar ambas as possíveis camadas
            layer_names = ["Medidor", "Medidores"]
            layer = None
            
            for layer_name in layer_names:
                layers = QgsProject.instance().mapLayersByName(layer_name)
                if layers:
                    layer = layers[0]
                    print(f"Usando camada: {layer_name}")
                    break
            
            if not layer:
                print("ERRO: Nenhuma camada de intervenções encontrada")
                QMessageBox.information(self, "Mapa", "Camada de intervenções não foi carregada.")
                return
            
            # Verificar CRS e reprojetar se necessário
            layer_crs = layer.crs()
            canvas_crs = iface.mapCanvas().mapSettings().destinationCrs()
            
            print(f"CRS da camada: {layer_crs.authid()}")
            print(f"CRS do canvas: {canvas_crs.authid()}")
            
            # Verificar features
            feature_count = layer.featureCount()
            print(f"Features na camada: {feature_count}")
            
            if feature_count == 0:
                QMessageBox.information(self, "Mapa", 
                    "Nenhum medidor encontrado com os filtros aplicados.")
                return
            
            # Obter extensão das features
            extent = QgsRectangle()
            extent.setMinimal()
            
            features_with_geom = 0
            for feature in layer.getFeatures():
                if feature.hasGeometry():
                    geom = feature.geometry()
                    if not geom.isNull():
                        # Reprojetar geometria se necessário
                        if layer_crs != canvas_crs:
                            transform = QgsCoordinateTransform(
                                layer_crs, 
                                canvas_crs, 
                                QgsProject.instance()
                            )
                            geom.transform(transform)
                        
                        extent.combineExtentWith(geom.boundingBox())
                        features_with_geom += 1
            
            if features_with_geom == 0:
                QMessageBox.information(self, "Mapa", 
                    "Medidor(es) localizado(s), mas sem geometria válida.")
                return
            
            print(f"Features com geometria: {features_with_geom}")
            print(f"Extensão calculada: {extent.toString()}")
            
            # Adicionar margem de 20% (menos que 50% para melhor enquadramento)
            extent.scale(1.2)
            
            # Aplicar extensão
            iface.mapCanvas().setExtent(extent)
            iface.mapCanvas().refresh()
            
            print("Zoom aplicado, configurando escala 1:50.000...")
            
            # Configurar escala 1:50.000 após pequeno delay
            QTimer.singleShot(300, lambda: self.configurar_escala_final(layer))
            
        except Exception as e:
            print(f"Erro ao dar zoom: {e}")
            import traceback
            traceback.print_exc()

    def configurar_escala_final(self, layer):
            """Configura a escala após o zoom estar aplicado.
            
            - Única interferência  → escala fixa 1:100.000 (comportamento original).
            - Múltiplas interferências → zoom dinâmico pelo retângulo envolvente
              da camada (equivalente a 'Aproximar para camada(s)' do QGIS).
            """
            try:
                from qgis.utils import iface
                from qgis.PyQt.QtCore import QTimer

                canvas = iface.mapCanvas()

                # Verificar se há mais de uma interferência nos dados selecionados
                codigos_interf = set(d[5] for d in self.lista_dados_selecionados)
                eh_multipla    = len(codigos_interf) > 1

                if eh_multipla:
                    # === ZOOM DINÂMICO: retângulo envolvente de todas as features ===
                    if not layer:
                        print("[WARN] Camada não disponível para zoom dinâmico")
                        return

                    layer_crs  = layer.crs()
                    canvas_crs = canvas.mapSettings().destinationCrs()

                    extent = QgsRectangle()
                    extent.setMinimal()

                    for feature in layer.getFeatures():
                        if feature.hasGeometry():
                            geom = feature.geometry()
                            if not geom.isNull():
                                if layer_crs != canvas_crs:
                                    transform = QgsCoordinateTransform(
                                        layer_crs, canvas_crs, QgsProject.instance()
                                    )
                                    geom.transform(transform)
                                extent.combineExtentWith(geom.boundingBox())

                    if extent.isNull() or extent.isEmpty():
                        print("[WARN] Extensão vazia para zoom dinâmico")
                        return

                    # Margem de 20 % para não encostar nas bordas
                    extent.scale(1.2)
                    canvas.setExtent(extent)
                    canvas.refresh()
                    print(f"✓ Zoom dinâmico aplicado para {len(codigos_interf)} interferências | "
                          f"Extensão: {extent.toString()}")

                else:
                    # === ESCALA FIXA 1:100.000 (comportamento original) ===
                    extent = canvas.extent()
                    center = extent.center()

                    target_scale = 100000
                    canvas.zoomScale(target_scale)
                    canvas.setCenter(center)

                    current_scale = canvas.scale()
                    print(f"✓ Escala configurada: 1:{int(current_scale)} (alvo: 1:{target_scale})")

                # Refresh final garantido
                canvas.refreshAllLayers()
                if layer:
                    layer.triggerRepaint()

                QTimer.singleShot(200, canvas.refresh)

            except Exception as e:
                print(f"Erro ao configurar escala: {e}")
                import traceback
                traceback.print_exc()

    def center(self):
        """Centraliza a janela na tela."""
        screen_geometry = QDesktopWidget().screenGeometry()
        center_point = screen_geometry.center()
        frame_geometry = self.frameGeometry()
        frame_geometry.moveCenter(center_point)
        self.move(frame_geometry.topLeft())

    def closeEvent(self, event):
        for janela in list(self._janelas_abertas):
            try:
                if janela and janela.isVisible():
                    janela.close()
            except RuntimeError:
                pass
        self._janelas_abertas.clear()

        if hasattr(self, 'canvas_mensal') and self.canvas_mensal:
            if self.cid_hover_mensal is not None:
                self.canvas_mensal.mpl_disconnect(self.cid_hover_mensal)
                self.cid_hover_mensal = None
        
        if hasattr(self, 'canvas_diario') and self.canvas_diario:
            if self.cid_hover_diario is not None:
                self.canvas_diario.mpl_disconnect(self.cid_hover_diario)
                self.cid_hover_diario = None
        
        super().closeEvent(event)

