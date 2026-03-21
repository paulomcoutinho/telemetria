# -*- coding: utf-8 -*-
"""
Módulo: janela_monitoramento.py
=================================
Janela de busca e seleção de medidores para o fluxo de monitoramento.

Funcionalidades:
  - Busca com autocompletar por CNARH, Usuário, Unidade de Automonitoramento e Sistema Hídrico;
  - Listagem em QTreeWidget com seleção múltipla;
  - Detecção automática de múltiplas interferências para modo de agregação;
  - Verificação assíncrona consumo vs. outorgado via VerificacaoOutorgadoThread
    (disparada 300 ms após abertura por QTimer);
  - Minimiza esta janela ao abrir JanelaGraficosMedidor ou
    JanelaMonitoramentoDetalhes e restaura ao retornar;
  - Flag is_selecao_total propagada às janelas filhas.

Autor : SFI/ANA
Versão: 2.0 – Março/2026
"""

# ---------------------------------------------------------------------------
# Imports Qt – widgets, layout e utilitários de UI
# ---------------------------------------------------------------------------
from qgis.PyQt.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QComboBox,
    QTreeWidget, QTreeWidgetItem, QFrame, QMessageBox,
    QSizePolicy, QAbstractItemView, QHeaderView,
    QDateEdit, QCheckBox, QListWidget, QListWidgetItem,
    QDialog, QFrame, QProgressBar, QTableWidget, QTableWidgetItem,
    QTabWidget, QApplication, QGraphicsDropShadowEffect, QDesktopWidget,
    QRadioButton,
)
from qgis.PyQt.QtCore import Qt, QDate, QTimer

from qgis.PyQt.QtGui import QColor, QFont

# ---------------------------------------------------------------------------
# Imports QGIS core / canvas
# ---------------------------------------------------------------------------
from qgis.core import (
    QgsProject,
    QgsRasterLayer,
    QgsCoordinateTransform,
    QgsCoordinateReferenceSystem,
    QgsPointXY,
    QgsRectangle,
)
from qgis.utils import iface
from qgis.gui import QgsMapCanvas

# ---------------------------------------------------------------------------
# Imports Python padrão
# ---------------------------------------------------------------------------
import psycopg2
import os
from datetime import date

# ---------------------------------------------------------------------------
# Sistema de design centralizado
# ---------------------------------------------------------------------------
try:
    from . import ui_tema
except ImportError:
    import ui_tema

# ---------------------------------------------------------------------------
# Módulos filhos instanciados por esta janela
# ---------------------------------------------------------------------------
from .verificacao_outorgado_thread  import VerificacaoOutorgadoThread
from .janela_graficos_medidor       import JanelaGraficosMedidor
from .janela_monitoramento_detalhes import JanelaMonitoramentoDetalhes


class JanelaMonitoramento(QWidget):
    """Janela de busca e seleção de medidores para acesso ao monitoramento de telemetria.

    É o ponto de entrada do fluxo de monitoramento do plugin. Ao ser aberta,
    apresenta um formulário de busca com suporte a autocompletar e, opcionalmente,
    dispara uma verificação automática de consumo versus volume outorgado para
    o mês anterior (via ``VerificacaoOutorgadoThread``), alertando o usuário
    sobre eventuais excedentes.

    O ciclo de uso completo é:
        1. **Busca**: o usuário escolhe o critério (CNARH, Usuário, UAM ou Sistema
           Hídrico) e digita o termo; a busca acontece em tempo real via
           ``buscar_medidores_autocomplete``, exibindo sugestões no
           ``combo_sugestoes``.
        2. **Listagem**: ao confirmar a busca, os medidores encontrados são
           exibidos em ``QTreeWidget`` com colunas de identificação, rótulo e
           sistema hídrico.
        3. **Seleção múltipla**: o usuário pode marcar um ou mais medidores;
           o sistema detecta automaticamente se a seleção abrange múltiplas
           interferências CNARH, habilitando o modo de agregação quando pertinente.
        4. **Navegação**: ao clicar em "Ver Gráficos" ou "Ver Detalhes", a
           janela se minimiza e abre, respectivamente, ``JanelaGraficosMedidor``
           ou ``JanelaMonitoramentoDetalhes``, repassando todos os IDs e metadados
           selecionados.
        5. **Verificação de outorgado**: no primeiro carregamento, um ``QTimer``
           de 300 ms dispara ``_perguntar_verificacao_consumo``, que pergunta ao
           usuário se deseja executar a verificação assíncrona de excedência de
           consumo para o mês anterior; o resultado é exibido em um diálogo
           tabular ao término da thread.

    Attributes:
        tela_inicial (QWidget): Referência à ``TelaInicial``; restaurada ao
            fechar esta janela.
        conn (psycopg2.connection): Conexão ao PostgreSQL em autocommit.
        usuario_logado (str | None): Usuário autenticado na sessão.
        senha (str | None): Credencial repassada à thread de verificação
            para abertura de conexão independente.
        is_selecao_total (bool): ``True`` quando o usuário seleciona todos os
            medidores retornados pela busca (atalho "Selecionar tudo").
        lista_ids_selecionados (list[int]): IDs dos medidores marcados.
        lista_dados_selecionados (list[tuple]): Tuplas com metadados
            completos de cada medidor selecionado.
        combo_criterio_busca (QComboBox): Seletor do critério de pesquisa.
        input_busca (QLineEdit): Campo de busca com disparo automático de
            autocompletar a cada keystroke.
        combo_sugestoes (QComboBox): Dropdown de sugestões de autocompletar,
            visível somente durante a digitação.
        tree_medidores (QTreeWidget): Tabela de resultados da busca com
            suporte a seleção múltipla.
    """
    
    def __init__(self, tela_inicial, conexao, usuario=None, senha=None):
        super().__init__()
        self.tela_inicial = tela_inicial
        self.setWindowTitle("Monitoramento - DURH Diária por Telemetria")
        self.conn = conexao
        self.usuario_logado = usuario
        self.senha = senha
        self._janelas_abertas = []
        self.is_selecao_total = False
        
        try:
            self.conn.rollback()
        except:
            pass        
        self.conn.autocommit = True
        
        self.setFixedSize(730, 600) 
        self.center()
        
        # Armazena listas para seleção múltipla
        self.lista_ids_selecionados = []
        self.lista_dados_selecionados = []

        try:
            self.primary_color = ui_tema.StyleConfig.PRIMARY_COLOR
            self.secondary_color = ui_tema.StyleConfig.SECONDARY_COLOR
            self.text_color = ui_tema.StyleConfig.TEXT_DARK
            self.bg_color = ui_tema.StyleConfig.BACKGROUND_WHITE
            self.border_color = ui_tema.StyleConfig.BORDER_COLOR
        except:
            self.primary_color = "#175cc3"
            self.secondary_color = "#5474b8"
            self.text_color = "#343a40"
            self.bg_color = "#ffffff"
            self.border_color = "#dee2e6"

        self.initUI()
        QTimer.singleShot(300, self._perguntar_verificacao_consumo) 
        
    def initUI(self):
        """Configura a interface de busca de medidores."""
        ui_tema.aplicar_tema_arredondado(self)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        # === BUSCA DE MEDIDOR ===
        busca_container = QWidget()
        busca_container.setStyleSheet("""
            QWidget {
                background-color: white;
                border:1px solid #dee2e6;
                border-radius:5px;
                padding: 10px;
            }
        """)
        busca_layout = QVBoxLayout(busca_container)
        busca_layout.setContentsMargins(10, 10, 10, 10)
        busca_layout.setSpacing(10)

        # Layout superior de busca
        layout_busca_superior = QHBoxLayout()

        lbl_busca = QLabel("Buscar medidor por:")
        lbl_busca.setStyleSheet("font-weight: bold; font-size: 12px;")

        self.combo_criterio_busca = QComboBox()
        self.combo_criterio_busca.addItems(["CNARH", "Usuário", "UAM", "Sistema Hídrico"])
        self.combo_criterio_busca.setFixedWidth(140)
        self.combo_criterio_busca.setStyleSheet("""
            QComboBox {
                padding:5px;
                border:1px solid #ccc;
                border-radius:3px;
                background-color: white;
            }
        """)
        self.combo_criterio_busca.currentIndexChanged.connect(self.atualizar_placeholder_busca)

        self.input_busca = QLineEdit()
        self.input_busca.setPlaceholderText("Digite o termo de busca...")
        self.input_busca.setStyleSheet("""
            QLineEdit {
                border:1px solid #ccc;
                border-radius:3px;
                padding:5px;
                background-color: white;
            }
        """)
        self.input_busca.textChanged.connect(self.buscar_medidores_autocomplete)
        self.input_busca.returnPressed.connect(self.buscar_ou_limpar)

        # ComboBox para sugestões de autocompletar
        self.combo_sugestoes = QComboBox()
        self.combo_sugestoes.setStyleSheet("""
            QComboBox {
                border:1px solid #4CAF50;
                border-radius:3px;
                padding:3px;
                background-color: #f9fff9;
            }
        """)
        self.combo_sugestoes.setVisible(False)
        self.combo_sugestoes.activated.connect(self.selecionar_sugestao_busca)

        self.btn_buscar = QPushButton("Buscar")
        self.btn_buscar.setFixedSize(100, 33)
        self.btn_buscar.setStyleSheet("""
            QPushButton {
                background-color: #5474b8;
                color: white;
                font-size: 11px;
                font-weight: bold;
                border: none;
                border-radius:3px;
            }
            QPushButton:hover {
                background-color: #2050b8;
            }
        """)
        self.btn_buscar.clicked.connect(self.buscar_ou_limpar)

        layout_busca_superior.addWidget(lbl_busca)
        layout_busca_superior.addWidget(self.combo_criterio_busca)
        layout_busca_superior.addWidget(self.input_busca, stretch=1)
        layout_busca_superior.addWidget(self.combo_sugestoes, stretch=1)
        layout_busca_superior.addWidget(self.btn_buscar)

        busca_layout.addLayout(layout_busca_superior)

        # === LISTA DE RESULTADOS DA BUSCA ===
        resultados_label = QLabel("Resultado da busca (selecione um ou mais medidor):")
        resultados_label.setStyleSheet("font-weight: bold; font-size: 12px; margin-top: 10px;")
        busca_layout.addWidget(resultados_label)

        self.lista_resultados = QListWidget()
        self.lista_resultados.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.lista_resultados.setMinimumHeight(150)
        self.lista_resultados.setStyleSheet("""
            QListWidget {
                border:1px solid #ccc;
                border-radius:3px;
                background-color: white;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom:1px solid #f0f0f0;
            }
            QListWidget::item:selected {
                background-color: #5474b8;
                color: white;
            }
            QListWidget::item:hover {
                background-color: #e0f0ff;
            }
        """)
        self.lista_resultados.itemClicked.connect(self.selecionar_medidor_da_lista)

        busca_layout.addWidget(self.lista_resultados)

        # === BOTÕES DE AÇÃO (RODAPÉ) ===
        btn_container = QWidget()
        btn_layout = QHBoxLayout(btn_container)
        btn_layout.setContentsMargins(0, 0, 0, 0)

        action_btn_style = """
            QPushButton {
                background-color: #175cc3;
                color: white;
                font-size: 12px;
                font-weight: bold;
                padding: 8px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
            QPushButton:hover {
                background-color: #5474b8;
            }
        """

        self.btn_selecionar_tudo = QPushButton("Selecionar tudo")
        self.btn_selecionar_tudo.setStyleSheet(action_btn_style)
        self.btn_selecionar_tudo.setEnabled(False)
        self.btn_selecionar_tudo.clicked.connect(self.selecionar_todos_medidores)

        self.btn_abrir_monitoramento = QPushButton("Telemetria")
        self.btn_abrir_monitoramento.setStyleSheet(action_btn_style)
        self.btn_abrir_monitoramento.clicked.connect(self.abrir_janela_detalhes)
        self.btn_abrir_monitoramento.setEnabled(False)

        self.btn_estatisticas = QPushButton("Estatísticas")
        self.btn_estatisticas.setStyleSheet(action_btn_style)
        self.btn_estatisticas.clicked.connect(self.abrir_monitoramento_detalhes)
        self.btn_estatisticas.setEnabled(False)

        btn_layout.addWidget(self.btn_selecionar_tudo)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_abrir_monitoramento)
        btn_layout.addWidget(self.btn_estatisticas)

        busca_layout.addWidget(btn_container)

        layout.addWidget(busca_container)

        # === BOTÃO VOLTAR ===
        btn_voltar = QPushButton("Voltar para Menu")
        btn_voltar.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {ui_tema.StyleConfig.SECONDARY_COLOR};
                border: 1px solid {ui_tema.StyleConfig.SECONDARY_COLOR};
                padding: 8px 20px;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background-color: #f0f0f0;
            }}
        """)
        btn_voltar.clicked.connect(self.voltar)

        layout.addWidget(btn_voltar)

        self.setLayout(layout)

    def _perguntar_verificacao_consumo(self):
        """Diálogo de seleção de período — 12 meses (checkboxes) OU intervalo livre.
 
        Apresenta dois RadioButtons para alternar o modo de verificação:
          • **12 últimos meses** — checkboxes com os 12 meses recentes (padrão).
          • **Por período** — dois QDateEdit (início / fim) com calendário popup.
 
        Ao confirmar, define:
          - ``self.meses_para_verificar`` — lista de itens de fila.
          - ``self._modo_verificacao``    — ``'mensal'`` ou ``'por_periodo'``.
        """
 
        hoje        = date.today()
        nomes_meses = [
            "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
        ]
 
        # Gera 12 meses: atual primeiro, depois os 11 anteriores
        self._meses_opcoes = []
        for i in range(12):
            mes_num = hoje.month - i
            ano_num = hoje.year
            while mes_num < 1:
                mes_num += 12
                ano_num -= 1
            sufixo = " (Mês Atual)" if i == 0 else (" (Mês Anterior)" if i == 1 else "")
            self._meses_opcoes.append({
                'mes':   mes_num,
                'ano':   ano_num,
                'nome':  nomes_meses[mes_num - 1],
                'label': f"{nomes_meses[mes_num - 1]}/{ano_num}{sufixo}",
                'tipo':  'atual' if i == 0 else f"anterior_{i}",
            })
 
        dialog = QDialog(self)
        dialog.setWindowTitle("Verificação de Consumo")
        dialog.setMinimumSize(440, 740)
        dialog.setModal(True)
 
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)
 
        # ── Container principal ────────────────────────────────────────────
        container = QFrame()
        container.setObjectName("ContainerBranco")
        container.setStyleSheet(f"""
            QFrame#ContainerBranco {{
                background-color: {ui_tema.StyleConfig.BACKGROUND_WHITE};
                border-radius: 15px;
                border: 1px solid {ui_tema.StyleConfig.BORDER_COLOR};
                padding: 20px;
            }}
        """)
        container_layout = QVBoxLayout(container)
        container_layout.setSpacing(12)
 
        # Ícone / título / mensagem
        lbl_icon = QLabel("❓")
        lbl_icon.setAlignment(Qt.AlignCenter)
        lbl_icon.setStyleSheet(f"""
            font-size: 36px; color: {ui_tema.StyleConfig.PRIMARY_COLOR};
            background: transparent;
        """)
        container_layout.addWidget(lbl_icon)
 
        titulo = QLabel("Verificação de Consumo Total")
        titulo.setAlignment(Qt.AlignCenter)
        titulo.setStyleSheet(f"""
            font-weight: bold; font-size: 16px;
            color: {ui_tema.StyleConfig.PRIMARY_COLOR};
            background: transparent; font-family: 'Segoe UI', sans-serif;
        """)
        container_layout.addWidget(titulo)
 
        mensagem = QLabel(
            "Deseja verificar se o consumo total de alguma interferência "
            "está acima do volume outorgado?"
        )
        mensagem.setAlignment(Qt.AlignCenter)
        mensagem.setWordWrap(True)
        mensagem.setStyleSheet(f"""
            font-size: 12px; color: {ui_tema.StyleConfig.TEXT_DARK};
            line-height: 1.5; background: transparent;
            font-family: 'Segoe UI', sans-serif;
        """)
        container_layout.addWidget(mensagem)
 
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet(f"background-color: {ui_tema.StyleConfig.BORDER_COLOR};")
        container_layout.addWidget(line)
 
        lbl_selecao = QLabel("Selecione o método de análise:")
        lbl_selecao.setStyleSheet(
            "font-weight: bold; font-size: 12px; color: #495057; margin-top: 5px;"
        )
        container_layout.addWidget(lbl_selecao)
 
        # ── RadioButtons ───────────────────────────────────────────────────
        estilo_radio = f"""
            QRadioButton {{
                font-size: 12px; font-weight: bold;
                color: {ui_tema.StyleConfig.PRIMARY_COLOR}; spacing: 8px;
            }}
            QRadioButton::indicator {{ width: 16px; height: 16px; }}
        """
        radio_layout = QHBoxLayout()
        radio_layout.setSpacing(20)
 
        self._radio_meses   = QRadioButton("12 últimos meses")
        self._radio_periodo = QRadioButton("Por período")
        self._radio_meses.setStyleSheet(estilo_radio)
        self._radio_periodo.setStyleSheet(estilo_radio)
        self._radio_meses.setChecked(True)  # padrão
 
        radio_layout.addWidget(self._radio_meses)
        radio_layout.addWidget(self._radio_periodo)
        radio_layout.addStretch()
        container_layout.addLayout(radio_layout)
 
        # ── QStackedWidget ─────────────────────────────────────────────────
        #stacked = QStackedWidget()
        #stacked.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)
 
        # ── Painel 0: checkboxes de meses ─────────────────────────────────
        painel_meses = QWidget()
        pm_layout = QVBoxLayout(painel_meses)
        pm_layout.setContentsMargins(0, 4, 0, 0)
        pm_layout.setSpacing(6)

        estilo_chk = f"""
            QCheckBox {{
                font-size: 12px; color: #343a40; spacing: 8px;
            }}
            QCheckBox::indicator {{
                width: 18px; height: 18px; border-radius: 3px;
                border: 2px solid {ui_tema.StyleConfig.PRIMARY_COLOR};
            }}
            QCheckBox::indicator:checked {{
                background-color: {ui_tema.StyleConfig.PRIMARY_COLOR};
                border: 2px solid {ui_tema.StyleConfig.PRIMARY_COLOR};
            }}
        """
        checkbox_layout = QVBoxLayout()
        checkbox_layout.setSpacing(8)
        checkbox_layout.setContentsMargins(20, 5, 20, 5)

        self._checkboxes_meses = []
        for opcao in self._meses_opcoes:
            chk = QCheckBox(opcao['label'])
            chk.setStyleSheet(estilo_chk)
            chk.setChecked(opcao['tipo'] in ('atual', 'anterior_1'))
            checkbox_layout.addWidget(chk)
            self._checkboxes_meses.append(chk)

        pm_layout.addLayout(checkbox_layout)

        btn_sel_todos = QPushButton("Selecionar todos")
        btn_sel_todos.setCursor(Qt.PointingHandCursor)
        btn_sel_todos.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {ui_tema.StyleConfig.PRIMARY_COLOR};
                border: 1px solid {ui_tema.StyleConfig.PRIMARY_COLOR};
                border-radius: 6px; padding: 5px 14px;
                font-size: 11px; font-weight: bold;
            }}
            QPushButton:hover {{ background-color: #e8f0fe; }}
        """)

        def _toggle_selecao():
            todos = all(c.isChecked() for c in self._checkboxes_meses)
            for c in self._checkboxes_meses:
                c.setChecked(not todos)
            btn_sel_todos.setText("Selecionar todos" if todos else "Limpar seleção")

        def _sync_label():
            todos = all(c.isChecked() for c in self._checkboxes_meses)
            btn_sel_todos.setText("Limpar seleção" if todos else "Selecionar todos")

        for chk in self._checkboxes_meses:
            chk.stateChanged.connect(lambda _: _sync_label())
        btn_sel_todos.clicked.connect(_toggle_selecao)

        sel_layout = QHBoxLayout()
        sel_layout.addStretch()
        sel_layout.addWidget(btn_sel_todos)
        pm_layout.addLayout(sel_layout)

        container_layout.addWidget(painel_meses)

        # ── Painel de período (QDateEdit) ──────────────────────────────────
        painel_periodo = QWidget()
        pp_layout = QGridLayout(painel_periodo)
        pp_layout.setContentsMargins(20, 10, 20, 10)
        pp_layout.setSpacing(10)

        estilo_date = """
            QDateEdit {
                border: 1px solid #ccc; border-radius: 4px;
                padding: 5px; font-size: 12px;
            }
        """

        lbl_di = QLabel("Data início:")
        lbl_di.setStyleSheet("font-size: 12px; color: #343a40;")
        pp_layout.addWidget(lbl_di, 0, 0)

        self._date_inicio = QDateEdit()
        self._date_inicio.setCalendarPopup(True)
        self._date_inicio.setDisplayFormat("dd/MM/yyyy")
        self._date_inicio.setDate(QDate(hoje.year, hoje.month, 1).addMonths(-1))
        self._date_inicio.setStyleSheet(estilo_date)
        pp_layout.addWidget(self._date_inicio, 0, 1)

        lbl_df = QLabel("Data fim:")
        lbl_df.setStyleSheet("font-size: 12px; color: #343a40;")
        pp_layout.addWidget(lbl_df, 1, 0)

        self._date_fim = QDateEdit()
        self._date_fim.setCalendarPopup(True)
        self._date_fim.setDisplayFormat("dd/MM/yyyy")
        self._date_fim.setDate(QDate(hoje.year, hoje.month, 1).addDays(-1))
        self._date_fim.setStyleSheet(estilo_date)
        pp_layout.addWidget(self._date_fim, 1, 1)

        painel_periodo.setVisible(False)   # oculto por padrão
        container_layout.addWidget(painel_periodo)

        # ── Conectar RadioButtons → show/hide ─────────────────────────────
        def _on_radio_meses(checked):
            if checked:
                painel_periodo.setVisible(False)
                painel_meses.setVisible(True)
                dialog.setMinimumSize(0, 0)
                dialog.setMinimumSize(440, 740)
                dialog.resize(440, 740)

        def _on_radio_periodo(checked):
            if checked:
                painel_meses.setVisible(False)
                painel_periodo.setVisible(True)
                dialog.setMinimumSize(0, 0)
                dialog.resize(440, 420)
                dialog.setMinimumSize(440, 420)

        self._radio_meses.toggled.connect(_on_radio_meses)
        self._radio_periodo.toggled.connect(_on_radio_periodo)
 
        # Aviso de validação
        self.lbl_aviso_selecao = QLabel("⚠️ Selecione pelo menos um período!")
        self.lbl_aviso_selecao.setAlignment(Qt.AlignCenter)
        self.lbl_aviso_selecao.setStyleSheet(
            "color: #dc3545; font-size: 11px; font-weight: bold; margin-top: 5px;"
        )
        self.lbl_aviso_selecao.setVisible(False)
        container_layout.addWidget(self.lbl_aviso_selecao)
 
        layout.addWidget(container)
 
        # ── Botões Não / Sim ───────────────────────────────────────────────
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
 
        btn_nao = QPushButton("Não")
        btn_nao.setCursor(Qt.PointingHandCursor)
        btn_nao.setStyleSheet(f"""
            QPushButton {{
                background-color: {ui_tema.StyleConfig.BORDER_COLOR};
                color: {ui_tema.StyleConfig.TEXT_DARK};
                border-radius: 8px; padding: 10px 30px;
                font-weight: bold; font-size: 12px; border: none;
            }}
            QPushButton:hover {{ background-color: #d0d0d0; }}
        """)
        btn_nao.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_nao)
 
        btn_sim = QPushButton("Sim, verificar")
        btn_sim.setCursor(Qt.PointingHandCursor)
        btn_sim.setStyleSheet(f"""
            QPushButton {{
                background-color: {ui_tema.StyleConfig.PRIMARY_COLOR};
                color: white; border-radius: 8px; padding: 10px 30px;
                font-weight: bold; font-size: 12px; border: none;
            }}
            QPushButton:hover {{ background-color: {ui_tema.StyleConfig.SECONDARY_COLOR}; }}
            QPushButton:disabled {{ background-color: #cccccc; color: #666666; }}
        """)
 
        def validar_e_prosseguir():
            # ── Modo: 6 últimos meses ──────────────────────────────────────
            if self._radio_meses.isChecked():
                selecionados = [
                    self._meses_opcoes[i]
                    for i, chk in enumerate(self._checkboxes_meses)
                    if chk.isChecked()
                ]
                if not selecionados:
                    self.lbl_aviso_selecao.setText("⚠️ Selecione pelo menos um período!")
                    self.lbl_aviso_selecao.setVisible(True)
                    container.setStyleSheet(f"""
                        QFrame#ContainerBranco {{
                            background-color: {ui_tema.StyleConfig.BACKGROUND_WHITE};
                            border-radius: 15px; border: 2px solid #dc3545; padding: 20px;
                        }}
                    """)
                    QTimer.singleShot(200, lambda: container.setStyleSheet(f"""
                        QFrame#ContainerBranco {{
                            background-color: {ui_tema.StyleConfig.BACKGROUND_WHITE};
                            border-radius: 15px;
                            border: 1px solid {ui_tema.StyleConfig.BORDER_COLOR};
                            padding: 20px;
                        }}
                    """))
                    return
                self.lbl_aviso_selecao.setVisible(False)
                self.meses_para_verificar = selecionados
                self._modo_verificacao    = 'mensal'
 
            # ── Modo: por período ──────────────────────────────────────────
            else:
                import calendar as _cal
                from datetime import date as _date
                qi = self._date_inicio.date()
                qf = self._date_fim.date()
                di = _date(qi.year(), qi.month(), qi.day())
                df = _date(qf.year(), qf.month(), qf.day())
                if di > df:
                    self.lbl_aviso_selecao.setText(
                        "⚠️ Data início deve ser anterior à data fim!"
                    )
                    self.lbl_aviso_selecao.setVisible(True)
                    return
                self.lbl_aviso_selecao.setVisible(False)

                nomes_meses = [
                    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
                ]
                # Expandir o intervalo em itens mensais independentes
                self.meses_para_verificar = []
                cur = _date(di.year, di.month, 1)
                while cur <= df:
                    ultimo_dia  = _cal.monthrange(cur.year, cur.month)[1]
                    inicio_mes  = max(di, _date(cur.year, cur.month, 1))
                    fim_mes     = min(df, _date(cur.year, cur.month, ultimo_dia))
                    nome_mes    = f"{nomes_meses[cur.month - 1]}/{cur.year}"
                    sufixo_dias = (
                        f" ({inicio_mes.strftime('%d/%m')} a {fim_mes.strftime('%d/%m')})"
                        if (inicio_mes.day != 1 or fim_mes.day != ultimo_dia)
                        else ""
                    )
                    self.meses_para_verificar.append({
                        'mes':         cur.month,
                        'ano':         cur.year,
                        'nome':        nome_mes + sufixo_dias,
                        'label':       nome_mes + sufixo_dias,
                        'tipo':        f"periodo_{cur.year}_{cur.month:02d}",
                        'data_inicio': inicio_mes,
                        'data_fim':    fim_mes,
                    })
                    # Avança para o 1º do próximo mês
                    cur = (_date(cur.year + 1, 1, 1) if cur.month == 12
                           else _date(cur.year, cur.month + 1, 1))

                self._modo_verificacao = 'por_periodo'
 
            dialog.accept()
            self._confirmar_tempo_processamento_selecao()
 
        btn_sim.clicked.connect(validar_e_prosseguir)
        btn_layout.addWidget(btn_sim)
        layout.addLayout(btn_layout)
 
        shadow = QGraphicsDropShadowEffect(dialog)
        shadow.setBlurRadius(15)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 30))
        dialog.setGraphicsEffect(shadow)
 
        if dialog.exec_() == QDialog.Rejected:
            print("[INFO] Usuário optou por não verificar consumo outorgado")

    @staticmethod
    def _rotulo_periodo(item, campo_nome='nome_mes', campo_ano='ano'):
        """Retorna o rótulo correto para exibição, sem duplicar o ano no modo
        por_periodo.
 
        No modo 'por_periodo' o campo nome já carrega o intervalo completo
        (ex.: '01/02/2026 a 28/02/2026'); concatenar /{ano} duplicaria o ano.
        No modo mensal o comportamento original é mantido (ex.: 'Fevereiro/2026').
        """
        nome = item.get(campo_nome, '')
        ano  = item.get(campo_ano, 0)
        if item.get('tipo') == 'periodo' or item.get('modo') == 'por_periodo':
            return nome
        return f"{nome}/{ano}" if ano else nome
          
    def _confirmar_tempo_processamento_selecao(self):
        """Segunda confirmação: lista os períodos selecionados e pede confirmação."""
        meses      = self.meses_para_verificar
        quantidade = len(meses)
        texto_meses = "<br>• ".join(
            m['nome'] if (m.get('tipo', '').startswith('periodo') or str(m['ano']) in m['nome'])
            else f"{m['nome']}/{m['ano']}"
            for m in meses
        )
 
        dialog = QDialog(self)
        dialog.setWindowTitle("Atenção")
        dialog.setMinimumSize(340, 260)
        dialog.setModal(True)
 
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)
 
        container = QFrame()
        container.setStyleSheet("""
            QFrame {
                background-color: #fff3cd;
                border-radius: 15px; border: 1px solid #ffc107; padding: 20px;
            }
        """)
        cl = QVBoxLayout(container)
        cl.setSpacing(10)
 
        lbl_icon = QLabel("⚠️")
        lbl_icon.setAlignment(Qt.AlignCenter)
        lbl_icon.setStyleSheet("font-size: 36px; color: #856404; background: transparent;")
        cl.addWidget(lbl_icon)
 
        mensagem = QLabel(
            f"A verificação analisará todas as interferências do(s) período(s):<br>"
            f"<b>• {texto_meses}</b><br><br>"
            f"comparando consumo real vs. outorgado.<br><br>"
            f"<span style='color: #856404;'>Deseja continuar?</span>"
        )
        mensagem.setAlignment(Qt.AlignCenter)
        mensagem.setWordWrap(True)
        mensagem.setStyleSheet("""
            font-size: 12px; color: #856404; line-height: 1.5;
            background: transparent; font-family: 'Segoe UI', sans-serif;
        """)
        cl.addWidget(mensagem)
        layout.addWidget(container)
 
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
 
        btn_nao = QPushButton("Não")
        btn_nao.setCursor(Qt.PointingHandCursor)
        btn_nao.setStyleSheet("""
            QPushButton {
                background-color: #6c757d; color: white;
                border-radius: 8px; padding: 10px 30px;
                font-weight: bold; font-size: 12px; border: none;
            }
            QPushButton:hover { background-color: #5a6268; }
        """)
        btn_nao.clicked.connect(dialog.reject)
        btn_layout.addWidget(btn_nao)
 
        btn_sim = QPushButton("Sim, continuar")
        btn_sim.setCursor(Qt.PointingHandCursor)
        btn_sim.setStyleSheet("""
            QPushButton {
                background-color: #28a745; color: white;
                border-radius: 8px; padding: 10px 30px;
                font-weight: bold; font-size: 12px; border: none;
            }
            QPushButton:hover { background-color: #218838; }
        """)
        btn_sim.clicked.connect(dialog.accept)
        btn_layout.addWidget(btn_sim)
        layout.addLayout(btn_layout)
 
        shadow = QGraphicsDropShadowEffect(dialog)
        shadow.setBlurRadius(15); shadow.setXOffset(0); shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 30))
        dialog.setGraphicsEffect(shadow)
 
        if dialog.exec_() == QDialog.Accepted:
            self._iniciar_verificacao_selecionados()
        else:
            print("[INFO] Usuário cancelou após alerta de tempo")
            
    def _iniciar_verificacao_selecionados(self):
        """Monta a fila e inicia o processamento."""
        self._cancelado = False
        self.fila_processamento = [
            {
                'mes':         m['mes'],
                'ano':         m['ano'],
                'nome_mes':    m['nome'],
                'tipo':        m['tipo'],
                'modo':        getattr(self, '_modo_verificacao', 'mensal'),
                'data_inicio': m.get('data_inicio'),
                'data_fim':    m.get('data_fim'),
            }
            for m in self.meses_para_verificar
        ]
        self.resultados_processamento = {}
 
        n = len(self.fila_processamento)
 
        self.progress_dialog = QDialog(self)
        self.progress_dialog.setWindowTitle("Processando...")
        self.progress_dialog.setMinimumSize(450, 240 + max(0, n - 2) * 24)
        self.progress_dialog.setModal(True)
 
        try:
            ui_tema.aplicar_tema_arredondado(self.progress_dialog)
        except Exception:
            pass
 
        layout = QVBoxLayout(self.progress_dialog)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)
 
        container = QFrame()
        container.setObjectName("ContainerBranco")
        container.setStyleSheet(f"""
            QFrame#ContainerBranco {{
                background-color: {ui_tema.StyleConfig.BACKGROUND_WHITE};
                border-radius: 15px;
                border: 1px solid {ui_tema.StyleConfig.BORDER_COLOR};
                padding: 20px;
            }}
        """)
        cl = QVBoxLayout(container)
        cl.setSpacing(12)
 
        self.lbl_icon = QLabel("⏳")
        self.lbl_icon.setAlignment(Qt.AlignCenter)
        self.lbl_icon.setStyleSheet(f"""
            font-size: 36px; color: {ui_tema.StyleConfig.PRIMARY_COLOR};
            background: transparent; padding: 10px;
        """)
        cl.addWidget(self.lbl_icon)
 
        titulo = QLabel("Verificando Interferências")
        titulo.setAlignment(Qt.AlignCenter)
        titulo.setStyleSheet(f"""
            font-weight: bold; font-size: 16px;
            color: {ui_tema.StyleConfig.PRIMARY_COLOR};
            background: transparent; font-family: 'Segoe UI', sans-serif;
        """)
        cl.addWidget(titulo)
 
        self.lbl_mes_atual_processando = QLabel()
        self.lbl_mes_atual_processando.setAlignment(Qt.AlignCenter)
        self.lbl_mes_atual_processando.setStyleSheet(f"""
            font-size: 14px; font-weight: bold; color: #175cc3;
            background-color: #e3f2fd; border-radius: 8px;
            padding: 8px 15px; border: 2px solid #175cc3;
            font-family: 'Segoe UI', sans-serif;
        """)
        cl.addWidget(self.lbl_mes_atual_processando)
 
        # Rótulo correto na fila de progresso
        linhas_fila = "<b>Fila de processamento:</b><br>" + "".join(
            f"<span style='color:#6c757d;'>• "
            + (m['nome_mes'] if (m.get('modo') == 'por_periodo'
               or str(m['ano']) in m['nome_mes'])
               else f"{m['nome_mes']}/{m['ano']}")
            + "</span><br>"
            for m in self.fila_processamento
        )
        self.lbl_mensagem = QLabel(linhas_fila)
        self.lbl_mensagem.setAlignment(Qt.AlignCenter)
        self.lbl_mensagem.setWordWrap(True)
        self.lbl_mensagem.setStyleSheet(f"""
            font-size: 12px; color: {ui_tema.StyleConfig.TEXT_DARK};
            line-height: 1.6; background: transparent;
            font-family: 'Segoe UI', sans-serif;
        """)
        cl.addWidget(self.lbl_mensagem)
 
        if n > 1:
            self.progress_bar = QProgressBar()
            self.progress_bar.setRange(0, n)
            self.progress_bar.setValue(0)
            self.progress_bar.setStyleSheet("""
                QProgressBar {
                    border: 1px solid #dee2e6; border-radius: 5px;
                    background-color: #f8f9fa; text-align: center;
                }
                QProgressBar::chunk { background-color: #175cc3; border-radius: 5px; }
            """)
            cl.addWidget(self.progress_bar)
        else:
            self.progress_bar = None
 
        layout.addWidget(container)
 
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setCursor(Qt.PointingHandCursor)
        btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: transparent; color: #999999;
                border: 1px solid #cccccc; border-radius: 6px;
                padding: 6px 20px; font-size: 11px;
            }
            QPushButton:hover { color: #dc3545; border-color: #dc3545; }
        """)
        btn_cancelar.clicked.connect(self._on_verificacao_cancelada)
        layout.addWidget(btn_cancelar, alignment=Qt.AlignCenter)
 
        shadow = QGraphicsDropShadowEffect(self.progress_dialog)
        shadow.setBlurRadius(15); shadow.setXOffset(0); shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 30))
        self.progress_dialog.setGraphicsEffect(shadow)
 
        self.progress_dialog.show()
        QApplication.setOverrideCursor(Qt.WaitCursor)
        self._processar_proximo_da_fila()
        return True
  
    def _processar_proximo_da_fila(self):
        """Processa o próximo item da fila de verificação."""
 
        if not self.fila_processamento or self._cancelado:
            self._finalizar_processamento_fila()
            return
 
        item_atual = self.fila_processamento.pop(0)
        self.item_atual_processamento = item_atual
 
        self._atualizar_mensagem_progresso(item_atual, 'processando')
 
        # Repassa modo e datas à thread
        self.worker_thread_atual = VerificacaoOutorgadoThread(
            self.conn,
            item_atual['mes'],
            item_atual['ano'],
            item_atual['nome_mes'],
            self.senha,
            modo=item_atual.get('modo', 'mensal'),
            data_inicio=item_atual.get('data_inicio'),
            data_fim=item_atual.get('data_fim'),
        )
 
        self.worker_thread_atual.resultado_signal.connect(
            lambda dados, mes, ano: self._on_item_fila_concluido(dados, mes, ano, item_atual['tipo'])
        )
        self.worker_thread_atual.erro_signal.connect(self._on_verificacao_erro)
        self.worker_thread_atual.progresso_signal.connect(self._on_progresso_atualizado)
        self.worker_thread_atual.finished.connect(self._processar_proximo_da_fila)
 
        self.worker_thread_atual.start()
 
    def _atualizar_mensagem_progresso(self, item_atual, status):
        """Atualiza o indicador visual do período em processamento."""
        # Rótulo sem duplicar o ano no modo por_periodo
        rotulo = (
            item_atual['nome_mes'] if item_atual.get('modo') == 'por_periodo'
            else f"{item_atual['nome_mes']}/{item_atual['ano']}"
        )
 
        if hasattr(self, 'lbl_mes_atual_processando') and self.lbl_mes_atual_processando:
            if status == 'processando':
                self.lbl_mes_atual_processando.setText(f"▶ PROCESSANDO: {rotulo}")
                self.lbl_mes_atual_processando.setStyleSheet(f"""
                    font-size: 14px; font-weight: bold;
                    color: #ffffff; background-color: #175cc3;
                    border-radius: 8px; padding: 8px 15px;
                    border: 2px solid #175cc3; font-family: 'Segoe UI', sans-serif;
                """)
            else:
                self.lbl_mes_atual_processando.setText(f"✓ CONCLUÍDO: {rotulo}")
                self.lbl_mes_atual_processando.setStyleSheet(f"""
                    font-size: 14px; font-weight: bold;
                    color: #155724; background-color: #d4edda;
                    border-radius: 8px; padding: 8px 15px;
                    border: 2px solid #28a745; font-family: 'Segoe UI', sans-serif;
                """)
 
        if hasattr(self, 'progress_bar') and self.progress_bar and status == 'concluido':
            self.progress_bar.setValue(self.progress_bar.value() + 1)
 
        meses     = getattr(self, 'meses_para_verificar', [])
        concluidos = set(getattr(self, 'resultados_processamento', {}).keys())
 
        if len(meses) <= 1:
            if status == 'processando':
                msg = (
                    f"<b>Analisando:</b> {rotulo}<br>"
                    f"<span style='color:#175cc3;font-weight:bold;'>▶ Em andamento...</span><br><br>"
                    f"<span style='color:#666;font-size:11px;'>Aguarde...</span>"
                )
            else:
                msg = (
                    f"<b>Concluído:</b> {rotulo}<br>"
                    f"<span style='color:#28a745;'>✓ Verificação finalizada</span><br><br>"
                    f"<span style='color:#666;font-size:11px;'>Preparando resultados...</span>"
                )
            if hasattr(self, 'lbl_mensagem') and self.lbl_mensagem:
                self.lbl_mensagem.setText(msg)
            return
 
        linhas = ["<b>Fila de processamento:</b><br>"]
        for pos, opcao in enumerate(meses, 1):
            tipo        = opcao['tipo']
            # Rótulo correto para cada item da fila
            label_opcao = (
                opcao['nome'] if (tipo.startswith('periodo') or str(opcao['ano']) in opcao['nome'])
                else f"{opcao['nome']}/{opcao['ano']}"
            )
            if tipo == item_atual['tipo']:
                cor    = "#175cc3" if status == 'processando' else "#28a745"
                icone  = "▶" if status == 'processando' else "✓"
                estado = "processando..." if status == 'processando' else "concluído"
                linhas.append(
                    f"<span style='color:{cor};font-weight:bold;'>"
                    f"{icone} {pos}. {label_opcao} ({estado})</span><br>"
                )
            elif tipo in concluidos:
                linhas.append(
                    f"<span style='color:#28a745;'>✓ {pos}. {label_opcao} (concluído)</span><br>"
                )
            else:
                linhas.append(
                    f"<span style='color:#6c757d;'>{pos}. {label_opcao} (aguardando...)</span><br>"
                )
 
        linhas.append(
            "<br><span style='color:#666;font-size:11px;'>Processando em sequência. Aguarde...</span>"
        )
        if hasattr(self, 'lbl_mensagem') and self.lbl_mensagem:
            self.lbl_mensagem.setText("".join(linhas))
            
    def _on_item_fila_concluido(self, resultados, nome_mes, ano, tipo):
        """Callback quando um item da fila termina."""
        # Recuperar mes e data_inicio do item atual para permitir ordenação
        item = getattr(self, 'item_atual_processamento', {})

        # Separar alertas (exibição na tela) de informativos (somente Excel)
        alertas      = [r for r in resultados if len(r) <= 9 or r[9]]
        informativos = [r for r in resultados if len(r) > 9 and not r[9]]

        self.resultados_processamento[tipo] = {
            'resultados':      alertas,       # apenas alertas → tela
            'informativos':    informativos,  # dentro do limite → só Excel
            'nome_mes':        nome_mes,
            'ano':             ano,
            'mes':             item.get('mes', 0),
            'data_inicio':     item.get('data_inicio'),
        }

        self._atualizar_mensagem_progresso(self.item_atual_processamento, 'concluido')

        if alertas:
            print(f"[INFO] {tipo}: {len(alertas)} alertas, {len(informativos)} informativos")
        else:
            print(f"[INFO] {tipo}: Nenhum alerta. {len(informativos)} registros informativos.")

    def _finalizar_processamento_fila(self):
        """Finaliza a fila, ordena cronologicamente e por consumo, e exibe."""
        QApplication.restoreOverrideCursor()

        if self.progress_dialog:
            self.progress_dialog.close()
            self.progress_dialog = None

        self.worker_thread_atual = None

        meses_com_resultado = []
        for opcao in getattr(self, 'meses_para_verificar', []):
            dados = self.resultados_processamento.get(opcao['tipo'])
            if dados is not None:
                meses_com_resultado.append(dados)

        # ── 1. Ordenação cronológica dos meses ────────────────────────────────
        # Mensal    → chave = (ano, mes, 0)
        # Por_periodo → chave = (ano, mes, dia) de data_inicio
        def _chave_ordem(d):
            di = d.get('data_inicio')
            if di:
                return (di.year, di.month, di.day)
            return (d.get('ano', 0), d.get('mes', 0), 0)

        meses_com_resultado.sort(key=_chave_ordem)

        # ── 2. Ordenação das interferências por consumo ───────────────────────
        # Referência: consumo do mês mais antigo que contém a interferência.
        # Fallback: próximo mês disponível em ordem cronológica.
        # Resultado: todas as listas de meses ficam na mesma ordem global,
        # garantindo alinhamento entre abas e colunas do Excel.
        if meses_com_resultado:
            # Índice por mês (já em ordem cronológica): {cod_interf: consumo}
            consumo_por_mes = [
                {r[0]: float(r[6]) if len(r) > 6 else 0.0
                 for r in (d.get('resultados') or [])}
                for d in meses_com_resultado
            ]

            # União de todas as interferências encontradas em qualquer mês
            todas_interf = {
                r[0]
                for d in meses_com_resultado
                for r in (d.get('resultados') or [])
            }

            def _chave_consumo(cod_interf):
                """Retorna o consumo negativo do mês mais antigo com dados
                para esta interferência (negativo → sort descendente)."""
                for idx_c in consumo_por_mes:
                    if cod_interf in idx_c:
                        return -idx_c[cod_interf]
                return 0.0

            # Ordem global: maior consumo (mês mais antigo) primeiro
            ordem_interf = sorted(todas_interf, key=_chave_consumo)
            pos          = {cod: i for i, cod in enumerate(ordem_interf)}

            # Aplicar a mesma ordem a todos os meses
            for d in meses_com_resultado:
                if d.get('resultados'):
                    d['resultados'].sort(key=lambda r: pos.get(r[0], 999999))

        # ── 3. Exibir resultado ───────────────────────────────────────────────
        if any(d.get('resultados') for d in meses_com_resultado):
            self.mostrar_alerta_consumo_abas(meses_com_resultado)
        else:
            QMessageBox.information(
                self,
                "Verificação Concluída",
                "Nenhuma interferência com consumo acima do volume outorgado "
                "foi encontrada no(s) período(s) selecionado(s)."
            )

        self.resultados_processamento = {}
        self.fila_processamento       = []
        
    def _on_progresso_atualizado(self, mensagem):
        """Atualiza a mensagem de progresso."""
        if hasattr(self, 'lbl_mensagem') and self.lbl_mensagem:
            self.lbl_mensagem.setText(
                f"{mensagem}<br><br>"
                f"<span style='color: #666666; font-size: 11px;'>"
                f"Esta operação pode levar alguns minutos</span>"
            )

    def _on_verificacao_erro(self, mensagem_erro):
        """Callback quando ocorre erro na verificação."""
        QMessageBox.critical(self, "Erro na Verificação", 
            f"Falha ao verificar consumo outorgado:\n\n{mensagem_erro}")
        print(f"[ERRO] Verificação: {mensagem_erro}")

    def _on_verificacao_cancelada(self):
        """Callback quando usuário clica em Cancelar."""
        print("[INFO] Usuário solicitou cancelamento...")
        
        # Marcar flag de cancelamento
        self._cancelado = True
        
        # Atualizar mensagem
        if hasattr(self, 'lbl_mensagem') and self.lbl_mensagem:
            self.lbl_mensagem.setText(
                "<b>Cancelando operação...</b><br>"
                "Aguardando resposta do banco de dados"
            )
        
        # Desabilitar botão
        botao = self.progress_dialog.sender()
        if botao:
            botao.setEnabled(False)
            botao.setText("Cancelando...")
               
        # Cancelar thread atual se existir
        if hasattr(self, 'worker_thread_atual') and self.worker_thread_atual:
            if self.worker_thread_atual.isRunning():
                self.worker_thread_atual.cancelar()
                if not self.worker_thread_atual.wait(3000):
                    self.worker_thread_atual.terminate()
                    self.worker_thread_atual.wait(1000)
        
        # Limpar fila para não processar mais nada
        self.fila_processamento = []
        
        self._finalizar_processamento_fila()
        
        QMessageBox.information(
            self,
            "Verificação Cancelada",
            "A verificação de consumo foi cancelada."
        )
                  
    def mostrar_alerta_consumo_abas(self, lista_meses):
        """Exibe resultados em abas, uma por mês/período."""
        dialog = QDialog(self)
        dialog.setWindowTitle("⚠️ Alerta - Consumo Acima do Outorgado")
        dialog.setMinimumSize(1100, 600)
        dialog.setModal(False)
 
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
 
        # Cabeçalho
        header = QWidget()
        header.setStyleSheet("""
            QWidget {
                background-color: #fff3cd; border: 1px solid #ffc107;
                border-radius: 8px; padding: 15px;
            }
        """)
        hl = QVBoxLayout(header)
 
        titulo = QLabel("⚠️ Interferências com Consumo Acima do Volume Outorgado")
        titulo.setStyleSheet("font-size: 16px; font-weight: bold; color: #856404;")
        hl.addWidget(titulo)
 
        # Rótulo correto no subtítulo: sem duplicar o ano no modo por_periodo
        partes = [
            f"<b>"
            + (d['nome_mes'] if not d.get('ano') else f"{d['nome_mes']}/{d['ano']}")
            + f":</b> {len(d['resultados'])} alerta(s)"
            for d in lista_meses if d and d.get('resultados') is not None
        ]
        subtitulo = QLabel(
            " | ".join(partes) + "<br>"
            "<span style='font-size:11px;'>💡 Clique no cabeçalho para ordenar | "
            "Duplo clique em uma linha para visualizar os gráficos</span>"
        )
        subtitulo.setStyleSheet("font-size: 12px; color: #856404; margin-top: 5px;")
        hl.addWidget(subtitulo)
        layout.addWidget(header)
 
        # Abas
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #dee2e6; border-radius: 8px; background-color: white;
            }
            QTabBar::tab {
                background-color: #f8f9fa; border: 1px solid #dee2e6;
                border-bottom: none; border-top-left-radius: 8px;
                border-top-right-radius: 8px; padding: 10px 20px;
                min-width: 180px; font-weight: bold; font-size: 12px; color: #6c757d;
            }
            QTabBar::tab:selected {
                background-color: #ffffff; color: #343a40;
                border-bottom: 2px solid #175cc3;
            }
            QTabBar::tab:hover:!selected { background-color: #e9ecef; color: #495057; }
        """)
 
        self.tabelas_abas = {}
        self.dados_abas   = {}
        self.ordem_abas   = {}
 
        # Oculta a barra de abas quando há apenas um resultado
        lista_validos = [d for d in lista_meses if d and d.get('resultados') is not None]
        if len(lista_validos) == 1:
            self.tab_widget.tabBar().setVisible(False)

        MESES_ABREV = {
            "Janeiro": "Jan", "Fevereiro": "Fev", "Março": "Mar",
            "Abril": "Abr", "Maio": "Mai", "Junho": "Jun",
            "Julho": "Jul", "Agosto": "Ago", "Setembro": "Set",
            "Outubro": "Out", "Novembro": "Nov", "Dezembro": "Dez",
        }

        def _abreviar_rotulo(rotulo):
            """Jan/2026 (14/01 a 31/01) · 6  →  cabe em min-width:150px"""
            for ext, abrev in MESES_ABREV.items():
                rotulo = rotulo.replace(ext, abrev)
            return rotulo
 
        for idx, dados_mes in enumerate(lista_meses):
            if dados_mes and dados_mes.get('resultados') is not None:
                identificador = f"mes_{idx}"
                n = len(dados_mes['resultados'])

                # ano=0 ocorre no modo por_periodo — usar data_inicio como fonte
                ano_aba = dados_mes.get('ano') or 0
                mes_aba = dados_mes.get('mes') or 0
                di      = dados_mes.get('data_inicio')
                if (not ano_aba or not mes_aba) and di:
                    ano_aba = di.year
                    mes_aba = di.month

                aba = self._criar_aba_mes(
                    dados_mes['resultados'],
                    dados_mes['nome_mes'],
                    ano_aba,
                    mes_aba,
                    identificador,
                )
                rotulo_aba = (
                    dados_mes['nome_mes']
                    if (not ano_aba or str(ano_aba) in dados_mes['nome_mes'])
                    else f"{dados_mes['nome_mes']}/{ano_aba}"
                )
                self.tab_widget.addTab(aba, _abreviar_rotulo(f"{rotulo_aba} ({n})"))
 
        layout.addWidget(self.tab_widget)
 
        # Legenda
        leg_layout = QHBoxLayout()
        leg_critico = QLabel("⬤ Alerta Crítico (>100% do outorgado)")
        leg_critico.setStyleSheet("color: #dc3545; font-size: 11px;")
        leg_layout.addWidget(leg_critico)
        leg_layout.addStretch()
        layout.addLayout(leg_layout)
 
        # Botões
        btn_layout = QHBoxLayout()
 
        btn_exportar = QPushButton("Exportar para Excel")
        btn_exportar.setStyleSheet(f"""
            QPushButton {{
                background-color: {self.primary_color}; color: white;
                border: none; border-radius: 5px;
                padding: 10px 20px; font-weight: bold;
            }}
            QPushButton:hover {{ background-color: {self.secondary_color}; }}
        """)
        btn_exportar.clicked.connect(
            lambda: self.exportar_alerta_excel_meses(lista_meses)
        )
        btn_layout.addWidget(btn_exportar)
        btn_layout.addStretch()
 
        btn_fechar = QPushButton("Fechar")
        btn_fechar.setStyleSheet("""
            QPushButton {
                background-color: #6c757d; color: white; border: none;
                border-radius: 5px; padding: 10px 20px;
            }
            QPushButton:hover { background-color: #5a6268; }
        """)
        btn_fechar.clicked.connect(dialog.close)
        btn_layout.addWidget(btn_fechar)
 
        layout.addLayout(btn_layout)
        dialog.finished.connect(lambda: self._limpar_referencias_alerta_abas())
        self.dialog_alerta = dialog
        dialog.show()
        self._janelas_abertas.append(self.dialog_alerta)
        
    def _criar_aba_mes(self, dados, nome_mes, ano, mes_num, identificador_aba):
        """Cria aba com tabela de 9 colunas para um mês/período."""
        aba = QWidget()
        al  = QVBoxLayout(aba)
        al.setContentsMargins(10, 10, 10, 10)
        al.setSpacing(5)
 
        tabela = QTableWidget()
        tabela.setColumnCount(9)
        tabela.setHorizontalHeaderLabels([
            "INT_CD", "CNARH", "Empreendimento", "Usuário", "Operador",
            "Medidor(es)", "Consumo (m³)", "Outorgado (m³)", "% Acrescido",
        ])
        tabela.setStyleSheet("""
            QTableWidget { border: none; gridline-color: #dee2e6; }
            QHeaderView::section {
                background-color: #175cc3; color: white;
                padding: 8px; font-weight: bold; border: none; cursor: pointer;
            }
            QHeaderView::section:hover   { background-color: #2050b8; }
            QHeaderView::section:pressed { background-color: #5474b8; }
            QTableWidget::item { padding: 6px; border-bottom: 1px solid #dee2e6; }
            QTableWidget::item:selected { background-color: #e3f2fd; color: #175cc3; }
        """)
 
        self.dados_abas[identificador_aba] = {
            'original': list(dados), 'atual': list(dados),
            'nome_mes': nome_mes, 'ano': ano,
            'mes_num':  mes_num,
        }
        self.ordem_abas[identificador_aba]   = {'coluna': None, 'decrescente': False}
        self.tabelas_abas[identificador_aba] = tabela
 
        tabela.horizontalHeader().setSectionsClickable(True)
        tabela.horizontalHeader().sectionClicked.connect(
            lambda col, id_aba=identificador_aba: self._ordenar_tabela_aba(col, id_aba)
        )
        tabela.itemDoubleClicked.connect(
            lambda item, id_aba=identificador_aba, nm=nome_mes, a=ano:
                self._on_interferencia_clicada_aba(item, id_aba, nm, a)
        )
 
        self._preencher_tabela_aba(identificador_aba)
 
        hdr = tabela.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # INT_CD
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # CNARH
        hdr.setSectionResizeMode(2, QHeaderView.Stretch)           # Empreendimento
        hdr.setSectionResizeMode(3, QHeaderView.Stretch)           # Usuário
        hdr.setSectionResizeMode(4, QHeaderView.Stretch)           # Operador
        hdr.setSectionResizeMode(5, QHeaderView.Stretch)           # Medidor(es)
        hdr.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Consumo
        hdr.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # Outorgado
        hdr.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # % Acrescido
 
        tabela.setAlternatingRowColors(True)
        tabela.setSelectionBehavior(QAbstractItemView.SelectRows)
        tabela.setEditTriggers(QAbstractItemView.NoEditTriggers)
 
        al.addWidget(tabela)
        return aba
        
    def _preencher_tabela_aba(self, identificador_aba):
        """Preenche a tabela de uma aba (9 colunas, incluindo Empreendimento
        e % Acrescido)."""
        tabela = self.tabelas_abas.get(identificador_aba)
        dados  = self.dados_abas.get(identificador_aba, {}).get('atual', [])
        if not tabela or not dados:
            return
 
        fmt = lambda v: f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        tabela.setRowCount(len(dados))
 
        for row_idx, row_data in enumerate(dados):
            # Tupla da thread:
            #   [0] cod_interf  [1] cnarh  [2] nome_empreendimento
            #   [3] usuario     [4] operador  [5] rotulos
            #   [6] consumo     [7] outorgado [8] percentual
            cod_interf        = row_data[0]
            cnarh             = row_data[1]
            nome_empreend     = row_data[2] if len(row_data) > 2 else None
            usuario           = row_data[3] if len(row_data) > 3 else None
            operador          = row_data[4] if len(row_data) > 4 else None
            rotulos           = row_data[5] if len(row_data) > 5 else None
            consumo           = row_data[6] if len(row_data) > 6 else 0
            outorgado         = row_data[7] if len(row_data) > 7 else 0
            outorgado_f       = float(outorgado) if outorgado else 0.0
            consumo_f         = float(consumo)   if consumo   else 0.0
            percentual        = float(row_data[8]) if len(row_data) > 8 else (
                round((consumo_f / outorgado_f - 1.0) * 100.0, 1) if outorgado_f > 0 else 0.0
            )
 
            # Colunas de texto: cols 0-5
            for col, val in enumerate([cod_interf, cnarh, nome_empreend,
                                        usuario, operador, rotulos]):
                item = QTableWidgetItem(str(val) if val else "N/A")
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setData(Qt.UserRole,
                    int(val) if col == 0 and val and str(val).isdigit()
                    else str(val).lower() if val else "")
                tabela.setItem(row_idx, col, item)
 
            # Consumo (col 6)
            item = QTableWidgetItem(fmt(consumo))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setForeground(QColor("#dc3545"))
            item.setFont(QFont("Arial", 9, QFont.Bold))
            item.setData(Qt.UserRole, consumo_f)
            tabela.setItem(row_idx, 6, item)
 
            # Outorgado (col 7)
            item = QTableWidgetItem(fmt(outorgado))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setForeground(QColor("#28a745"))
            item.setData(Qt.UserRole, outorgado_f)
            tabela.setItem(row_idx, 7, item)
 
            # % Acrescido (col 8)
            pct_txt = f"+{percentual:.1f}%".replace(".", ",")
            item = QTableWidgetItem(pct_txt)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setTextAlignment(Qt.AlignCenter)
            item.setForeground(QColor("#dc3545"))
            item.setFont(QFont("Arial", 9, QFont.Bold))
            item.setData(Qt.UserRole, percentual)
            tabela.setItem(row_idx, 8, item)
 
            # Destaque de linha: consumo > outorgado (limiar 100%)
            if outorgado_f > 0 and consumo_f > outorgado_f:
                for col in range(tabela.columnCount()):
                    tabela.item(row_idx, col).setBackground(QColor("#f8d7da"))
                    
    def _ordenar_tabela_aba(self, coluna, identificador_aba):
        dados_info = self.dados_abas.get(identificador_aba)
        if not dados_info:
            return
 
        # [0]cod_interf  [1]cnarh  [2]empreendimento  [3]usuario  [4]operador
        # [5]rotulos      [6]consumo  [7]outorgado  [8]percentual
        chaves = {
            0: lambda x: int(x[0]) if x[0] and str(x[0]).isdigit() else 0,
            1: lambda x: str(x[1] or "").lower(),
            2: lambda x: str(x[2] or "").lower(),
            3: lambda x: str(x[3] or "").lower(),
            4: lambda x: str(x[4] or "").lower(),
            5: lambda x: str(x[5] or "").lower(),
            6: lambda x: float(x[6]) if len(x) > 6 else 0.0,
            7: lambda x: float(x[7]) if len(x) > 7 else 0.0,
            8: lambda x: float(x[8]) if len(x) > 8 else 0.0,
        }
        ordem = self.ordem_abas.get(identificador_aba, {'coluna': None, 'decrescente': False})
        if ordem['coluna'] == coluna:
            ordem['decrescente'] = not ordem['decrescente']
        else:
            ordem['coluna'] = coluna; ordem['decrescente'] = False
        self.ordem_abas[identificador_aba] = ordem
 
        dados_info['atual'].sort(
            key=chaves.get(coluna, lambda x: str(x[0]).lower()),
            reverse=ordem['decrescente']
        )
        self._atualizar_cabecalho_aba(identificador_aba, coluna, ordem['decrescente'])
        self._preencher_tabela_aba(identificador_aba)

    def _atualizar_cabecalho_aba(self, identificador_aba, coluna, decrescente):
        tabela = self.tabelas_abas.get(identificador_aba)
        if not tabela:
            return
        titulos = [
            "INT_CD", "CNARH", "Empreendimento", "Usuário", "Operador",
            "Medidor(es)", "Consumo (m³)", "Outorgado (m³)", "% Acrescido",
        ]
        ind = " ▼" if decrescente else " ▲"
        tabela.setHorizontalHeaderLabels([
            t + ind if i == coluna else t for i, t in enumerate(titulos)
        ])
        
    def _on_interferencia_clicada_aba(self, item, identificador_aba, nome_mes, ano):
        """Callback para duplo clique em uma interferência de uma aba."""
        if not item:
            return
        
        dados_atual = self.dados_abas.get(identificador_aba, {}).get('atual', [])
        row = item.row()
        if row < 0 or row >= len(dados_atual):
            return
        
        dados_interf = dados_atual[row]
        cod_interf   = dados_interf[0]
        rotulos      = dados_interf[4]

        # Recuperar mês e ano desta aba para pré-selecionar o gráfico
        info_aba = self.dados_abas.get(identificador_aba, {})
        ano_aba  = info_aba.get('ano') or ano
        mes_aba  = info_aba.get('mes_num')  # definido em _criar_aba_mes

        self._abrir_graficos_interferencia(cod_interf, rotulos,
                                           ano_inicial=ano_aba,
                                           mes_inicial=mes_aba)

    def _limpar_referencias_alerta_abas(self):
        """Limpa as referências das abas de alertas."""
        self.tabelas_abas = {}
        self.dados_abas = {}
        self.ordem_abas = {}
        self.tab_widget = None
    
    def _preencher_tabela_alerta(self):
        if not self.table_alerta or not self.dados_alerta_atual:
            return
 
        fmt = lambda v: f"{float(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
        self.table_alerta.setRowCount(len(self.dados_alerta_atual))
 
        for row_idx, row_data in enumerate(self.dados_alerta_atual):
            cod_interf    = row_data[0]
            cnarh         = row_data[1]
            nome_empreend = row_data[2] if len(row_data) > 2 else None
            usuario       = row_data[3] if len(row_data) > 3 else None
            operador      = row_data[4] if len(row_data) > 4 else None
            rotulos       = row_data[5] if len(row_data) > 5 else None
            consumo       = row_data[6] if len(row_data) > 6 else 0
            outorgado     = row_data[7] if len(row_data) > 7 else 0
            outorgado_f   = float(outorgado) if outorgado else 0.0
            consumo_f     = float(consumo)   if consumo   else 0.0
            percentual    = float(row_data[8]) if len(row_data) > 8 else (
                round((consumo_f / outorgado_f - 1.0) * 100.0, 1) if outorgado_f > 0 else 0.0
            )
 
            for col, val in enumerate([cod_interf, cnarh, nome_empreend,
                                        usuario, operador, rotulos]):
                item = QTableWidgetItem(str(val) if val else "N/A")
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setData(Qt.UserRole,
                    int(val) if col == 0 and val and str(val).isdigit()
                    else str(val).lower() if val else "")
                self.table_alerta.setItem(row_idx, col, item)
 
            # Consumo (col 6)
            item = QTableWidgetItem(fmt(consumo))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setForeground(QColor("#dc3545"))
            item.setFont(QFont("Arial", 9, QFont.Bold))
            item.setData(Qt.UserRole, consumo_f)
            self.table_alerta.setItem(row_idx, 6, item)
 
            # Outorgado (col 7)
            item = QTableWidgetItem(fmt(outorgado))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setForeground(QColor("#28a745"))
            item.setData(Qt.UserRole, outorgado_f)
            self.table_alerta.setItem(row_idx, 7, item)
 
            # % Acrescido (col 8)
            pct_txt = f"+{percentual:.1f}%".replace(".", ",")
            item = QTableWidgetItem(pct_txt)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setTextAlignment(Qt.AlignCenter)
            item.setForeground(QColor("#dc3545"))
            item.setFont(QFont("Arial", 9, QFont.Bold))
            item.setData(Qt.UserRole, percentual)
            self.table_alerta.setItem(row_idx, 8, item)
 
            # Destaque de linha: consumo > outorgado (100%)
            if outorgado_f > 0 and consumo_f > outorgado_f:
                for col in range(self.table_alerta.columnCount()):
                    self.table_alerta.item(row_idx, col).setBackground(QColor("#f8d7da"))

    def exportar_alerta_excel_meses(self, lista_meses):
        """Exporta todos os meses verificados em um único arquivo Excel
        com colunas de consumo/outorgado/% Acrescido lado a lado."""
        import os, sys
        import unicodedata
        from datetime import datetime

        meses_para_exportar = [d for d in lista_meses if d and d.get('resultados')]
        if not meses_para_exportar:
            QMessageBox.warning(self, "Aviso", "Nenhum dado para exportar.")
            return

        # Workaround openpyxl + lxml no Python 3.12 / QGIS 3.3x (Windows)
        try:
            import sys as _sys
            _lxml    = _sys.modules.pop('lxml',       None)
            _lxml_et = _sys.modules.pop('lxml.etree', None)
            from openpyxl import Workbook  # noqa
        finally:
            if _lxml    is not None: _sys.modules['lxml']       = _lxml
            if _lxml_et is not None: _sys.modules['lxml.etree'] = _lxml_et

        try:
            downloads_path = (
                os.path.join(os.environ['USERPROFILE'], 'Downloads')
                if sys.platform == "win32"
                else os.path.join(os.path.expanduser('~'), 'Downloads')
            )

            ts = datetime.now().strftime('%Y%m%d_%H%M')

            def _san(txt):
                txt = unicodedata.normalize('NFKD', str(txt))
                txt = txt.encode('ascii', 'ignore').decode('ascii')
                return "".join(c if c.isalnum() or c in ('_', '-') else '_' for c in txt)

            if len(meses_para_exportar) == 1:
                d   = meses_para_exportar[0]
                rot = _san(d['nome_mes']) + (f"_{d['ano']}" if d.get('ano') else "")
            else:
                rot = f"{len(meses_para_exportar)}_periodos"

            nome_arquivo = f"ALERTA_CONSUMO_{rot}_{ts}.xlsx"
            caminho      = os.path.join(downloads_path, nome_arquivo)

            wb = self._criar_workbook_alerta(meses_para_exportar)
            wb.save(caminho)

            QMessageBox.information(
                self, "Exportação Concluída",
                f"Arquivo salvo em:\n{caminho}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar Excel:\n{str(e)}")
            import traceback; traceback.print_exc()

    def _criar_workbook_alerta(self, meses_para_exportar):
        """Cria o Workbook consolidado com todos os meses em uma única aba.

        Estrutura de colunas:
            Fixas (6):  INT_CD | CNARH | Empreendimento | Usuário |
                        Operador | Medidor(es)
            Por mês (3 × N):
                        Consumo (m³) [abrev] | Outorgado (m³) [abrev] |
                        % Acrescido [abrev]

        Larguras calculadas dinamicamente a partir do conteúdo real de cada
        coluna (cabeçalho + dados), com margem de 2 caracteres e teto de 60.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        from collections import OrderedDict

        # ── Abreviações de mês ────────────────────────────────────────────────
        ABREV_MES = {
            'janeiro': 'jan', 'fevereiro': 'fev', 'março': 'mar',
            'abril': 'abr', 'maio': 'mai', 'junho': 'jun',
            'julho': 'jul', 'agosto': 'ago', 'setembro': 'set',
            'outubro': 'out', 'novembro': 'nov', 'dezembro': 'dez',
        }

        def _abreviar_rotulo(nome_mes, ano):
            """Converte rótulo completo em abreviação curta para cabeçalho Excel.

            Exemplos:
              'Janeiro/2026'                      → 'jan/26'
              'Fevereiro/2026 (10/02 a 28/02)'    → 'fev/26 (10/02-28/02)'
              'Março/2026 (01/03 a 18/03)'        → 'mar/26 (01/03-18/03)'
            """
            # Separar nome do mês do sufixo de dias, se houver
            # Formato: "NomeMês/Ano" ou "NomeMês/Ano (DD/MM a DD/MM)"
            sufixo = ""
            base   = nome_mes

            if '(' in nome_mes:
                base, resto = nome_mes.split('(', 1)
                base   = base.strip()
                # resto = "DD/MM a DD/MM)"
                resto  = resto.rstrip(')')
                partes = [p.strip() for p in resto.split(' a ')]
                if len(partes) == 2:
                    sufixo = f" ({partes[0]}-{partes[1]})"

            # base = "NomeMês/Ano" ou apenas "NomeMês" (modo mensal sem ano no nome)
            if '/' in base:
                nome_parte, ano_parte = base.rsplit('/', 1)
            else:
                nome_parte = base
                ano_parte  = str(ano) if ano else ""

            nome_lower = nome_parte.strip().lower()
            abrev      = ABREV_MES.get(nome_lower, nome_parte.strip()[:3].lower())
            ano_str    = str(ano_parte).strip()[-2:] if ano_parte else (str(ano)[-2:] if ano else "")

            return f"{abrev}/{ano_str}{sufixo}"

        # ── Estilos — padrão institucional ────────────────────────────────────
        fill_azul   = PatternFill("solid", fgColor="175cc3")
        fill_alt    = PatternFill("solid", fgColor="eaf2ff")
        fill_alerta = PatternFill("solid", fgColor="f8d7da")
        fill_info = PatternFill("solid", fgColor="fff9e6")
        font_titulo = Font(bold=True, size=13, color="175cc3")
        font_sub    = Font(size=10, italic=True, color="495057")
        font_cab    = Font(bold=True, size=10, color="ffffff")
        font_normal = Font(size=10)
        font_num    = Font(bold=True, size=10, color="dc3545")
        font_out    = Font(size=10, color="28a745")
        font_rod    = Font(size=8, italic=True, color="888888")
        ali_centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ali_esq     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        ali_dir     = Alignment(horizontal="right",  vertical="center")
        borda       = Border(
            left=Side(style='thin'),  right=Side(style='thin'),
            top=Side(style='thin'),   bottom=Side(style='thin'),
        )

        # ── Workbook ──────────────────────────────────────────────────────────
        import sys as _sys
        _lxml    = _sys.modules.pop('lxml',       None)
        _lxml_et = _sys.modules.pop('lxml.etree', None)
        try:
            wb = Workbook()
        finally:
            if _lxml    is not None: _sys.modules['lxml']       = _lxml
            if _lxml_et is not None: _sys.modules['lxml.etree'] = _lxml_et
            
        ws = wb.active
        ws.title = "Alerta Consumo"

        # ── Definição de colunas ──────────────────────────────────────────────
        N_FIXAS = 6
        N_MES   = 3

        colunas_fixas = [
            ("INT_CD",          ali_esq),
            ("CNARH",           ali_esq),
            ("Empreendimento",  ali_esq),
            ("Usuário",         ali_esq),
            ("Operador",        ali_esq),
            ("Medidor(es)",     ali_esq),
        ]

        # Abreviações para os cabeçalhos das colunas dinâmicas
        abrevs_meses = []
        rotulos_meses_completos = []
        for d in meses_para_exportar:
            rot_completo = (d['nome_mes'] if not d.get('ano')
                            else f"{d['nome_mes']}/{d['ano']}")
            abrev        = _abreviar_rotulo(d['nome_mes'], d.get('ano', 0))
            abrevs_meses.append(abrev)
            rotulos_meses_completos.append(rot_completo)

        n_total_cols = N_FIXAS + N_MES * len(meses_para_exportar)
        col_last     = get_column_letter(n_total_cols)

        # ── Consolidar dados por cod_interf ───────────────────────────────────
        interf_order     = OrderedDict()
        dados_por_interf = {}

        # 1. Identificar quais interferências tiveram alerta em PELO MENOS um mês
        interfs_com_alerta = set()
        for d in meses_para_exportar:
            for row_data in (d.get('resultados') or []):
                interfs_com_alerta.add(row_data[0])

        # 2. Consolidar apenas essas interferências (alertas + informativos)
        for d in meses_para_exportar:
            todos = list(d.get('resultados') or []) + list(d.get('informativos') or [])
            for row_data in todos:
                cod = row_data[0]
                if cod not in interfs_com_alerta:
                    continue   # ← descarta informativos puros
                if cod not in interf_order:
                    interf_order[cod]     = row_data
                    dados_por_interf[cod] = {}

        for mi, d in enumerate(meses_para_exportar):
            todos = list(d.get('resultados') or []) + list(d.get('informativos') or [])
            for row_data in todos:
                if row_data[0] not in interfs_com_alerta:
                    continue   # ← descarta informativos puros
                dados_por_interf[row_data[0]][mi] = row_data

        # ── Pré-calcular larguras dinâmicas ───────────────────────────────────
        # Inicializa com comprimento do cabeçalho de cada coluna
        larguras = {}

        # Colunas fixas: inicializa com nome do cabeçalho
        for ci, (nome, _) in enumerate(colunas_fixas, 1):
            larguras[ci] = len(nome)

        # Colunas dinâmicas: cabeçalho = "Consumo (m³)\nabrev" etc.
        # Usa a linha mais longa do cabeçalho (quebrado em \n)
        for mi, abrev in enumerate(abrevs_meses):
            base = N_FIXAS + mi * N_MES + 1
            for offset, prefixo in enumerate(
                    [f"Consumo (m³)\n{abrev}",
                     f"Outorgado (m³)\n{abrev}",
                     f"% Acrescido\n{abrev}"]):
                max_linha = max(len(l) for l in prefixo.split('\n'))
                larguras[base + offset] = max_linha

        # Iterar dados para ajustar larguras
        fmt_num = lambda v: f"{v:,.2f}" if v is not None else "—"

        for cod_interf, row_base in interf_order.items():
            cnarh         = str(row_base[1]) if len(row_base) > 1 and row_base[1] else "N/A"
            nome_empreend = str(row_base[2]) if len(row_base) > 2 and row_base[2] else "N/A"
            usuario       = str(row_base[3]) if len(row_base) > 3 and row_base[3] else "N/A"
            operador      = str(row_base[4]) if len(row_base) > 4 and row_base[4] else "N/A"
            rotulos       = str(row_base[5]) if len(row_base) > 5 and row_base[5] else "N/A"

            valores_fixos = [str(cod_interf), cnarh, nome_empreend,
                             usuario, operador, rotulos]
            for ci, val in enumerate(valores_fixos, 1):
                larguras[ci] = max(larguras[ci], len(val))

            for mi in range(len(meses_para_exportar)):
                base    = N_FIXAS + mi * N_MES + 1
                row_mes = dados_por_interf[cod_interf].get(mi)
                if row_mes:
                    consumo   = float(row_mes[6]) if len(row_mes) > 6 and row_mes[6] else 0.0
                    outorgado = float(row_mes[7]) if len(row_mes) > 7 and row_mes[7] else 0.0
                    pct       = float(row_mes[8]) if len(row_mes) > 8 else 0.0
                    larguras[base]     = max(larguras[base],     len(fmt_num(consumo)))
                    larguras[base + 1] = max(larguras[base + 1], len(fmt_num(outorgado)))
                    larguras[base + 2] = max(larguras[base + 2], len(f"+{pct:.1f}%"))

        # ── Linha 1 — Título ──────────────────────────────────────────────────
        ws.merge_cells(f"A1:{col_last}1")
        ws["A1"]           = "RELATÓRIO DE ALERTA — CONSUMO ACIMA DO VOLUME OUTORGADO"
        ws["A1"].font      = font_titulo
        ws["A1"].alignment = ali_centro
        ws.row_dimensions[1].height = 24

        # ── Linha 2 — Subtítulo ───────────────────────────────────────────────
        ws.merge_cells(f"A2:{col_last}2")
        total_alertas = sum(len(d['resultados']) for d in meses_para_exportar)
        periodos_txt  = "  |  ".join(rotulos_meses_completos)
        ws["A2"] = (
            f"Períodos: {periodos_txt}  —  "
            f"Total de alertas: {total_alertas}  —  "
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}"
        )
        ws["A2"].font      = font_sub
        ws["A2"].alignment = ali_centro
        ws.row_dimensions[2].height = 16

        # ── Linha 3 — Cabeçalho ───────────────────────────────────────────────
        for ci, (nome, _) in enumerate(colunas_fixas, 1):
            cell           = ws.cell(row=3, column=ci, value=nome)
            cell.font      = font_cab
            cell.fill      = fill_azul
            cell.alignment = ali_centro
            cell.border    = borda

        for mi, abrev in enumerate(abrevs_meses):
            base = N_FIXAS + mi * N_MES + 1
            for offset, rotulo_col in enumerate([
                f"Consumo (m³)\n{abrev}",
                f"Outorgado (m³)\n{abrev}",
                f"% Acrescido\n{abrev}",
            ]):
                cell           = ws.cell(row=3, column=base + offset, value=rotulo_col)
                cell.font      = font_cab
                cell.fill      = fill_azul
                cell.alignment = ali_centro
                cell.border    = borda

        ws.row_dimensions[3].height = 30

        # ── Linhas de dados ───────────────────────────────────────────────────
        for idx, (cod_interf, row_base) in enumerate(interf_order.items(), start=4):
            cnarh         = row_base[1] if len(row_base) > 1 else None
            nome_empreend = row_base[2] if len(row_base) > 2 else None
            usuario       = row_base[3] if len(row_base) > 3 else None
            operador      = row_base[4] if len(row_base) > 4 else None
            rotulos       = row_base[5] if len(row_base) > 5 else None

            tem_excesso = any(
                float(r[6] if len(r) > 6 else 0) > float(r[7] if len(r) > 7 else 0)
                for r in (dados_por_interf[cod_interf].get(mi, ())
                          for mi in range(len(meses_para_exportar)))
                if r
            )

            fill_linha = fill_alt if idx % 2 == 0 else None

            # Colunas fixas
            for ci, (val, (_, ali)) in enumerate(zip(
                    [str(cod_interf)    if cod_interf    else "N/A",
                     str(cnarh)         if cnarh         else "N/A",
                     str(nome_empreend) if nome_empreend else "N/A",
                     str(usuario)       if usuario       else "N/A",
                     str(operador)      if operador      else "N/A",
                     str(rotulos)       if rotulos       else "N/A"],
                    colunas_fixas), 1):
                cell           = ws.cell(row=idx, column=ci, value=val)
                cell.font      = font_normal
                cell.alignment = ali
                cell.border    = borda
                cell.fill      = fill_alerta if tem_excesso else (fill_linha or PatternFill())

            # Colunas dinâmicas
            for mi in range(len(meses_para_exportar)):
                base    = N_FIXAS + mi * N_MES + 1
                row_mes = dados_por_interf[cod_interf].get(mi)

                if row_mes:
                    consumo    = float(row_mes[6]) if len(row_mes) > 6 and row_mes[6] else 0.0
                    outorgado  = float(row_mes[7]) if len(row_mes) > 7 and row_mes[7] else 0.0
                    percentual = float(row_mes[8]) if len(row_mes) > 8 else (
                        round((consumo / outorgado - 1.0) * 100.0, 1)
                        if outorgado > 0 else 0.0
                    )
                    excesso_mes = consumo > outorgado
                else:
                    consumo = outorgado = percentual = None
                    excesso_mes = False

                eh_alerta_cel = bool(row_mes and len(row_mes) > 9 and row_mes[9])
                fill_cel = (fill_alerta   if eh_alerta_cel
                            else fill_info if row_mes
                            else (fill_alt if idx % 2 == 0 else PatternFill()))

                # Consumo
                c1 = ws.cell(row=idx, column=base,
                             value=consumo if consumo is not None else "—")
                c1.font      = font_num if excesso_mes else font_normal
                c1.alignment = ali_dir
                c1.border    = borda
                c1.fill      = fill_cel
                if consumo is not None:
                    c1.number_format = '#,##0.00'

                # Outorgado
                c2 = ws.cell(row=idx, column=base + 1,
                             value=outorgado if outorgado is not None else "—")
                c2.font      = font_out
                c2.alignment = ali_dir
                c2.border    = borda
                c2.fill      = fill_cel
                if outorgado is not None:
                    c2.number_format = '#,##0.00'

                # % Acrescido
                pct_val = (percentual / 100.0) if percentual is not None else None
                c3 = ws.cell(row=idx, column=base + 2,
                             value=pct_val if pct_val is not None else "—")
                c3.font      = font_num if excesso_mes else font_normal
                c3.alignment = ali_centro
                c3.border    = borda
                c3.fill      = fill_cel
                if pct_val is not None:
                    c3.number_format = '0.0%'

            ws.row_dimensions[idx].height = 15

        # ── Aplicar larguras dinâmicas (conteúdo + margem 2, teto 100) ─────────
        for ci, larg in larguras.items():
            ws.column_dimensions[get_column_letter(ci)].width = min(larg + 2, 100)

        # ── Freeze e rodapé ───────────────────────────────────────────────────
        ws.freeze_panes = "A4"

        rodape_row = len(interf_order) + 5
        ws.merge_cells(f"A{rodape_row}:{col_last}{rodape_row}")
        cell_rod           = ws.cell(
            row=rodape_row, column=1,
            value="Sistema DURH Diária por Telemetria (SFI/ANA) – Relatório gerado automaticamente"
        )
        cell_rod.font      = font_rod
        cell_rod.alignment = ali_centro

        return wb
        
    def atualizar_placeholder_busca(self):
        """Atualiza o placeholder do campo de busca conforme o critério selecionado."""
        criterio = self.combo_criterio_busca.currentText()
        if criterio == "CNARH":
            self.input_busca.setPlaceholderText("Digite número CNARH...")
        elif criterio == "Usuário":
            self.input_busca.setPlaceholderText("Digite nome do usuário...")
        elif criterio == "UAM":
            self.input_busca.setPlaceholderText("Digite a Unidade de Automonitoramento...")            
        elif criterio == "Sistema Hídrico":
            self.input_busca.setPlaceholderText("Digite o Sistema Hídrico...")
    
    def buscar_medidores_autocomplete(self, texto):
        """Busca medidores em tempo real para preencher o autocompletar."""
        texto = texto.strip()
        if len(texto) < 2:
            self.combo_sugestoes.setVisible(False)
            return

        criterio = self.combo_criterio_busca.currentText()
        cursor = None

        try:
            cursor = self.conn.cursor()
            termo_busca = f"%{texto}%"

            if criterio == "CNARH":
                query = "SELECT DISTINCT nu_cnarh FROM view_ft_intervencao WHERE LOWER(nu_cnarh) LIKE LOWER(%s) AND vazao_nominal > 0 ORDER BY nu_cnarh LIMIT 10;"
            elif criterio == "Usuário":
                query = "SELECT DISTINCT nome_usuario FROM view_ft_intervencao WHERE LOWER(nome_usuario) LIKE LOWER(%s) AND vazao_nominal > 0 ORDER BY nome_usuario LIMIT 10;"
            elif criterio == "UAM":
                query = "SELECT DISTINCT nmautomonit FROM ft_uam_buffer WHERE LOWER(nmautomonit) LIKE LOWER(%s) ORDER BY nmautomonit LIMIT 10;"            
            elif criterio == "Sistema Hídrico":
                query = "SELECT DISTINCT bafnm FROM ft_sishidrico_buffer WHERE LOWER(bafnm) LIKE LOWER(%s) ORDER BY bafnm LIMIT 10;"
            else:
                return

            cursor.execute(query, (termo_busca,))
            resultados = cursor.fetchall()

            self.combo_sugestoes.clear()
            if resultados:
                for resultado in resultados:
                    if resultado[0]:
                        self.combo_sugestoes.addItem(str(resultado[0]))
                self.combo_sugestoes.setVisible(True)
            else:
                self.combo_sugestoes.setVisible(False)

        except Exception as e:
            print(f"Erro no autocompletar: {e}")
        finally:
            if cursor:
                cursor.close()
                
    def buscar_ou_limpar(self):
        """Alterna entre buscar e limpar conforme o estado atual do botão."""
        if self.btn_buscar.text() == "Limpar busca":
            self.limpar_busca()
        else:
            self.buscar_medidores_avancada()

    def buscar_medidores_avancada(self):
        """Executa a busca de medidores e atualiza o estado dos botões de ação."""
        criterio = self.combo_criterio_busca.currentText()
        termo = self.input_busca.text().strip()

        if not termo:
            QMessageBox.warning(self, "Campo vazio", "Digite um termo para busca.")
            return

        cursor = None
        try:
            cursor = self.conn.cursor()
            termo_busca = f"%{termo}%"

            if criterio == "Sistema Hídrico":
                QApplication.setOverrideCursor(Qt.WaitCursor)
                query = """
                SELECT DISTINCT v.id, v.rotulo_medidor, v.vazao_nominal,
                COALESCE(v.nome_usuario, 'Não informado') as usuario,
                COALESCE(v.nome_operador, 'Não informado') as operador,
                v.nu_cnarh as cnarh, v.nu_interferencia_cnarh as codigo_interferencia
                FROM view_ft_intervencao v
                JOIN ft_sishidrico_buffer s ON ST_Intersects(v.geom, ST_Transform(s.geom, 4326))
                WHERE LOWER(s.bafnm) LIKE LOWER(%s) AND v.vazao_nominal > 0
                ORDER BY v.rotulo_medidor;
                """
                try:
                    cursor.execute(query, (termo_busca,))
                    resultados = cursor.fetchall()
                finally:
                    QApplication.restoreOverrideCursor()

            elif criterio == "UAM":
                # -- Busca Espacial --
                QApplication.setOverrideCursor(Qt.WaitCursor)                
                query = """
                SELECT DISTINCT v.id, v.rotulo_medidor, v.vazao_nominal,
                COALESCE(v.nome_usuario, 'Não informado') as usuario,
                COALESCE(v.nome_operador, 'Não informado') as operador,
                v.nu_cnarh as cnarh, v.nu_interferencia_cnarh as codigo_interferencia
                FROM view_ft_intervencao v
                JOIN ft_uam_buffer s ON ST_Intersects(v.geom, ST_Transform(s.geom, 4326))
                WHERE LOWER(s.nmautomonit) LIKE LOWER(%s) AND v.vazao_nominal > 0
                ORDER BY v.rotulo_medidor;
                """                
                try:
                    cursor.execute(query, (termo_busca,))
                    resultados = cursor.fetchall()
                finally:
                    QApplication.restoreOverrideCursor()
                    
            else:
                base_query = """
                SELECT DISTINCT id, rotulo_medidor, vazao_nominal,
                COALESCE(nome_usuario, 'Não informado') as usuario,
                COALESCE(nome_operador, 'Não informado') as operador,
                nu_cnarh as cnarh, nu_interferencia_cnarh as codigo_interferencia
                FROM view_ft_intervencao WHERE vazao_nominal > 0
                """
                if criterio == "CNARH":
                    query = base_query + " AND LOWER(nu_cnarh) LIKE LOWER(%s) ORDER BY rotulo_medidor;"
                elif criterio == "Usuário":
                    query = base_query + " AND LOWER(nome_usuario) LIKE LOWER(%s) ORDER BY rotulo_medidor;"
                else:
                    return

                cursor.execute(query, (termo_busca,))
                resultados = cursor.fetchall()

            self.lista_resultados.clear()

            if not resultados:
                self.lista_resultados.addItem("Nenhum resultado encontrado.")
                self.limpar_selecao()
                self.btn_selecionar_tudo.setEnabled(False)
                self.btn_buscar.setText("Limpar busca")
                return

            self.btn_selecionar_tudo.setEnabled(True)

            for row in resultados:
                id_med, rotulo, vazao, usuario, operador, cnarh, codigo_interferencia = row
                texto = f"{rotulo} - {usuario}"
                if cnarh:
                    texto += f" (CNARH: {cnarh}"
                if codigo_interferencia:
                    texto += f" | Interferência: {codigo_interferencia}"
                if cnarh or codigo_interferencia:
                    texto += ")"

                item = QListWidgetItem(texto)
                item.setData(Qt.UserRole, (id_med, rotulo, usuario, operador, cnarh, codigo_interferencia))
                self.lista_resultados.addItem(item)

            self.btn_buscar.setText("Limpar busca")

        except Exception as e:
            if criterio in ["Sistema Hídrico", "UAM"]:
                QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Erro", f"Erro na busca: {e}")
        finally:
            if cursor:
                cursor.close()
                
    def selecionar_sugestao_busca(self, index):
        """Seleciona uma sugestão do combobox, preenche o input e executa a busca."""
        if index >= 0:
            texto = self.combo_sugestoes.itemText(index)
            self.input_busca.blockSignals(True)
            self.input_busca.setText(texto)
            self.input_busca.blockSignals(False)
            self.combo_sugestoes.setVisible(False)
            self.buscar_ou_limpar()

    def selecionar_todos_medidores(self):
        """Seleciona ou limpa todos os itens da lista de resultados."""
        if self.lista_resultados.count() == 0:
            return

        if self.lista_resultados.count() == 1:
            item = self.lista_resultados.item(0)
            if item.text() == "Nenhum resultado encontrado.":
                return

        tudo_selecionado = self.btn_selecionar_tudo.text() == "Limpar seleção"

        if tudo_selecionado:
            self.lista_resultados.clearSelection()
            self.btn_selecionar_tudo.setText("Selecionar tudo")
            self.limpar_selecao()
            return

        self.lista_resultados.selectAll()

        self.lista_ids_selecionados = []
        self.lista_dados_selecionados = []
        codigos_interferencia = []

        for i in range(self.lista_resultados.count()):
            item = self.lista_resultados.item(i)
            dados = item.data(Qt.UserRole)
            if dados:
                self.lista_ids_selecionados.append(dados[0])
                self.lista_dados_selecionados.append(dados)
                codigos_interferencia.append(dados[5])

        codigos_unicos = set(codigos_interferencia)
        self.eh_multipla_interferencia = len(codigos_unicos) > 1
        self.codigos_interferencia_unicos = codigos_unicos
        self.is_selecao_total = True

        self.btn_abrir_monitoramento.setEnabled(True)
        self.btn_estatisticas.setEnabled(True)
        self.btn_selecionar_tudo.setText("Limpar seleção")

    def selecionar_medidor_da_lista(self, item=None):
        """
        Gerencia a seleção de medidores.
        Sempre habilita Telemetria e Estatísticas, mas armazena info sobre múltiplas interferências.
        """
        selected_items = self.lista_resultados.selectedItems()

        if not selected_items:
            self.limpar_selecao()
            return

        self.lista_ids_selecionados = []
        self.lista_dados_selecionados = []
        codigos_interferencia = []

        for it in selected_items:
            dados = it.data(Qt.UserRole)
            if dados:
                self.lista_ids_selecionados.append(dados[0])
                self.lista_dados_selecionados.append(dados)
                codigos_interferencia.append(dados[5])

        codigos_unicos = set(codigos_interferencia)
        self.eh_multipla_interferencia = len(codigos_unicos) > 1
        self.codigos_interferencia_unicos = codigos_unicos
        self.is_selecao_total = False

        self.btn_abrir_monitoramento.setEnabled(True)
        self.btn_estatisticas.setEnabled(True)

        # Se a seleção manual não cobre todos os itens, restaurar label do botão
        total_itens = self.lista_resultados.count()
        total_selecionados = len(self.lista_resultados.selectedItems())
        if total_selecionados < total_itens:
            self.btn_selecionar_tudo.setText("Selecionar tudo")

    def mostrar_dialogo_agregacao(self, codigos_interferencia):
        """
        Mostra diálogo para escolher modo de agregação dos dados.
        Retorna True se usuário escolheu, False se cancelou.
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("Seleção Múltipla - Agregação de Dados")
        dialog.setModal(True)
        dialog.setMinimumWidth(450)
        
        layout = QVBoxLayout(dialog)
        
        # Mensagem principal
        criterio = self.combo_criterio_busca.currentText()
        termo_busca = self.input_busca.text().strip()
        
        # === Capturar nome completo do termo de busca ===
        if criterio == "Sistema Hídrico":
            # Usar o nome completo do bafnm selecionado no combo_sugestoes ou input
            termo_display = termo_busca
            # Tentar encontrar o nome completo nas sugestões
            for i in range(self.combo_sugestoes.count()):
                item_text = self.combo_sugestoes.itemText(i)
                if termo_busca.lower() in item_text.lower():
                    termo_display = f"Sistema Hídrico {item_text}"  # Nome completo do SH
                    break
            btn_texto_alt = "Sistema Hídrico"
            self.nome_sh_completo = termo_display
            self.nome_uam_completo = None
            self.nome_cnarh_completo = None
            self.nome_usuario_completo = None

        elif criterio == "UAM":
            # Tentar encontrar a UAM completa nas sugestões
            termo_display = termo_busca
            for i in range(self.combo_sugestoes.count()):
                item_text = self.combo_sugestoes.itemText(i)
                if termo_busca.lower() in item_text.lower():
                    termo_display = f"CANRH {item_text}"
                    break
            btn_texto_alt = "UAM"
            self.nome_sh_completo = None
            self.nome_uam_completo = termo_display
            self.nome_cnarh_completo = None
            self.nome_usuario_completo = None
            
        elif criterio == "CNARH":
            # Tentar encontrar o CNARH completo nas sugestões
            termo_display = termo_busca
            for i in range(self.combo_sugestoes.count()):
                item_text = self.combo_sugestoes.itemText(i)
                if termo_busca.lower() in item_text.lower():
                    termo_display = f"CANRH {item_text}"
                    break
            btn_texto_alt = "CNARH"
            self.nome_sh_completo = None
            self.nome_uam_completo = None            
            self.nome_cnarh_completo = termo_display
            self.nome_usuario_completo = None
            
        else:  # Usuário
            # Tentar encontrar o nome completo do usuário nas sugestões
            termo_display = termo_busca
            for i in range(self.combo_sugestoes.count()):
                item_text = self.combo_sugestoes.itemText(i)
                if termo_busca.lower() in item_text.lower():
                    termo_display = f"Usuário {item_text}"
                    break
            btn_texto_alt = "Usuário"
            self.nome_sh_completo = None
            self.nome_uam_completo = None            
            self.nome_cnarh_completo = None
            self.nome_usuario_completo = termo_display  # Armazenar para uso posterior

        lbl_mensagem = QLabel(
            f"Você selecionou medidores de {len(codigos_interferencia)} interferências diferentes.<br><br>"
            f"Deseja agregar os dados por <b>código de interferência</b> ou pelo <b>{termo_display}</b>?"
        )
        lbl_mensagem.setWordWrap(True)
        lbl_mensagem.setStyleSheet("font-size: 12px; padding: 10px;")
        layout.addWidget(lbl_mensagem)
        
        # Layout dos botões
        btn_layout = QHBoxLayout()
        
        resultado = {"escolha": None}  # Para capturar resultado
        
        def escolher_interferencia():
            resultado["escolha"] = "interferencia"
            dialog.accept()
        
        def escolher_criterio():
            resultado["escolha"] = "criterio_busca"
            dialog.accept()
        
        btn_agregar_interferencia = QPushButton("Interferência")
        btn_agregar_interferencia.setStyleSheet("""
            QPushButton {
                background-color: #175cc3;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #5474b8;
            }
        """)
        btn_agregar_interferencia.clicked.connect(escolher_interferencia)
        
        btn_agregar_alternativo = QPushButton(btn_texto_alt)
        btn_agregar_alternativo.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        btn_agregar_alternativo.clicked.connect(escolher_criterio)

        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 8px 16px;
                border-radius: 5px;
            }
        """)
        btn_cancelar.clicked.connect(dialog.reject)
        
        btn_layout.addWidget(btn_agregar_interferencia)
        btn_layout.addWidget(btn_agregar_alternativo)
        btn_layout.addWidget(btn_cancelar)        
        
        layout.addLayout(btn_layout)    
        
        # Executa diálogo
        if dialog.exec_() == QDialog.Accepted and resultado["escolha"]:
            # Define o modo de agregação
            self.modo_agregacao = resultado["escolha"]
            self.criterio_busca_atual = criterio
            self.termo_busca_atual = termo_busca
            return True
        else:
            # Usuário cancelou
            return False

    def escolher_agregacao(self, modo, dialog):
        """Define o modo de agregação e fecha o diálogo."""
        self.modo_agregacao = modo  # "interferencia" ou "criterio_busca"
        self.criterio_busca_atual = self.combo_criterio_busca.currentText()
        self.termo_busca_atual = self.input_busca.text().strip()
        
        # Sempre habilita Estatísticas, desabilita Telemetria para múltiplas interferências
        self.btn_abrir_monitoramento.setEnabled(False)
        self.btn_estatisticas.setEnabled(True)
        
        dialog.accept()

    def limpar_selecao(self):
        """Limpa as variáveis de seleção e desabilita botões de ação."""
        self.lista_ids_selecionados = []
        self.lista_dados_selecionados = []
        self.modo_agregacao = None
        self.criterio_busca_atual = None
        self.termo_busca_atual = None
        self.eh_multipla_interferencia = False
        self.codigos_interferencia_unicos = set()
        self.is_selecao_total = False
        self.btn_abrir_monitoramento.setEnabled(False)
        self.btn_estatisticas.setEnabled(False)

    def _verificar_selecao_completa_interferencias(self, codigos_interf):
        """Verifica se todos os medidores válidos de cada interferência foram selecionados.
        
        Consulta o total de intervencao_id disponíveis por interferência na view
        (aplicando os mesmos filtros da busca) e compara com a quantidade
        efetivamente selecionada. Retorna o código da primeira interferência
        incompleta encontrada, ou None se todas estiverem completas.
        
        Args:
            codigos_interf (set): Códigos de interferência presentes na seleção.
            
        Returns:
            str | None: Código da interferência incompleta, ou None se ok.
        """
        if not codigos_interf:
            return None

        cursor = None
        try:
            cursor = self.conn.cursor()

            # Contar medidores válidos disponíveis no banco para cada interferência.
            # Usa os mesmos filtros aplicados em buscar_medidores_avancada para
            # garantir consistência com o que o analista vê na lista de resultados.
            query = """
                SELECT nu_interferencia_cnarh,
                       COUNT(DISTINCT id) AS total_medidores
                FROM view_ft_intervencao
                WHERE nu_interferencia_cnarh = ANY(%s)
                  AND vazao_nominal > 0
                GROUP BY nu_interferencia_cnarh;
            """
            cursor.execute(query, (list(codigos_interf),))
            total_por_interferencia = {str(row[0]): int(row[1]) for row in cursor.fetchall()}

            # Contar quantos medidores de cada interferência foram selecionados
            selecionados_por_interferencia = {}
            for dados in self.lista_dados_selecionados:
                cod = str(dados[5])
                selecionados_por_interferencia[cod] = \
                    selecionados_por_interferencia.get(cod, 0) + 1

            # Comparar: retornar o primeiro código com contagem divergente
            for cod in sorted(codigos_interf):
                cod_str    = str(cod)
                total_db   = total_por_interferencia.get(cod_str, 0)
                total_sel  = selecionados_por_interferencia.get(cod_str, 0)
                if total_sel < total_db:
                    return cod_str

            return None  # Todas as interferências estão completas

        except Exception as e:
            print(f"[ERRO] Verificação de seleção completa: {e}")
            # Em caso de erro na consulta, permite prosseguir sem bloquear
            return None
        finally:
            if cursor:
                cursor.close()

    def _abrir_graficos_interferencia(self, cod_interf, rotulos, callback_fechar=None,
                                      ano_inicial=None, mes_inicial=None):
        """Abre JanelaGraficosMedidor para UMA interferência específica (sem agregação).
        
        Args:
            cod_interf: Código da interferência a visualizar.
            rotulos: Rótulos dos medidores (informativo).
            callback_fechar: Função opcional chamada ao fechar a janela de gráficos.
                             Usado para reabrir a tabela de alertas quando esta é
                             a origem da chamada.
        """
        from PyQt5.QtWidgets import QApplication
        
        QApplication.setOverrideCursor(Qt.WaitCursor)
        
        try:
            cursor = self.conn.cursor()
            
            # Buscar todos os medidores desta interferência usando intervencao_id,
            # que é o campo que faz JOIN com tb_telemetria_intervencao_diaria.
            # ATENÇÃO: usar i.intervencao_id (e não i.id) garante que os IDs
            # retornados correspondam a intervencao_id na tabela de telemetria.
            query = """
                SELECT DISTINCT 
                    i.intervencao_id,
                    i.rotulo_intervencao_medidor as rotulo,
                    COALESCE(i.usuario, 'Não informado') as usuario,
                    COALESCE(i.operador, 'Não informado') as operador,
                    i.nu_cnarh as cnarh,
                    i.nu_interferencia_cnarh as codigo_interferencia
                FROM view_usuario_operador_id_rotulo i
                WHERE i.nu_interferencia_cnarh = %s
                  AND i.nu_interferencia_cnarh <> 'TESTE'
                  AND i.rotulo_intervencao_medidor !~~ '%%999%%'
                  AND i.rotulo_intervencao_medidor !~~ '%%VERDE GRANDE%%'
                  AND i.rotulo_intervencao_medidor !~~ '%%#'
                ORDER BY i.rotulo_intervencao_medidor;
            """
            
            cursor.execute(query, (str(cod_interf),))
            resultados = cursor.fetchall()
            cursor.close()
            
            if not resultados:
                QMessageBox.warning(self, "Aviso", 
                    f"Nenhum medidor encontrado para a interferência {cod_interf}")
                QApplication.restoreOverrideCursor()
                return
            
            # Preparar dados no formato esperado por JanelaGraficosMedidor.
            # Mesmo formato usado em selecionar_medidor_da_lista():
            # (id, rotulo, usuario, operador, cnarh, codigo_interferencia)
            lista_ids  = []
            lista_dados = []
            
            for row in resultados:
                intervencao_id, rotulo, usuario, operador, cnarh, cod_interf_db = row
                lista_ids.append(intervencao_id)
                lista_dados.append((intervencao_id, rotulo, usuario, operador, cnarh, cod_interf_db))
            
            # ABRIR JanelaGraficosMedidor MODO SIMPLES (sem agregação)
            # Similar ao comportamento de seleção única em selecionar_medidor_da_lista()
            self.janela_graficos = JanelaGraficosMedidor(
                self,                      # janela_anterior
                self.conn,                 # conexao
                lista_ids,                 # lista_ids_selecionados
                lista_dados,               # lista_dados_selecionados
                self.usuario_logado,       # usuario
                self.senha,                # senha
                lista_dados[0][2],         # nome_usuario (do primeiro medidor)
                None,                      # modo_agregacao = None (sem agregação!)
                None,                      # criterio_busca
                None,                      # termo_busca
                None,                      # nome_completo
                False,                     # eh_multipla_interferencia = False
                {str(cod_interf)},         # codigos_interf (set com um elemento)
                ano_inicial=ano_inicial,   # ano de busca
                mes_inicial=mes_inicial,   # mês de busca             
            )
                        
            self.janela_graficos.show()
            self._janelas_abertas.append(self.janela_graficos)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", 
                f"Falha ao abrir gráficos:\n\n{str(e)}")
            import traceback
            traceback.print_exc()
            
        finally:
            QApplication.restoreOverrideCursor()
     
    def abrir_monitoramento_detalhes(self):
        """Abre a janela de gráficos (Estatísticas) diretamente.
        
        Quando há medidores de mais de uma interferência na seleção, verifica
        previamente se todos os medidores de cada interferência foram incluídos,
        exibindo alerta caso algum esteja faltando.
        """
        if not self.lista_ids_selecionados:
            QMessageBox.warning(self, "Aviso", "Selecione pelo menos um medidor.")
            return

        codigos_interf         = set([d[5] for d in self.lista_dados_selecionados])
        eh_multipla_interferencia = len(codigos_interf) > 1

        # Validação: só quando há mais de uma interferência na seleção.
        # Para medidor(es) de uma única interferência (total ou parcial),
        # a abertura segue normalmente sem nenhuma verificação.
        if eh_multipla_interferencia:
            interferencia_incompleta = self._verificar_selecao_completa_interferencias(
                codigos_interf
            )
            if interferencia_incompleta:
                QMessageBox.warning(
                    self,
                    "Seleção Incompleta",
                    f"Por favor, selecione os demais medidores da interferência "
                    f"{interferencia_incompleta} para prosseguir com as estatísticas."
                )
                return

        # Definir modo de agregação padrão baseado no critério de busca
        criterio = self.combo_criterio_busca.currentText()
        termo    = self.input_busca.text().strip()

        # Tentar encontrar nome completo nas sugestões
        nome_completo = termo
        for i in range(self.combo_sugestoes.count()):
            item_text = self.combo_sugestoes.itemText(i)
            if termo.lower() in item_text.lower():
                nome_completo = item_text
                break

        # Modo padrão: agregar por critério de busca (SH/UAM/CNARH/Usuário)
        modo_agreg   = "criterio_busca" if eh_multipla_interferencia else None
        nome_usuario = self.lista_dados_selecionados[0][2] if self.lista_dados_selecionados else None

        # Abrir janela de estatísticas repassando o flag de seleção total
        self.janela_graficos = JanelaGraficosMedidor(
            self, self.conn, self.lista_ids_selecionados,
            self.lista_dados_selecionados, self.usuario_logado, self.senha,
            nome_usuario, modo_agreg, criterio, termo, nome_completo,
            eh_multipla_interferencia, codigos_interf,
            is_selecao_total=self.is_selecao_total
        )
        self.janela_graficos.show()
        self._janelas_abertas.append(self.janela_graficos)
  
    def abrir_janela_detalhes(self):
        """
        Abre a janela de monitoramento detalhado (Telemetria).
        Sempre abre, independente do tipo de seleção.
        Repassa contexto de busca para suporte a múltiplas interferências.
        """
        if not self.lista_ids_selecionados:
            QMessageBox.warning(self, "Aviso", "Selecione pelo menos um medidor.")
            return

        QApplication.setOverrideCursor(Qt.WaitCursor)

        try:
            criterio  = self.combo_criterio_busca.currentText()
            termo     = self.input_busca.text().strip()

            # Tentar encontrar nome completo nas sugestões
            nome_completo = termo
            for i in range(self.combo_sugestoes.count()):
                item_text = self.combo_sugestoes.itemText(i)
                if termo.lower() in item_text.lower():
                    nome_completo = item_text
                    break

            codigos_interf        = set(d[5] for d in self.lista_dados_selecionados)
            eh_multipla_interf    = len(codigos_interf) > 1

            self.janela_detalhes = JanelaMonitoramentoDetalhes(
                self,
                self.conn,
                self.lista_ids_selecionados,
                self.lista_dados_selecionados,
                self.usuario_logado,
                criterio_busca = criterio,
                nome_completo = nome_completo,
                eh_multipla_interferencia = eh_multipla_interf,
                codigos_interf = codigos_interf,
                is_selecao_total = self.is_selecao_total,
            )

            self.janela_detalhes.show()
            self._janelas_abertas.append(self.janela_detalhes)

        finally:
            QApplication.restoreOverrideCursor()
    
    def limpar_busca(self):
        """Limpa todos os campos de busca e resultados."""
        self.input_busca.blockSignals(True)
        self.input_busca.clear()
        self.input_busca.blockSignals(False)
        self.combo_sugestoes.clear()
        self.combo_sugestoes.setVisible(False)
        self.lista_resultados.clear()
        self.btn_selecionar_tudo.setEnabled(False)
        self.btn_selecionar_tudo.setText("Selecionar tudo")
        self.btn_buscar.setText("Buscar")
        self.limpar_selecao()
    
    def center(self):
        """Centraliza a janela na tela."""
        screen_geometry = QDesktopWidget().screenGeometry()
        center_point = screen_geometry.center()
        frame_geometry = self.frameGeometry()
        frame_geometry.moveCenter(center_point)
        self.move(frame_geometry.topLeft())

    def closeEvent(self, event):
        """Fecha sub-janelas ao fechar o monitoramento."""
        for janela in list(self._janelas_abertas):
            try:
                if janela and janela.isVisible():
                    janela.close()
            except RuntimeError:
                pass
        self._janelas_abertas.clear()
        event.accept()
        
    def voltar(self):
        """Volta para a tela inicial."""
        self.close()
        self.tela_inicial.show()
